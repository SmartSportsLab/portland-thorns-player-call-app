#!/usr/bin/env python3
"""
Create Player Progression Report tracking performance over time.
Analyzes players with 2+ years of data to identify trends and patterns.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def load_player_history(base_dir):
    """Load all player data across all years and positions."""
    
    exports_dir = base_dir / "Exports"
    historical_dir = exports_dir / "Past Seasons"
    
    # Define position mappings
    positions = {
        'AM Advanced Playmaker': 'Attacking Midfielder',
        'CB Hybrid': 'Center Back',
        'DM Box-To-Box': 'Centre Midfielder',
        'W Touchline Winger': 'Winger'
    }
    
    conferences = ['ACC', 'BIG10', 'BIG12', 'IVY', 'SEC']
    years = [2021, 2022, 2023, 2024, 2025]
    
    player_history = defaultdict(lambda: {'data': [], 'position': None})
    
    print("📚 Loading all historical data...")
    
    for position_file_prefix, position_name in positions.items():
        for conference in conferences:
            for year in years:
                if year == 2025:
                    file_path = exports_dir / f"{position_file_prefix} {conference} {year}.xlsx"
                else:
                    file_path = historical_dir / f"{position_file_prefix} {conference} {year}.xlsx"
                
                if file_path.exists():
                    try:
                        df = pd.read_excel(file_path, sheet_name=0)
                        
                        # Add metadata
                        df['Year'] = year
                        df['Conference'] = conference
                        df['Position'] = position_name
                        
                        # Store player data
                        for _, row in df.iterrows():
                            player_name = row['Player']
                            # Track all positions a player has appeared in
                            if player_history[player_name]['position'] is None:
                                player_history[player_name]['position'] = position_name
                            elif player_history[player_name]['position'] != position_name:
                                # Player appears in multiple positions - keep the first one
                                pass
                            player_history[player_name]['data'].append(row.to_dict())
                    except Exception as e:
                        print(f"  ⚠️  Error loading {file_path.name}: {e}")
    
    print(f"  ✅ Loaded data for {len(player_history)} unique players")
    
    return player_history

def calculate_progression_metrics(player_data):
    """Calculate progression metrics for a player."""
    
    # Sort by year
    years_data = sorted(player_data, key=lambda x: x.get('Year', 0))
    
    if len(years_data) < 2:
        return None
    
    progression = {
        'first_year': years_data[0].get('Year'),
        'last_year': years_data[-1].get('Year'),
        'years_active': len(set(d.get('Year') for d in years_data)),
        'first_team': years_data[0].get('Team'),
        'current_team': years_data[-1].get('Team'),
        'team_changes': len(set(d.get('Team') for d in years_data)),
        'metrics': {}
    }
    
    # Get ALL available metrics from the data
    all_available_metrics = set()
    for year_data in years_data:
        all_available_metrics.update(year_data.keys())
    
    # Remove metadata columns
    excluded_cols = {
        'Player', 'Team', 'Position', 'Year', 'Conference', 'Season', 
        'Team within selected timeframe', 'NCAA_Seasons', 'Age', 'Market value',
        'Contract expires', 'Birth country', 'Passport country', 'Foot', 
        'Height', 'Weight', 'On loan', 'Matches played', 'Yellow cards', 'Red cards',
        'Clean sheets', 'Non-penalty goals', 'Head goals', 'Shots', 'Fouls',
        # Goalkeeper-specific metrics (not relevant for outfield players)
        'Conceded goals', 'Conceded goals per 90', 'Shots against', 'Shots against per 90',
        'Save rate, %', 'xG against', 'xG against per 90'
    }
    
    # Track all numeric metrics except excluded ones
    metrics_to_track = [m for m in all_available_metrics if m not in excluded_cols]
    
    # Calculate progression for each metric
    for metric in metrics_to_track:
        # Only process if both years have this metric
        if metric not in years_data[0] or metric not in years_data[-1]:
            continue
        
        first_value = years_data[0][metric]
        last_value = years_data[-1][metric]
        
        # Skip non-numeric values
        if not isinstance(first_value, (int, float)) or not isinstance(last_value, (int, float)):
            continue
        
        if pd.notna(first_value) and pd.notna(last_value):
            # Calculate change
            abs_change = last_value - first_value
            if first_value != 0:
                pct_change = (abs_change / first_value) * 100
            else:
                pct_change = 0 if abs_change == 0 else 100
            
            progression['metrics'][metric] = {
                'first': first_value,
                'last': last_value,
                'abs_change': abs_change,
                'pct_change': pct_change
            }
    
    # Calculate year-over-year (YoY) changes from last season if player played in 2024
    if len(years_data) >= 2:
        # Find the most recent year and previous year
        sorted_years = sorted(set(d.get('Year') for d in years_data))
        
        if len(sorted_years) >= 2:
            current_year = sorted_years[-1]
            previous_year = sorted_years[-2]
            
            # Only calculate YoY if current year is 2025
            if current_year == 2025:
                progression['yoy_from_2024'] = {}
                
                # Get data for both years
                current_year_data = [d for d in years_data if d.get('Year') == current_year]
                previous_year_data = [d for d in years_data if d.get('Year') == previous_year]
                
                if current_year_data and previous_year_data:
                    # Use the last entry for each year (in case of duplicates)
                    current_row = current_year_data[-1]
                    previous_row = previous_year_data[-1]
                    
                    # Track all metrics for YoY
                    for metric in metrics_to_track:
                        if metric not in current_row or metric not in previous_row:
                            continue
                        
                        current_value = current_row[metric]
                        previous_value = previous_row[metric]
                        
                        # Skip non-numeric values
                        if not isinstance(current_value, (int, float)) or not isinstance(previous_value, (int, float)):
                            continue
                        
                        if pd.notna(current_value) and pd.notna(previous_value):
                            # Calculate change
                            abs_change = current_value - previous_value
                            if previous_value != 0:
                                pct_change = (abs_change / previous_value) * 100
                            else:
                                pct_change = 0 if abs_change == 0 else 100
                            
                            metric_data = {
                                'previous': previous_value,
                                'current': current_value,
                                'abs_change': abs_change,
                                'pct_change': pct_change
                            }
                            
                            # If this is already a "per 90" metric, use it as per_90_change
                            if 'per 90' in metric.lower():
                                metric_data['per_90_change'] = abs_change
                            else:
                                # For non-per-90 metrics like Goals, Assists, xG, xA,
                                # calculate per-90 change if we have Minutes played
                                if 'Minutes played' in current_row and 'Minutes played' in previous_row:
                                    curr_minutes = current_row['Minutes played']
                                    prev_minutes = previous_row['Minutes played']
                                    
                                    if pd.notna(curr_minutes) and pd.notna(prev_minutes) and curr_minutes > 0 and prev_minutes > 0:
                                        curr_per_90 = (current_value / curr_minutes) * 90
                                        prev_per_90 = (previous_value / prev_minutes) * 90
                                        metric_data['per_90_change'] = curr_per_90 - prev_per_90
                            
                            progression['yoy_from_2024'][metric] = metric_data
    
    return progression

def analyze_progressions(player_history):
    """Analyze all player progressions."""
    
    all_progressions = []
    multi_year_players = 0
    
    print("\n📈 Analyzing player progressions...")
    
    for player_name, player_info in player_history.items():
        if len(player_info['data']) >= 2:
            progression = calculate_progression_metrics(player_info['data'])
            
            if progression:
                # Filter out players who only appear in one unique year (duplicates)
                if progression['years_active'] >= 2:
                    multi_year_players += 1
                    progression['player_name'] = player_name
                    progression['position'] = player_info['position']
                    all_progressions.append(progression)
    
    print(f"  ✅ Analyzed {multi_year_players} players with 2+ unique years of data")
    
    return all_progressions

def create_progression_report(all_progressions, output_path):
    """Create Excel report with progression analysis."""
    
    wb = Workbook()
    
    # Get base_dir from output_path
    base_dir = output_path.parent
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create summary sheet
    create_summary_sheet(wb, all_progressions)
    
    # Create detailed progression sheet
    create_detailed_progression_sheet(wb, all_progressions)
    
    # Create risers/fallers sheet
    create_risers_fallers_sheet(wb, all_progressions, base_dir)
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✅ Report saved: {output_path}")

def create_summary_sheet(wb, all_progressions):
    """Create summary statistics sheet."""
    
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "Player Progression Summary"
    ws['A1'].font = Font(size=16, bold=True)
    
    # Statistics
    total_analyzed = len(all_progressions)
    active_2025 = len([p for p in all_progressions if p.get('last_year') == 2025])
    
    ws['A3'] = "Total Players Analyzed (2+ years of data):"
    ws['B3'] = total_analyzed
    ws['B3'].font = Font(bold=True)
    
    ws['A4'] = "Players Active in 2025 (shown in Player Progressions tab):"
    ws['B4'] = active_2025
    ws['B4'].font = Font(bold=True)
    
    ws['A5'] = "Players Not Active in 2025 (historical only):"
    ws['B5'] = total_analyzed - active_2025
    
    # Count by position (all players)
    position_counts = defaultdict(int)
    position_2025_counts = defaultdict(int)
    for p in all_progressions:
        position_counts[p['position']] += 1
        if p.get('last_year') == 2025:
            position_2025_counts[p['position']] += 1
    
    ws['A7'] = "All Players (2+ years) by Position:"
    row = 7
    for position, count in sorted(position_counts.items()):
        ws[f'A{row}'] = f"  {position}:"
        ws[f'B{row}'] = count
        row += 1
    
    ws[f'A{row}'] = "Active in 2025 by Position:"
    row += 1
    for position, count in sorted(position_2025_counts.items()):
        ws[f'A{row}'] = f"  {position}:"
        ws[f'B{row}'] = count
        row += 1
    
    # Years of experience distribution (all players)
    years_active_counts = defaultdict(int)
    years_active_2025_counts = defaultdict(int)
    for p in all_progressions:
        years_active_counts[p['years_active']] += 1
        if p.get('last_year') == 2025:
            years_active_2025_counts[p['years_active']] += 1
    
    ws[f'A{row+2}'] = "All Players (2+ years) by Years Active:"
    row = row + 2
    for years, count in sorted(years_active_counts.items()):
        ws[f'A{row}'] = f"  {years} years:"
        ws[f'B{row}'] = count
        row += 1
    
    ws[f'A{row}'] = "Active in 2025 by Years Active:"
    row += 1
    for years, count in sorted(years_active_2025_counts.items()):
        ws[f'A{row}'] = f"  {years} years:"
        ws[f'B{row}'] = count
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15

def create_detailed_progression_sheet(wb, all_progressions):
    """Create detailed player progression sheet."""
    
    ws = wb.create_sheet("Player Progressions", 0)
    
    # First, collect all unique metrics across all players
    all_metrics = set()
    for p in all_progressions:
        all_metrics.update(p['metrics'].keys())
    
    # Sort metrics alphabetically
    sorted_metrics = sorted(all_metrics)
    
    print(f"\n📊 Creating progression sheet with {len(sorted_metrics)} metrics...")
    
    # Build headers: Basic info, then all metrics with Change/Pct columns, then YoY
    headers = [
        'Player', 'Position', 'First Year', 'Last Year', 'Years Active',
        'Team Changes', 'First Team', 'Current Team'
    ]
    
    # Add overall progression columns for each metric (2 columns: abs_change, pct_change)
    for metric in sorted_metrics:
        headers.append(f'{metric}_Change')
        headers.append(f'{metric}_%_Change')
    
    # Add YoY columns for metrics that have YoY data
    yoy_metrics = set()
    for p in all_progressions:
        if 'yoy_from_2024' in p:
            yoy_metrics.update(p['yoy_from_2024'].keys())
    
    sorted_yoy_metrics = sorted(yoy_metrics)
    for metric in sorted_yoy_metrics:
        headers.append(f'YoY_{metric}_Change')
        headers.append(f'YoY_{metric}_%_Change')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data - filter to only players active in 2025
    row = 2
    for p in sorted(all_progressions, key=lambda x: (x['position'], x['player_name'])):
        # Only include players who were active in 2025
        if p.get('last_year') != 2025:
            continue
            
        col = 1
        
        # Basic info
        ws.cell(row=row, column=col, value=p['player_name']); col += 1
        ws.cell(row=row, column=col, value=p['position']); col += 1
        ws.cell(row=row, column=col, value=p['first_year']); col += 1
        ws.cell(row=row, column=col, value=p['last_year']); col += 1
        ws.cell(row=row, column=col, value=p['years_active']); col += 1
        ws.cell(row=row, column=col, value=p['team_changes']); col += 1
        ws.cell(row=row, column=col, value=p['first_team']); col += 1
        ws.cell(row=row, column=col, value=p['current_team']); col += 1
        
        # Overall progression metrics
        for metric in sorted_metrics:
            if metric in p['metrics']:
                m = p['metrics'][metric]
                ws.cell(row=row, column=col, value=m['abs_change']); col += 1
                ws.cell(row=row, column=col, value=m['pct_change']); col += 1
            else:
                col += 2  # Empty columns
        
        # YoY metrics
        for metric in sorted_yoy_metrics:
            if 'yoy_from_2024' in p and metric in p['yoy_from_2024']:
                m = p['yoy_from_2024'][metric]
                ws.cell(row=row, column=col, value=m['abs_change']); col += 1
                ws.cell(row=row, column=col, value=m['pct_change']); col += 1
            else:
                col += 2  # Empty columns
        
        row += 1
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def load_2025_scores(base_dir):
    """Load 2025 Total_Score_1_10 and Total_Grade from By_Team files."""
    
    scores_dict = {}
    
    # Position mappings
    position_map = {
        'AM Advanced Playmaker': 'Attacking Midfielder',
        'CB Hybrid': 'Center Back',
        'DM Box-To-Box': 'Centre Midfielder',
        'W Touchline Winger': 'Winger'
    }
    
    folder_map = {
        'Attacking Midfielder': '03_Advanced_Playmaker',
        'Center Back': '01_Hybrid_CB',
        'Centre Midfielder': '02_DM_Box_To_Box',
        'Winger': '04_Touchline_Winger'
    }
    
    conferences = ['ACC', 'BIG10', 'BIG12', 'IVY', 'SEC']
    
    for pos_file_prefix, position_name in position_map.items():
        folder_name = folder_map[position_name]
        
        for conference in conferences:
            # Load By_Team file for scores
            by_team_file = base_dir / folder_name / f"{conference}_2025_{position_name.replace(' ', '_')}_Intent_Focused_By_Team.xlsx"
            
            if by_team_file.exists():
                try:
                    df = pd.read_excel(by_team_file, sheet_name='All Players')
                    
                    for _, row in df.iterrows():
                        player_name = row['Player']
                        # Store for all possible positions (in case of multi-position players)
                        scores_dict[player_name] = {
                            'Total_Score_1_10': row.get('Total_Score_1_10'),
                            'Total_Grade': row.get('Total_Grade'),
                            'Position': position_name  # Track which position this score is for
                        }
                except Exception as e:
                    print(f"  ⚠️  Error loading scores from {by_team_file.name}: {e}")
    
    return scores_dict

def create_risers_fallers_sheet(wb, all_progressions, base_dir):
    """Create risers and fallers sheet."""
    
    ws = wb.create_sheet("Risers & Fallers", 0)
    
    # Load 2025 scores and grades
    scores_data = load_2025_scores(base_dir)
    
    # Calculate composite improvement scores
    for p in all_progressions:
        improvement_score = 0
        
        # Weight per-90 improvements
        weights = {'Goals': 0.3, 'Assists': 0.3, 'xA': 0.2, 'xG': 0.2}
        
        # Check YoY data for per_90_change
        if 'yoy_from_2024' in p:
            for metric, weight in weights.items():
                if metric in p['yoy_from_2024'] and 'per_90_change' in p['yoy_from_2024'][metric]:
                    improvement_score += p['yoy_from_2024'][metric]['per_90_change'] * weight
        
        p['improvement_score'] = improvement_score
        
        # Add 2025 scores if available (lookup by player name only for multi-position support)
        if p['player_name'] in scores_data:
            p['current_score'] = scores_data[p['player_name']]['Total_Score_1_10']
            p['current_grade'] = scores_data[p['player_name']]['Total_Grade']
        else:
            p['current_score'] = None
            p['current_grade'] = None
    
    # Filter to only players who played in 2025 (have YoY data)
    players_with_2025_data = [p for p in all_progressions if 'yoy_from_2024' in p and p.get('improvement_score', 0) != 0]
    
    # Sort by improvement
    sorted_by_improvement = sorted(players_with_2025_data, key=lambda x: x.get('improvement_score', 0), reverse=True)
    
    # Group by position and experience level
    from collections import defaultdict
    groups = defaultdict(list)
    for p in sorted_by_improvement:
        position = p['position']
        years = p.get('years_active', 0)
        if years == 2:
            exp_level = 'Sophomore'
        elif years == 3:
            exp_level = 'Junior'
        elif years >= 4:
            exp_level = 'Senior'
        else:
            exp_level = 'Other'
        
        groups[(position, exp_level)].append(p)
    
    row = 1
    
    # Top 20 risers overall
    ws['A1'] = "TOP 20 RISERS (OVERALL)"
    ws['A1'].font = Font(size=14, bold=True)
    row += 1
    
    headers = ['Player', 'Position', 'Experience', 'Current Team', 'Current Score', 'Current Grade', 'Years Active', 'Improvement Score']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    row += 1
    
    # Top 20 risers
    for i, p in enumerate(sorted_by_improvement[:20], row):
        years = p.get('years_active', 0)
        exp_level = 'Sophomore' if years == 2 else 'Junior' if years == 3 else 'Senior' if years >= 4 else 'Other'
        ws.cell(row=i, column=1, value=p['player_name'])
        ws.cell(row=i, column=2, value=p['position'])
        ws.cell(row=i, column=3, value=exp_level)
        ws.cell(row=i, column=4, value=p.get('current_team', ''))
        ws.cell(row=i, column=5, value=f"{p['current_score']:.2f}" if p['current_score'] else "")
        ws.cell(row=i, column=6, value=p['current_grade'] if p['current_grade'] else "")
        ws.cell(row=i, column=7, value=p['years_active'])
        ws.cell(row=i, column=8, value=f"{p['improvement_score']:.2f}")
    row += 21
    
    # Top 20 fallers overall
    ws.cell(row=row, column=1, value="TOP 20 FALLERS (OVERALL)")
    ws.cell(row=row, column=1).font = Font(size=14, bold=True)
    row += 1
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    row += 1
    
    # Top 20 fallers
    for i, p in enumerate(sorted_by_improvement[-20:], row):
        years = p.get('years_active', 0)
        exp_level = 'Sophomore' if years == 2 else 'Junior' if years == 3 else 'Senior' if years >= 4 else 'Other'
        ws.cell(row=i, column=1, value=p['player_name'])
        ws.cell(row=i, column=2, value=p['position'])
        ws.cell(row=i, column=3, value=exp_level)
        ws.cell(row=i, column=4, value=p.get('current_team', ''))
        ws.cell(row=i, column=5, value=f"{p['current_score']:.2f}" if p['current_score'] else "")
        ws.cell(row=i, column=6, value=p['current_grade'] if p['current_grade'] else "")
        ws.cell(row=i, column=7, value=p['years_active'])
        ws.cell(row=i, column=8, value=f"{p['improvement_score']:.2f}")
    row += 21
    
    # Grouped by position and experience - sort by position order, then by Sophomore > Junior > Senior
    pos_order = {'Center Back': 1, 'Centre Midfielder': 2, 'Winger': 3, 'Attacking Midfielder': 4}
    exp_order = {'Sophomore': 1, 'Junior': 2, 'Senior': 3, 'Other': 4}
    for (position, exp_level), players in sorted(groups.items(), key=lambda x: (pos_order.get(x[0][0], 99), exp_order.get(x[0][1], 99))):
        # Sort this group
        players_sorted = sorted(players, key=lambda x: x.get('improvement_score', 0), reverse=True)
        
        # Risers for this group
        ws.cell(row=row, column=1, value=f"TOP {exp_level.upper()} {position.upper()} RISERS")
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        row += 1
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        row += 1
        
        for p in players_sorted[:10]:  # Top 10 in each category
            ws.cell(row=row, column=1, value=p['player_name'])
            ws.cell(row=row, column=2, value=p['position'])
            ws.cell(row=row, column=3, value=exp_level)
            ws.cell(row=row, column=4, value=p.get('current_team', ''))
            ws.cell(row=row, column=5, value=f"{p['current_score']:.2f}" if p['current_score'] else "")
            ws.cell(row=row, column=6, value=p['current_grade'] if p['current_grade'] else "")
            ws.cell(row=row, column=7, value=p['years_active'])
            ws.cell(row=row, column=8, value=f"{p['improvement_score']:.2f}")
            row += 1
        
        row += 2
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

def main():
    """Main function."""
    
    base_dir = Path(__file__).parent.parent.parent
    output_path = base_dir / "Player_Progression_Report.xlsx"
    
    print("=" * 70)
    print("🏈 PLAYER PROGRESSION ANALYSIS")
    print("=" * 70)
    
    # Load player history
    player_history = load_player_history(base_dir)
    
    # Analyze progressions
    all_progressions = analyze_progressions(player_history)
    
    # Create report
    if all_progressions:
        create_progression_report(all_progressions, output_path)
        print(f"\n✅ Report created with {len(all_progressions)} player progressions")
    else:
        print("\n⚠️  No player progressions found to analyze")
    
    print("=" * 70)

if __name__ == "__main__":
    main()

