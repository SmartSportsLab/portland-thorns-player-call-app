#!/usr/bin/env python3
"""
Check that every team tab in every conference report has at least one player
in every position profile.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Position profiles that should be present
POSITION_PROFILES = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']

def check_team_tab_coverage(file_path, conference):
    """Check a single conference report file."""
    print(f"\n{'='*80}")
    print(f"Checking {conference} Report")
    print(f"{'='*80}")
    
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        
        # Filter to team sheets (exclude position profile sheets and Data Summary)
        team_sheets = [s for s in sheet_names if s not in POSITION_PROFILES + ['Data Summary']]
        
        print(f"\nFound {len(team_sheets)} team sheets")
        
        issues_found = []
        teams_checked = 0
        
        for team_sheet in team_sheets:
            teams_checked += 1
            ws = wb[team_sheet]
            
            # Read the sheet into a dataframe
            data = []
            headers = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    headers = [str(cell) if cell else '' for cell in row]
                    continue
                if not any(cell for cell in row):  # Skip empty rows
                    continue
                data.append([cell for cell in row])
            
            if not headers or not data:
                print(f"  ⚠️  {team_sheet}: Empty sheet or no data")
                issues_found.append({
                    'team': team_sheet,
                    'issue': 'Empty sheet or no data',
                    'missing_profiles': POSITION_PROFILES.copy()
                })
                continue
            
            # Create dataframe
            df = pd.DataFrame(data, columns=headers[:len(data[0])] if data else headers)
            
            # Check if Position_Profile column exists
            if 'Position_Profile' not in df.columns:
                # Try to find it with different case/spacing
                pos_cols = [col for col in df.columns if 'position' in str(col).lower() and 'profile' in str(col).lower()]
                if pos_cols:
                    df = df.rename(columns={pos_cols[0]: 'Position_Profile'})
                else:
                    print(f"  ⚠️  {team_sheet}: No Position_Profile column found")
                    issues_found.append({
                        'team': team_sheet,
                        'issue': 'No Position_Profile column',
                        'missing_profiles': POSITION_PROFILES.copy()
                    })
                    continue
            
            # Get unique position profiles in this team from Position_Profile column
            if 'Position_Profile' in df.columns and df['Position_Profile'].notna().any():
                present_profiles = df['Position_Profile'].dropna().unique().tolist()
                present_profiles = [str(p).strip() for p in present_profiles if str(p).strip()]
            else:
                present_profiles = []
            
            # Also check for position profile headers in the sheet (like "Hybrid CB Players")
            # This catches cases where the header exists but there are no players
            for row_idx in range(1, min(ws.max_row + 1, 200)):  # Check first 200 rows
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value:
                    cell_str = str(cell_value)
                    for profile in POSITION_PROFILES:
                        if profile in cell_str and ("Players" in cell_str or "players" in cell_str.lower()):
                            if profile not in present_profiles:
                                present_profiles.append(profile)
            
            # Check which profiles are missing
            missing_profiles = [p for p in POSITION_PROFILES if p not in present_profiles]
            
            if missing_profiles:
                print(f"  ❌ {team_sheet}: Missing {len(missing_profiles)} position profile(s)")
                print(f"      Missing: {', '.join(missing_profiles)}")
                print(f"      Present: {', '.join(present_profiles) if present_profiles else 'None'}")
                issues_found.append({
                    'team': team_sheet,
                    'issue': f'Missing {len(missing_profiles)} position profile(s)',
                    'missing_profiles': missing_profiles,
                    'present_profiles': present_profiles
                })
            else:
                print(f"  ✅ {team_sheet}: All position profiles present")
                if present_profiles:
                    profile_counts = df['Position_Profile'].value_counts().to_dict()
                    counts_str = ', '.join([f"{p}: {profile_counts.get(p, 0)}" for p in POSITION_PROFILES])
                    print(f"      Counts: {counts_str}")
        
        wb.close()
        
        return issues_found, teams_checked
        
    except Exception as e:
        print(f"  ❌ Error reading {file_path}: {e}")
        import traceback
        traceback.print_exc()
        return [], 0


def main():
    """Main function."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
    
    all_issues = {}
    total_teams_checked = 0
    
    for conference in conferences:
        report_file = base_dir / f"Portland Thorns 2025 {conference} Championship Scouting Report.xlsx"
        
        if not report_file.exists():
            print(f"\n⚠️  Report not found: {report_file.name}")
            continue
        
        issues, teams_checked = check_team_tab_coverage(report_file, conference)
        all_issues[conference] = issues
        total_teams_checked += teams_checked
    
    # Summary
    print(f"\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}")
    
    total_issues = sum(len(issues) for issues in all_issues.values())
    
    print(f"\nTotal teams checked: {total_teams_checked}")
    print(f"Total issues found: {total_issues}")
    
    if total_issues > 0:
        print(f"\n{'='*80}")
        print("ISSUES BY CONFERENCE")
        print(f"{'='*80}")
        
        for conference, issues in all_issues.items():
            if issues:
                print(f"\n{conference}: {len(issues)} teams with issues")
                for issue in issues:
                    print(f"  • {issue['team']}: {issue['issue']}")
                    if 'missing_profiles' in issue:
                        print(f"    Missing profiles: {', '.join(issue['missing_profiles'])}")
            else:
                print(f"\n{conference}: ✅ All teams have all position profiles")
    else:
        print(f"\n✅ SUCCESS: All teams in all conferences have at least one player in every position profile!")
    
    print(f"\n{'='*80}")


if __name__ == "__main__":
    main()

