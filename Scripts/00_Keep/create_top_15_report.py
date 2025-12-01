#!/usr/bin/env python3
"""
Create a shortlist report with grade B and above players from each position profile across all five conferences.
Extracts data from existing conference reports.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import sys

# Add path for imports
sys.path.insert(0, str(Path(__file__).parent))

# Grade colors (matching update_mike_norris_reports.py)
GRADE_COLORS = {
    'A': '8B0000',  # Dark red
    'B': 'C5504B',  # Red
    'C': 'F2A2A2',  # Light red
    'D': '8FAADC',  # Light blue
    'F': '1F4E79'   # Dark blue
}

# Position profile to internal name mapping
POSITION_PROFILE_MAP = {
    'Hybrid CB': 'Center Back',
    'DM Box-To-Box': 'Centre Midfielder',
    'AM Advanced Playmaker': 'Attacking Midfielder',
    'Right Touchline Winger': 'Winger'
}

# Position profile to initials mapping (for Changed Position column)
POSITION_INITIALS = {
    'Hybrid CB': 'HCB',
    'DM Box-To-Box': 'DM',
    'AM Advanced Playmaker': 'AM',
    'Right Touchline Winger': 'RTW'
}

def get_relevant_metrics_for_position(position_config):
    """Extract all relevant metric column names from position config."""
    relevant_metrics = set()
    
    if 'metrics' in position_config:
        for category in ['Core', 'Specific']:
            if category in position_config['metrics']:
                for metric_name, metric_value in position_config['metrics'][category].items():
                    if isinstance(metric_value, dict) and 'components' in metric_value:
                        # Composite metric - add all components
                        for comp_name in metric_value['components'].keys():
                            relevant_metrics.add(comp_name)
                    else:
                        # Simple metric
                        relevant_metrics.add(metric_name)
    
    return relevant_metrics

def load_players_from_report(report_file, position_profile, base_dir):
    """Load players from a position profile sheet in a conference report and enrich with raw data."""
    try:
        wb = load_workbook(report_file, data_only=True)
        
        if position_profile not in wb.sheetnames:
            wb.close()
            return pd.DataFrame()
        
        ws = wb[position_profile]
        
        # Read headers from row 1 and row 2 (merged headers)
        # Row 1 has base names, row 2 has suffixes like "per 90" or "% better than position"
        # Handle merged cells: if header1 spans multiple columns, header2 will have values in those columns
        headers = []
        last_header1 = None
        
        for col_idx in range(1, ws.max_column + 1):
            header1 = ws.cell(row=1, column=col_idx).value
            header2 = ws.cell(row=2, column=col_idx).value
            
            # If header1 exists, use it (might be from merged cell)
            if header1:
                last_header1 = str(header1).strip()
            
            # Combine headers intelligently
            header1_str = last_header1 if last_header1 else None
            header2_str = str(header2).strip() if header2 else None
            
            # Determine the full header name
            if header1_str and header2_str:
                # Check if header2 is a suffix
                if header2_str.lower() in ['per 90', 'per90', '% better than position', '%', 'won, %', 'won %', 'accurate, %', 'accurate %']:
                    full_header = f"{header1_str} {header2_str}"
                # Check if header1 is a base column (doesn't need suffix)
                elif header1_str.lower() in ['player', 'team', 'position', 'conference grade', 'power five grade', '2025 total score', 
                                              'previous year', 'previous score', 'change from previous', 'total minutes', 
                                              '% of team minutes', 'top 15s (power five)', 'seasons played', 'changed position']:
                    full_header = header1_str
                # If header2 is the same as header1 (merged cell), use header1
                elif header1_str.lower() == header2_str.lower():
                    full_header = header1_str
                else:
                    # Combine them
                    full_header = f"{header1_str} {header2_str}".strip()
            elif header1_str:
                full_header = header1_str
            elif header2_str:
                full_header = header2_str
            else:
                full_header = f"Column_{col_idx}"
            
            headers.append(full_header)
        
        # Read data starting from row 3 (row 1 = headers, row 2 = sub-headers, row 3+ = data)
        data = []
        for row_idx in range(3, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, len(headers) + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            
            # Skip empty rows
            if not any(cell for cell in row_data if cell not in [None, '', ' ']):
                continue
            
            data.append(row_data)
        
        wb.close()
        
        if not data:
            return pd.DataFrame()
        
        # Create dataframe
        df = pd.DataFrame(data, columns=headers[:len(data[0])] if data else headers)
        
        # Get conference name from filename
        conference = report_file.stem.replace('Portland Thorns 2025 ', '').replace(' Championship Scouting Report', '')
        df['Conference'] = conference
        
        # Load raw data to get percentage metrics
        # Map position profiles to file prefixes
        position_to_prefix = {
            'Hybrid CB': 'CB Hybrid',
            'DM Box-To-Box': 'DM Box-To-Box',
            'AM Advanced Playmaker': 'AM Advanced Playmaker',
            'Right Touchline Winger': 'W Touchline Winger'
        }
        
        file_prefix = position_to_prefix.get(position_profile)
        if file_prefix:
            raw_file = base_dir / "Exports" / "Players Stats By Position" / f"{file_prefix} {conference} 2025.xlsx"
            if raw_file.exists():
                try:
                    df_raw = pd.read_excel(raw_file)
                    # Merge ALL relevant metrics from raw data (both per 90 and percentage)
                    if 'Player' in df_raw.columns and 'Player' in df.columns:
                        # Get all metric columns from raw data (exclude non-metric columns)
                        exclude_cols = ['Team', 'Position', 'Minutes played', 'Duration']
                        metric_cols = [col for col in df_raw.columns if col not in exclude_cols and col != 'Player']
                        # Don't include Player in merge_cols - we'll preserve the report's Player column
                        merge_cols = metric_cols
                        df_raw_subset = df_raw[['Player'] + metric_cols].copy()
                        
                        # Merge with main dataframe, keeping report columns but adding raw metrics
                        # Use suffixes to avoid conflicts - prefer raw data for metrics
                        # Try fuzzy matching on player names (case-insensitive, handle variations)
                        df['Player_normalized'] = df['Player'].astype(str).str.strip().str.lower()
                        df_raw_subset['Player_normalized'] = df_raw_subset['Player'].astype(str).str.strip().str.lower()
                        
                        # Merge on normalized player names AND Team to ensure correct matching
                        # First, normalize team names too
                        df['Team_normalized'] = df['Team'].astype(str).str.strip().str.lower()
                        if 'Team' in df_raw_subset.columns:
                            df_raw_subset['Team_normalized'] = df_raw_subset['Team'].astype(str).str.strip().str.lower()
                            merge_on = ['Player_normalized', 'Team_normalized']
                        else:
                            merge_on = ['Player_normalized']
                        
                        # Merge with suffixes - raw data columns (no suffix) will overwrite report columns (with _report suffix)
                        # IMPORTANT: Exclude 'Player' from the merge to preserve the report's Player column
                        df_merged = df.merge(df_raw_subset[merge_on + metric_cols], on=merge_on, how='left', suffixes=('_report', ''))
                        
                        # Drop the normalized columns
                        df_merged = df_merged.drop(columns=['Player_normalized'])
                        if 'Team_normalized' in df_merged.columns:
                            df_merged = df_merged.drop(columns=['Team_normalized'])
                        
                        # If merge failed for some players, try fallback: match on player name only
                        if metric_cols:
                            # Check which players didn't get merged (have NaN in first metric column)
                            first_metric_col = metric_cols[0]
                            if first_metric_col in df_merged.columns:
                                unmatched_mask = df_merged[first_metric_col].isna()
                                unmatched = df_merged[unmatched_mask]
                                
                                if len(unmatched) > 0 and len(df_raw_subset) > 0:
                                    print(f"     ⚠️  {len(unmatched)} players didn't merge, trying fallback matching...")
                                    # Try to match unmatched players by name only
                                    for idx, row in unmatched.iterrows():
                                        player_name = str(row.get('Player', '')).strip()
                                        team_name = str(row.get('Team', '')).strip()
                                        
                                        # Try to find in raw data by name match
                                        for raw_idx, raw_row in df_raw_subset.iterrows():
                                            raw_player = str(raw_row.get('Player', '')).strip()
                                            raw_team = str(raw_row.get('Team', '')).strip() if 'Team' in raw_row.index else ''
                                            
                                            # Check if names match (exact or similar)
                                            if player_name.lower() == raw_player.lower():
                                                # Also check team if available
                                                if not raw_team or team_name.lower() == raw_team.lower():
                                                    # Copy ALL metrics from raw data
                                                    for col in metric_cols:
                                                        if col in df_raw_subset.columns:
                                                            df_merged.at[idx, col] = raw_row[col]
                                                    break
                        
                        df = df_merged
                except Exception as e:
                    print(f"     ⚠️  Could not load raw data for {position_profile}: {e}")
        
        return df
        
    except Exception as e:
        print(f"  ⚠️  Error loading {position_profile} from {report_file.name}: {e}")
        return pd.DataFrame()


def load_seasons_data(base_dir):
    """Load historical data to count seasons played and get most recent previous season scores."""
    from update_mike_norris_reports import load_historical_data, filter_by_position, POSITION_PROFILE_MAP
    from update_mike_norris_reports import calculate_with_historical_normalization
    import json
    
    seasons_count = {}  # {player_name: set of years}
    previous_scores = {}  # {player_name: {position_profile: (year, score)}}
    
    # Load config for scoring
    config_file = base_dir / "Scripts" / "00_Keep" / "position_metrics_config.json"
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
    position_profiles = ['Center Back', 'Centre Midfielder', 'Attacking Midfielder', 'Winger']
    
    # Map position profiles to file prefixes
    position_to_prefix = {
        'Center Back': 'CB Hybrid',
        'Centre Midfielder': 'DM Box-To-Box',
        'Attacking Midfielder': 'AM Advanced Playmaker',
        'Winger': 'W Touchline Winger'
    }
    
    for conference in conferences:
        try:
            df_historical = load_historical_data(base_dir, conference)
            if df_historical.empty or 'Player' not in df_historical.columns or 'Year' not in df_historical.columns:
                continue
            
            # Count seasons
            for _, row in df_historical.iterrows():
                player = str(row.get('Player', '')).strip()
                year = row.get('Year')
                if player and pd.notna(year):
                    if player not in seasons_count:
                        seasons_count[player] = set()
                    seasons_count[player].add(int(year))
            
            # Calculate scores for each year and position (2021-2024)
            for year in range(2021, 2025):  # 2021-2024 (before 2025)
                df_year = df_historical[df_historical['Year'] == year].copy()
                if df_year.empty:
                    continue
                
                for position_name in position_profiles:
                    if position_name not in config['position_profiles']:
                        continue
                    
                    position_config = config['position_profiles'][position_name]
                    display_name = POSITION_PROFILE_MAP.get(position_name, position_name)
                    
                    # Filter by position
                    df_year_pos = filter_by_position(df_year, position_name)
                    if len(df_year_pos) == 0:
                        continue
                    
                    # Try to add PAdj metrics if possible, but continue without them if not available
                    try:
                        from update_mike_norris_reports import load_team_possessions, add_padj_metrics_to_dataframe
                        team_possessions, league_avg_possessions = load_team_possessions(base_dir, conference, year)
                        df_year_pos = add_padj_metrics_to_dataframe(
                            df_year_pos, team_possessions, league_avg_possessions, team_col='Team'
                        )
                        # Update config to use PAdj
                        from update_mike_norris_reports import update_config_to_use_padj
                        position_config_for_scoring = update_config_to_use_padj(position_config, df_year_pos)
                    except (ImportError, AttributeError, Exception):
                        # If PAdj not available, use regular config
                        position_config_for_scoring = position_config
                    
                    # Get historical data for scoring (years before current year)
                    df_all_pos = filter_by_position(df_historical[df_historical['Year'] < year], position_name)
                    if len(df_all_pos) == 0:
                        # If no historical data before this year, use this year's data as reference
                        df_all_pos = df_year_pos.copy()
                    
                    # Score players
                    pass_attempt_weight = 0.8
                    pass_accuracy_weight = 0.2
                    
                    try:
                        df_scored = calculate_with_historical_normalization(
                            df_year_pos, df_all_pos, position_config_for_scoring,
                            pass_attempt_weight, pass_accuracy_weight, position_name
                        )
                        
                        # Store scores
                        for _, row in df_scored.iterrows():
                            player = str(row.get('Player', '')).strip()
                            score = row.get('Total_Score_1_10')
                            
                            if player and pd.notna(score):
                                if player not in previous_scores:
                                    previous_scores[player] = {}
                                
                                # Store the most recent score for this position profile
                                # Keep the highest year (most recent) score
                                if display_name not in previous_scores[player]:
                                    previous_scores[player][display_name] = (year, score)
                                else:
                                    # Keep the most recent year
                                    existing_year, existing_score = previous_scores[player][display_name]
                                    if year > existing_year:
                                        previous_scores[player][display_name] = (year, score)
                                
                                # Also store the most recent score across ALL positions (for position changers)
                                # This allows us to find previous scores even if player changed positions
                                if '_most_recent' not in previous_scores[player]:
                                    previous_scores[player]['_most_recent'] = (year, score, display_name)
                                else:
                                    existing_year, existing_score, existing_pos = previous_scores[player]['_most_recent']
                                    if year > existing_year:
                                        previous_scores[player]['_most_recent'] = (year, score, display_name)
                    except Exception as e:
                        # Skip if scoring fails for this year/position
                        print(f"     ⚠️  Could not calculate scores for {conference} {year} {position_name}: {e}")
                        pass
                        
        except Exception as e:
            print(f"  ⚠️  Could not load historical data for {conference}: {e}")
    
    # Convert to count
    seasons_dict = {player: len(years) for player, years in seasons_count.items()}
    return seasons_dict, previous_scores


def load_all_players_from_reports(base_dir):
    """Load all players from all conference reports."""
    conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
    position_profiles = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    all_data = {}
    
    # Load seasons data and previous scores
    print("\n📅 Loading historical data to count seasons played and get previous scores...")
    seasons_dict, previous_scores_dict = load_seasons_data(base_dir)
    print(f"  ✅ Loaded seasons data for {len(seasons_dict)} players")
    print(f"  ✅ Loaded previous scores for {len(previous_scores_dict)} players")
    
    for position_profile in position_profiles:
        all_data[position_profile] = []
        
        for conference in conferences:
            report_file = base_dir / f"Portland Thorns 2025 {conference} Championship Scouting Report.xlsx"
            
            if not report_file.exists():
                print(f"  ⚠️  Report not found: {report_file.name}")
                continue
            
            print(f"  📄 Loading {position_profile} from {conference}...")
            df = load_players_from_report(report_file, position_profile, base_dir)
            
            if len(df) > 0:
                # Add seasons played column
                df['Seasons Played'] = df['Player'].apply(
                    lambda x: seasons_dict.get(str(x).strip(), 1) if pd.notna(x) else 1
                )
                
                # Add previous season score
                # First, check what previous score columns exist in the dataframe
                # Note: After merge with suffixes=('_report', ''), report columns get '_report' suffix if there's a conflict
                # So we need to check for both the original name and the suffixed name
                prev_score_cols = [
                        '2024 Total Score_report',  # Check suffixed version first (from report)
                        '2024 Total Score',  # Then check original (might be from raw data or no conflict)
                        '2024_Total_Score', '2024 Total Score_1_10',
                        'Previous Score_report', 'Previous Score',
                        'Previous_Score', 'Previous Year Score',
                        '2023 Total Score_report', '2023 Total Score',
                        '2023_Total_Score'
                    ]
                
                # Find which previous score column exists in this dataframe
                available_prev_col = None
                prev_year = None
                for col in prev_score_cols:
                    if col in df.columns:
                        available_prev_col = col
                        if '2024' in col:
                            prev_year = 2024
                        elif '2023' in col:
                            prev_year = 2023
                        break
                
                def get_previous_score(row):
                    player = str(row.get('Player', '')).strip()
                    if not player:
                        return None, None, None
                    
                    seasons = row.get('Seasons Played', 1)
                    prev_position = None
                    
                    # PRIORITY 1: Check if player has previous scores from calculated historical data
                    # This looks at all years 2021-2024, not just 2024
                    if player in previous_scores_dict:
                        # First, check if they have a score for the current position profile
                        if position_profile in previous_scores_dict[player]:
                            year, score = previous_scores_dict[player][position_profile]
                            prev_position = position_profile  # Same position
                            return year, score, prev_position
                        
                        # If not, check for most recent score across ALL positions (for position changers)
                        if '_most_recent' in previous_scores_dict[player]:
                            year, score, prev_position = previous_scores_dict[player]['_most_recent']
                            return year, score, prev_position
                        
                        # Fallback: check all position profiles for this player
                        # Get the most recent score from any position
                        most_recent = None
                        most_recent_year = 0
                        most_recent_pos = None
                        for pos_name, value in previous_scores_dict[player].items():
                            if pos_name != '_most_recent':
                                if isinstance(value, tuple) and len(value) >= 2:
                                    pos_year, pos_score = value[0], value[1]
                                    if isinstance(pos_year, int):
                                        if pos_year > most_recent_year:
                                            most_recent_year = pos_year
                                            most_recent = (pos_year, pos_score)
                                            most_recent_pos = pos_name
                        if most_recent:
                            return most_recent[0], most_recent[1], most_recent_pos
                    
                    # PRIORITY 2: Check if the report already has previous score columns (2024 fallback)
                    # This is a fallback if historical calculation didn't work
                    if available_prev_col:
                        score_val = row.get(available_prev_col)
                        if pd.notna(score_val) and score_val != '':
                            try:
                                score_val = float(score_val)
                                if score_val > 0:
                                    return prev_year, score_val, None  # Don't know previous position from fallback
                            except (ValueError, TypeError):
                                pass
                    
                    # Check if 2025 is their first season
                    if seasons == 1:
                        return 'Rookie', None, None  # Rookie
                    
                    # If they have multiple seasons but no previous score found
                    # This could mean: they didn't play in previous years, transferred, or data is missing
                    # Return None to indicate missing data (not a rookie, but no score available)
                    return None, None, None
                
                previous_data = df.apply(get_previous_score, axis=1, result_type='expand')
                df['Previous Year'] = previous_data[0]
                df['Previous Score'] = previous_data[1]
                df['Previous Position'] = previous_data[2]
                
                # Determine if position changed and include previous position initials
                def format_changed_position(row):
                    prev_pos = row.get('Previous Position')
                    if pd.notna(prev_pos) and prev_pos != position_profile and prev_pos != '':
                        # Get initials for previous position
                        prev_initials = POSITION_INITIALS.get(prev_pos, prev_pos[:3].upper())
                        return f'✓ ({prev_initials})'
                    return ''
                
                df['Changed Position'] = df.apply(format_changed_position, axis=1)
                
                all_data[position_profile].append(df)
                print(f"     ✅ Loaded {len(df)} players")
    
    # Combine all conferences for each position profile and remove duplicates
    combined_data = {}
    for position_profile, dataframes in all_data.items():
        if dataframes:
            combined = pd.concat(dataframes, ignore_index=True)
            
            # Remove duplicates based on Player + Team combination
            # Keep the first occurrence (which should have the best data)
            before_dedup = len(combined)
            combined = combined.drop_duplicates(subset=['Player', 'Team'], keep='first').reset_index(drop=True)
            after_dedup = len(combined)
            
            if before_dedup != after_dedup:
                print(f"\n📊 {position_profile}: {before_dedup} total players, {before_dedup - after_dedup} duplicates removed, {after_dedup} unique players")
            else:
                print(f"\n📊 {position_profile}: {after_dedup} total players across all conferences (no duplicates)")
            
            combined_data[position_profile] = combined
        else:
            print(f"\n⚠️  {position_profile}: No data loaded")
    
    return combined_data


def create_top_15_report(base_dir):
    """Create Excel report with grade B and above players per position profile."""
    print("="*80)
    print("CREATING PORTLAND THORNS 2025 SHORTLIST REPORT")
    print("="*80)
    
    # Load config to get relevant metrics
    config_file = base_dir / "Scripts" / "00_Keep" / "position_metrics_config.json"
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    # Load all players from existing reports
    all_data = load_all_players_from_reports(base_dir)
    
    if not all_data:
        print("❌ No data loaded!")
        return
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Position profiles in order
    position_profiles = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    note_font = Font(italic=True, size=9)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for position_profile in position_profiles:
        if position_profile not in all_data:
            print(f"  ⚠️  No data for {position_profile}")
            continue
        
        df_profile = all_data[position_profile]
        if df_profile is None:
            print(f"  ⚠️  No data for {position_profile}")
            continue
        
        if isinstance(df_profile, pd.DataFrame):
            if df_profile.empty:
                print(f"  ⚠️  No data for {position_profile}")
                continue
            df = df_profile.copy()
        else:
            print(f"  ⚠️  Invalid data type for {position_profile}")
            continue
        
        # Filter by 70%+ team minutes
        minutes_pct_col = None
        for col in ['% of Team Minutes', 'Pct_Of_Team_Minutes', '% of team minutes']:
            if col in df.columns:
                minutes_pct_col = col
                break
        
        if minutes_pct_col:
            # Convert to numeric
            df[minutes_pct_col] = pd.to_numeric(df[minutes_pct_col], errors='coerce')
            before_filter = len(df)
            df = df[df[minutes_pct_col] >= 70.0].copy()
            after_filter = len(df)
            print(f"  📊 Filtered to {after_filter} players with 70%+ team minutes (from {before_filter})")
        else:
            print(f"  ⚠️  Could not find minutes percentage column for {position_profile}")
        
        # Find the score column
        score_col = None
        for col in ['Total_Score_1_10', '2025 Total Score', 'Total Score', 'Score']:
            if col in df.columns:
                score_col = col
                break
        
        if score_col is None:
            print(f"  ⚠️  No score column found for {position_profile}")
            continue
        
        # Convert score column to numeric
        df[score_col] = pd.to_numeric(df[score_col], errors='coerce')
        df = df[df[score_col].notna()].copy()
        
        # Note: Regression filter removed - keeping all players regardless of score change
        
        # Find both grade columns
        conf_grade_col = None
        power_grade_col = None
        
        for col in ['Conference_Grade', 'Conference Grade']:
            if col in df.columns:
                conf_grade_col = col
                break
        
        for col in ['Power_Five_Grade', 'Power Five Grade']:
            if col in df.columns:
                power_grade_col = col
                break
        
        if conf_grade_col is None or power_grade_col is None:
            print(f"  ⚠️  Missing grade columns for {position_profile} (Conference: {conf_grade_col}, Power Five: {power_grade_col})")
            continue
        
        # Filter to grades A and B for BOTH Conference Grade AND Power Five Grade
        df['Conf_Grade'] = df[conf_grade_col].astype(str).str.strip().str.upper()
        df['Power_Grade'] = df[power_grade_col].astype(str).str.strip().str.upper()
        filtered_players = df[(df['Conf_Grade'].isin(['A', 'B'])) & (df['Power_Grade'].isin(['A', 'B']))].copy()
        
        # Sort by grade (A first), then by score (descending)
        grade_order = {'A': 0, 'B': 1}
        filtered_players['_Conf_Grade_Sort'] = filtered_players['Conf_Grade'].map(grade_order).fillna(999)
        filtered_players['_Power_Grade_Sort'] = filtered_players['Power_Grade'].map(grade_order).fillna(999)
        # Sort by best grade first (lower sort value), then by score (descending)
        filtered_players['_Best_Grade_Sort'] = filtered_players[['_Conf_Grade_Sort', '_Power_Grade_Sort']].min(axis=1)
        filtered_players = filtered_players.sort_values(['_Best_Grade_Sort', score_col], ascending=[True, False])
        filtered_players = filtered_players.drop(columns=['_Conf_Grade_Sort', '_Power_Grade_Sort', '_Best_Grade_Sort', 'Conf_Grade', 'Power_Grade'])
        
        # Add rank
        filtered_players['Rank'] = range(1, len(filtered_players) + 1)
        
        print(f"  📊 Found {len(filtered_players)} players with grade B+ in both Conference and Power Five (70%+ minutes)")
        
        # Get relevant metrics for this position
        position_name = POSITION_PROFILE_MAP.get(position_profile, position_profile)
        if position_name in config['position_profiles']:
            position_config = config['position_profiles'][position_name]
            relevant_metrics = get_relevant_metrics_for_position(position_config)
        else:
            relevant_metrics = set()
        
        # Create sheet
        ws = wb.create_sheet(title=position_profile)
        
        # Title and note
        ws.merge_cells('A1:Z1')
        title_cell = ws.cell(row=1, column=1, value=f"{position_profile} Players - Grade B+ in Both Conference & Power Five (70%+ Minutes)")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Note about scoring methodology
        ws.merge_cells('A2:Z2')
        note_cell = ws.cell(row=2, column=1, value="Note: Filters: 70%+ team minutes, Grade B+ in both Conference & Power Five. Scores calculated against Past Seasons (2021-2025 Power Five conferences)")
        note_cell.font = note_font
        note_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Base headers
        base_headers = [
            'Rank', 'Player', 'Team', 'Conference', 'Position', 
            'Total Score', 'Conference Grade', 'Power Five Grade',
            'Previous Year', 'Previous Score', 'Change From Previous', 'Changed Position', 'Total Minutes', '% of Team Minutes', 'Seasons Played', 'Top 15s (Power Five)',
            'Consistency Score', 'Metrics Above Avg', 'Metrics Below Avg', 'Metrics At Avg', 'Consistency %', 'Style Fits'
        ]
        
        # Define the desired metric order (for Hybrid CB - will be filtered per position)
        desired_metric_order = [
            'Defensive duels per 90',
            'Defensive duels won, %',
            'Aerial duels per 90',
            'Aerial duels won, %',
            'Interceptions + Sliding Tackles',  # Combined metric
            'Shots blocked per 90',
            'Passes per 90',
            'Accurate passes, %',
            'Progressive passes per 90',
            'Accurate progressive passes, %',
            'Long passes per 90',
            'Accurate long passes, %',
            'Short / medium passes per 90',
            'Accurate short / medium passes, %'
        ]
        
        # Add relevant metrics as headers - match exact metric names from config
        metric_headers = []
        metric_column_map = {}  # Map metric name -> actual column name in dataframe
        padj_interceptions_col = None
        padj_sliding_tackles_col = None
        
        def normalize_metric_name(name):
            """Normalize metric name for matching - remove suffixes and normalize"""
            name = str(name).lower()
            # Remove common suffixes
            name = name.replace(' per 90', '').replace(' per90', '')
            name = name.replace(' won, %', '').replace(' won %', '').replace(' won', '')
            name = name.replace('accurate, %', '').replace('accurate %', '').replace(' accurate', '')
            name = name.replace(', %', '').replace(' %', '')
            name = name.replace('padj ', '').replace('padj', '')
            name = ' '.join(name.split())  # Clean up spaces
            return name.strip()
        
        def is_percentage_metric(metric_name):
            """Check if a metric is a percentage metric"""
            metric_lower = str(metric_name).lower()
            return '%' in metric_lower or 'won' in metric_lower or 'accurate' in metric_lower
        
        def is_per_90_metric(metric_name):
            """Check if a metric is a per 90 metric"""
            metric_lower = str(metric_name).lower()
            return 'per 90' in metric_lower or 'per90' in metric_lower
        
        def get_base_metric_name(name):
            """Get base metric name without modifiers"""
            name = normalize_metric_name(name)
            # Remove common prefixes
            if name.startswith('padj '):
                name = name[5:]
            return name.strip()
        
        for metric in sorted(relevant_metrics):
            found_col = None
            metric_normalized = normalize_metric_name(metric)
            metric_base = get_base_metric_name(metric)
            
            # Check if it's a PAdj metric
            is_padj = metric.startswith("PAdj") or "padj" in metric.lower()
            
            # Check if it's a percentage or per 90 metric
            is_percentage = is_percentage_metric(metric)
            is_per90 = is_per_90_metric(metric)
            
            # Track PAdj Interceptions and Sliding tackles for combining
            if metric == "PAdj Interceptions":
                padj_interceptions_col = None
            elif metric == "PAdj Sliding tackles":
                padj_sliding_tackles_col = None
            
            # Try exact match first
            if metric in df.columns:
                found_col = metric
            else:
                # Try case-insensitive exact match
                for col in df.columns:
                    if col.lower() == metric.lower():
                        found_col = col
                        break
                
                # Try normalized match - but be careful about percentage vs per 90
                if not found_col:
                    for col in df.columns:
                        col_normalized = normalize_metric_name(col)
                        if col_normalized == metric_normalized:
                            # Verify type matches (percentage vs per 90)
                            col_is_percentage = is_percentage_metric(col)
                            col_is_per90 = is_per_90_metric(col)
                            
                            if (is_percentage and col_is_percentage) or (is_per90 and col_is_per90) or (not is_percentage and not is_per90):
                                found_col = col
                                break
                
                # Try base name match with type checking
                if not found_col:
                    for col in df.columns:
                        col_base = get_base_metric_name(col)
                        if col_base == metric_base:
                            col_is_percentage = is_percentage_metric(col)
                            col_is_per90 = is_per_90_metric(col)
                            
                            # Match type (percentage vs per 90)
                            if (is_percentage and col_is_percentage) or (is_per90 and col_is_per90) or (not is_percentage and not is_per90 and not col_is_percentage and not col_is_per90):
                                # Prefer PAdj version if metric is PAdj
                                if is_padj and ("padj" in col.lower()):
                                    found_col = col
                                    break
                                elif not is_padj and ("padj" not in col.lower()):
                                    found_col = col
                                    break
                
                # Try word-based matching with type checking
                if not found_col:
                    metric_words = set(metric_normalized.split())
                    for col in df.columns:
                        col_normalized = normalize_metric_name(col)
                        col_words = set(col_normalized.split())
                        # Check if all metric words are in column (or vice versa for short metrics)
                        if metric_words.issubset(col_words) or col_words.issubset(metric_words):
                            # Verify type matches
                            col_is_percentage = is_percentage_metric(col)
                            col_is_per90 = is_per_90_metric(col)
                            
                            if (is_percentage and col_is_percentage) or (is_per90 and col_is_per90) or (not is_percentage and not is_per90 and not col_is_percentage and not col_is_per90):
                                # Prefer PAdj version if metric is PAdj
                                if is_padj and ("padj" in col.lower()):
                                    found_col = col
                                    break
                                elif not is_padj and ("padj" not in col.lower()):
                                    if not found_col:  # Only set if we haven't found one yet
                                        found_col = col
            
            # Track PAdj columns for combining
            if metric == "PAdj Interceptions" and found_col:
                padj_interceptions_col = found_col
            elif metric == "PAdj Sliding tackles" and found_col:
                padj_sliding_tackles_col = found_col
            
            # Skip individual PAdj Interceptions and Sliding tackles - we'll combine them
            if metric in ["PAdj Interceptions", "PAdj Sliding tackles"]:
                continue
            
            # Always add metric from config, even if not found (will show empty)
            if metric not in metric_column_map:
                metric_headers.append(metric)  # Use metric name from config as header
                metric_column_map[metric] = found_col if found_col else None
        
        # Add combined "Interceptions + Sliding Tackles" metric
        if padj_interceptions_col or padj_sliding_tackles_col:
            metric_headers.append("Interceptions + Sliding Tackles")
            metric_column_map["Interceptions + Sliding Tackles"] = {
                'interceptions': padj_interceptions_col,
                'sliding_tackles': padj_sliding_tackles_col
            }
        
        # Reorder metrics according to desired order (only include metrics that exist)
        ordered_metric_headers = []
        for desired_metric in desired_metric_order:
            if desired_metric in metric_headers:
                ordered_metric_headers.append(desired_metric)
        
        # Add any remaining metrics that weren't in the desired order
        for metric in metric_headers:
            if metric not in ordered_metric_headers:
                ordered_metric_headers.append(metric)
        
        metric_headers = ordered_metric_headers
        
        # Calculate consistency rankings BEFORE creating the sheet
        from calculate_consistency_ranking import calculate_consistency_for_shortlist
        filtered_players = calculate_consistency_for_shortlist(
            filtered_players, position_profile, metric_headers, metric_column_map, base_dir
        )
        
        # Calculate style fits for each player
        print(f"  📊 Calculating style fits for {position_profile}...")
        from generate_player_overviews import (
            load_nwsl_data, 
            load_player_data_from_conference_reports,
            calculate_thorns_style_fit
        )
        
        # Load NWSL data and position configs
        nwsl_dir = base_dir / 'Exports' / 'Team Stats By Conference' / 'NWSL'
        if nwsl_dir.exists():
            league_df, league_avg, thorns_data, thorns_ranks = load_nwsl_data(nwsl_dir)
            
            # Filter to top 3 metrics only
            top3_thorns_ranks = {k: v for k, v in (thorns_ranks or {}).items() if v['rank'] <= 3 and 'Unnamed' not in str(k)}
            
            # Load all players for ranking
            all_players_df = load_player_data_from_conference_reports(base_dir)
            
            # Load position configs
            config_file = base_dir / 'Scripts' / '00_Keep' / 'position_metrics_config.json'
            position_configs = {}
            if config_file.exists():
                with open(config_file, 'r') as f:
                    position_configs = json.load(f)
            
            # Calculate style fits for each player
            style_fits_list = []
            for idx, player_row in filtered_players.iterrows():
                style_fits = calculate_thorns_style_fit(
                    player_row, top3_thorns_ranks, all_players_df, position_profile, position_configs
                )
                style_fits_list.append(len(style_fits))
            
            filtered_players['Style Fits'] = style_fits_list
            print(f"     ✅ Calculated style fits: {sum(style_fits_list)} total matches across {len([x for x in style_fits_list if x > 0])} players")
        else:
            print(f"     ⚠️  NWSL data directory not found, skipping style fit calculation")
            filtered_players['Style Fits'] = 0
        
        all_headers = base_headers + metric_headers
        
        # Write headers
        header_row = 3
        for col_idx, header in enumerate(all_headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Data rows
        for row_idx, (idx, player_row) in enumerate(filtered_players.iterrows(), 4):
            # Base columns
            ws.cell(row=row_idx, column=1, value=player_row['Rank']).border = border
            ws.cell(row=row_idx, column=2, value=player_row.get('Player', '')).border = border
            ws.cell(row=row_idx, column=3, value=player_row.get('Team', '')).border = border
            ws.cell(row=row_idx, column=4, value=player_row.get('Conference', '')).border = border
            ws.cell(row=row_idx, column=5, value=player_row.get('Position', '')).border = border
            
            # Total Score
            score = player_row.get('Total_Score_1_10', player_row.get('2025 Total Score', player_row.get('Total Score', '')))
            score_cell = ws.cell(row=row_idx, column=6)
            if pd.notna(score) and score != '':
                try:
                    score_cell.value = round(float(score), 2)
                except:
                    score_cell.value = score
            score_cell.border = border
            
            # Conference Grade (with color)
            conf_grade = player_row.get('Conference_Grade', player_row.get('Conference Grade', ''))
            conf_grade_cell = ws.cell(row=row_idx, column=7, value=str(conf_grade) if pd.notna(conf_grade) else '')
            if conf_grade and str(conf_grade).strip().upper() in GRADE_COLORS:
                conf_grade_cell.fill = PatternFill(start_color=GRADE_COLORS[str(conf_grade).strip().upper()], 
                                                   end_color=GRADE_COLORS[str(conf_grade).strip().upper()], 
                                                   fill_type='solid')
                conf_grade_cell.font = Font(bold=True, color="FFFFFF")
            conf_grade_cell.border = border
            
            # Power Five Grade (with color)
            power_grade = player_row.get('Power_Five_Grade', player_row.get('Power Five Grade', ''))
            power_grade_cell = ws.cell(row=row_idx, column=8, value=str(power_grade) if pd.notna(power_grade) else '')
            if power_grade and str(power_grade).strip().upper() in GRADE_COLORS:
                power_grade_cell.fill = PatternFill(start_color=GRADE_COLORS[str(power_grade).strip().upper()], 
                                                   end_color=GRADE_COLORS[str(power_grade).strip().upper()], 
                                                   fill_type='solid')
                power_grade_cell.font = Font(bold=True, color="FFFFFF")
            power_grade_cell.border = border
            
            # Previous Year and Previous Score (from our calculated column or fallback to 2024)
            previous_year = player_row.get('Previous Year')
            previous_score = player_row.get('Previous Score')
            
            # Fallback to 2024 score if our calculation didn't work
            if pd.isna(previous_score) or previous_score is None:
                # Check for 2024 score in various possible column names (including after merge)
                for col_name in ['2024 Total Score_report', '2024 Total Score', '2024_Total_Score']:
                    if col_name in player_row.index:
                        fallback_score = player_row.get(col_name)
                        if pd.notna(fallback_score) and fallback_score != '':
                            try:
                                previous_score = float(fallback_score)
                                previous_year = 2024
                                break
                            except (ValueError, TypeError):
                                pass
            
            # Write Previous Year (column 9)
            previous_year_cell = ws.cell(row=row_idx, column=9)
            if previous_year == 'Rookie' or previous_year == 'R':
                previous_year_cell.value = 'Rookie'
            elif pd.notna(previous_year) and previous_year is not None:
                try:
                    # Convert float to int if it's a year
                    if isinstance(previous_year, float):
                        previous_year_cell.value = int(previous_year)
                    else:
                        previous_year_cell.value = previous_year
                except:
                    previous_year_cell.value = previous_year
            else:
                previous_year_cell.value = ''
            previous_year_cell.border = border
            
            # Write Previous Score (column 10)
            previous_score_cell = ws.cell(row=row_idx, column=10)
            
            # Check if rookie
            if previous_year == 'Rookie' or previous_year == 'R':
                previous_score_cell.value = 'N/A'
            elif pd.notna(previous_score) and previous_score != '':
                try:
                    previous_score_cell.value = round(float(previous_score), 2)
                except:
                    previous_score_cell.value = previous_score
            else:
                previous_score_cell.value = ''
            previous_score_cell.border = border
            
            # Change From Previous (calculate if we have both scores)
            current_score = player_row.get('Total_Score_1_10', player_row.get('2025 Total Score', player_row.get('Total Score', '')))
            change_cell = ws.cell(row=row_idx, column=11)
            
            if pd.notna(current_score) and pd.notna(previous_score) and previous_score != '' and previous_year != 'Rookie' and previous_year != 'R':
                try:
                    change = float(current_score) - float(previous_score)
                    change_cell.value = round(change, 2)
                except:
                    # Fallback to existing change value
                    change = player_row.get('Change_From_2024', player_row.get('Change From 2024', ''))
                    if pd.notna(change) and change != '':
                        try:
                            change_cell.value = round(float(change), 2)
                        except:
                            change_cell.value = change
                    else:
                        change_cell.value = ''
            elif previous_year == 'Rookie' or previous_year == 'R':
                change_cell.value = 'N/A'  # Rookie - no comparison
            else:
                # Fallback to existing change value
                change = player_row.get('Change_From_2024', player_row.get('Change From 2024', ''))
                if pd.notna(change) and change != '':
                    try:
                        change_cell.value = round(float(change), 2)
                    except:
                        change_cell.value = change
                else:
                    change_cell.value = ''
            change_cell.border = border
            
            # Changed Position (column 12)
            changed_position = player_row.get('Changed Position', '')
            changed_pos_cell = ws.cell(row=row_idx, column=12)
            changed_pos_cell.value = changed_position if pd.notna(changed_position) else ''
            changed_pos_cell.alignment = Alignment(horizontal='center', vertical='center')
            changed_pos_cell.border = border
            
            # Minutes
            minutes = player_row.get('Minutes played', player_row.get('Total Minutes', player_row.get('Minutes', '')))
            minutes_cell = ws.cell(row=row_idx, column=13)
            if pd.notna(minutes) and minutes != '':
                try:
                    minutes_cell.value = int(float(minutes))
                except:
                    minutes_cell.value = minutes
            minutes_cell.border = border
            
            # % of Team Minutes
            pct_minutes = player_row.get('% of Team Minutes', player_row.get('Pct_Of_Team_Minutes', ''))
            pct_cell = ws.cell(row=row_idx, column=14)
            if pd.notna(pct_minutes) and pct_minutes != '':
                try:
                    pct_cell.value = round(float(pct_minutes), 1)
                except:
                    pct_cell.value = pct_minutes
            pct_cell.border = border
            
            # Seasons Played
            seasons = player_row.get('Seasons Played', 1)
            seasons_cell = ws.cell(row=row_idx, column=15)
            if pd.notna(seasons):
                try:
                    seasons_cell.value = int(float(seasons))
                except:
                    seasons_cell.value = seasons
            else:
                seasons_cell.value = 1  # Default to 1 if not found
            seasons_cell.border = border
            
            # Top 15s (Power Five) column (column 16)
            # Calculate how many metrics from the shortlist (metric_headers) the player ranks top 15 in
            top15_count = 0
            try:
                # Load from raw data files (same as conference reports) to get all metrics including percentages
                from update_mike_norris_reports import load_conference_season_data
                
                # Load all players for ranking (only once, cache it)
                cache_key = f'_all_players_raw_cache_{position_profile}'
                if not hasattr(create_top_15_report, cache_key):
                    print(f"    📊 Loading all Power Five players from raw data for Top 15s calculation...")
                    raw_data_frames = []
                    power_five_conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
                    
                    for conference in power_five_conferences:
                        df = load_conference_season_data(base_dir, conference, 2025)
                        if not df.empty:
                            df['Conference'] = conference
                            df['Year'] = 2025
                            raw_data_frames.append(df)
                    
                    if len(raw_data_frames) > 0:
                        all_players_raw = pd.concat(raw_data_frames, ignore_index=True)
                        # Filter to same position profile
                        pos_players_raw = all_players_raw[all_players_raw['Position_Profile'] == position_profile].copy()
                        setattr(create_top_15_report, cache_key, pos_players_raw)
                        print(f"    ✅ Loaded {len(pos_players_raw)} Power Five players from raw data for ranking")
                    else:
                        setattr(create_top_15_report, cache_key, pd.DataFrame())
                
                power_five_players = getattr(create_top_15_report, cache_key)
                
                # Get actual column names by matching metric headers to raw data columns
                # This ensures we check all configured metrics, not just those in the shortlist dataframe
                metric_cols_to_check = []
                for metric_header in metric_headers:
                    # Skip combined metrics - handled separately
                    if metric_header == "Interceptions + Sliding Tackles":
                        continue
                    
                    # Try to find matching column in raw data (power_five_players)
                    matching_col = None
                    metric_lower = str(metric_header).lower()
                    
                    # Try exact match first (case-insensitive)
                    for col in power_five_players.columns:
                        if str(col).lower().strip() == metric_lower.strip():
                            matching_col = col
                            break
                    
                    # If no exact match, try base name matching (same logic as conference report)
                    if not matching_col:
                        # For per 90 metrics
                        if 'per 90' in metric_lower and '%' not in metric_lower:
                            metric_base = metric_lower.replace(' per 90', '').replace(' per90', '').strip()
                            for col in power_five_players.columns:
                                col_lower = str(col).lower()
                                if 'per 90' in col_lower and '%' not in col_lower and '% better than position' not in col_lower:
                                    col_base = col_lower.replace(' per 90', '').replace(' per90', '').strip()
                                    metric_base_norm = ' '.join(metric_base.split())
                                    col_base_norm = ' '.join(col_base.split())
                                    if metric_base_norm == col_base_norm:
                                        matching_col = col
                                        break
                        
                        # For percentage metrics
                        elif ('%' in metric_lower or 'percent' in metric_lower) and '% better than position' not in metric_lower:
                            metric_base = metric_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                            for col in power_five_players.columns:
                                col_lower = str(col).lower()
                                if ('%' in col_lower or 'percent' in col_lower) and 'per 90' not in col_lower and '% better than position' not in col_lower:
                                    col_base = col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                                    metric_base_norm = ' '.join(metric_base.split())
                                    col_base_norm = ' '.join(col_base.split())
                                    if metric_base_norm == col_base_norm:
                                        matching_col = col
                                        break
                    
                    if matching_col:
                        metric_cols_to_check.append((metric_header, matching_col))
                
                if len(power_five_players) > 0 and len(metric_cols_to_check) > 0:
                    
                    # Handle combined metric "Interceptions + Sliding Tackles" separately
                    combined_metric_handled = False
                    for metric_header in metric_headers:
                        if metric_header == "Interceptions + Sliding Tackles":
                            combined_cols = metric_column_map.get(metric_header)
                            if isinstance(combined_cols, dict):
                                interceptions_col = combined_cols.get('interceptions')
                                if interceptions_col and interceptions_col in player_row.index:
                                    player_val = player_row.get(interceptions_col)
                                    if pd.notna(player_val) and player_val != '':
                                        try:
                                            player_val = float(player_val)
                                            # Find matching column in power_five_players
                                            matching_col = None
                                            for col in power_five_players.columns:
                                                if 'interception' in str(col).lower() and 'per 90' in str(col).lower():
                                                    matching_col = col
                                                    break
                                            
                                            if matching_col is not None and matching_col in power_five_players.columns:
                                                pf_values = pd.to_numeric(power_five_players[matching_col], errors='coerce').dropna()
                                                if len(pf_values) > 0:
                                                    # Count how many players have >= this value
                                                    pf_rank = (pf_values >= player_val).sum()
                                                    # Get the 15th highest value (for handling ties)
                                                    sorted_values = pf_values.sort_values(ascending=False)
                                                    if len(sorted_values) >= 15:
                                                        value_at_15th = sorted_values.iloc[14]  # 0-indexed, so 14 = 15th
                                                        # Top 15 if: rank <= 15 OR value >= 15th highest value (handles ties)
                                                        if pf_rank <= 15 or player_val >= value_at_15th:
                                                            top15_count += 1
                                                            combined_metric_handled = True
                                                    elif pf_rank <= 15:
                                                        # If less than 15 players total, just check rank
                                                        top15_count += 1
                                                        combined_metric_handled = True
                                        except (ValueError, TypeError):
                                            pass
                            break
                    
                    # Check ALL metrics that appear in the shortlist (using actual column names from raw data)
                    # These are the metrics displayed in the shortlist for this position
                    for metric_header, raw_data_col in metric_cols_to_check:
                        # Skip if this is part of the combined metric we already handled
                        if combined_metric_handled and ('interception' in str(metric_header).lower() or 'sliding' in str(metric_header).lower()):
                            continue
                        
                        # Get player's value from shortlist (try metric_column_map first, then raw_data_col)
                        player_value = None
                        shortlist_col = metric_column_map.get(metric_header)
                        if shortlist_col and shortlist_col in player_row.index:
                            player_value = player_row.get(shortlist_col)
                        
                        # If not found in shortlist, try to get from raw data by matching player
                        if pd.isna(player_value) or player_value == '':
                            # Try to find player in raw data
                            player_name = str(player_row.get('Player', '')).strip()
                            player_team = str(player_row.get('Team', '')).strip()
                            if player_name and player_team:
                                player_raw = power_five_players[
                                    (power_five_players['Player'].str.contains(player_name, case=False, na=False)) &
                                    (power_five_players['Team'].str.contains(player_team, case=False, na=False))
                                ]
                                if len(player_raw) > 0 and raw_data_col in player_raw.columns:
                                    player_value = player_raw[raw_data_col].iloc[0]
                        
                        if pd.isna(player_value) or player_value == '':
                            continue
                        
                        try:
                            player_val = float(player_value)
                            
                            # Use the raw_data_col we already found
                            matching_col = raw_data_col
                            
                            if matching_col is not None and matching_col in power_five_players.columns:
                                # Calculate rank (how many players have >= this value)
                                pf_values = pd.to_numeric(power_five_players[matching_col], errors='coerce').dropna()
                                if len(pf_values) > 0:
                                    # Count how many players have >= this value
                                    pf_rank = (pf_values >= player_val).sum()
                                    # Get the 15th highest value (for handling ties)
                                    sorted_values = pf_values.sort_values(ascending=False)
                                    if len(sorted_values) >= 15:
                                        value_at_15th = sorted_values.iloc[14]  # 0-indexed, so 14 = 15th
                                        # Top 15 if: rank <= 15 OR value >= 15th highest value (handles ties)
                                        if pf_rank <= 15 or player_val >= value_at_15th:
                                            top15_count += 1
                                    elif pf_rank <= 15:
                                        # If less than 15 players total, just check rank
                                        top15_count += 1
                        except (ValueError, TypeError):
                            continue
            except Exception as e:
                # If calculation fails, just set to 0
                import traceback
                print(f"    ⚠️  Error calculating Top 15s for {player_row.get('Player', 'Unknown')}: {e}")
                # Uncomment for debugging:
                # traceback.print_exc()
                pass
            
            top15_cell = ws.cell(row=row_idx, column=16, value=top15_count)
            top15_cell.border = border
            top15_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Consistency Score columns (17-21)
            consistency_score = player_row.get('Consistency Score', 0)
            metrics_above = player_row.get('Metrics Above Avg', 0)
            metrics_below = player_row.get('Metrics Below Avg', 0)
            metrics_at = player_row.get('Metrics At Avg', 0)
            consistency_pct = player_row.get('Consistency %', 0)
            total_metrics = player_row.get('Total Metrics Checked', 0)
            
            ws.cell(row=row_idx, column=17, value=round(float(consistency_score), 1) if pd.notna(consistency_score) else 0)
            
            # Format as "X/Y" for metrics above, below, and at average
            metrics_above_val = int(metrics_above) if pd.notna(metrics_above) else 0
            metrics_below_val = int(metrics_below) if pd.notna(metrics_below) else 0
            metrics_at_val = int(metrics_at) if pd.notna(metrics_at) else 0
            total_metrics_val = int(total_metrics) if pd.notna(total_metrics) else 0
            
            if total_metrics_val > 0:
                ws.cell(row=row_idx, column=18, value=f"{metrics_above_val}/{total_metrics_val}")
                ws.cell(row=row_idx, column=19, value=f"{metrics_below_val}/{total_metrics_val}")
                ws.cell(row=row_idx, column=20, value=f"{metrics_at_val}/{total_metrics_val}")
            else:
                ws.cell(row=row_idx, column=18, value=f"{metrics_above_val}/0")
                ws.cell(row=row_idx, column=19, value=f"{metrics_below_val}/0")
                ws.cell(row=row_idx, column=20, value=f"{metrics_at_val}/0")
            
            ws.cell(row=row_idx, column=21, value=round(float(consistency_pct), 1) if pd.notna(consistency_pct) else 0)
            
            for col_idx in range(17, 22):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style Fits column (22)
            style_fits_count = player_row.get('Style Fits', 0)
            ws.cell(row=row_idx, column=22, value=int(style_fits_count) if pd.notna(style_fits_count) else 0)
            style_fits_cell = ws.cell(row=row_idx, column=22)
            style_fits_cell.border = border
            style_fits_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Relevant metrics - use the mapped column names (starting from column 23, after consistency and style fit columns)
            for col_idx, metric_header in enumerate(metric_headers, 23):
                metric_cell = ws.cell(row=row_idx, column=col_idx)
                
                # Handle combined metric
                if metric_header == "Interceptions + Sliding Tackles":
                    combined_value = None
                    combined_cols = metric_column_map.get(metric_header)
                    if isinstance(combined_cols, dict):
                        interceptions_val = None
                        sliding_tackles_val = None
                        
                        if combined_cols.get('interceptions') and combined_cols['interceptions'] in player_row.index:
                            interceptions_val = player_row.get(combined_cols['interceptions'])
                        if combined_cols.get('sliding_tackles') and combined_cols['sliding_tackles'] in player_row.index:
                            sliding_tackles_val = player_row.get(combined_cols['sliding_tackles'])
                        
                        # Sum the values if both exist
                        if pd.notna(interceptions_val) and pd.notna(sliding_tackles_val):
                            try:
                                combined_value = float(interceptions_val) + float(sliding_tackles_val)
                            except:
                                pass
                        elif pd.notna(interceptions_val):
                            combined_value = interceptions_val
                        elif pd.notna(sliding_tackles_val):
                            combined_value = sliding_tackles_val
                    
                    if combined_value is not None:
                        try:
                            metric_cell.value = round(float(combined_value), 2) if abs(float(combined_value)) < 1000 else round(float(combined_value), 0)
                        except:
                            metric_cell.value = combined_value
                    else:
                        metric_cell.value = 0  # Fill with 0 if combined metric can't be calculated
                elif metric_header == "xG per 90 - Goals per 90":
                    # Calculate difference between xG per 90 and Goals per 90
                    xg_val = None
                    goals_val = None
                    
                    # Find xG per 90 column
                    for col in player_row.index:
                        col_lower = str(col).lower()
                        if 'xg per 90' in col_lower and 'goals' not in col_lower:
                            xg_val = player_row.get(col)
                            break
                    
                    # If not found, try alternative names
                    if pd.isna(xg_val) or xg_val == '':
                        for col in player_row.index:
                            col_lower = str(col).lower()
                            if ('xg' in col_lower or 'expected goals' in col_lower) and 'per 90' in col_lower and 'goals' not in col_lower:
                                xg_val = player_row.get(col)
                                break
                    
                    # Find Goals per 90 column
                    for col in player_row.index:
                        col_lower = str(col).lower()
                        if 'goals per 90' in col_lower and 'xg' not in col_lower:
                            goals_val = player_row.get(col)
                            break
                    
                    # Calculate difference if both exist
                    if pd.notna(xg_val) and pd.notna(goals_val) and xg_val != '' and goals_val != '':
                        try:
                            diff_value = round(float(xg_val) - float(goals_val), 2)
                            metric_cell.value = diff_value
                        except:
                            metric_cell.value = 0  # Fill with 0 if calculation fails
                    else:
                        metric_cell.value = 0  # Fill with 0 if either value is missing
                else:
                    actual_col_name = metric_column_map.get(metric_header)
                    metric_value = None
                    
                    # Try the mapped column first
                    if actual_col_name and actual_col_name in player_row.index:
                        metric_value = player_row.get(actual_col_name, '')
                    
                    # If not found or empty, try alternative column names (handle _report suffix and variations)
                    if pd.isna(metric_value) or metric_value == '':
                        metric_lower = str(metric_header).lower()
                        
                        # Strategy 1: Try with _report suffix
                        if actual_col_name:
                            report_col = actual_col_name + '_report'
                            if report_col in player_row.index:
                                test_val = player_row.get(report_col, '')
                                if pd.notna(test_val) and test_val != '':
                                    metric_value = test_val
                        
                        # Strategy 2: Try base name without "per 90" or "%" (e.g., "Goals" for "Goals per 90")
                        if pd.isna(metric_value) or metric_value == '':
                            base_name = metric_lower.replace(' per 90', '').replace(' per90', '').replace(', %', '').replace(' %', '').replace(' won', '').replace('accurate', '').strip()
                            for col in player_row.index:
                                col_lower = str(col).lower()
                                col_base = col_lower.replace(' per 90', '').replace(' per90', '').replace(', %', '').replace(' %', '').replace(' won', '').replace('accurate', '').replace('_report', '').strip()
                                if base_name == col_base and '_vs_' not in col_lower:
                                    test_val = player_row.get(col, '')
                                    if pd.notna(test_val) and test_val != '':
                                        metric_value = test_val
                                        break
                        
                        # Strategy 3: Try searching for similar column names (fuzzy match)
                        if pd.isna(metric_value) or metric_value == '':
                            for col in player_row.index:
                                col_lower = str(col).lower()
                                # Match if the metric name is contained in column name or vice versa
                                if (metric_lower in col_lower or col_lower in metric_lower) and '_vs_' not in col_lower:
                                    test_val = player_row.get(col, '')
                                    if pd.notna(test_val) and test_val != '':
                                        metric_value = test_val
                                        break
                    
                    # Fill missing values with 0
                    if pd.notna(metric_value) and metric_value != '':
                        try:
                            # Try to format as number if possible
                            if isinstance(metric_value, (int, float)):
                                metric_cell.value = round(float(metric_value), 2) if abs(float(metric_value)) < 1000 else round(float(metric_value), 0)
                            else:
                                metric_cell.value = metric_value
                        except:
                            metric_cell.value = metric_value
                    else:
                        metric_cell.value = 0  # Fill with 0 if no data
                
                metric_cell.border = border
        
        # Auto-adjust column widths
        column_widths = {
            'A': 8,   # Rank
            'B': 20,  # Player
            'C': 25,  # Team
            'D': 12,  # Conference
            'E': 15,  # Position
            'F': 12,  # Total Score
            'G': 18,  # Conference Grade
            'H': 18,  # Power Five Grade
            'I': 15,  # Previous Year
            'J': 12,  # Previous Score
            'K': 15,  # Change
            'L': 18,  # Changed Position (wider to accommodate initials)
            'M': 15,  # Minutes
            'N': 18,  # % of Team Minutes
            'O': 15,  # Seasons Played
            'P': 10   # Top 15s
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set metric column widths (starting from column 17, after Top 15s)
        for col_idx in range(17, len(all_headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 30
        
        print(f"  ✅ Created sheet for {position_profile}: {len(filtered_players)} players with {len(metric_headers)} metrics")
    
    # Create summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    
    # Title
    ws_summary.merge_cells('A1:E1')
    title_cell = ws_summary.cell(row=1, column=1, value="Portland Thorns 2025 Shortlist - Summary")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Note
    ws_summary.merge_cells('A2:E2')
    note_cell = ws_summary.cell(row=2, column=1, value="Note: Scores and grades are calculated against players from Past Seasons (2021-2025 Power Five conferences)")
    note_cell.font = note_font
    note_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Summary data
    summary_data = [
        ['Position Profile', 'Total Players Analyzed', 'Grade A Players', 'Grade B Players', 'Total (A+B, 70%+ min, B+ both grades)'],
        ['', '', '', '', ''],
    ]
    
    for position_profile in position_profiles:
        if position_profile in all_data:
            df = all_data[position_profile].copy()
            
            # Apply same filters as main sheets
            minutes_pct_col = None
            for col in ['% of Team Minutes', 'Pct_Of_Team_Minutes', '% of team minutes']:
                if col in df.columns:
                    minutes_pct_col = col
                    break
            
            if minutes_pct_col:
                df[minutes_pct_col] = pd.to_numeric(df[minutes_pct_col], errors='coerce')
                df = df[df[minutes_pct_col] >= 70.0].copy()
            
            # Note: Regression filter removed - keeping all players regardless of score change
            
            total = len(df)
            
            # Check both grade columns
            conf_grade_col = None
            power_grade_col = None
            
            for col in ['Conference_Grade', 'Conference Grade']:
                if col in df.columns:
                    conf_grade_col = col
                    break
            
            for col in ['Power_Five_Grade', 'Power Five Grade']:
                if col in df.columns:
                    power_grade_col = col
                    break
            
            if conf_grade_col and power_grade_col:
                df['Conf_Grade'] = df[conf_grade_col].astype(str).str.strip().str.upper()
                df['Power_Grade'] = df[power_grade_col].astype(str).str.strip().str.upper()
                # Filter to B+ in both
                df_filtered = df[(df['Conf_Grade'].isin(['A', 'B'])) & (df['Power_Grade'].isin(['A', 'B']))].copy()
                grade_a = len(df_filtered[df_filtered['Conf_Grade'] == 'A'])
                grade_b = len(df_filtered[df_filtered['Conf_Grade'] == 'B'])
                total_ab = len(df_filtered)
            else:
                grade_a = 'N/A'
                grade_b = 'N/A'
                total_ab = 'N/A'
            
            summary_data.append([position_profile, total, grade_a, grade_b, total_ab])
    
    # Write summary
    for row_idx, row_data in enumerate(summary_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    # Auto-adjust summary column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 18
    ws_summary.column_dimensions['E'].width = 25
    
    # Save file
    output_file = base_dir / "Portland Thorns 2025 Shortlist.xlsx"
    wb.save(output_file)
    print(f"\n✅ Report saved: {output_file.name}")
    
    return output_file


def main():
    """Main function."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    create_top_15_report(base_dir)


if __name__ == "__main__":
    main()
