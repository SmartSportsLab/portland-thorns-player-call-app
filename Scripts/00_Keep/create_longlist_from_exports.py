"""
Create a combined longlist Excel file from Wyscout position exports.
Combines all position/conference files into one unified longlist matching the template format.
"""

import pandas as pd
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def load_metrics_config():
    """Load metrics from JSON config file."""
    config_path = Path(__file__).parent / "radar_chart_metrics_short.json"
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
        return config
    except Exception as e:
        print(f"Error loading JSON config: {e}")
        return {}

def get_all_metrics(metrics_config):
    """Get all unique metrics across all positions."""
    all_metrics = set()
    for position, metrics in metrics_config.items():
        all_metrics.update(metrics)
    return sorted(list(all_metrics))

def map_position_profile(filename):
    """Map filename to position profile name used in template."""
    filename_lower = filename.lower()
    
    if 'cb hybrid' in filename_lower or 'hybrid cb' in filename_lower:
        return 'Hybrid CB'
    elif 'dm box-to-box' in filename_lower or 'box to box' in filename_lower:
        return 'DM Box-To-Box'
    elif 'am advanced playmaker' in filename_lower or 'advanced playmaker' in filename_lower:
        return 'AM Advanced Playmaker'
    elif 'w touchline winger' in filename_lower or 'touchline winger' in filename_lower:
        return 'Right Touchline Winger'
    else:
        # Try to extract from filename
        return filename.split()[0] if filename.split() else 'Unknown'

def extract_conference(filename):
    """Extract conference name from filename."""
    conferences = ['ACC', 'BIG10', 'BIG12', 'IVY', 'SEC']
    for conf in conferences:
        if conf in filename.upper():
            return conf
    return 'Unknown'

def read_wyscout_export(file_path):
    """Read a Wyscout export Excel file and return dataframe."""
    try:
        # Wyscout exports have header in row 0, data starts in row 1
        # Read first to check structure
        df_check = pd.read_excel(file_path, header=0, nrows=2)
        
        # Check if first row looks like headers (contains 'Player' or similar)
        first_row = df_check.iloc[0] if len(df_check) > 0 else None
        
        # Try header=0 first (standard Wyscout format)
        df = pd.read_excel(file_path, header=0)
        
        # Clean up the dataframe
        df = df.dropna(how='all')  # Remove completely empty rows
        df = df.dropna(how='all', axis=1)  # Remove completely empty columns
        
        # Remove rows where Player/Name is empty
        player_col = None
        for col in ['Player', 'Name', 'Player Name']:
            if col in df.columns:
                player_col = col
                break
        
        if player_col:
            df = df[df[player_col].notna()]
            # Also remove rows where player name is just a number or looks like a header
            df = df[~df[player_col].astype(str).str.match(r'^\d+$')]
        
        return df
    except Exception as e:
        print(f"   ⚠️  Error reading {file_path.name}: {e}")
        return None

def standardize_column_names(df):
    """Standardize column names to match template format."""
    # Common column name mappings
    column_mappings = {
        'Name': 'Player',
        'Player Name': 'Player',
        'Team': 'Team',
        'Club': 'Team',
        'Conference': 'Conference',
        'League': 'Conference',
    }
    
    # Rename columns
    df = df.rename(columns=column_mappings)
    
    return df

def combine_all_exports(exports_dir, output_path):
    """Combine all Wyscout export files into one longlist."""
    
    exports_path = Path(exports_dir)
    if not exports_path.exists():
        print(f"❌ Exports directory not found: {exports_dir}")
        return
    
    # Load metrics config
    metrics_config = load_metrics_config()
    all_metrics = get_all_metrics(metrics_config)
    
    # Find all Excel files
    excel_files = list(exports_path.glob("*.xlsx"))
    excel_files = [f for f in excel_files if not f.name.startswith('.')]
    
    if not excel_files:
        print(f"❌ No Excel files found in {exports_dir}")
        return
    
    print(f"📊 Found {len(excel_files)} Excel files to process")
    print()
    
    all_dataframes = []
    file_stats = []
    
    # Process each file
    for file_path in sorted(excel_files):
        print(f"📄 Processing: {file_path.name}")
        
        # Extract position and conference from filename
        position_profile = map_position_profile(file_path.name)
        conference = extract_conference(file_path.name)
        
        # Read the file
        df = read_wyscout_export(file_path)
        
        if df is None or df.empty:
            print(f"   ⚠️  Skipping (empty or unreadable)")
            continue
        
        # Standardize column names
        df = standardize_column_names(df)
        
        # Add position profile and conference if not present
        if 'Position Profile' not in df.columns:
            df['Position Profile'] = position_profile
        if 'Conference' not in df.columns:
            df['Conference'] = conference
        
        # Ensure Player column exists (try common variations)
        if 'Player' not in df.columns:
            for col in ['Name', 'Player Name', 'Full Name']:
                if col in df.columns:
                    df['Player'] = df[col]
                    break
        
        # Record stats
        file_stats.append({
            'file': file_path.name,
            'position': position_profile,
            'conference': conference,
            'players': len(df),
            'columns': len(df.columns)
        })
        
        all_dataframes.append(df)
        print(f"   ✅ Loaded {len(df)} players")
    
    if not all_dataframes:
        print("❌ No data to combine!")
        return
    
    print()
    print("=" * 80)
    print("📊 COMBINING DATA")
    print("=" * 80)
    
    # Combine all dataframes
    combined_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
    
    print(f"✅ Combined {len(combined_df)} total players from {len(all_dataframes)} files")
    print()
    
    # Get all columns
    all_columns = combined_df.columns.tolist()
    
    # Define base columns (prioritize these)
    base_columns = ['Player', 'Team', 'Conference', 'Position Profile', 'Grade', 'Position Rank', 'Minutes played', 'Minutes']
    
    # Reorder columns: base columns first, then metrics, then others
    ordered_columns = []
    used_columns = set()
    
    # Add base columns that exist
    for col in base_columns:
        if col in all_columns:
            ordered_columns.append(col)
            used_columns.add(col)
    
    # Add all metrics from JSON (if they exist in data)
    for metric in all_metrics:
        if metric in all_columns and metric not in used_columns:
            ordered_columns.append(metric)
            used_columns.add(metric)
    
    # Add any remaining columns (only those that exist in dataframe)
    remaining_cols = [col for col in all_columns if col not in used_columns]
    # Sort by converting to string for comparison
    remaining_cols_sorted = sorted(remaining_cols, key=lambda x: str(x))
    ordered_columns.extend(remaining_cols_sorted)
    
    # Only use columns that actually exist in the dataframe
    final_columns = [col for col in ordered_columns if col in combined_df.columns]
    
    # Reorder the dataframe
    combined_df = combined_df[final_columns]
    
    # Remove duplicate players (same name, team, position)
    initial_count = len(combined_df)
    # Only use columns that exist for duplicate checking
    dup_cols = []
    for col in ['Player', 'Team', 'Position Profile']:
        if col in combined_df.columns:
            dup_cols.append(col)
    
    if dup_cols:
        combined_df = combined_df.drop_duplicates(subset=dup_cols, keep='first')
        duplicates_removed = initial_count - len(combined_df)
    else:
        duplicates_removed = 0
    
    if duplicates_removed > 0:
        print(f"⚠️  Removed {duplicates_removed} duplicate player entries")
    
    print()
    print("=" * 80)
    print("📋 FILE STATISTICS")
    print("=" * 80)
    for stat in file_stats:
        print(f"  {stat['file']:50} | {stat['position']:25} | {stat['conference']:8} | {stat['players']:4} players")
    
    print()
    print("=" * 80)
    print("📊 FINAL DATASET")
    print("=" * 80)
    print(f"Total players: {len(combined_df)}")
    print(f"Total columns: {len(combined_df.columns)}")
    print()
    print("Position breakdown:")
    if 'Position Profile' in combined_df.columns:
        position_counts = combined_df['Position Profile'].value_counts()
        for pos, count in position_counts.items():
            print(f"  {pos}: {count}")
    print()
    print("Conference breakdown:")
    if 'Conference' in combined_df.columns:
        conf_counts = combined_df['Conference'].value_counts()
        for conf, count in conf_counts.items():
            print(f"  {conf}: {count}")
    
    # Check which metrics are present
    print()
    print("=" * 80)
    print("📊 METRIC AVAILABILITY")
    print("=" * 80)
    metrics_found = []
    metrics_missing = []
    
    for metric in all_metrics:
        if metric in combined_df.columns:
            metrics_found.append(metric)
            # Check if metric has any data
            non_null_count = combined_df[metric].notna().sum()
            print(f"  ✅ {metric:40} ({non_null_count} players have data)")
        else:
            metrics_missing.append(metric)
            print(f"  ❌ {metric:40} (missing)")
    
    print()
    if metrics_missing:
        print(f"⚠️  {len(metrics_missing)} metrics from JSON are missing in the data")
        print("   These columns will need to be added manually or calculated")
    else:
        print("✅ All metrics from JSON are present in the data!")
    
    # Save to Excel
    print()
    print("=" * 80)
    print("💾 SAVING TO EXCEL")
    print("=" * 80)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write main data
        combined_df.to_excel(writer, sheet_name='Player Data', index=False)
        
        # Get the worksheet to format
        worksheet = writer.sheets['Player Data']
        
        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Saved to: {output_path}")
    print()
    print("🎉 Longlist creation complete!")
    print()
    print("Next steps:")
    print("  1. Review the file to ensure data looks correct")
    print("  2. Add any missing metrics manually if needed")
    print("  3. Upload to the Performance Metrics page in the app")

def main():
    """Main function."""
    # Set paths
    exports_dir = "/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search/Exports/Players Stats By Position"
    output_path = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search/Portland Thorns 2025 Long Shortlist.xlsx")
    
    print("🚀 Creating Longlist from Wyscout Exports")
    print("=" * 80)
    print(f"📁 Source: {exports_dir}")
    print(f"📁 Output: {output_path}")
    print("=" * 80)
    print()
    
    combine_all_exports(exports_dir, output_path)

if __name__ == "__main__":
    main()

