#!/usr/bin/env python3
"""
Batch verification tool for Top 15s calculation.
Check multiple players at once and compare with shortlist values.
"""

import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Add path for imports
sys.path.insert(0, str(Path(__file__).parent))

from verify_top15s import verify_player_top15s

def batch_verify_from_shortlist(base_dir, position_profile=None, limit=10):
    """
    Verify Top 15s for players in the shortlist.
    
    Args:
        base_dir: Base directory
        position_profile: Specific position to check (None = all)
        limit: Maximum number of players to check per position
    """
    shortlist_file = base_dir / "Portland Thorns 2025 Shortlist.xlsx"
    
    if not shortlist_file.exists():
        print(f"❌ Shortlist file not found: {shortlist_file}")
        return
    
    wb = load_workbook(shortlist_file, data_only=True)
    
    positions_to_check = [position_profile] if position_profile else ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    mismatches = []
    
    for position in positions_to_check:
        if position not in wb.sheetnames:
            continue
        
        print(f"\n{'='*80}")
        print(f"CHECKING {position}")
        print(f"{'='*80}\n")
        
        ws = wb[position]
        
        # Find Top 15s column
        top15_col = None
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=3, column=col_idx).value
            if header and 'Top 15s' in str(header):
                top15_col = col_idx
                break
        
        if not top15_col:
            print(f"⚠️  Top 15s column not found for {position}")
            continue
        
        # Check first N players
        checked = 0
        for row_idx in range(4, min(4 + limit, ws.max_row + 1)):
            player_name = ws.cell(row=row_idx, column=2).value  # Player column
            team_name = ws.cell(row=row_idx, column=3).value    # Team column
            shortlist_top15 = ws.cell(row=row_idx, column=top15_col).value
            
            if not player_name or not team_name:
                continue
            
            print(f"\n{'─'*80}")
            print(f"Player {checked + 1}: {player_name} ({team_name})")
            print(f"Shortlist Top 15s: {shortlist_top15}")
            print(f"{'─'*80}")
            
            # Quick verification (simplified)
            try:
                from update_mike_norris_reports import load_conference_season_data, get_relevant_metrics_for_position
                import json
                
                # Load data
                all_data = []
                power_five_conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
                for conference in power_five_conferences:
                    df = load_conference_season_data(base_dir, conference, 2025)
                    if not df.empty:
                        df['Conference'] = conference
                        all_data.append(df)
                
                if not all_data:
                    print("⚠️  Could not load data")
                    continue
                
                all_players_raw = pd.concat(all_data, ignore_index=True)
                power_five_players = all_players_raw[
                    (all_players_raw['Position_Profile'] == position) &
                    (all_players_raw['Conference'].isin(power_five_conferences))
                ].copy()
                
                # Find player
                player_mask = (
                    power_five_players['Player'].str.contains(str(player_name), case=False, na=False) &
                    power_five_players['Team'].str.contains(str(team_name), case=False, na=False)
                )
                player_data = power_five_players[player_mask]
                
                if len(player_data) == 0:
                    print(f"⚠️  Player not found in raw data")
                    continue
                
                player_row = player_data.iloc[0]
                
                # Get config metrics
                config_file = Path(__file__).parent / "position_metrics_config.json"
                with open(config_file, 'r') as f:
                    config = json.load(f)
                
                POSITION_PROFILE_MAP = {
                    'Hybrid CB': 'Center Back',
                    'DM Box-To-Box': 'Centre Midfielder',
                    'AM Advanced Playmaker': 'Attacking Midfielder',
                    'Right Touchline Winger': 'Winger'
                }
                
                position_name = POSITION_PROFILE_MAP.get(position, position)
                if position_name in config.get('position_profiles', {}):
                    position_config = config['position_profiles'][position_name]
                    relevant_metrics = get_relevant_metrics_for_position(position_config)
                else:
                    print("⚠️  Config not found")
                    continue
                
                # Quick count
                calculated_top15 = 0
                top15_details = []
                
                for metric in sorted(relevant_metrics):
                    if metric in ["PAdj Interceptions", "PAdj Sliding tackles"]:
                        continue
                    
                    # Find column (simplified)
                    matching_col = None
                    for col in player_row.index:
                        if str(col).lower() == str(metric).lower():
                            matching_col = col
                            break
                    
                    if not matching_col:
                        continue
                    
                    player_value = player_row.get(matching_col)
                    if pd.isna(player_value):
                        continue
                    
                    try:
                        player_val = float(player_value)
                    except:
                        continue
                    
                    # Find in power_five_players
                    pf_col = None
                    for col in power_five_players.columns:
                        if str(col).lower() == str(matching_col).lower():
                            pf_col = col
                            break
                    
                    if not pf_col:
                        continue
                    
                    pf_values = pd.to_numeric(power_five_players[pf_col], errors='coerce').dropna()
                    if len(pf_values) == 0:
                        continue
                    
                    pf_rank = (pf_values >= player_val).sum()
                    if pf_rank <= 15:
                        calculated_top15 += 1
                        top15_details.append(f"{metric} (rank {pf_rank})")
                
                print(f"Calculated Top 15s: {calculated_top15}")
                
                if shortlist_top15 != calculated_top15:
                    print(f"⚠️  MISMATCH! Difference: {abs(shortlist_top15 - calculated_top15)}")
                    mismatches.append({
                        'player': player_name,
                        'team': team_name,
                        'position': position,
                        'shortlist': shortlist_top15,
                        'calculated': calculated_top15
                    })
                    if top15_details:
                        print(f"   Top 15 metrics: {', '.join(top15_details[:5])}")
                else:
                    print(f"✅ MATCH!")
                
                checked += 1
                
            except Exception as e:
                print(f"⚠️  Error: {e}")
                import traceback
                traceback.print_exc()
                continue
    
    wb.close()
    
    # Summary
    print(f"\n{'='*80}")
    print("BATCH VERIFICATION SUMMARY")
    print(f"{'='*80}\n")
    
    if mismatches:
        print(f"⚠️  Found {len(mismatches)} mismatches:")
        for m in mismatches:
            print(f"   {m['player']} ({m['team']}) - {m['position']}: Shortlist={m['shortlist']}, Calculated={m['calculated']}")
    else:
        print("✅ All checked players match!")


if __name__ == "__main__":
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    position = sys.argv[1] if len(sys.argv) > 1 else None
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else 10
    
    batch_verify_from_shortlist(base_dir, position, limit)

