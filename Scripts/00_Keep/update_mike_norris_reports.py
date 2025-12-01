#!/usr/bin/env python3
"""
Update Mike Norris reports with latest 2025 intra-conference data.
This script loads new data, calculates scores against historical data, and regenerates reports.
"""

import sys
import re
from pathlib import Path
import pandas as pd
import numpy as np
import json
import subprocess
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# Import scoring functions from process_any_position.py
CURRENT_DIR = Path(__file__).parent
sys.path.insert(0, str(CURRENT_DIR))
from process_any_position import (
    calculate_with_historical_normalization,
    calculate_percentile_against_distribution,
    convert_percentile_to_1_10,
    assign_grade_single,
    get_relevant_columns
)

# Import header fix helper for team stats
UTILITIES_DIR = CURRENT_DIR.parents[3] / "Scripts" / "08_utilities"
if str(UTILITIES_DIR) not in sys.path:
    sys.path.insert(0, str(UTILITIES_DIR))
try:
    from fix_wyscout_headers import fix_team_headers
except Exception:
    fix_team_headers = None

# Position profile mapping
POSITION_PROFILE_MAP = {
    'Center Back': 'Hybrid CB',
    'Centre Midfielder': 'DM Box-To-Box',
    'Attacking Midfielder': 'AM Advanced Playmaker',
    'Winger': 'Right Touchline Winger'
}

PLAYER_FILE_PREFIXES = {
    'CB Hybrid': 'Center Back',
    'DM Box-To-Box': 'Centre Midfielder',
    'AM Advanced Playmaker': 'Attacking Midfielder',
    'W Touchline Winger': 'Winger'
}


def get_relevant_metrics_for_position(position_config):
    """
    Extract all relevant metric column names from position config.
    
    Args:
        position_config: Dictionary containing position metrics config (Core and Specific)
    
    Returns:
        Set of metric column names that are used in scoring for this position
    """
    relevant_metrics = set()
    
    # Process both Core and Specific metrics
    for metric_type in ['Core', 'Specific']:
        if metric_type not in position_config.get('metrics', {}):
            continue
        
        metrics = position_config['metrics'][metric_type]
        
        for metric_key, metric_value in metrics.items():
            # If it's a simple metric (just a weight, no components)
            if isinstance(metric_value, (int, float)):
                # The metric_key itself is the column name
                relevant_metrics.add(metric_key)
            # If it's a composite metric (has components)
            elif isinstance(metric_value, dict) and 'components' in metric_value:
                # Add all component metric names
                components = metric_value['components']
                for component_name in components.keys():
                    relevant_metrics.add(component_name)
    
    return relevant_metrics


def filter_players_with_metrics(df, relevant_metrics):
    """
    Filter dataframe to only include players who have at least one relevant metric populated.
    
    Args:
        df: DataFrame with player data
        relevant_metrics: Set of metric column names that should be checked
    
    Returns:
        Filtered DataFrame with only players who have at least one metric populated
    """
    if not relevant_metrics:
        return df
    
    # Check which columns exist in the dataframe
    available_metrics = [col for col in relevant_metrics if col in df.columns]
    
    if not available_metrics:
        # If none of the relevant metrics exist in the dataframe, return all players
        return df
    
    # For each player, check if they have at least one non-null, non-zero metric value
    def has_metrics(row):
        for metric in available_metrics:
            value = row.get(metric)
            # Check if value exists and is not null/NaN/empty/zero
            if pd.notna(value) and value != '' and value != 0:
                return True
        return False
    
    # Filter rows where at least one metric is populated
    mask = df.apply(has_metrics, axis=1)
    filtered_df = df[mask].copy()
    
    removed_count = len(df) - len(filtered_df)
    if removed_count > 0:
        print(f"    🗑️  Removed {removed_count} players with no metric data")
    
    return filtered_df

# Position filters for Wyscout positions
POSITION_FILTERS = {
    'Center Back': ['CB', 'LCB', 'RCB'],
    'Centre Midfielder': ['DMF', 'CMF', 'LCMF', 'RCMF'],
    'Attacking Midfielder': ['AMF', 'LAMF', 'RAMF'],
    'Winger': ['LWB', 'RWB', 'LWF', 'RWF', 'LW', 'RW']  # Added LW and RW to catch actual wingers
}

# ---------------------------------------------------------------------------
# Helper paths & loaders
# ---------------------------------------------------------------------------
def get_player_stats_dir(base_dir, year):
    """Return directory for player stats for a given year."""
    exports_dir = base_dir / "Exports"
    if year >= 2025:
        return exports_dir / "Players Stats By Position"
    return exports_dir / "Past Seasons"


def get_team_stats_dir(base_dir, conference):
    """Return directory containing team stats for a conference."""
    return base_dir / "Exports" / "Team Stats By Conference" / conference


def clean_team_name_from_file(file_path):
    """Extract clean team name from file stem."""
    name = file_path.stem.replace("Team Stats ", "").strip()
    name = re.sub(r'\s*\(\d+\)$', '', name).strip()
    return name


def clean_team_stats_dataframe(df):
    """Remove header rows and fix column names for team stats exports."""
    if df.empty:
        return df
    # Remove first row if it contains team metadata rather than match data
    if isinstance(df.iloc[0].get('Date', None), str) and not str(df.iloc[0].get('Date')).startswith('20'):
        df = df.iloc[1:].reset_index(drop=True)
    if fix_team_headers:
        df = fix_team_headers(df)
    return df


def load_conference_season_data(base_dir, conference, year):
    """Load player data for a specific conference and season by combining position files."""
    stats_dir = get_player_stats_dir(base_dir, year)
    if not stats_dir.exists():
        print(f"  ⚠️  Player stats directory not found: {stats_dir}")
        return pd.DataFrame()
    
    frames = []
    for file_prefix, position_name in PLAYER_FILE_PREFIXES.items():
        file_name = f"{file_prefix} {conference} {year}.xlsx"
        file_path = stats_dir / file_name
        if not file_path.exists():
            print(f"  ⚠️  Missing file: {file_name}")
            continue
        try:
            df = pd.read_excel(file_path)
            df = df[df['Player'].notna()].copy()
            df['Conference'] = conference
            df['Year'] = year
            display_profile = POSITION_PROFILE_MAP.get(position_name, position_name)
            df['Position_Profile'] = display_profile
            frames.append(df)
        except Exception as exc:
            print(f"  ⚠️  Could not load {file_name}: {exc}")
    
    if not frames:
        return pd.DataFrame()
    
    combined = pd.concat(frames, ignore_index=True)
    return combined

# Grade colors
GRADE_COLORS = {
    'A': '8B0000',  # Dark red
    'B': 'C5504B',  # Red
    'C': 'F2A2A2',  # Light red
    'D': '8FAADC',  # Light blue
    'F': '1F4E79'   # Dark blue
}


def assign_grade_from_percentile(percentile):
    """Assign grade based on percentile (90+=A, 80+=B, 70+=C, 60+=D, <60=F)."""
    if pd.isna(percentile):
        return 'F'
    elif percentile >= 90:
        return 'A'
    elif percentile >= 80:
        return 'B'
    elif percentile >= 70:
        return 'C'
    elif percentile >= 60:
        return 'D'
    else:
        return 'F'


def calculate_team_grade(player_row, df_all_players):
    """Calculate Team Grade by comparing player against others in same team and position profile."""
    player_team = player_row.get('Team', '')
    player_position_profile = player_row.get('Position_Profile', '')
    player_score = player_row.get('Total_Score_1_10', None)
    
    if pd.isna(player_team) or pd.isna(player_position_profile) or pd.isna(player_score):
        return ''
    
    # Filter to same team and position profile
    team_pos_df = df_all_players[
        (df_all_players['Team'] == player_team) & 
        (df_all_players['Position_Profile'] == player_position_profile)
    ].copy()
    
    # Need at least 2 players for comparison
    if len(team_pos_df) < 2:
        return ''
    
    # Calculate percentile
    scores = team_pos_df['Total_Score_1_10'].dropna()
    if len(scores) < 2:
        return ''
    
    # Calculate percentile (higher score = higher percentile)
    percentile = (scores < player_score).sum() / len(scores) * 100
    
    return assign_grade_from_percentile(percentile)


def calculate_conference_grade(player_row, df_all_players, conference):
    """Calculate Conference Grade by comparing player against others in same conference and position profile."""
    player_position_profile = player_row.get('Position_Profile', '')
    player_score = player_row.get('Total_Score_1_10', None)
    
    if pd.isna(player_position_profile) or pd.isna(player_score):
        return ''
    
    # Filter to same conference and position profile
    conf_pos_df = df_all_players[
        (df_all_players['Position_Profile'] == player_position_profile)
    ].copy()
    
    # Need at least 2 players for comparison
    if len(conf_pos_df) < 2:
        return ''
    
    # Calculate percentile
    scores = conf_pos_df['Total_Score_1_10'].dropna()
    if len(scores) < 2:
        return ''
    
    # Calculate percentile (higher score = higher percentile)
    percentile = (scores < player_score).sum() / len(scores) * 100
    
    return assign_grade_from_percentile(percentile)


def load_all_power_five_historical_data(base_dir):
    """Load all historical data from all Power Five conferences (2021-2025)."""
    conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
    all_data = []
    
    for conference in conferences:
        for year in range(2021, 2026):
            df = load_conference_season_data(base_dir, conference, year)
            if df.empty:
                continue
            df['Year'] = year
            df['Conference'] = conference
            all_data.append(df)
            print(f"  ✅ Loaded {conference} {year} data: {len(df)} rows")
    
    if not all_data:
        return pd.DataFrame()
    
    return pd.concat(all_data, ignore_index=True)


def calculate_power_five_grade(player_row, df_all_power_five_scored, position_profile):
    """Calculate Power Five Grade by comparing player against all Power Five players (2021-25) in same position profile."""
    player_score = player_row.get('Total_Score_1_10', None)
    
    if pd.isna(player_score):
        return ''
    
    # Filter to same position profile
    pos_df = df_all_power_five_scored[
        (df_all_power_five_scored['Position_Profile'] == position_profile)
    ].copy()
    
    # Need at least 2 players for comparison
    if len(pos_df) < 2:
        return ''
    
    # Calculate percentile
    scores = pos_df['Total_Score_1_10'].dropna()
    if len(scores) < 2:
        return ''
    
    # Calculate percentile (higher score = higher percentile)
    percentile = (scores < player_score).sum() / len(scores) * 100
    
    return assign_grade_from_percentile(percentile)


def load_team_total_minutes(base_dir, conference):
    """Load team total minutes from team stats files.
    
    The 'Duration' column in team stats files represents the total minutes
    the team has played across all matches. We sum all Duration values to get
    the total team minutes.
    
    Automatically loads all available team stats files in the directory.
    """
    team_stats_dir = get_team_stats_dir(base_dir, conference)
    team_minutes = {}
    
    # Find all team stats files in the directory
    team_stats_files = list(team_stats_dir.glob("Team Stats *.xlsx"))
    
    if not team_stats_files:
        print(f"  ⚠️  No team stats files found in {team_stats_dir}")
        return team_minutes
    
    print(f"  📁 Found {len(team_stats_files)} team stats files")
    
    for file_path in sorted(team_stats_files):
        try:
            team_name = clean_team_name_from_file(file_path)
            df = pd.read_excel(file_path)
            df = clean_team_stats_dataframe(df)
            if 'Duration' not in df.columns:
                continue
            duration_series = pd.to_numeric(df['Duration'], errors='coerce')
            total_minutes = duration_series.dropna().sum()
            team_minutes[team_name] = total_minutes
            print(f"  ✅ {team_name}: {total_minutes:.0f} total minutes (sum of Duration)")
        except Exception as e:
            print(f"  ⚠️  Could not load {file_path.name}: {e}")
    
    return team_minutes


def calculate_player_percentage_of_team_minutes(df_all_scored, team_minutes_dict):
    """Calculate each player's percentage of their team's total minutes."""
    percentages = []
    
    for idx, row in df_all_scored.iterrows():
        player_team = row.get('Team', '')
        player_minutes = row.get('Minutes played', 0)
        
        if pd.isna(player_team) or pd.isna(player_minutes) or player_minutes == 0:
            percentages.append(None)
            continue
        
        # Find matching team in minutes dict (handle variations)
        team_total_minutes = None
        player_team_str = str(player_team).strip()
        
        # Try exact match first
        if player_team_str in team_minutes_dict:
            team_total_minutes = team_minutes_dict[player_team_str]
        else:
            # Try case-insensitive match
            player_team_lower = player_team_str.lower()
            for team_name, total_mins in team_minutes_dict.items():
                team_name_lower = str(team_name).strip().lower()
                
                # Exact case-insensitive match
                if player_team_lower == team_name_lower:
                    team_total_minutes = total_mins
                    break
                
                # Try partial matching - check if one name contains the other
                # This handles cases like "Notre Dame Fighting" vs "Notre Dame Fighting Irish"
                if player_team_lower in team_name_lower or team_name_lower in player_team_lower:
                    # Additional validation: check if first 2-3 words match
                    player_parts = player_team_lower.split()
                    team_parts = team_name_lower.split()
                    
                    if len(player_parts) >= 2 and len(team_parts) >= 2:
                        # If first 2 words match, it's likely the same team
                        if player_parts[0] == team_parts[0] and player_parts[1] == team_parts[1]:
                            team_total_minutes = total_mins
                            break
        
        if team_total_minutes and team_total_minutes > 0:
            percentage = (player_minutes / team_total_minutes) * 100
            # Cap at 100% (can happen due to stoppage time or rounding differences)
            percentage = min(percentage, 100.0)
            percentages.append(percentage)
        else:
            percentages.append(None)
    
    return percentages


def load_historical_data(base_dir, conference):
    """Load all historical data for a conference (2021-2025)."""
    all_data = []
    
    for year in range(2021, 2026):
        df = load_conference_season_data(base_dir, conference, year)
        if df.empty:
            continue
        df['Year'] = year
        all_data.append(df)
        print(f"  ✅ Loaded {year} data: {len(df)} rows")
    
    if not all_data:
        return pd.DataFrame()
    
    return pd.concat(all_data, ignore_index=True)


def filter_by_position(df, position_profile):
    """Filter dataframe by position profile."""
    if position_profile not in POSITION_FILTERS:
        return df
    
    filters = POSITION_FILTERS[position_profile]
    
    # Filter by Position column (can be comma-separated)
    mask = df['Position'].astype(str).apply(
        lambda x: any(filt in str(x) for filt in filters)
    )
    
    return df[mask].copy()


def load_2024_data_for_comparison(base_dir, conference):
    """Load 2024 data for year-over-year comparison."""
    return load_conference_season_data(base_dir, conference, 2024)


def get_championship_teams(conference):
    """Get the list of teams involved in championship games for a conference."""
    championship_teams = {
        # Return an empty list to keep ALL teams for ACC/SEC per Mike's latest request
        'ACC': [],
        'SEC': [],
        # Other conferences can still use specific shortlists if desired
    }
    return championship_teams.get(conference, [])


def filter_to_championship_teams(df, allowed_teams):
    """Filter dataframe to only include players from championship teams."""
    if not allowed_teams:
        return df
    
    all_teams = df['Team'].dropna().unique()
    matching_teams = []
    
    for team in all_teams:
        team_str = str(team).strip()
        # Check exact match or if team contains allowed team name
        for allowed_team in allowed_teams:
            allowed_lower = allowed_team.lower()
            team_lower = team_str.lower()
            
            # Exact match
            if allowed_lower == team_lower:
                matching_teams.append(team)
                break
            # Partial match (but exclude Virginia Tech when looking for Virginia)
            elif allowed_lower in team_lower:
                # Special case: exclude Virginia Tech when looking for Virginia
                if 'virginia' in allowed_lower and 'virginia tech' in team_lower:
                    continue
                # Special case: handle Mississippi State variations
                if 'mississippi' in allowed_lower:
                    if 'mississippi st' in team_lower or 'mississippi state' in team_lower:
                        matching_teams.append(team)
                        break
                else:
                    matching_teams.append(team)
                    break
    
    return df[df['Team'].isin(matching_teams)].copy() if matching_teams else df


def create_report_for_conference(base_dir, conference):
    """Create Mike Norris report for a conference."""
    print(f"\n{'='*80}")
    print(f"PROCESSING {conference}")
    print(f"{'='*80}")
    
    # Load config
    config_file = base_dir / "Scripts" / "00_Keep" / "position_metrics_config.json"
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    # Get championship teams for this conference
    championship_teams = get_championship_teams(conference)

    # Calculate team possessions for PAdj metrics
    print(f"\n📊 Calculating team possessions for PAdj metrics...")
    team_stats_dir = get_team_stats_dir(base_dir, conference)
    team_possessions = calculate_team_possessions(team_stats_dir)
    league_avg_possessions = calculate_league_avg_possessions(team_possessions)
    if team_possessions:
        print(f"  ✅ Calculated possessions for {len(team_possessions)} teams")
        if league_avg_possessions is not None:
            print(f"  ✅ League average possessions: {league_avg_possessions:.2f} per game")
    else:
        print(f"  ⚠️  No team possession data available for {conference}")
    
    # Load and score all Power Five historical data (2021-2025) for Power Five Grade comparison
    print(f"\n🌐 Loading and scoring all Power Five historical data (2021-2025)...")
    df_all_power_five = load_all_power_five_historical_data(base_dir)
    
    all_power_five_scored = {}
    if len(df_all_power_five) > 0:
        print(f"  ✅ Loaded {len(df_all_power_five)} total Power Five rows")
        
        # Score all Power Five data by position profile
        position_profiles = ['Center Back', 'Centre Midfielder', 'Attacking Midfielder', 'Winger']
        pass_attempt_weight = 0.8
        pass_accuracy_weight = 0.2
        
        for position_name in position_profiles:
            if position_name not in config['position_profiles']:
                continue
            
            position_config = config['position_profiles'][position_name]
            display_name = POSITION_PROFILE_MAP[position_name]
            
            # Filter Power Five data by position
            df_power_five_pos = filter_by_position(df_all_power_five, position_name)
            if len(df_power_five_pos) == 0:
                continue

            # Add PAdj metrics and update config to use them when available
            df_power_five_pos = add_padj_metrics_to_dataframe(
                df_power_five_pos, team_possessions, league_avg_possessions, team_col='Team'
            )
            position_config_for_scoring = update_config_to_use_padj(position_config, df_power_five_pos)
            
            # Score all Power Five data for this position (using historical normalization)
            # We'll score each year separately to maintain historical context
            scored_by_year = []
            for year in range(2021, 2026):
                df_year = df_power_five_pos[df_power_five_pos['Year'] == year].copy()
                if len(df_year) == 0:
                    continue
                
                # Use historical data up to that year for normalization
                df_historical = df_power_five_pos[df_power_five_pos['Year'] < year].copy()
                if len(df_historical) == 0:
                    # For 2021, use that year's data as reference
                    df_historical = df_year.copy()
                
                df_scored = calculate_with_historical_normalization(
                    df_year, df_historical, position_config_for_scoring,
                    pass_attempt_weight, pass_accuracy_weight, position_name
                )
                scored_by_year.append(df_scored)
            
            if scored_by_year:
                df_power_five_scored = pd.concat(scored_by_year, ignore_index=True)
                # Add Position_Profile column
                df_power_five_scored['Position_Profile'] = display_name
                all_power_five_scored[display_name] = df_power_five_scored
                print(f"  ✅ Scored {display_name}: {len(all_power_five_scored[display_name])} players")
    else:
        print(f"  ⚠️  No Power Five historical data found")
    
    # Load 2025 data directly from exports
    print(f"📄 Loading {conference} 2025 player data...")
    df_2025 = load_conference_season_data(base_dir, conference, 2025)
    if df_2025.empty:
        print(f"❌ No 2025 data found for {conference}")
        return
    print(f"  ✅ Loaded {len(df_2025)} players")
    
    # Load historical data (2021-2025)
    print(f"\n📚 Loading historical data (2021-2025)...")
    df_all = load_historical_data(base_dir, conference)
    
    if len(df_all) == 0:
        print(f"  ⚠️  No historical data found")
        return
    
    print(f"  ✅ Total historical rows: {len(df_all)}")
    
    # Load 2024 data for comparison
    print(f"\n📊 Loading 2024 data for comparison...")
    df_2024 = load_2024_data_for_comparison(base_dir, conference)
    
    # Load team total minutes from team stats files
    print(f"\n⏱️  Loading team total minutes from team stats files...")
    team_minutes_dict = load_team_total_minutes(base_dir, conference)
    
    # Load team averages for team-relative analysis
    print(f"\n📊 Loading team averages for team-relative analysis...")
    team_averages_dict = load_team_averages(base_dir, conference)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Process each position profile
    position_profiles = ['Center Back', 'Centre Midfielder', 'Attacking Midfielder', 'Winger']
    
    all_scored_data = []
    
    for position_name in position_profiles:
        if position_name not in config['position_profiles']:
            continue
        
        position_config = config['position_profiles'][position_name]
        display_name = POSITION_PROFILE_MAP[position_name]
        
        print(f"\n📋 Processing {display_name}...")
        
        # Filter 2025 data by position
        df_2025_pos = filter_by_position(df_2025, position_name)
        if len(df_2025_pos) == 0:
            print(f"  ⚠️  No players found for {position_name}")
            continue
        
        # Deduplicate players: if a player appears multiple times (from different position files),
        # keep only the row where Position_Profile matches the current position profile
        if 'Position_Profile' in df_2025_pos.columns:
            before_dedup = len(df_2025_pos)
            # Create a mask: prefer rows where Position_Profile matches display_name
            df_2025_pos['_keep_priority'] = (df_2025_pos['Position_Profile'] == display_name).astype(int)
            # Sort by priority (1 first) and Player name, then drop duplicates keeping first
            df_2025_pos = df_2025_pos.sort_values(['_keep_priority', 'Player'], ascending=[False, True])
            df_2025_pos = df_2025_pos.drop_duplicates(subset=['Player'], keep='first')
            df_2025_pos = df_2025_pos.drop(columns=['_keep_priority'])
            after_dedup = len(df_2025_pos)
            if before_dedup > after_dedup:
                print(f"  🔍 Removed {before_dedup - after_dedup} duplicate player entries")
        
        # Filter to only championship teams
        df_2025_pos = filter_to_championship_teams(df_2025_pos, championship_teams)
        if len(df_2025_pos) == 0:
            print(f"  ⚠️  No players from championship teams for {position_name}")
            continue
        print(f"  📋 Filtered to {len(df_2025_pos)} players from championship teams")
        
        # Debug: Show which teams have players before filtering by metrics
        if position_name == 'Winger':
            teams_before_filter = df_2025_pos['Team'].value_counts()
            print(f"  📊 Teams with winger-positioned players (before metric filtering):")
            for team, count in teams_before_filter.items():
                print(f"      - {team}: {count} players")
        
        # Filter historical data by position
        df_all_pos = filter_by_position(df_all, position_name)
        if len(df_all_pos) == 0:
            print(f"  ⚠️  No historical data for {position_name}")
            continue
        
        # Deduplicate historical data: if a player appears multiple times for the same year,
        # keep only the row where Position_Profile matches the current position profile
        if 'Position_Profile' in df_all_pos.columns and 'Year' in df_all_pos.columns:
            before_dedup = len(df_all_pos)
            # Create a mask: prefer rows where Position_Profile matches display_name
            df_all_pos['_keep_priority'] = (df_all_pos['Position_Profile'] == display_name).astype(int)
            # Sort by priority (1 first), Year, and Player name, then drop duplicates keeping first
            df_all_pos = df_all_pos.sort_values(['_keep_priority', 'Year', 'Player'], ascending=[False, False, True])
            df_all_pos = df_all_pos.drop_duplicates(subset=['Player', 'Year'], keep='first')
            df_all_pos = df_all_pos.drop(columns=['_keep_priority'])
            after_dedup = len(df_all_pos)
            if before_dedup > after_dedup:
                print(f"  🔍 Removed {before_dedup - after_dedup} duplicate historical player entries")

        # Add PAdj metrics before scoring
        df_2025_pos = add_padj_metrics_to_dataframe(
            df_2025_pos, team_possessions, league_avg_possessions, team_col='Team'
        )
        df_all_pos = add_padj_metrics_to_dataframe(
            df_all_pos, team_possessions, league_avg_possessions, team_col='Team'
        )

        # Update config to use PAdj metrics when available
        position_config_for_scoring = update_config_to_use_padj(position_config, df_2025_pos)
        
        # Get weights from config
        pass_attempt_weight = 0.8
        pass_accuracy_weight = 0.2
        
        # Calculate scores
        print(f"  🧮 Calculating scores...")
        df_scored = calculate_with_historical_normalization(
            df_2025_pos, df_all_pos, position_config_for_scoring,
            pass_attempt_weight, pass_accuracy_weight, position_name
        )
        
        # Calculate percentiles against conference distribution
        print(f"  📊 Calculating percentiles...")
        reference_distribution = df_scored['Total_Score_1_10']
        df_scored['Total_Percentile'] = df_scored['Total_Score_1_10'].apply(
            lambda x: calculate_percentile_against_distribution(x, reference_distribution)
        )
        
        # Assign grades
        print(f"  🎓 Assigning grades...")
        df_scored['Total_Grade'] = df_scored['Total_Score_1_10'].apply(assign_grade_single)
        
        # Calculate 2024 scores for comparison
        if len(df_2024) > 0:
            df_2024_pos = filter_by_position(df_2024, position_name)
            if len(df_2024_pos) > 0:
                # Deduplicate 2024 data: if a player appears multiple times,
                # keep only the row where Position_Profile matches the current position profile
                if 'Position_Profile' in df_2024_pos.columns:
                    before_dedup = len(df_2024_pos)
                    df_2024_pos['_keep_priority'] = (df_2024_pos['Position_Profile'] == display_name).astype(int)
                    df_2024_pos = df_2024_pos.sort_values(['_keep_priority', 'Player'], ascending=[False, True])
                    df_2024_pos = df_2024_pos.drop_duplicates(subset=['Player'], keep='first')
                    df_2024_pos = df_2024_pos.drop(columns=['_keep_priority'])
                    after_dedup = len(df_2024_pos)
                    if before_dedup > after_dedup:
                        print(f"  🔍 Removed {before_dedup - after_dedup} duplicate 2024 player entries")
                df_2024_pos = add_padj_metrics_to_dataframe(
                    df_2024_pos, team_possessions, league_avg_possessions, team_col='Team'
                )
                df_2024_scored = calculate_with_historical_normalization(
                    df_2024_pos, df_all_pos[df_all_pos['Year'] < 2025], position_config_for_scoring,
                    pass_attempt_weight, pass_accuracy_weight, position_name
                )
                # Create lookup dict for 2024 scores
                scores_2024_dict = dict(zip(
                    df_2024_scored['Player'].astype(str),
                    df_2024_scored['Total_Score_1_10']
                ))
                
                # Debug: Check which players aren't matching
                print(f"  🔍 Debugging 2024 score matching for {display_name}...")
                players_2025 = set(df_scored['Player'].astype(str).str.strip())
                players_2024_in_dict = set(scores_2024_dict.keys())
                players_2024_all = set(df_2024_pos['Player'].astype(str).str.strip())
                
                matched = 0
                not_matched = []
                
                for _, player_row in df_scored.iterrows():
                    player_name = str(player_row['Player']).strip()
                    player_team = player_row.get('Team', '')
                    
                    # Check if player exists in 2024 dict
                    if player_name in scores_2024_dict:
                        matched += 1
                    else:
                        # Check if player exists in 2024 data at all (different team or position issue)
                        player_in_2024_all = player_name in players_2024_all
                        if player_in_2024_all:
                            # Player exists in 2024 but not in scored dict (might be scoring issue)
                            player_2024_data = df_2024_pos[df_2024_pos['Player'].astype(str).str.strip() == player_name]
                            if len(player_2024_data) > 0:
                                team_2024 = player_2024_data['Team'].iloc[0] if 'Team' in player_2024_data.columns else 'Unknown'
                                reason = f"exists in 2024 data but not scored (2024 team: {team_2024})"
                            else:
                                reason = "exists in 2024 data but not scored (data issue)"
                        else:
                            reason = "not found in 2024 data (likely freshman/new player)"
                        
                        not_matched.append({
                            'player': player_name,
                            'team': str(player_team),
                            'reason': reason
                        })
                
                print(f"    ✅ Matched: {matched}/{len(df_scored)} players")
                print(f"    ❌ Not matched: {len(not_matched)}/{len(df_scored)} players")
                
                if not_matched:
                    print(f"    📋 Missing players breakdown:")
                    freshmen_count = sum(1 for p in not_matched if 'freshman' in p['reason'] or 'not found' in p['reason'])
                    transfer_count = sum(1 for p in not_matched if 'exists in 2024' in p['reason'])
                    print(f"      - Freshmen/new players: {freshmen_count}")
                    print(f"      - Transfers/scoring issues: {transfer_count}")
                    print(f"    📝 Sample missing players (first 5):")
                    for p in not_matched[:5]:
                        print(f"      • {p['player']} ({p['team']}): {p['reason']}")
                
                # Add 2024 scores as fallback (will be replaced by previous score logic below)
                df_scored['2024_Total_Score'] = df_scored['Player'].astype(str).map(scores_2024_dict)
            else:
                print(f"  ⚠️  No 2024 players found for {position_name}")
                df_scored['2024_Total_Score'] = None
        else:
            print(f"  ⚠️  No 2024 data available")
            df_scored['2024_Total_Score'] = None
        
        # Now add Previous Score and Previous Year using the same logic as shortlist
        # This looks at all years 2021-2024, not just 2024
        from create_top_15_report import load_seasons_data
        seasons_dict, previous_scores_dict = load_seasons_data(base_dir)
        
        def get_previous_info(row):
            player = str(row.get('Player', '')).strip()
            if not player:
                return None, None, None
            
            # PRIORITY 1: Check if player has previous scores from calculated historical data (2021-2024)
            if player in previous_scores_dict:
                # First, check if they have a score for the current position profile
                if display_name in previous_scores_dict[player]:
                    year, score = previous_scores_dict[player][display_name]
                    return year, score, None
                
                # If not, check for most recent score across ALL positions (for position changers)
                if '_most_recent' in previous_scores_dict[player]:
                    year, score, prev_position = previous_scores_dict[player]['_most_recent']
                    return year, score, prev_position
                
                # Fallback: check all position profiles for this player
                # Get the most recent score from any position
                most_recent = None
                most_recent_year = 0
                most_recent_pos = None
                for pos_name, value in previous_scores_dict[player].items():
                    if pos_name != '_most_recent':
                        if isinstance(value, tuple) and len(value) >= 2:
                            pos_year, pos_score = value[0], value[1]
                            if isinstance(pos_year, int):
                                if pos_year > most_recent_year:
                                    most_recent_year = pos_year
                                    most_recent = (pos_year, pos_score)
                                    most_recent_pos = pos_name
                if most_recent:
                    return most_recent[0], most_recent[1], most_recent_pos
            
            # PRIORITY 2: Fallback to 2024 score if available
            if '2024_Total_Score' in row.index:
                score_2024 = row.get('2024_Total_Score')
                if pd.notna(score_2024) and score_2024 != '':
                    try:
                        score_val = float(score_2024)
                        if score_val > 0:
                            return 2024, score_val, None
                    except:
                        pass
            
            # Check if 2025 is their first season
            seasons = seasons_dict.get(player, 1)
            if seasons == 1:
                return 'Rookie', None, None
            
            # If they have multiple seasons but no previous score found
            # This could mean: they didn't play in previous years, transferred, or data is missing
            return None, None, None
        
        previous_data = df_scored.apply(get_previous_info, axis=1, result_type='expand')
        df_scored['Previous Year'] = previous_data[0]
        df_scored['Previous Score'] = previous_data[1]
        df_scored['Previous Position'] = previous_data[2]
        
        # Calculate change from previous
        df_scored['Change From Previous'] = None
        df_scored['Total_Score_1_10'] = pd.to_numeric(df_scored['Total_Score_1_10'], errors='coerce')
        df_scored['Previous Score'] = pd.to_numeric(df_scored['Previous Score'], errors='coerce')
        df_scored['Change From Previous'] = df_scored['Total_Score_1_10'] - df_scored['Previous Score']
        
        # Drop the temporary 2024_Total_Score column (we have Previous Score now)
        if '2024_Total_Score' in df_scored.columns:
            df_scored = df_scored.drop(columns=['2024_Total_Score'])
        
        # Add position profile column
        df_scored['Position_Profile'] = display_name
        
        # Ensure Total_Grade is not None/NaN - fill with empty string if missing
        if 'Total_Grade' in df_scored.columns:
            df_scored['Total_Grade'] = df_scored['Total_Grade'].fillna('')
        else:
            df_scored['Total_Grade'] = ''
        
        # Calculate Team Grade
        print(f"  🏆 Calculating Team Grades...")
        df_scored['Team_Grade'] = df_scored.apply(
            lambda row: calculate_team_grade(row, df_scored), axis=1
        )
        
        # Calculate Conference Grade
        print(f"  🏆 Calculating Conference Grades...")
        df_scored['Conference_Grade'] = df_scored.apply(
            lambda row: calculate_conference_grade(row, df_scored, conference), axis=1
        )
        
        # Calculate Power Five Grade
        print(f"  🏆 Calculating Power Five Grades...")
        if display_name in all_power_five_scored:
            df_power_five_pos = all_power_five_scored[display_name]
            df_scored['Power_Five_Grade'] = df_scored.apply(
                lambda row: calculate_power_five_grade(row, df_power_five_pos, display_name), axis=1
            )
        else:
            df_scored['Power_Five_Grade'] = ''
            print(f"    ⚠️  No Power Five data available for {display_name}")
        
        # Get relevant metrics for this position profile
        relevant_metrics = get_relevant_metrics_for_position(position_config)
        
        # Filter out players who don't have any relevant metrics populated
        print(f"  🔍 Filtering players with no metric data...")
        teams_before = df_scored['Team'].value_counts() if 'Team' in df_scored.columns else pd.Series()
        df_scored = filter_players_with_metrics(df_scored, relevant_metrics)
        teams_after = df_scored['Team'].value_counts() if 'Team' in df_scored.columns and len(df_scored) > 0 else pd.Series()
        
        # Debug output for winger position
        if position_name == 'Winger':
            print(f"  📊 Teams after metric filtering:")
            if len(teams_after) > 0:
                for team, count in teams_after.items():
                    before_count = teams_before.get(team, 0)
                    if before_count > count:
                        print(f"      - {team}: {count} players (removed {before_count - count} with no metrics)")
                    else:
                        print(f"      - {team}: {count} players")
            else:
                print(f"      ⚠️  No teams remaining after filtering")
        
        if len(df_scored) == 0:
            print(f"  ⚠️  No players with metric data for {display_name}, skipping sheet")
            continue
        
        # Calculate percentage of team minutes for this position profile's players
        percentages_pos = calculate_player_percentage_of_team_minutes(df_scored, team_minutes_dict)
        df_scored['Pct_Of_Team_Minutes'] = percentages_pos
        
        # Calculate Top 15s BEFORE appending to all_scored_data (so team tabs have the values)
        # Get all_power_five_for_position for Top 15s calculation
        all_power_five_for_position = all_power_five_scored.get(display_name, pd.DataFrame())
        
        # Calculate Top 15s using the same logic as create_position_profile_sheet
        print(f"  📊 Calculating Top 15s (Power Five) for {display_name} (for team tabs)...")
        df_scored = calculate_top_15s_for_team_tabs(df_scored, display_name, all_power_five_for_position, position_config)
        
        # Store for team sheets (after Top 15s calculation)
        all_scored_data.append(df_scored)
        
        # Create position profile sheet (pass position_config to filter metrics)
        # Pass all_power_five_scored to calculate position profile averages across all conferences
        create_position_profile_sheet(wb, df_scored, display_name, team_minutes_dict, position_config, team_averages_dict, all_power_five_for_position)
        
        print(f"  ✅ {display_name}: {len(df_scored)} players processed")
    
    # Create team summary sheets
    if all_scored_data:
        df_all_scored = pd.concat(all_scored_data, ignore_index=True)
        
        # Calculate percentage of team minutes for each player
        print(f"\n📊 Calculating player percentage of team minutes...")
        percentages = calculate_player_percentage_of_team_minutes(df_all_scored, team_minutes_dict)
        df_all_scored['Pct_Of_Team_Minutes'] = percentages
        
        create_team_summary_sheets(wb, df_all_scored, conference, team_minutes_dict, config, team_averages_dict)
    
    # Create Data Summary sheet
    create_data_notes_sheet(wb, conference)
    
    # Set print settings for all sheets before saving
    print(f"  📄 Setting print areas and page layout for PDF conversion...")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Set print area to all used cells
        if ws.max_row > 0 and ws.max_column > 0:
            print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            ws.print_area = print_area
        
        # Set page setup for landscape orientation and fit to width
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = 100
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
    
    # Save workbook
    output_file = base_dir / f"Portland Thorns 2025 {conference} Championship Scouting Report.xlsx"
    wb.save(output_file)
    print(f"\n✅ Report saved: {output_file.name}")
    
    # PDF conversion disabled for now - focusing on Excel quality
    # print(f"\n📄 Converting to PDF...")
    # pdf_file = convert_excel_to_pdf(output_file, base_dir)
    # if pdf_file:
    #     print(f"✅ PDF saved: {pdf_file.name}")


def calculate_team_possessions(team_data_dir):
    """
    Calculate average possessions per game for each team.
    
    Args:
        team_data_dir: Directory containing team stat Excel files
        
    Returns:
        Dictionary mapping team names to average possessions per game
    """
    team_possessions = {}
    files = sorted(team_data_dir.glob('Team Stats *.xlsx'))
    
    for file in files:
        try:
            df = pd.read_excel(file)
            df = clean_team_stats_dataframe(df)
            if 'Passes' not in df.columns or 'Average passes per possession' not in df.columns:
                continue
            df['Passes'] = pd.to_numeric(df['Passes'], errors='coerce')
            df['Average passes per possession'] = pd.to_numeric(df['Average passes per possession'], errors='coerce')
            df_clean = df.dropna(subset=['Passes', 'Average passes per possession']).copy()
            if df_clean.empty:
                continue
            df_clean['Possessions'] = df_clean['Passes'] / df_clean['Average passes per possession']
            avg_possessions = df_clean['Possessions'].mean()
            team_name = clean_team_name_from_file(file)
            team_possessions[team_name] = avg_possessions
        except Exception as e:
            print(f"  ⚠️  Error processing {file.name}: {e}")
    
    return team_possessions


def calculate_league_avg_possessions(team_possessions):
    """Calculate league average possessions per game."""
    if not team_possessions:
        return None
    return np.mean(list(team_possessions.values()))


def get_padj_candidate_metrics():
    """
    Return list of metric base names that should be PAdj-adjusted.
    
    These are count-based metrics (not percentages) that depend on possession.
    """
    return [
        # Defensive actions
        'Interceptions',
        'Sliding tackles',
        'Clearances',
        'Shots blocked',
        'Defensive duels',
        'Aerial duels',
        
        # Passing actions
        'Passes',
        'Long passes',
        'Forward passes',
        'Back passes',
        'Lateral passes',
        'Progressive passes',
        'Smart passes',
        'Through passes',
        'Passes to final third',
        'Passes to penalty area',
        'Deep completed passes',
        
        # Attacking actions
        'Shots',
        'Key passes',
        'Touches in penalty area',
        'Offensive duels',
        
        # Other actions
        'Recoveries',
        'Dribbles',
        'Progressive runs',
        'Received passes',
    ]


def update_config_to_use_padj(config, df):
    """
    Update position config to use PAdj versions of metrics where available.
    
    This function modifies the config to prefer PAdj versions when they exist in the dataframe.
    
    Args:
        config: Position metrics configuration dictionary
        df: DataFrame to check for available PAdj metrics
        
    Returns:
        Updated config dictionary
    """
    import copy
    config_padj = copy.deepcopy(config)
    
    # Get all available PAdj columns in dataframe
    padj_columns = set([col for col in df.columns if 'PAdj' in str(col)])
    
    def update_metric_name(metric_name):
        """Try to find PAdj version of a metric"""
        # Check if PAdj version exists
        padj_version = f"PAdj {metric_name}"
        if padj_version in padj_columns:
            return padj_version
        return metric_name
    
    def update_components(components_dict):
        """Update component metric names to use PAdj versions"""
        updated = {}
        for comp_name, weight in components_dict.items():
            # Skip percentages - they don't get PAdj-adjusted
            if '%' in str(comp_name) or 'percent' in str(comp_name).lower():
                updated[comp_name] = weight
            else:
                # Try to use PAdj version
                updated_comp = update_metric_name(comp_name)
                updated[updated_comp] = weight
        return updated
    
    # Update Core metrics
    if 'metrics' in config_padj and 'Core' in config_padj['metrics']:
        # Collect items to update (avoid modifying dict during iteration)
        items_to_update = list(config_padj['metrics']['Core'].items())
        keys_to_remove = []
        keys_to_add = {}
        
        for metric_name, metric_config in items_to_update:
            if isinstance(metric_config, dict) and 'components' in metric_config:
                metric_config['components'] = update_components(metric_config['components'])
            elif isinstance(metric_config, (int, float)):
                # Simple metric - try to update the key name
                updated_key = update_metric_name(metric_name)
                if updated_key != metric_name and updated_key in padj_columns:
                    # Mark for replacement
                    keys_to_remove.append(metric_name)
                    keys_to_add[updated_key] = metric_config
        
        # Apply replacements
        for key in keys_to_remove:
            del config_padj['metrics']['Core'][key]
        config_padj['metrics']['Core'].update(keys_to_add)
    
    # Update Specific metrics
    if 'metrics' in config_padj and 'Specific' in config_padj['metrics']:
        # Collect items to update (avoid modifying dict during iteration)
        items_to_update = list(config_padj['metrics']['Specific'].items())
        keys_to_remove = []
        keys_to_add = {}
        
        for metric_name, metric_config in items_to_update:
            if isinstance(metric_config, dict) and 'components' in metric_config:
                metric_config['components'] = update_components(metric_config['components'])
            elif isinstance(metric_config, (int, float)):
                # Simple metric - try to update the key name
                updated_key = update_metric_name(metric_name)
                if updated_key != metric_name and updated_key in padj_columns:
                    # Mark for replacement
                    keys_to_remove.append(metric_name)
                    keys_to_add[updated_key] = metric_config
        
        # Apply replacements
        for key in keys_to_remove:
            del config_padj['metrics']['Specific'][key]
        config_padj['metrics']['Specific'].update(keys_to_add)
    
    return config_padj


def add_padj_metrics_to_dataframe(df, team_possessions, league_avg_possessions, team_col='Team'):
    """
    Add PAdj versions of metrics to a dataframe.
    
    Args:
        df: DataFrame with player data
        team_possessions: Dictionary mapping team names to possessions per game
        league_avg_possessions: League average possessions per game
        team_col: Name of the team column in the dataframe
        
    Returns:
        DataFrame with PAdj metrics added
    """
    df = df.copy()
    
    # Get candidate metrics
    candidate_bases = get_padj_candidate_metrics()
    
    # Find matching columns in dataframe
    padj_metrics_added = []
    
    for base_name in candidate_bases:
        # Look for columns that match this base name
        matching_cols = []
        for col in df.columns:
            col_lower = str(col).lower()
            base_lower = base_name.lower()
            
            # Skip if it's a percentage or already PAdj
            if '%' in col_lower or 'percent' in col_lower or 'padj' in col_lower:
                continue
            
            # Check if base name matches
            if base_lower in col_lower or col_lower in base_lower:
                # More precise matching
                if base_lower == col_lower or \
                   base_lower + ' per 90' in col_lower or \
                   base_lower + 's per 90' in col_lower or \
                   col_lower.endswith(' ' + base_lower) or \
                   col_lower.endswith(' ' + base_lower + 's'):
                    matching_cols.append(col)
        
        # Calculate PAdj for each matching column
        for col in matching_cols:
            padj_col_name = f"PAdj {col}"
            
            # Skip if already exists
            if padj_col_name in df.columns:
                continue
            
            # Calculate PAdj for each row
            padj_values = []
            for idx, row in df.iterrows():
                team_name = row.get(team_col, '')
                metric_value = row.get(col, 0)
                
                # Get team possessions
                team_poss = team_possessions.get(team_name, None)
                
                if pd.notna(metric_value) and team_poss is not None and league_avg_possessions is not None:
                    # PAdj = Metric × (League Avg Possessions) / (Team Possessions)
                    padj_value = metric_value * (league_avg_possessions / team_poss)
                    padj_value = round(padj_value, 2)
                else:
                    padj_value = metric_value
                
                padj_values.append(padj_value)
            
            # Add PAdj column
            df[padj_col_name] = padj_values
            padj_metrics_added.append(padj_col_name)
    
    if padj_metrics_added:
        print(f"  ✅ Added {len(padj_metrics_added)} PAdj metrics")
    
    return df


def load_team_averages(base_dir, conference):
    """Compute team averages directly from team stats exports for a given conference."""
    team_stats_dir = get_team_stats_dir(base_dir, conference)
    if not team_stats_dir.exists():
        print(f"  ⚠️  Team stats directory not found: {team_stats_dir}")
        return {}
    
    team_averages_dict = {}
    files = sorted(team_stats_dir.glob("Team Stats *.xlsx"))
    if not files:
        print(f"  ⚠️  No team stats files found in {team_stats_dir}")
        return {}
    
    skip_columns = {'Date', 'Match', 'Competition', 'Scheme', 'Coach', 'Round', 'Result'}
    
    for file_path in files:
        try:
            team_name = clean_team_name_from_file(file_path)
            df = pd.read_excel(file_path)
            df = clean_team_stats_dataframe(df)
            if df.empty:
                continue
            
            team_metrics = {}
            for col in df.columns:
                if col in skip_columns or col.startswith('Unnamed'):
                    continue
                series = pd.to_numeric(df[col], errors='coerce')
                if series.notna().any():
                    avg_value = series.mean()
                    if pd.notna(avg_value):
                        team_metrics[col] = float(avg_value)
            
            if team_metrics:
                team_averages_dict[team_name] = team_metrics
        except Exception as e:
            print(f"  ⚠️  Error loading team averages from {file_path.name}: {e}")
    
    print(f"  ✅ Loaded team averages for {len(team_averages_dict)} teams")
    return team_averages_dict

def load_standardized_metrics():
    """Load the CSV that defines which metrics are already standardized"""
    csv_path = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Wyscout/Team Metrics Standardized Groups - Team Metrics Standardized Yes or No.csv")
    
    if not csv_path.exists():
        print(f"  ⚠️  Standardization CSV not found: {csv_path}")
        return set()
    
    try:
        df_std = pd.read_csv(csv_path)
        # Get metrics that are already standardized (Yes) or are per 90 or percentages
        standardized_metrics = set()
        
        for _, row in df_std.iterrows():
            metric = row['Metric']
            is_standardized = row.get('Already Standardized', 'No') == 'Yes'
            
            # Also check if metric name indicates standardization
            metric_lower = str(metric).lower()
            is_per_90 = 'per 90' in metric_lower
            is_percentage = '%' in metric or 'percent' in metric_lower
            
            if is_standardized or is_per_90 or is_percentage:
                standardized_metrics.add(metric)
        
        print(f"      Loaded {len(standardized_metrics)} standardized metrics")
        return standardized_metrics
    except Exception as e:
        print(f"  ⚠️  Error loading standardization CSV: {e}")
        return set()

def calculate_team_relative_metrics(df, team_averages_dict):
    """Calculate team-relative metrics for each player - only for standardized metrics"""
    df = df.copy()
    
    # Load list of standardized metrics
    standardized_metrics = load_standardized_metrics()
    
    print(f"      Processing {len(df.columns)} columns in dataframe")
    if 'Team' in df.columns:
        team_values = [str(team).strip() for team in df['Team'].dropna().unique()]
        team_list = sorted([team for team in team_values if team])
    else:
        team_list = []
    print(f"      Teams in dataframe: {team_list if team_list else 'No Team column'}")
    
    # Normalize metric names for comparison
    def normalize_metric_name(name):
        return str(name).lower().replace(',', '').replace('  ', ' ').strip()
    
    # Create a normalized set of standardized metrics for quick lookup
    normalized_std_metrics = {normalize_metric_name(m): m for m in standardized_metrics}
    
    # Helper: robust team name matching
    def match_team_name(team_name):
        if not team_name or pd.isna(team_name):
            return None
        team_name_str = str(team_name).strip()
        if team_name_str in team_averages_dict:
            return team_name_str
        team_lower = team_name_str.lower()
        for existing in team_averages_dict.keys():
            existing_str = str(existing).strip()
            if existing_str.lower() == team_lower:
                return existing_str
        for existing in team_averages_dict.keys():
            existing_str = str(existing).strip()
            existing_lower = existing_str.lower()
            if team_lower in existing_lower or existing_lower in team_lower:
                return existing_str
        return None

    # For each metric, calculate vs team average (only if standardized)
    for metric_col in df.columns:
        if metric_col in ['Player', 'Team', 'Position', 'Total_Score_1_10', '2024_Total_Score', 
                         'Change_From_2024', 'Total_Grade', 'Team_Grade', 'Conference_Grade', 
                         'Power_Five_Grade', 'Total_Percentile', 'Position_Profile', 'Year', 
                         'Season', 'Conference', 'Minutes played', 'Pct_Of_Team_Minutes']:
            continue
        
        # Check if this metric exists in team averages
        # Try case-insensitive matching and handle punctuation differences
        team_metric = None
        if not team_averages_dict:
            continue
        
        # Check if this metric is standardized and comparable before calculating team-relative values
        normalized_player_metric = normalize_metric_name(metric_col)
        
        # Team-relative comparisons only work for:
        # 1. Percentages (player success rate vs team success rate)
        # 2. Ratios/Averages (e.g., "Average pass length", "PPDA")
        # NOT for per-90 metrics (team totals vs individual player totals are incomparable)
        
        is_comparable = False
        
        # Check if it's a percentage (these are comparable)
        if '%' in metric_col or 'percent' in normalized_player_metric:
            is_comparable = True
        
        # Check if it's a ratio/average that's already standardized (not per 90)
        elif normalized_player_metric in normalized_std_metrics:
            # But exclude per-90 metrics
            if 'per 90' not in normalized_player_metric:
                is_comparable = True
        
        # Also check for known ratio/average metrics by name pattern
        elif any(keyword in normalized_player_metric for keyword in ['average', 'ratio', 'ppda', 'tempo']):
            if 'per 90' not in normalized_player_metric:
                is_comparable = True
        
        # Skip metrics that aren't comparable (per-90 or non-standardized)
        if not is_comparable:
            continue
        
        # Get available metrics from first team (they should all have the same metrics)
        sample_team_metrics = list(team_averages_dict.values())[0].keys() if team_averages_dict else []
        
        # Try exact match first
        for team_metric_name in sample_team_metrics:
            if normalize_metric_name(team_metric_name) == normalized_player_metric:
                team_metric = team_metric_name
                break
        
        # If no exact match, try removing word order differences (e.g., "Accurate passes" vs "Passes Accurate")
        if not team_metric:
            # Remove "accurate" from both and compare
            # Also handle "successful" and other modifiers
            player_words_base = normalized_player_metric
            for word in ['accurate', 'successful', 'won']:
                player_words_base = player_words_base.replace(word, '')
            player_words = set(player_words_base.split())
            player_words.discard('')  # Remove empty strings
            
            for team_metric_name in sample_team_metrics:
                normalized_team = normalize_metric_name(team_metric_name)
                team_words_base = normalized_team
                for word in ['accurate', 'successful', 'won']:
                    team_words_base = team_words_base.replace(word, '')
                team_words = set(team_words_base.split())
                team_words.discard('')  # Remove empty strings
                # Check if word sets match (order-independent)
                if player_words == team_words and len(player_words) > 0:
                    team_metric = team_metric_name
                    break
        
        # If still no match, try matching with location/type modifiers removed
        # (e.g., "Crosses from right flank" -> "Crosses", "Through passes" -> "Through Passes")
        if not team_metric:
            # Remove location/type modifiers like "from right flank", "to final third", etc.
            player_base = normalized_player_metric
            location_modifiers = ['from right flank', 'from left flank', 'to final third', 'in final third', 'in penalty area', 'to goalie box']
            for modifier in location_modifiers:
                player_base = player_base.replace(modifier, '')
            player_base = ' '.join(player_base.split())  # Clean up spaces
            
            # Also handle "short / medium" -> generic "passes"
            if 'short' in player_base and 'medium' in player_base:
                player_base = player_base.replace('short', '').replace('medium', '').replace('/', '').strip()
                player_base = 'passes' if 'pass' in player_base else player_base
            
            # Handle "through passes" - try to match to generic "passes" if no specific match
            if 'through' in player_base and 'pass' in player_base:
                # First try to match "through passes" specifically
                for team_metric_name in sample_team_metrics:
                    normalized_team = normalize_metric_name(team_metric_name)
                    if 'through' in normalized_team and 'pass' in normalized_team:
                        is_player_pct = '%' in metric_col or 'percent' in normalized_player_metric
                        is_team_pct = '%' in team_metric_name or 'percent' in normalized_team.lower()
                        if is_player_pct == is_team_pct:
                            team_metric = team_metric_name
                            break
                # If no specific match, fall through to generic "passes" match below
            
            # If still no match, try generic matching
            if not team_metric:
                for team_metric_name in sample_team_metrics:
                    normalized_team = normalize_metric_name(team_metric_name)
                    # Remove "accurate" and compare
                    team_base = normalized_team.replace('accurate', '').strip()
                    team_base = ' '.join(team_base.split())
                    
                    # For passes: if player has "through" or "short/medium", try matching to generic "passes"
                    if 'pass' in player_base:
                        # Remove modifiers from player_base
                        player_pass_base = player_base
                        for word in ['through', 'short', 'medium', 'long', 'smart', 'progressive', 'forward', 'back', 'lateral']:
                            player_pass_base = player_pass_base.replace(word, '')
                        player_pass_base = ' '.join(player_pass_base.split())
                        
                        # Match to generic "passes" or "pass" metrics
                        if 'pass' in team_base and (player_pass_base == 'passes' or player_pass_base == 'pass' or 'pass' in player_pass_base):
                            is_player_pct = '%' in metric_col or 'percent' in normalized_player_metric
                            is_team_pct = '%' in team_metric_name or 'percent' in normalized_team.lower()
                            if is_player_pct == is_team_pct:
                                team_metric = team_metric_name
                                break
                    
                    # For crosses: match location-specific crosses to generic "crosses"
                    elif 'cross' in player_base:
                        if 'cross' in team_base:
                            is_player_pct = '%' in metric_col or 'percent' in normalized_player_metric
                            is_team_pct = '%' in team_metric_name or 'percent' in normalized_team.lower()
                            if is_player_pct == is_team_pct:
                                team_metric = team_metric_name
                                break
                    
                    # For dribbles: match "successful dribbles" to "dribbles" if available
                    elif 'dribble' in player_base:
                        if 'dribble' in team_base:
                            is_player_pct = '%' in metric_col or 'percent' in normalized_player_metric
                            is_team_pct = '%' in team_metric_name or 'percent' in normalized_team.lower()
                            if is_player_pct == is_team_pct:
                                team_metric = team_metric_name
                                break
                    
                    # Generic fallback: check if bases match
                    elif player_base == team_base or team_base in player_base or player_base in team_base:
                        is_player_pct = '%' in metric_col or 'percent' in normalized_player_metric
                        is_team_pct = '%' in team_metric_name or 'percent' in normalized_team.lower()
                        if is_player_pct == is_team_pct:
                            team_metric = team_metric_name
                            break
        
        # If still no match, try with " per 90" suffix
        if not team_metric:
            for team_metric_name in sample_team_metrics:
                normalized_team = normalize_metric_name(team_metric_name)
                if normalized_team == normalized_player_metric + " per 90":
                    team_metric = team_metric_name
                    break
        
        # Debug for specific metrics
        if metric_col in ['Defensive duels per 90', 'Defensive duels won, %']:
            print(f"    🔍 Checking '{metric_col}' (normalized: '{normalized_player_metric}')")
            print(f"       Available team metrics: {[m for m in sample_team_metrics if 'defensive' in m.lower() and 'duel' in m.lower()]}")
            if team_metric:
                print(f"       ✅ Matched to: '{team_metric}'")
            else:
                print(f"       ❌ No match found")
        
        if team_metric:
            # Create new column for team-relative value
            team_rel_col = f"{metric_col}_vs_team"
            team_pos_rel_col = f"{metric_col}_vs_position"
            team_avg_col = f"{metric_col}_team_avg"
            team_pos_avg_col = f"{metric_col}_team_position_avg"
            
            def calc_vs_team(row):
                team_name = row.get('Team', '')
                player_value = row.get(metric_col, None)
                
                if pd.isna(player_value) or player_value == '' or team_name == '':
                    return None
                
                matched_team_name = match_team_name(team_name)
                team_data = team_averages_dict.get(matched_team_name, {})
                if not team_data:
                    return None
                
                team_avg = team_data.get(team_metric, None)
                if team_avg is None or team_avg == 0:
                    return None
                
                # Calculate percentage difference
                try:
                    player_val = float(player_value)
                    pct_diff = ((player_val - team_avg) / team_avg) * 100
                    return round(pct_diff, 1)
                except (ValueError, TypeError) as e:
                    return None
            
            df[team_rel_col] = df.apply(calc_vs_team, axis=1)

            # Populate absolute team average column
            def get_team_average_value(team_name):
                matched_name = match_team_name(team_name)
                if not matched_name:
                    return None
                team_data = team_averages_dict.get(matched_name, {})
                value = team_data.get(team_metric, None)
                try:
                    return float(value) if value is not None else None
                except (TypeError, ValueError):
                    return None

            df[team_avg_col] = df['Team'].apply(get_team_average_value)

            # Populate team-position average column (mean across players on same team + position)
            if 'Position_Profile' in df.columns:
                metric_numeric = pd.to_numeric(df[metric_col], errors='coerce')
                group_index = pd.MultiIndex.from_arrays([df['Team'], df['Position_Profile']])
                pos_avg_series = metric_numeric.groupby(group_index).transform('mean')
                df[team_pos_avg_col] = pos_avg_series
                
                # Calculate "% better than position" (player vs team-position average)
                def calc_vs_position(row):
                    player_value = row.get(metric_col, None)
                    pos_avg_value = row.get(team_pos_avg_col, None)
                    
                    if pd.isna(player_value) or pd.isna(pos_avg_value) or pos_avg_value == 0:
                        return None
                    
                    try:
                        player_val = float(player_value)
                        pos_avg = float(pos_avg_value)
                        pct_diff = ((player_val - pos_avg) / pos_avg) * 100
                        return round(pct_diff, 1)
                    except (ValueError, TypeError):
                        return None
                
                df[team_pos_rel_col] = df.apply(calc_vs_position, axis=1)
            else:
                df[team_pos_avg_col] = None
                df[team_pos_rel_col] = None
            
            # Debug: Print matching results
            non_null_count = len(df[df[team_rel_col].notna()])
            if non_null_count > 0:
                print(f"       ✅ Matched '{metric_col}' -> '{team_metric}' (calculated {non_null_count}/{len(df)} values)")
                # Show sample calculation
                sample_row = df[df[team_rel_col].notna()].iloc[0]
                sample_team = sample_row.get('Team', '')
                sample_player_val = sample_row.get(metric_col, '')
                sample_team_avg = team_averages_dict.get(sample_team, {}).get(team_metric, 'N/A')
                sample_pct = sample_row.get(team_rel_col, '')
                print(f"          Example: {sample_team} - Player: {sample_player_val}, Team Avg: {sample_team_avg}, % diff: {sample_pct}%")
            else:
                print(f"       ⚠️  Matched '{metric_col}' -> '{team_metric}' but calculated 0 values")
                # Debug why no values
                if len(df) > 0:
                    sample_row = df.iloc[0]
                    sample_team = sample_row.get('Team', '')
                    sample_player_val = sample_row.get(metric_col, None)
                    team_has_data = sample_team in team_averages_dict
                    team_has_metric = team_averages_dict.get(sample_team, {}).get(team_metric, None) is not None if team_has_data else False
                    print(f"          Debug: Team '{sample_team}' in dict: {team_has_data}, Has metric: {team_has_metric}, Player value: {sample_player_val}")
    
    return df

def calculate_top_15s_for_team_tabs(df, display_name, all_power_five_for_position, position_config):
    """
    Calculate Top 15s for a dataframe before appending to all_scored_data.
    This ensures team tabs have Top_15s values.
    """
    # Get relevant metrics for this position profile
    if position_config:
        relevant_metrics = get_relevant_metrics_for_position(position_config)
        # Filter to only metrics that exist in the dataframe and are in the relevant set
        exclude_cols = ['Player', 'Team', 'Position', 'Total_Score_1_10', 'Previous Year', 'Previous Score', 'Change From Previous', 'Previous Position',
                         'Total_Grade', 'Team_Grade', 'Conference_Grade', 'Power_Five_Grade', 
                        'Total_Percentile', 'Position_Profile', 'Year', 'Season', 'Conference', 
                        'Minutes played', 'Pct_Of_Team_Minutes', 'Top_15s']
        metric_cols = [col for col in df.columns if col not in exclude_cols 
                      and not col.endswith('_vs_team') and not col.endswith('_vs_position_profile')
                      and not col.endswith('_team_avg') and not col.endswith('_team_position_avg')
                      and 'PAdj' not in col and 'padj' not in col.lower()
                      and col in relevant_metrics]
    else:
        exclude_cols = ['Player', 'Team', 'Position', 'Total_Score_1_10', 'Previous Year', 'Previous Score', 'Change From Previous', 'Previous Position',
                         'Total_Grade', 'Team_Grade', 'Conference_Grade', 'Power_Five_Grade', 
                        'Total_Percentile', 'Position_Profile', 'Year', 'Season', 'Conference', 
                        'Minutes played', 'Pct_Of_Team_Minutes', 'Top_15s']
        metric_cols = [col for col in df.columns if col not in exclude_cols 
                      and not col.endswith('_vs_team') and not col.endswith('_vs_position_profile')
                      and not col.endswith('_team_avg') and not col.endswith('_team_position_avg')
                      and 'PAdj' not in col and 'padj' not in col.lower()]
    
    # Initialize Top_15s column
    df['Top_15s'] = 0
    
    try:
        # Use the all_power_five_for_position data that's already loaded
        if all_power_five_for_position is not None and len(all_power_five_for_position) > 0:
            all_players_df = all_power_five_for_position.copy()
        else:
            # Fallback: try loading from reports
            from generate_player_overviews import load_player_data_from_conference_reports
            from pathlib import Path
            base_dir = Path(__file__).parent.parent
            if base_dir.exists():
                all_players_df = load_player_data_from_conference_reports(base_dir)
            else:
                df['Top_15s'] = 0
                return df
        
        if len(all_players_df) == 0:
            df['Top_15s'] = 0
            return df
        
        # Power Five conferences
        power_five = {'ACC', 'SEC', 'BIG10', 'Big Ten', 'BIG12', 'Big 12', 'PAC12', 'Pac-12'}
        
        # Filter to same position profile and Power Five
        position_profile_col = None
        for col in ['Position Profile', 'Position_Profile', 'position_profile', 'position profile']:
            if col in all_players_df.columns:
                position_profile_col = col
                break
        
        if not position_profile_col:
            df['Top_15s'] = 0
            return df
        
        pos_players = all_players_df[all_players_df[position_profile_col] == display_name].copy()
        power_five_players = pos_players[pos_players['Conference'].isin(power_five)].copy()
        
        # Filter to only 2025 data
        if 'Year' in power_five_players.columns:
            power_five_players = power_five_players[power_five_players['Year'] == 2025].copy()
        elif 'Season' in power_five_players.columns:
            power_five_players = power_five_players[power_five_players['Season'] == 2025].copy()
        
        if len(power_five_players) == 0:
            df['Top_15s'] = 0
            return df
        
        # Reset index to ensure alignment
        df = df.reset_index(drop=True)
        
        # Calculate Top 15s for each player
        top15s_list = []
        matches_found = 0
        
        for idx, (_, player_row) in enumerate(df.iterrows()):
            top15_count = 0
            
            # Check each metric
            for col in metric_cols:
                player_value = None
                if col in player_row.index:
                    player_value = player_row[col]
                elif col in df.columns:
                    player_value = player_row.get(col)
                
                if pd.isna(player_value) or player_value == '':
                    continue
                
                try:
                    player_val = float(player_value)
                    
                    # Find matching column in power_five_players
                    matching_col = None
                    col_lower = str(col).lower().strip()
                    
                    # Try exact match first
                    for pf_col in power_five_players.columns:
                        if str(pf_col).lower().strip() == col_lower:
                            matching_col = pf_col
                            matches_found += 1
                            break
                    
                    # If no exact match, try base name matching
                    if not matching_col:
                        if 'per 90' in col_lower and '%' not in col_lower:
                            col_base = col_lower.replace(' per 90', '').replace(' per90', '').strip()
                            for pf_col in power_five_players.columns:
                                pf_col_lower = str(pf_col).lower()
                                if 'per 90' in pf_col_lower and '%' not in pf_col_lower and '% better than position' not in pf_col_lower:
                                    pf_col_base = pf_col_lower.replace(' per 90', '').replace(' per90', '').strip()
                                    col_base_norm = ' '.join(col_base.split())
                                    pf_col_base_norm = ' '.join(pf_col_base.split())
                                    if col_base_norm == pf_col_base_norm:
                                        matching_col = pf_col
                                        matches_found += 1
                                        break
                        elif ('%' in col_lower or 'percent' in col_lower) and '% better than position' not in col_lower:
                            col_base = col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                            for pf_col in power_five_players.columns:
                                pf_col_lower = str(pf_col).lower()
                                if ('%' in pf_col_lower or 'percent' in pf_col_lower) and 'per 90' not in pf_col_lower and '% better than position' not in pf_col_lower:
                                    pf_col_base = pf_col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                                    col_base_norm = ' '.join(col_base.split())
                                    pf_col_base_norm = ' '.join(pf_col_base.split())
                                    if col_base_norm == pf_col_base_norm:
                                        matching_col = pf_col
                                        matches_found += 1
                                        break
                    
                    if matching_col and matching_col in power_five_players.columns:
                        pf_values = pd.to_numeric(power_five_players[matching_col], errors='coerce').dropna()
                        if len(pf_values) > 0:
                            pf_rank = (pf_values >= player_val).sum()
                            sorted_values = pf_values.sort_values(ascending=False)
                            if len(sorted_values) >= 15:
                                value_at_15th = sorted_values.iloc[14]
                                if pf_rank <= 15 or player_val >= value_at_15th:
                                    top15_count += 1
                            elif pf_rank <= 15:
                                top15_count += 1
                except (ValueError, TypeError):
                    continue
            
            top15s_list.append(top15_count)
        
        df['Top_15s'] = top15s_list
        print(f"     ✅ Calculated Top 15s for {len(df)} players ({matches_found} metric matches found)")
        
    except Exception as e:
        import traceback
        print(f"     ⚠️  Error calculating Top 15s: {e}")
        traceback.print_exc()
        df['Top_15s'] = 0
    
    return df

def create_position_profile_sheet(wb, df, display_name, team_minutes_dict, position_config=None, team_averages_dict=None, all_power_five_for_position=None):
    """Create position profile sheet with grouped metric format."""
    ws = wb.create_sheet(title=display_name)
    
    # Calculate position profile averages across all five conferences
    position_profile_averages = {}
    if all_power_five_for_position is not None and len(all_power_five_for_position) > 0:
        print(f"  📊 Calculating position profile averages for {display_name}...")
        print(f"     Using {len(all_power_five_for_position)} players across all five conferences")
        
        # Calculate average for each metric column
        exclude_cols = ['Player', 'Team', 'Position', 'Total_Score_1_10', 'Previous Year', 'Previous Score', 'Change From Previous', 'Previous Position',
                        'Total_Grade', 'Team_Grade', 'Conference_Grade', 'Power_Five_Grade', 
                        'Total_Percentile', 'Position_Profile', 'Year', 'Season', 'Conference', 
                        'Minutes played', 'Pct_Of_Team_Minutes']
        metric_cols = [col for col in all_power_five_for_position.columns 
                      if col not in exclude_cols and not col.endswith('_vs_team') 
                      and not col.endswith('_vs_position') and not col.endswith('_team_avg')
                      and not col.endswith('_team_position_avg')]
        
        for metric_col in metric_cols:
            try:
                metric_numeric = pd.to_numeric(all_power_five_for_position[metric_col], errors='coerce')
                avg_value = metric_numeric.mean()
                if pd.notna(avg_value):
                    position_profile_averages[metric_col] = avg_value
            except Exception:
                pass
        
        print(f"     Calculated averages for {len(position_profile_averages)} metrics")
    else:
        print(f"  ⚠️  No Power Five data available for {display_name} position profile averages")
    
    # Calculate position-relative metrics (% better than position profile average)
    if position_profile_averages:
        print(f"  📊 Calculating position-relative metrics for {display_name}...")
        for metric_col, position_avg in position_profile_averages.items():
            if position_avg == 0 or pd.isna(position_avg):
                continue
            
            pos_rel_col = f"{metric_col}_vs_position_profile"
            
            def calc_vs_position_profile(row):
                player_value = row.get(metric_col, None)
                if pd.isna(player_value) or player_value == '':
                    return None
                
                try:
                    player_val = float(player_value)
                    pct_diff = ((player_val - position_avg) / position_avg) * 100
                    return round(pct_diff, 1)
                except (ValueError, TypeError):
                    return None
            
            df[pos_rel_col] = df.apply(calc_vs_position_profile, axis=1)
    
    # Calculate team-relative metrics if team averages are available (for reference, but we'll use position-relative in display)
    if team_averages_dict:
        print(f"  📊 Calculating team-relative metrics for {display_name}...")
        print(f"     Team averages dict has {len(team_averages_dict)} teams")
        if team_averages_dict:
            sample_team = list(team_averages_dict.keys())[0]
            print(f"     Sample team '{sample_team}' has {len(team_averages_dict[sample_team])} metrics")
        df = calculate_team_relative_metrics(df, team_averages_dict)
    else:
        print(f"  ⚠️  No team averages available for {display_name}")
    
    # Sort by score (highest first)
    df_sorted = df.sort_values(['Total_Score_1_10'], ascending=[False])
    
    # Base headers - before metrics
    base_headers = ['Player', 'Team', 'Position', 'Conference Grade', 'Power Five Grade', '2025 Total Score', 'Previous Year', 'Previous Score', 'Change From Previous', 
                   'Total Minutes', '% of Team Minutes', 'Top 15s (Power Five)']
    
    # Get relevant metrics for this position profile
    if position_config:
        relevant_metrics = get_relevant_metrics_for_position(position_config)
        # Filter to only metrics that exist in the dataframe and are in the relevant set
        exclude_cols = ['Player', 'Team', 'Position', 'Total_Score_1_10', 'Previous Year', 'Previous Score', 'Change From Previous', 'Previous Position',
                         'Total_Grade', 'Team_Grade', 'Conference_Grade', 'Power_Five_Grade',
                        'Total_Percentile', 'Position_Profile', 'Year', 'Season', 'Conference', 
                        'Minutes played', 'Pct_Of_Team_Minutes']
        all_metric_cols = [col for col in df_sorted.columns if col not in exclude_cols 
                          and not col.endswith('_vs_team') and not col.endswith('_vs_position_profile')
                          and not col.endswith('_team_avg') and not col.endswith('_team_position_avg')]
        
        # Helper function to normalize metric names for comparison (case-insensitive, handle spacing)
        def normalize_metric_name(name):
            """Normalize metric name for comparison"""
            return name.lower().strip().replace(' ', '').replace('_', '').replace('-', '').replace('/', '')
        
        # Create a mapping of normalized config metrics to their original names
        normalized_relevant = {normalize_metric_name(m): m for m in relevant_metrics}
        
        # Also create a mapping for partial matching (base names without modifiers)
        def get_base_for_matching(name):
            """Get base name for matching (removes common modifiers)"""
            base = name.lower()
            # Remove common suffixes (order matters - remove longer ones first)
            for suffix in [' per 90', ', %', ' %', ' won', ' accurate', ' successful', ' blocked']:
                base = base.replace(suffix, '')
            # Remove "padj " prefix if present (for PAdj metrics)
            if base.startswith('padj '):
                base = base[5:]  # Remove "padj "
            # Remove trailing punctuation
            base = base.strip(' ,')
            return normalize_metric_name(base)
        
        base_to_relevant = {}
        for metric in relevant_metrics:
            base = get_base_for_matching(metric)
            if base not in base_to_relevant:
                base_to_relevant[base] = []
            base_to_relevant[base].append(metric)
        
        # Match dataframe columns to relevant metrics (flexible but strict matching)
        metric_cols = []
        matched_config_metrics = set()
        col_to_config = {}  # Track which config metric each column matched
        
        for col in all_metric_cols:
            matched = False
            matched_config = None
            
            # First, try exact match (case-insensitive, normalized)
            normalized_col = normalize_metric_name(col)
            if normalized_col in normalized_relevant:
                matched_config = normalized_relevant[normalized_col]
                metric_cols.append(col)
                matched_config_metrics.add(matched_config)
                col_to_config[col] = matched_config
                matched = True
            
            # If not matched, try base name matching (more flexible)
            if not matched:
                col_base = get_base_for_matching(col)
                if col_base in base_to_relevant:
                    # Check if this column matches any of the relevant metrics with this base
                    for config_metric in base_to_relevant[col_base]:
                        config_base = get_base_for_matching(config_metric)
                        
                        # Base names must match exactly
                        if col_base == config_base:
                            # Check metric type indicators
                            col_has_per90 = 'per 90' in col.lower()
                            config_has_per90 = 'per 90' in config_metric.lower()
                            col_has_pct = '%' in col or 'percent' in col.lower()
                            config_has_pct = '%' in config_metric or 'percent' in config_metric.lower()
                            
                            # Match if types align (both per 90, both %, or compatible)
                            # Allow per 90 and % to match for same base (e.g., "duels per 90" and "duels won, %")
                            type_compatible = (
                                (col_has_per90 == config_has_per90 and col_has_pct == config_has_pct) or  # Exact type match
                                (col_base == config_base)  # Same base, different types are OK (e.g., per 90 vs %)
                            )
                            
                            if type_compatible:
                                matched_config = config_metric
                                metric_cols.append(col)
                                matched_config_metrics.add(config_metric)
                                col_to_config[col] = config_metric
                                matched = True
                                break
        
        # Remove duplicates while preserving order
        seen = set()
        metric_cols = [col for col in metric_cols if col not in seen and not seen.add(col)]
        
        # Filter out PAdj columns
        metric_cols = [col for col in metric_cols if 'PAdj' not in col and 'padj' not in col.lower()]
        
        print(f"  📊 Filtered to {len(metric_cols)} relevant metrics (from {len(all_metric_cols)} total, {len(relevant_metrics)} in config)")
        print(f"  📊 Matched {len(matched_config_metrics)}/{len(relevant_metrics)} config metrics")
        if len(metric_cols) == 0:
            print(f"  ⚠️  Warning: No metrics matched! Relevant metrics from config: {sorted(relevant_metrics)[:5]}...")
            print(f"  ⚠️  Sample dataframe columns: {sorted(all_metric_cols)[:5]}...")
        elif len(matched_config_metrics) < len(relevant_metrics):
            missing = set(relevant_metrics) - matched_config_metrics
            print(f"  ⚠️  Missing {len(missing)} config metrics: {sorted(missing)[:5]}...")
    else:
        # Fallback: include all metrics if no config provided
        exclude_cols = ['Player', 'Team', 'Position', 'Total_Score_1_10', 'Previous Year', 'Previous Score', 'Change From Previous', 'Previous Position',
                         'Total_Grade', 'Team_Grade', 'Conference_Grade', 'Power_Five_Grade',
                        'Total_Percentile', 'Position_Profile', 'Year', 'Season', 'Conference', 
                        'Minutes played', 'Pct_Of_Team_Minutes']
        metric_cols = [col for col in df_sorted.columns if col not in exclude_cols 
                      and not col.endswith('_vs_team') and not col.endswith('_vs_position_profile')
                      and not col.endswith('_team_avg') and not col.endswith('_team_position_avg')
                      and 'PAdj' not in col and 'padj' not in col.lower()]
    
    # Calculate Top 15s (Power Five) for each player (AFTER we know which metrics will be displayed)
    print(f"  📊 Calculating Top 15s (Power Five) for {display_name}...")
    print(f"     Will check {len(metric_cols)} metrics: {metric_cols[:5]}...")
    
    try:
        # Use the all_power_five_for_position data that's already loaded (no need to load from reports)
        if all_power_five_for_position is not None and len(all_power_five_for_position) > 0:
            all_players_df = all_power_five_for_position.copy()
            print(f"  📊 Using Power Five data already loaded ({len(all_players_df)} players)")
        else:
            # Fallback: try loading from reports if Power Five data not available
            from generate_player_overviews import load_player_data_from_conference_reports
            base_dir = Path(__file__).parent.parent
            if not base_dir.exists():
                print(f"  ⚠️  Base directory does not exist: {base_dir}")
                df_sorted['Top_15s'] = 0
            else:
                reports_dir = base_dir
                conference_reports = list(reports_dir.glob('Portland Thorns 2025 * Championship Scouting Report.xlsx'))
                if len(conference_reports) == 0:
                    print(f"  ⚠️  No Power Five data available and no reports found")
                    df_sorted['Top_15s'] = 0
                else:
                    print(f"  📊 Loading from reports as fallback...")
                    all_players_df = load_player_data_from_conference_reports(base_dir)
                    print(f"  ✅ Loaded {len(all_players_df)} players from reports")
        
        if len(all_players_df) == 0:
            print(f"  ⚠️  No players available for Top 15s calculation")
            df_sorted['Top_15s'] = 0
        else:
                # Power Five conferences
                power_five = {'ACC', 'SEC', 'BIG10', 'Big Ten', 'BIG12', 'Big 12', 'PAC12', 'Pac-12'}
                
                # Filter to same position profile and Power Five
                # Check which column name is used for position profile
                position_profile_col = None
                for col in ['Position Profile', 'Position_Profile', 'position_profile', 'position profile']:
                    if col in all_players_df.columns:
                        position_profile_col = col
                        break
                
                if not position_profile_col:
                    print(f"  ⚠️  Could not find Position Profile column in all_players_df")
                    print(f"     Available columns: {[c for c in all_players_df.columns if 'position' in c.lower() or 'profile' in c.lower()]}")
                    df_sorted['Top_15s'] = 0
                else:
                    pos_players = all_players_df[all_players_df[position_profile_col] == display_name].copy()
                    power_five_players = pos_players[pos_players['Conference'].isin(power_five)].copy()
                    
                    # Filter to only 2025 data (same as shortlist uses)
                    if 'Year' in power_five_players.columns:
                        power_five_players = power_five_players[power_five_players['Year'] == 2025].copy()
                    elif 'Season' in power_five_players.columns:
                        power_five_players = power_five_players[power_five_players['Season'] == 2025].copy()
                    
                    print(f"     Found {len(power_five_players)} Power Five players for {display_name} (2025 only)")
                    
                    if len(power_five_players) == 0:
                        print(f"  ⚠️  No Power Five players found for {display_name}")
                        df_sorted['Top_15s'] = 0
                    else:
                        # Reset index to ensure alignment between iteration and assignment
                        df_sorted = df_sorted.reset_index(drop=True)
                        
                        # Calculate Top 15s for each player using the metric_cols that will be displayed
                        top15s_list = []
                        matches_found = 0
                        
                        for idx, (_, player_row) in enumerate(df_sorted.iterrows()):
                            top15_count = 0
                            
                            # Check each metric that will be displayed (from metric_cols)
                            for col in metric_cols:
                                # Try multiple ways to get the value
                                player_value = None
                                if col in player_row.index:
                                    player_value = player_row[col]
                                elif col in df_sorted.columns:
                                    player_value = player_row.get(col)
                                
                                if pd.isna(player_value) or player_value == '':
                                    continue
                                
                                try:
                                    player_val = float(player_value)
                                    
                                    # Find matching column in power_five_players (case-insensitive)
                                    matching_col = None
                                    col_lower = str(col).lower().strip()
                                    
                                    # Try exact match first (case-insensitive)
                                    for pf_col in power_five_players.columns:
                                        if str(pf_col).lower().strip() == col_lower:
                                            matching_col = pf_col
                                            matches_found += 1
                                            break
                                    
                                    # If no exact match, try base name matching
                                    if not matching_col:
                                        # For per 90 metrics
                                        if 'per 90' in col_lower and '%' not in col_lower:
                                            col_base = col_lower.replace(' per 90', '').replace(' per90', '').strip()
                                            for pf_col in power_five_players.columns:
                                                pf_col_lower = str(pf_col).lower()
                                                if 'per 90' in pf_col_lower and '%' not in pf_col_lower and '% better than position' not in pf_col_lower:
                                                    pf_col_base = pf_col_lower.replace(' per 90', '').replace(' per90', '').strip()
                                                    # Normalize spaces
                                                    col_base_norm = ' '.join(col_base.split())
                                                    pf_col_base_norm = ' '.join(pf_col_base.split())
                                                    
                                                    if col_base_norm == pf_col_base_norm:
                                                        matching_col = pf_col
                                                        matches_found += 1
                                                        break
                                        
                                        # For percentage metrics (skip "% better than position" columns)
                                        elif ('%' in col_lower or 'percent' in col_lower) and '% better than position' not in col_lower:
                                            col_base = col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                                            for pf_col in power_five_players.columns:
                                                pf_col_lower = str(pf_col).lower()
                                                if ('%' in pf_col_lower or 'percent' in pf_col_lower) and 'per 90' not in pf_col_lower and '% better than position' not in pf_col_lower:
                                                    pf_col_base = pf_col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                                                    
                                                    col_base_norm = ' '.join(col_base.split())
                                                    pf_col_base_norm = ' '.join(pf_col_base.split())
                                                    
                                                    if col_base_norm == pf_col_base_norm:
                                                        matching_col = pf_col
                                                        matches_found += 1
                                                        break
                                    
                                    if matching_col and matching_col in power_five_players.columns:
                                        pf_values = pd.to_numeric(power_five_players[matching_col], errors='coerce').dropna()
                                        if len(pf_values) > 0:
                                            # Count how many players have >= this value
                                            pf_rank = (pf_values >= player_val).sum()
                                            # Get the 15th highest value (for handling ties)
                                            sorted_values = pf_values.sort_values(ascending=False)
                                            if len(sorted_values) >= 15:
                                                value_at_15th = sorted_values.iloc[14]  # 0-indexed, so 14 = 15th
                                                # Top 15 if: rank <= 15 OR value >= 15th highest value (handles ties)
                                                if pf_rank <= 15 or player_val >= value_at_15th:
                                                    top15_count += 1
                                            elif pf_rank <= 15:
                                                # If less than 15 players total, just check rank
                                                top15_count += 1
                                except (ValueError, TypeError) as e:
                                    continue
                            
                            top15s_list.append(top15_count)
                        
                        # Assign Top_15s (indices already aligned from reset_index above)
                        df_sorted['Top_15s'] = top15s_list
                        print(f"  ✅ Calculated Top 15s (Power Five) for {len(df_sorted)} players ({matches_found} metric matches found)")
                        print(f"     Sample Top 15s (Power Five) values: {top15s_list[:5]}")
                        
    except Exception as e:
        import traceback
        print(f"  ⚠️  Error calculating Top 15s: {e}")
        traceback.print_exc()
        df_sorted['Top_15s'] = 0
    
    # For position profile tabs, show base columns plus relevant metrics
    # metric_cols is already populated above with relevant metrics for this position profile
    
    # Automatically group metrics by their base name
    # Format: {group_name: {metric_name: [sub_headers]}}
    metric_groups = {}
    
    def get_base_name(metric_name):
        """Extract base name from metric for grouping"""
        # Remove common suffixes to get base name (case-insensitive)
        base = metric_name.lower()
        # Remove " per 90"
        base = base.replace(' per 90', '')
        # Remove " won, %" or " won %"
        base = base.replace(' won, %', '').replace(' won %', '').replace(' won', '')
        # Remove " accurate, %" or " accurate %" or just " accurate" (at start or with space before)
        # Do this before removing trailing ", %" to handle "Accurate passes, %" correctly
        base = base.replace('accurate, %', '').replace('accurate %', '').replace('accurate ', '').replace('accurate', '')
        # Clean up any double spaces that might result
        base = ' '.join(base.split())
        # Remove " successful, %" or " successful %"
        base = base.replace(' successful, %', '').replace(' successful %', '').replace(' successful', '')
        # Remove trailing ", %" or " %"
        base = base.replace(', %', '').replace(' %', '')
        # Normalize: strip and capitalize first letter of each word
        base = base.strip()
        if base:
            # Capitalize first letter of each word
            words = base.split()
            base = ' '.join(word.capitalize() for word in words)
        return base
    
    # Group metrics by base name
    for metric in metric_cols:
        base_name = get_base_name(metric)
        
        # Determine sub-headers based on metric type
        # Order: Per 90 first, then %, remove the third "%" column (% better than position)
        sub_headers = []
        
        # Check if this metric has a position-relative column (% better than position)
        pos_rel_col = f"{metric}_vs_position_profile"
        has_position_relative = pos_rel_col in df_sorted.columns
        
        # Check if it's a per-90 metric
        if 'per 90' in metric:
            sub_headers.append('per 90')
            # Replace "%" with "% better than position" if available, otherwise keep "%"
            if has_position_relative:
                sub_headers.append('% better than position')
            else:
                # Check if there's a corresponding percentage version
                base_for_pct = get_base_name(metric)
                pct_version = None
                for other_metric in metric_cols:
                    if other_metric != metric:
                        other_base = get_base_name(other_metric)
                        if base_for_pct.lower() == other_base.lower():
                            if '%' in other_metric or 'percent' in other_metric.lower():
                                pct_version = other_metric
                                break
                if pct_version:
                    sub_headers.append('%')
        
        # Check if it's a percentage (but not if we already handled it above)
        elif '%' in metric or 'percent' in metric.lower():
            # Check if this percentage metric has a corresponding per-90 version
            base_for_per90 = get_base_name(metric)
            per90_version = None
            for other_metric in metric_cols:
                if other_metric != metric:
                    other_base = get_base_name(other_metric)
                    # Match if base names are the same and other metric is per-90
                    if base_for_per90.lower() == other_base.lower():
                        if 'per 90' in other_metric:
                            per90_version = other_metric
                            break
            # If no per-90 version exists, add this percentage metric
            # Replace "%" with "% better than position" if available
            if not per90_version:
                if has_position_relative:
                    sub_headers.append('% better than position')
                else:
                    sub_headers.append('%')
        
        # Skip PAdj and other ratio metrics entirely
        elif any(keyword in metric.lower() for keyword in ['padj', 'average', 'ratio', 'tempo']):
            continue  # Skip PAdj metrics
        
        else:
            # Default: show metric name (for metrics that don't fit other categories)
            sub_headers.append(metric)
        
        # Only add to group if we have sub-headers
        if sub_headers:
            if base_name not in metric_groups:
                metric_groups[base_name] = {}
            metric_groups[base_name][metric] = sub_headers
    
    # Use all metric groups (no longer separating by position-relative since we removed that column)
    metric_groups_ordered = metric_groups
    
    # Row 1: Write base headers and group headers (merged)
    current_col = 1
    for header in base_headers:
        cell = ws.cell(row=1, column=current_col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Merge base headers to span rows 1-2 (no duplicate in row 2)
        # Note: Merged cells automatically hide the duplicate - no need to write to row 2
        ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
        current_col += 1
    
    # Write metric group headers (row 1, merged)
    # Use the ordered groups (team-relative first)
    for group_name, metrics_dict in metric_groups_ordered.items():
        # Count total columns for this group
        total_cols = sum(len(sub_headers) for sub_headers in metrics_dict.values())
        start_col = current_col
        end_col = current_col + total_cols - 1
        
        # Merge cells for group header (even if single column, merge for consistency)
        # Always merge, even for single columns (openpyxl allows this, creates consistent structure)
        if start_col < end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        # For single-column groups (start_col == end_col), we still treat it as a merged group
        # by setting the cell value and formatting
        cell = ws.cell(row=1, column=start_col, value=group_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Row 2: Write sub-headers for each metric
        sub_col = start_col
        for metric_name, sub_headers in metrics_dict.items():
            for sub_header in sub_headers:
                cell = ws.cell(row=2, column=sub_col, value=sub_header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sub_col += 1
        
        current_col = end_col + 1
    
    # Write data - Conference Grade in column 4, Power Five Grade in column 5
    # Data rows start at row 3 (rows 1-2 are headers)
    for row_idx, (_, player_row) in enumerate(df_sorted.iterrows(), 3):
        ws.cell(row=row_idx, column=1, value=player_row.get('Player', ''))
        ws.cell(row=row_idx, column=2, value=player_row.get('Team', ''))
        ws.cell(row=row_idx, column=3, value=player_row.get('Position', ''))
        ws.cell(row=row_idx, column=4, value=player_row.get('Conference_Grade', ''))  # Conference Grade
        ws.cell(row=row_idx, column=5, value=player_row.get('Power_Five_Grade', ''))  # Power Five Grade
        # Write numeric columns with right alignment
        score_cell = ws.cell(row=row_idx, column=6, value=player_row.get('Total_Score_1_10', ''))
        score_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Previous Year (column 7)
        prev_year = player_row.get('Previous Year', '')
        prev_year_cell = ws.cell(row=row_idx, column=7, value=prev_year)
        prev_year_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Previous Score (column 8)
        prev_score = player_row.get('Previous Score', '')
        if pd.notna(prev_score) and prev_score != '':
            try:
                prev_score_cell = ws.cell(row=row_idx, column=8, value=round(float(prev_score), 2))
            except:
                prev_score_cell = ws.cell(row=row_idx, column=8, value=prev_score)
        else:
            prev_score_cell = ws.cell(row=row_idx, column=8, value='')
        prev_score_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Change From Previous (column 9)
        change = player_row.get('Change From Previous', '')
        if pd.notna(change) and change != '':
            try:
                change_cell = ws.cell(row=row_idx, column=9, value=round(float(change), 2))
            except:
                change_cell = ws.cell(row=row_idx, column=9, value=change)
        else:
            change_cell = ws.cell(row=row_idx, column=9, value='')
        change_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        minutes_cell = ws.cell(row=row_idx, column=10, value=player_row.get('Minutes played', ''))  # Total Minutes
        minutes_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Write percentage of team minutes (format as percentage with 2 decimals, cap at 100%)
        pct_value = player_row.get('Pct_Of_Team_Minutes', None)
        if pd.notna(pct_value) and pct_value is not None:
            pct_value = min(float(pct_value), 100.0)  # Cap at 100%
            pct_cell = ws.cell(row=row_idx, column=11, value=round(pct_value, 2))  # % of Team Minutes
            pct_cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            pct_cell = ws.cell(row=row_idx, column=11, value='')
            pct_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Write Top 15s (Power Five) (column 12)
        top15s_value = player_row.get('Top_15s', 0)
        if pd.isna(top15s_value) or top15s_value == '':
            top15s_value = 0
        top15s_cell = ws.cell(row=row_idx, column=12, value=int(top15s_value))
        top15s_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write metric columns in grouped format (starting after base headers)
        metric_col_idx = len(base_headers) + 1
        
        # Write data for each metric group (use ordered groups - position-relative first)
        for group_name, metrics_dict in metric_groups_ordered.items():
            for metric_name, sub_headers in metrics_dict.items():
                # Get the metric value
                metric_value = player_row.get(metric_name, '')
                
                # Write columns based on sub_headers
                for sub_header in sub_headers:
                    if sub_header == 'per 90':
                        # Write the per-90 metric value
                        if pd.notna(metric_value) and metric_value != '':
                            try:
                                metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value=round(float(metric_value), 2))
                            except (ValueError, TypeError):
                                metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value=metric_value)
                        else:
                            metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value='')
                        metric_cell.alignment = Alignment(horizontal='right', vertical='center')
                        metric_col_idx += 1
                    elif sub_header == '%':
                        # For % column, we need to find the corresponding % metric if this is a per-90 metric
                        if 'per 90' in metric_name:
                            # Find the corresponding % version using base name matching
                            base_for_pct = get_base_name(metric_name)
                            pct_metric_name = None
                            for other_metric in metric_cols:
                                if other_metric != metric_name:
                                    other_base = get_base_name(other_metric)
                                    if base_for_pct.lower() == other_base.lower():
                                        if '%' in other_metric or 'percent' in other_metric.lower():
                                            pct_metric_name = other_metric
                                            break
                            if pct_metric_name:
                                pct_value = player_row.get(pct_metric_name, '')
                            else:
                                pct_value = ''
                        else:
                            # This metric itself is a percentage
                            pct_value = metric_value
                        
                        # Write the percentage value
                        if pd.notna(pct_value) and pct_value != '':
                            try:
                                pct_val = float(pct_value)
                                pct_val = min(pct_val, 100.0)  # Cap at 100%
                                metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value=round(pct_val, 1))
                            except (ValueError, TypeError):
                                metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value=pct_value)
                        else:
                            metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value='')
                        metric_cell.alignment = Alignment(horizontal='right', vertical='center')
                        metric_col_idx += 1
                    elif sub_header == '% better than position':
                        # Write position-relative value (% better than position profile average)
                        pos_rel_col = f"{metric_name}_vs_position_profile"
                        pos_rel_value = player_row.get(pos_rel_col, None)
                        if pd.notna(pos_rel_value) and pos_rel_value != '':
                            try:
                                pos_rel_cell = ws.cell(row=row_idx, column=metric_col_idx, value=round(float(pos_rel_value), 1))
                            except (ValueError, TypeError):
                                pos_rel_cell = ws.cell(row=row_idx, column=metric_col_idx, value=pos_rel_value)
                        else:
                            pos_rel_cell = ws.cell(row=row_idx, column=metric_col_idx, value='')
                        pos_rel_cell.alignment = Alignment(horizontal='right', vertical='center')
                        metric_col_idx += 1
                    else:
                        # Write the main metric value (for other metric types)
                        if pd.notna(metric_value) and metric_value != '':
                            try:
                                metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value=round(float(metric_value), 2))
                            except (ValueError, TypeError):
                                metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value=metric_value)
                        else:
                            metric_cell = ws.cell(row=row_idx, column=metric_col_idx, value='')
                        metric_cell.alignment = Alignment(horizontal='right', vertical='center')
                        metric_col_idx += 1
    
    # Apply conditional formatting to Grade columns (columns D, E)
    # Data starts at row 3 (rows 1-2 are headers)
    if ws.max_row > 2:
        # Conference Grade column (D)
        conf_grade_range = f'D3:D{ws.max_row}'
        for grade, color in GRADE_COLORS.items():
            rule = CellIsRule(
                operator='equal', formula=[f'"{grade}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
            )
            ws.conditional_formatting.add(conf_grade_range, rule)
        
        # Power Five Grade column (E)
        power_five_grade_range = f'E3:E{ws.max_row}'
        for grade, color in GRADE_COLORS.items():
            rule = CellIsRule(
                operator='equal', formula=[f'"{grade}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
            )
            ws.conditional_formatting.add(power_five_grade_range, rule)
    
    # Apply conditional formatting to % of Team Minutes column (column J / 10)
    # Data starts at row 3 (rows 1-2 are headers)
    if ws.max_row > 2:
        # Percentage of team minutes column (column J / 10)
        pct_range = f'J3:J{ws.max_row}'
        
        # Percentage ranges with color scheme:
        # Dark red for highest %, dark blue for lowest %
        pct_ranges = [
            (0, 20, '1F4E79'),      # 0-20%: Dark blue (very low usage)
            (20, 40, '8FAADC'),    # 20-40%: Light blue (low usage)
            (40, 60, 'F2A2A2'),    # 40-60%: Light red (medium usage)
            (60, 80, 'C5504B'),    # 60-80%: Medium red (high usage)
            (80, 100, '8B0000')    # 80-100%: Dark red (very high usage)
        ]
        
        for min_pct, max_pct, color in pct_ranges:
            if min_pct == 0:
                rule = CellIsRule(
                    operator='lessThanOrEqual', formula=[f'{max_pct}'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
            elif max_pct >= 100:
                rule = CellIsRule(
                    operator='greaterThanOrEqual', formula=[f'{min_pct}'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
            else:
                rule = CellIsRule(
                    operator='between', formula=[f'{min_pct}', f'{max_pct}'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
            ws.conditional_formatting.add(pct_range, rule)
    
    # Freeze the first two rows (headers)
    ws.freeze_panes = 'A3'
    
    # Auto-adjust column widths
    # Calculate total columns: base headers + metric columns
    total_metric_cols = sum(len(sub_headers) for group_dict in metric_groups_ordered.values() 
                           for sub_headers in group_dict.values())
    total_cols = len(base_headers) + total_metric_cols
    
    for col_idx in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def create_team_summary_sheets(wb, df_all_scored, conference, team_minutes_dict, config=None, team_averages_dict=None):
    """Create team summary sheets with grouped metric format."""
    # Get championship teams for this conference
    allowed_teams = get_championship_teams(conference)
    
    # Create a mapping of position profile to relevant metrics
    position_profile_to_metrics = {}
    if config:
        for position_name, position_config in config.get('position_profiles', {}).items():
            display_name = POSITION_PROFILE_MAP.get(position_name, position_name)
            relevant_metrics = get_relevant_metrics_for_position(position_config)
            position_profile_to_metrics[display_name] = relevant_metrics
    
    # Filter to only allowed teams
    if allowed_teams:
        all_teams = df_all_scored['Team'].dropna().unique()
        teams = []
        for team in all_teams:
            team_str = str(team).strip()
            # Check exact match or if team contains allowed team name (but exclude Virginia Tech)
            for allowed_team in allowed_teams:
                allowed_lower = allowed_team.lower()
                team_lower = team_str.lower()
                
                # Exact match
                if allowed_lower == team_lower:
                    teams.append(team)
                    break
                # Partial match (but exclude Virginia Tech if looking for Virginia)
                elif allowed_lower in team_lower:
                    # Special case: exclude Virginia Tech when looking for Virginia
                    if 'virginia' in allowed_lower and 'virginia tech' in team_lower:
                        continue
                    # Special case: handle Mississippi State variations
                    if 'mississippi' in allowed_lower:
                        if 'mississippi st' in team_lower or 'mississippi state' in team_lower:
                            teams.append(team)
                            break
                    else:
                        teams.append(team)
                        break
        teams = sorted(teams)
        print(f"  📋 Filtering to {len(teams)} teams: {[str(t) for t in teams]}")
    else:
        # If no filter specified, include all teams
        teams = sorted(df_all_scored['Team'].dropna().unique())
    
    for team in teams:
        team_df = df_all_scored[df_all_scored['Team'] == team].copy()
        
        # Calculate team-relative metrics for this team's players
        if team_averages_dict:
            team_df = calculate_team_relative_metrics(team_df, team_averages_dict)
        
        ws = wb.create_sheet(title=team)
        
        # Headers - start with base columns, then add metric columns
        base_headers = ['Player', 'Position', 'Position Profile', 'Conference Grade', 'Power Five Grade', '2025 Total Score', 
                       'Previous Year', 'Previous Score', 'Change From Previous', 'Total Minutes', '% of Team Minutes', 'Top 15s (Power Five)']
        
        # Get all unique relevant metrics across all position profiles in this team
        exclude_cols = ['Player', 'Team', 'Position', 'Total_Score_1_10', 'Previous Year', 'Previous Score', 'Change From Previous', 'Previous Position',
                         'Total_Grade', 'Team_Grade', 'Conference_Grade', 'Power_Five_Grade',
                        'Total_Percentile', 'Position_Profile', 'Year', 'Season', 'Conference', 
                        'Minutes played', 'Pct_Of_Team_Minutes', 'Top_15s']
        
        # Collect all relevant metrics from all position profiles in this team
        all_relevant_metrics = set()
        for position_profile in team_df['Position_Profile'].unique():
            if position_profile in position_profile_to_metrics:
                all_relevant_metrics.update(position_profile_to_metrics[position_profile])
        
        # Filter to only metrics that exist in dataframe and are relevant
        # Include both regular metrics and PAdj metrics
        all_metric_cols = [col for col in team_df.columns if col not in exclude_cols and not col.endswith('_vs_team')]
        
        # Separate PAdj metrics from regular metrics
        padj_cols = [col for col in all_metric_cols if 'PAdj' in col]
        regular_cols = [col for col in all_metric_cols if 'PAdj' not in col]
        
        if all_relevant_metrics:
            # Filter regular metrics to relevant ones
            metric_cols = [col for col in regular_cols if col in all_relevant_metrics]
            # Include PAdj versions of relevant metrics
            for col in regular_cols:
                if col in all_relevant_metrics:
                    padj_version = f"PAdj {col}"
                    if padj_version in padj_cols:
                        metric_cols.append(padj_version)
        else:
            # Fallback: include all metrics if no config (including PAdj)
            metric_cols = all_metric_cols
        
        # Automatically group metrics by their base name (same logic as position profile sheets)
        metric_groups = {}
        
        def get_base_name(metric_name):
            """Extract base name from metric for grouping"""
            base = metric_name.lower()
            # Remove PAdj prefix first
            if base.startswith('padj '):
                base = base[5:]  # Remove "padj " prefix
            base = base.replace(' per 90', '')
            base = base.replace(' won, %', '').replace(' won %', '').replace(' won', '')
            base = base.replace('accurate, %', '').replace('accurate %', '').replace('accurate ', '').replace('accurate', '')
            base = ' '.join(base.split())  # Clean up double spaces
            base = base.replace(' successful, %', '').replace(' successful %', '').replace(' successful', '')
            base = base.replace(', %', '').replace(' %', '')
            base = base.strip()
            if base:
                words = base.split()
                base = ' '.join(word.capitalize() for word in words)
            return base
        
        # For team tabs, only include specific metrics
        allowed_metrics = [
            'Defensive duels won, %',
            'Aerial duels won, %',
            'Offensive duels won, %',
            'Accurate passes, %',
            'Accurate long passes, %',
            'Accurate progressive passes, %'
        ]
        
        # Normalize metric names for comparison (case-insensitive, handle variations)
        def normalize_for_matching(name):
            """Normalize metric name for flexible matching"""
            return str(name).lower().replace(',', '').replace('  ', ' ').strip()
        
        normalized_allowed = {normalize_for_matching(m): m for m in allowed_metrics}
        
        # Group metrics by base name
        for metric in metric_cols:
            # Check if this metric is in the allowed list
            normalized_metric = normalize_for_matching(metric)
            is_allowed = False
            
            # Check exact match or partial match
            for allowed_norm, allowed_orig in normalized_allowed.items():
                if normalized_metric == allowed_norm:
                    is_allowed = True
                    break
                # Also check if metric contains key words from allowed metrics
                if 'defensive duels won' in normalized_metric and 'defensive duels won' in allowed_norm:
                    is_allowed = True
                    break
                elif 'aerial duels won' in normalized_metric and 'aerial duels won' in allowed_norm:
                    is_allowed = True
                    break
                elif 'offensive duels won' in normalized_metric and 'offensive duels won' in allowed_norm:
                    is_allowed = True
                    break
                elif 'accurate passes' in normalized_metric and 'accurate passes' in allowed_norm and 'long' not in normalized_metric and 'progressive' not in normalized_metric:
                    is_allowed = True
                    break
                elif 'accurate long passes' in normalized_metric and 'accurate long passes' in allowed_norm:
                    is_allowed = True
                    break
                elif 'accurate progressive passes' in normalized_metric and 'accurate progressive passes' in allowed_norm:
                    is_allowed = True
                    break
            
            if not is_allowed:
                continue  # Skip metrics not in the allowed list
            
            base_name = get_base_name(metric)
            
            # Check if this metric has a team-relative column
            team_rel_col = f"{metric}_vs_team"
            has_team_relative = team_rel_col in team_df.columns
            
            
            # For team tabs, require team-relative data
            if not has_team_relative:
                continue  # Skip this metric entirely for team tabs
            
            # Determine sub-headers - show relative values only
            sub_headers = ['% better than team', '% better than position']
            
            # Add to group
            if base_name not in metric_groups:
                metric_groups[base_name] = {}
            metric_groups[base_name][metric] = sub_headers
        
        # For team tabs, only include groups with team-relative columns
        # Filter out any groups that don't have team-relative data
        metric_groups_ordered = {}
        for group_name, metrics_dict in metric_groups.items():
            # Only include groups that have at least one metric with team-relative
            has_team_relative = False
            filtered_metrics = {}
            for metric_name, sub_headers in metrics_dict.items():
                if '% better than team' in sub_headers:
                    has_team_relative = True
                    filtered_metrics[metric_name] = sub_headers
            
            if has_team_relative and filtered_metrics:
                metric_groups_ordered[group_name] = filtered_metrics
        
        # Row 1: Write base headers and group headers (merged)
        current_col = 1
        for header in base_headers:
            cell = ws.cell(row=1, column=current_col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Merge base headers to span rows 1-2 (no duplicate in row 2)
            # Note: Merged cells automatically hide the duplicate - no need to write to row 2
            ws.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
            current_col += 1
        
        # Write metric group headers (row 1, merged)
        # Use the ordered groups (team-relative first)
        for group_name, metrics_dict in metric_groups_ordered.items():
            # Count total columns for this group
            total_cols = sum(len(sub_headers) for sub_headers in metrics_dict.values())
            start_col = current_col
            end_col = current_col + total_cols - 1
            
            # Generate a more descriptive header name based on the actual metric
            # For team tabs, we want to show what's actually being compared
            def get_display_name(metric_name):
                """Create a descriptive display name for the metric"""
                name = str(metric_name)
                
                # Map to more descriptive names
                if 'defensive duels won' in name.lower():
                    return 'Defensive Duels Won %'
                elif 'aerial duels won' in name.lower():
                    return 'Aerial Duels Won %'
                elif 'offensive duels won' in name.lower():
                    return 'Offensive Duels Won %'
                elif 'accurate crosses from right flank' in name.lower():
                    return 'Crosses From Right Flank Accurate %'
                elif 'accurate short / medium passes' in name.lower() or 'accurate short/medium passes' in name.lower():
                    return 'Short / Medium Passes Accurate %'
                elif 'accurate long passes' in name.lower():
                    return 'Long Passes Accurate %'
                elif 'accurate smart passes' in name.lower():
                    return 'Smart Passes Accurate %'
                elif 'accurate through passes' in name.lower():
                    return 'Through Passes Accurate %'
                elif 'accurate progressive passes' in name.lower():
                    return 'Progressive Passes Accurate %'
                elif 'accurate passes' in name.lower() and '%' in name:
                    return 'Passes Accurate %'
                else:
                    # Fallback: capitalize and format the metric name
                    name = name.replace('accurate, %', 'Accurate %')
                    name = name.replace('accurate %', 'Accurate %')
                    name = name.replace(' won, %', ' Won %')
                    name = name.replace(' won %', ' Won %')
                    # Capitalize words
                    words = name.split()
                    words = [w.capitalize() for w in words]
                    return ' '.join(words)
            
            # Get the first (and typically only) metric in this group for team tabs
            first_metric = list(metrics_dict.keys())[0] if metrics_dict else group_name
            display_name = get_display_name(first_metric)
            
            # Merge cells for group header
            if start_col < end_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            
            cell = ws.cell(row=1, column=start_col, value=display_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Row 2: Write sub-headers for each metric
            sub_col = start_col
            for metric_name, sub_headers in metrics_dict.items():
                for sub_header in sub_headers:
                    cell = ws.cell(row=2, column=sub_col, value=sub_header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sub_col += 1
            
            current_col = end_col + 1
        
        # Write data (grouped by position profile, sorted by grade)
        # Data rows start at row 3 (rows 1-2 are headers)
        row = 3
        position_profiles_list = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
        for position_profile in position_profiles_list:
            pos_df = team_df[team_df['Position_Profile'] == position_profile].copy()
            
            # Write position profile header (always show, even if no players)
            ws.cell(row=row, column=1, value=f"{position_profile} Players").font = Font(bold=True, size=12)
            row += 1
            
            if len(pos_df) == 0:
                # Show "No players found" message if section is empty
                ws.cell(row=row, column=1, value="No players found").font = Font(italic=True)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                row += 2  # Add extra spacing before next section
                continue
            
            # Deduplicate players: if a player appears multiple times, keep the first occurrence
            # (This can happen if a player appears in multiple position files)
            before_dedup = len(pos_df)
            pos_df = pos_df.drop_duplicates(subset=['Player'], keep='first')
            after_dedup = len(pos_df)
            if before_dedup > after_dedup:
                print(f"  🔍 Removed {before_dedup - after_dedup} duplicate player entries from {team} - {position_profile}")
            
            # Sort by Power Five Grade (A first, F last), then by score
            grade_order = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'F': 4}
            pos_df['_Grade_Sort'] = pos_df['Power_Five_Grade'].map(grade_order).fillna(999)
            pos_df = pos_df.sort_values(['_Grade_Sort', 'Total_Score_1_10'], ascending=[True, False])
            
            # Write players
            for idx, player_row in pos_df.iterrows():
                # Get Conference Grade value
                conf_grade_value = ''
                if 'Conference_Grade' in pos_df.columns:
                    try:
                        conf_grade_val = player_row.loc['Conference_Grade']
                        if pd.notna(conf_grade_val) and conf_grade_val is not None and str(conf_grade_val).strip():
                            conf_grade_value = str(conf_grade_val).strip()
                    except (KeyError, IndexError):
                        pass
                
                # Get Power Five Grade value
                power_five_grade_value = ''
                if 'Power_Five_Grade' in pos_df.columns:
                    try:
                        power_five_grade_val = player_row.loc['Power_Five_Grade']
                        if pd.notna(power_five_grade_val) and power_five_grade_val is not None and str(power_five_grade_val).strip():
                            power_five_grade_value = str(power_five_grade_val).strip()
                    except (KeyError, IndexError):
                        pass
                
                ws.cell(row=row, column=1, value=player_row.get('Player', ''))
                ws.cell(row=row, column=2, value=player_row.get('Position', ''))
                ws.cell(row=row, column=3, value=position_profile)
                ws.cell(row=row, column=4, value=conf_grade_value)  # Conference Grade
                ws.cell(row=row, column=5, value=power_five_grade_value)  # Power Five Grade
                # Write numeric columns with right alignment
                score_cell = ws.cell(row=row, column=6, value=player_row.get('Total_Score_1_10', ''))
                score_cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Previous Year (column 7)
                prev_year = player_row.get('Previous Year', '')
                prev_year_cell = ws.cell(row=row, column=7, value=prev_year)
                prev_year_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Previous Score (column 8)
                prev_score = player_row.get('Previous Score', '')
                if pd.notna(prev_score) and prev_score != '':
                    try:
                        prev_score_cell = ws.cell(row=row, column=8, value=round(float(prev_score), 2))
                    except:
                        prev_score_cell = ws.cell(row=row, column=8, value=prev_score)
                else:
                    prev_score_cell = ws.cell(row=row, column=8, value='')
                prev_score_cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Change From Previous (column 9)
                change = player_row.get('Change From Previous', '')
                if pd.notna(change) and change != '':
                    try:
                        change_cell = ws.cell(row=row, column=9, value=round(float(change), 2))
                    except:
                        change_cell = ws.cell(row=row, column=9, value=change)
                else:
                    change_cell = ws.cell(row=row, column=9, value='')
                change_cell.alignment = Alignment(horizontal='right', vertical='center')
                
                minutes_cell = ws.cell(row=row, column=10, value=player_row.get('Minutes played', ''))  # Total Minutes
                minutes_cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Write percentage of team minutes (format as percentage with 2 decimals, cap at 100%)
                pct_value = player_row.get('Pct_Of_Team_Minutes', None)
                if pd.notna(pct_value) and pct_value is not None:
                    pct_value = min(float(pct_value), 100.0)  # Cap at 100%
                    pct_cell = ws.cell(row=row, column=11, value=round(pct_value, 2))  # % of Team Minutes
                    pct_cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    pct_cell = ws.cell(row=row, column=11, value='')
                    pct_cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Write Top 15s (Power Five) (column 12)
                # Try multiple column name variations
                top15s_value = None
                for col_name in ['Top_15s', 'Top 15s (Power Five)', 'Top 15s', 'Top15s']:
                    if col_name in pos_df.columns:
                        try:
                            val = player_row.loc[col_name] if col_name in player_row.index else player_row.get(col_name, 0)
                            if pd.notna(val) and val != '':
                                top15s_value = val
                                break
                        except:
                            pass
                
                if top15s_value is None or pd.isna(top15s_value) or top15s_value == '':
                    top15s_value = 0
                top15s_cell = ws.cell(row=row, column=12, value=int(top15s_value))
                top15s_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Write metric columns in grouped format (starting after base headers)
                metric_col_idx = len(base_headers) + 1
                
                # Write data for each metric group (use ordered groups - team-relative first)
                for group_name, metrics_dict in metric_groups_ordered.items():
                    for metric_name, sub_headers in metrics_dict.items():
                        # Get the metric value and derived comparison values
                        metric_value = player_row.get(metric_name, '')
                        team_rel_col = f"{metric_name}_vs_team"
                        team_pos_rel_col = f"{metric_name}_vs_position"
                        team_rel_value = player_row.get(team_rel_col, None)
                        team_pos_rel_value = player_row.get(team_pos_rel_col, None)
                        team_avg_value = player_row.get(f"{metric_name}_team_avg", None)
                        team_pos_avg_value = player_row.get(f"{metric_name}_team_position_avg", None)
                        
                        # For team tabs, sub_headers = ['% better than team', '% better than position']
                        # For position profile tabs, sub_headers might have both metric and team-relative
                        # Write columns based on sub_headers
                        for sub_header in sub_headers:
                            if sub_header == '% better than team':
                                # Write team-relative value
                                if pd.notna(team_rel_value) and team_rel_value != '':
                                    try:
                                        team_rel_cell = ws.cell(row=row, column=metric_col_idx, value=round(float(team_rel_value), 1))
                                    except (ValueError, TypeError):
                                        team_rel_cell = ws.cell(row=row, column=metric_col_idx, value=team_rel_value)
                                else:
                                    team_rel_cell = ws.cell(row=row, column=metric_col_idx, value='')
                                team_rel_cell.alignment = Alignment(horizontal='right', vertical='center')
                                metric_col_idx += 1
                            elif sub_header == '% better than position':
                                # Write position-relative value
                                if pd.notna(team_pos_rel_value) and team_pos_rel_value != '':
                                    try:
                                        team_pos_rel_cell = ws.cell(row=row, column=metric_col_idx, value=round(float(team_pos_rel_value), 1))
                                    except (ValueError, TypeError):
                                        team_pos_rel_cell = ws.cell(row=row, column=metric_col_idx, value=team_pos_rel_value)
                                else:
                                    team_pos_rel_cell = ws.cell(row=row, column=metric_col_idx, value='')
                                team_pos_rel_cell.alignment = Alignment(horizontal='right', vertical='center')
                                metric_col_idx += 1
                            else:
                                # Write the main metric value (for position profile tabs)
                                if pd.notna(metric_value) and metric_value != '':
                                    if '%' in str(metric_name) or 'percent' in str(metric_name).lower():
                                        try:
                                            pct_val = float(metric_value)
                                            pct_val = min(pct_val, 100.0)  # Cap at 100%
                                            metric_cell = ws.cell(row=row, column=metric_col_idx, value=round(pct_val, 1))
                                        except (ValueError, TypeError):
                                            metric_cell = ws.cell(row=row, column=metric_col_idx, value=metric_value)
                                    else:
                                        try:
                                            metric_cell = ws.cell(row=row, column=metric_col_idx, value=round(float(metric_value), 2))
                                        except (ValueError, TypeError):
                                            metric_cell = ws.cell(row=row, column=metric_col_idx, value=metric_value)
                                else:
                                    metric_cell = ws.cell(row=row, column=metric_col_idx, value=0)
                                metric_cell.alignment = Alignment(horizontal='right', vertical='center')
                                metric_col_idx += 1
                
                row += 1
            
            # Add spacing after each position profile section
            row += 1
        
        # Apply conditional formatting
        # Data starts at row 3 (rows 1-2 are headers)
        if ws.max_row > 2:
            # Conference Grade column (D)
            conf_grade_range = f'D3:D{ws.max_row}'
            for grade, color in GRADE_COLORS.items():
                rule = CellIsRule(
                    operator='equal', formula=[f'"{grade}"'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
                ws.conditional_formatting.add(conf_grade_range, rule)
            
            # Power Five Grade column (E)
            power_five_grade_range = f'E3:E{ws.max_row}'
            for grade, color in GRADE_COLORS.items():
                rule = CellIsRule(
                    operator='equal', formula=[f'"{grade}"'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
                ws.conditional_formatting.add(power_five_grade_range, rule)
            
            # Percentage of team minutes column (discrete ranges) - now column J (10)
            # Data starts at row 3 (rows 1-2 are headers)
            pct_range = f'J3:J{ws.max_row}'
            
            # Percentage ranges with color scheme:
            # Dark red for highest %, dark blue for lowest %
            pct_ranges = [
                (0, 20, '1F4E79'),      # 0-20%: Dark blue (very low usage)
                (20, 40, '8FAADC'),    # 20-40%: Light blue (low usage)
                (40, 60, 'F2A2A2'),    # 40-60%: Light red (medium usage)
                (60, 80, 'C5504B'),    # 60-80%: Medium red (high usage)
                (80, 100, '8B0000')    # 80-100%: Dark red (very high usage)
            ]
            
            for min_pct, max_pct, color in pct_ranges:
                if min_pct == 0:
                    rule = CellIsRule(
                        operator='lessThanOrEqual', formula=[f'{max_pct}'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                    )
                elif max_pct >= 100:
                    rule = CellIsRule(
                        operator='greaterThanOrEqual', formula=[f'{min_pct}'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                    )
                else:
                    rule = CellIsRule(
                        operator='between', formula=[f'{min_pct}', f'{max_pct}'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                    )
                ws.conditional_formatting.add(pct_range, rule)
        
        # Freeze the first two rows (headers)
        ws.freeze_panes = 'A3'
        
        # Auto-adjust column widths
        total_cols = len(base_headers)
        for group_name, metrics_dict in metric_groups_ordered.items():
            total_cols += sum(len(sub_headers) for sub_headers in metrics_dict.values())
        
        for col_idx in range(1, total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18


def convert_excel_to_pdf(excel_file, output_dir):
    """
    Convert Excel file to PDF using LibreOffice headless mode (cross-platform).
    Falls back to alternative methods if LibreOffice is not available.
    
    Args:
        excel_file: Path to Excel file
        output_dir: Directory where PDF should be saved
    
    Returns:
        Path to PDF file if successful, None otherwise
    """
    pdf_file = output_dir / f"{excel_file.stem}.pdf"
    
    # Method 1: Try LibreOffice (most reliable, cross-platform)
    libreoffice_paths = [
        'libreoffice',
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',  # macOS default
        '/opt/homebrew/bin/libreoffice',  # Homebrew on Apple Silicon
        '/usr/local/bin/libreoffice',  # Homebrew on Intel
        'soffice',  # Linux/Windows
    ]
    
    libreoffice_found = None
    for path in libreoffice_paths:
        if shutil.which(path) or Path(path).exists():
            libreoffice_found = path
            break
    
    if libreoffice_found:
        try:
            print(f"  🔧 Using LibreOffice to convert to PDF...")
            # LibreOffice headless conversion
            # The Excel file already has print settings configured
            cmd = [
                libreoffice_found,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(output_dir),
                str(excel_file)
            ]
            
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=120  # Increased timeout for large files
            )
            
            if result.returncode == 0:
                # LibreOffice saves PDF with same name in output directory
                expected_pdf = output_dir / f"{excel_file.stem}.pdf"
                if expected_pdf.exists():
                    print(f"  ✅ PDF conversion successful")
                    return expected_pdf
                else:
                    print(f"  ⚠️  PDF file not found at expected location: {expected_pdf}")
            else:
                error_msg = result.stderr or result.stdout
                print(f"  ⚠️  LibreOffice conversion failed")
                if error_msg:
                    print(f"     Error: {error_msg[:200]}")
        except subprocess.TimeoutExpired:
            print(f"  ⚠️  LibreOffice conversion timed out (file may be too large)")
        except Exception as e:
            print(f"  ⚠️  Error using LibreOffice: {e}")
    
    # If conversion failed, provide helpful instructions
    if not pdf_file.exists():
        print(f"  ⚠️  PDF conversion not available. LibreOffice not found.")
        print(f"  💡 To enable PDF conversion, install LibreOffice:")
        print(f"     macOS: brew install --cask libreoffice")
        print(f"     Linux: sudo apt-get install libreoffice")
        print(f"     Windows: Download from https://www.libreoffice.org/")
        print(f"")
        print(f"  📝 Excel file saved at: {excel_file}")
        print(f"  💡 Alternative: You can manually convert the Excel file to PDF using:")
        print(f"     - Excel: File → Export → Create PDF/XPS")
        print(f"     - Google Sheets: File → Download → PDF")
        print(f"     - Online converters (e.g., ilovepdf.com, smallpdf.com)")
        return None
    
    return pdf_file


def create_data_notes_sheet(wb, conference):
    """Create Data Summary sheet with comprehensive explanations."""
    ws = wb.create_sheet(title="Data Summary", index=0)
    
    # Get teams for this conference
    if conference == "ACC":
        teams = ["Duke Blue Devils", "Notre Dame Fighting Irish", "Stanford Cardinal", "Virginia Cavaliers"]
    else:
        teams = ["Alabama Crimson Tide", "Arkansas Razorbacks", "Georgia Bulldogs", "Kentucky Wildcats", 
                 "LSU Tigers", "Mississippi St. Bulldogs", "Tennessee Vols", "Vanderbilt Commodores"]
    
    notes = [
        ["DATA SUMMARY", ""],
        ["", ""],
        ["REPORT OVERVIEW", ""],
        ["Conference:", conference],
        ["Data Source:", "Intra-Conference Games Only (2025 Season)"],
        ["Historical Reference:", "2021-2025 Intra-Conference Data from all Power Five conferences"],
        ["Teams Included:", f"{len(teams)} teams: {', '.join(teams)}"],
        ["", ""],
        ["POSITION PROFILES", ""],
        ["• Hybrid CB: Center Back position - Defensive and ball-playing abilities", ""],
        ["• DM Box-To-Box: Centre Midfielder position - Defensive and creative midfield play", ""],
        ["• AM Advanced Playmaker: Attacking Midfielder position - Creative attacking play", ""],
        ["• Right Touchline Winger: Winger position - Wide attacking play from right side", ""],
        ["", ""],
        ["UNDERSTANDING THE COLUMNS", ""],
        ["", ""],
        ["Player Information:", ""],
        ["• Player: Player's name", ""],
        ["• Team: Player's team name", ""],
        ["• Position: Player's specific position(s) on the field (e.g., CB, LCB, RCB)", ""],
        ["• Position Profile: The position profile category used for scoring (on team tabs only)", ""],
        ["", ""],
        ["Grades:", ""],
        ["• Conference Grade: Letter grade (A, B, C, D, F) comparing the player to others in the", ""],
        ["  same position profile within their conference. See 'How Grades Are Calculated' below.", ""],
        ["• Power Five Grade: Letter grade comparing the player to ALL Power Five players", ""],
        ["  (ACC, SEC, BIG10, BIG12, IVY) from 2021-2025 in the same position profile.", ""],
        ["  See 'How Grades Are Calculated' below for details.", ""],
        ["", ""],
        ["Scores:", ""],
        ["• 2025 Total Score: A numerical score (1-10 scale) representing overall performance", ""],
        ["  for the 2025 season. This score is calculated using weighted metrics specific to", ""],
        ["  each position profile, normalized against historical data (2021-2025).", ""],
        ["• Previous Year: The year of the player's most recent previous season (2021-2024).", ""],
        ["  Shows 'Rookie' if 2025 is their first season.", ""],
        ["• Previous Score: The player's total score from their most recent previous season.", ""],
        ["  Looks at all years 2021-2024, not just 2024, to find the most recent score.", ""],
        ["  See 'Missing Data' section for why some players don't have this value.", ""],
        ["• Change From Previous: The difference between 2025 and previous scores. Positive numbers", ""],
        ["  indicate improvement, negative numbers indicate decline.", ""],
        ["", ""],
        ["Minutes Played:", ""],
        ["• Total Minutes: The total number of minutes the player has played in 2025", ""],
        ["  intra-conference games.", ""],
        ["• % of Team Minutes: The percentage of the team's total minutes that the player", ""],
        ["  has played. This shows how much the player has been used by their team.", ""],
        ["  Color coding: Dark blue (0-20%), Light blue (20-40%), Light red (40-60%),", ""],
        ["  Medium red (60-80%), Dark red (80-100%).", ""],
        ["", ""],
        ["Performance Metrics:", ""],
        ["• Position Profile Tabs: Each position profile tab shows only basic player information", ""],
        ["  (columns A-J: Player, Position, Position Profile, Grades, Scores, Minutes).", ""],
        ["  These tabs focus on overall performance scores rather than individual metrics.", ""],
        ["", ""],
        ["• Team Tabs: Each team tab shows players grouped by position profile, with", ""],
        ["  team-relative performance metrics. The following 6 metrics are displayed:", ""],
        ["  - Defensive Duels Won %", ""],
        ["  - Aerial Duels Won %", ""],
        ["  - Offensive Duels Won %", ""],
        ["  - Passes Accurate %", ""],
        ["  - Long Passes Accurate %", ""],
        ["  - Progressive Passes Accurate %", ""],
        ["", ""],
        ["• % better than team: This column shows how much better (or worse) a player", ""],
        ["  performs compared to their team's average for that metric. For example,", ""],
        ["  a value of 17.5% means the player is 17.5% better than the team average.", ""],
        ["  Negative values indicate the player performs below the team average.", ""],
        ["  This helps identify standout players who excel despite their team's playing style.", ""],
        ["", ""],
        ["• % better than position: This column shows how much better (or worse) a player", ""],
        ["  performs compared to the average of players in the same position on their team.", ""],
        ["  For example, a value of 10.0% means the player is 10.0% better than the", ""],
        ["  average of other players in their position on the same team.", ""],
        ["  This helps identify the best players within each position group on a team.", ""],
        ["", ""],
        ["• PAdj (Possession-Adjusted) Metrics: Some metrics are adjusted for team", ""],
        ["  possession style. This accounts for the fact that players on high-possession", ""],
        ["  teams may have inflated stats, while players on low-possession teams may have", ""],
        ["  deflated stats. PAdj metrics normalize for this difference.", ""],
        ["", ""],
        ["• Metrics ending in 'per 90' are normalized to show performance per 90 minutes", ""],
        ["  of playing time, allowing fair comparison between players with different minutes.", ""],
        ["• Metrics ending in '%' show percentage success rates (e.g., pass accuracy,", ""],
        ["  duel success rate).", ""],
        ["• Empty cells in metric columns are filled with 0 to indicate no data available.", ""],
        ["", ""],
        ["HOW GRADES ARE CALCULATED", ""],
        ["", ""],
        ["All grades use percentile-based ranking, meaning players are graded based on", ""],
        ["where they rank compared to other players, not on absolute score thresholds.", ""],
        ["", ""],
        ["Conference Grade:", ""],
        ["• Compares each player against all other players in the same position profile", ""],
        ["  within their conference (2025 data only).", ""],
        ["• Uses the player's 2025 Total Score to determine their percentile rank.", ""],
        ["• Grade thresholds:", ""],
        ["  - A = 90th percentile or higher (top 10% of players)", ""],
        ["  - B = 80th to 89th percentile (top 11-20% of players)", ""],
        ["  - C = 70th to 79th percentile (top 21-30% of players)", ""],
        ["  - D = 60th to 69th percentile (top 31-40% of players)", ""],
        ["  - F = Below 60th percentile (bottom 60% of players)", ""],
        ["", ""],
        ["Power Five Grade:", ""],
        ["• Compares each player against ALL Power Five players (ACC, SEC, BIG10, BIG12, IVY)", ""],
        ["  from 2021-2025 in the same position profile.", ""],
        ["• Uses the same percentile thresholds as Conference Grade.", ""],
        ["• Note: Power Five Grade typically shows more A's than Conference Grade because:", ""],
        ["  - Larger comparison pool (5 conferences × 5 years = more data points)", ""],
        ["  - Stronger conferences like SEC and ACC tend to rank higher when compared", ""],
        ["    against all Power Five conferences, not just their own conference.", ""],
        ["", ""],
        ["HOW SCORES ARE CALCULATED", ""],
        ["", ""],
        ["The 2025 Total Score is calculated using a multi-step process:", ""],
        ["", ""],
        ["1. Position-Specific Metrics:", ""],
        ["   Each position profile has a specific set of metrics that are weighted according", ""],
        ["   to their importance for that position. For example, Center Backs are evaluated", ""],
        ["   more heavily on defensive metrics, while Attacking Midfielders are evaluated", ""],
        ["   more on creative and offensive metrics.", ""],
        ["", ""],
        ["2. Historical Normalization:", ""],
        ["   Each metric is normalized against historical data from 2021-2025. This means", ""],
        ["   a player's performance is compared to how players have performed in that metric", ""],
        ["   over the past 5 years, not just in the current season.", ""],
        ["", ""],
        ["3. Weighted Scoring:", ""],
        ["   The metrics are combined using specific weights that reflect their importance", ""],
        ["   for each position profile. These weights are applied to create a final score", ""],
        ["   on a 1-10 scale. The current model uses a 70/30 weighting (70% intent,", ""],
        ["   30% accuracy), meaning players are rewarded more for attempting actions", ""],
        ["   (passes, dribbles, duels) than for their success rate.", ""],
        ["", ""],
        ["4. Possession-Adjusted (PAdj) Metrics:", ""],
        ["   Where applicable, metrics are adjusted for team possession style. This", ""],
        ["   ensures fair comparison between players on teams with different playing", ""],
        ["   styles (high-possession vs. low-possession teams).", ""],
        ["", ""],
        ["5. Important Note - Minutes Not Weighted:", ""],
        ["   The scores and grades are NOT weighted by minutes played. This means a player", ""],
        ["   who played 100 minutes will be evaluated the same way as a player who played", ""],
        ["   1000 minutes, as long as they meet the minimum threshold for inclusion.", ""],
        ["   This ensures that players with limited playing time are not penalized, but", ""],
        ["   also means that volume of play is not considered in the scoring system.", ""],
        ["", ""],
        ["MISSING DATA", ""],
        ["", ""],
        ["Previous Score:", ""],
        ["• Missing values in the 'Previous Score' column are expected and can occur for:", ""],
        ["  - Freshmen/rookies (first year players)", ""],
        ["  - Players who only played non-conference games in previous years (this report uses", ""],
        ["    intra-conference data only, so non-conference games are not included)", ""],
        ["  - Transfers from other conferences or teams not included in this report", ""],
        ["  - Players who took a year off or were injured for entire previous seasons", ""],
        ["  - Players who didn't meet minimum playing time in previous years (2021-2024)", ""],
        ["• When the Previous Score is missing, the 'Change From Previous' column will also", ""],
        ["  be blank, as we cannot calculate the change without a previous baseline.", ""],
        ["• Note: Further verification would be needed to determine the exact reason for", ""],
        ["  each missing value on a player-by-player basis.", ""],
        ["", ""],
        ["Players with No Metrics:", ""],
        ["• Players who have no recorded data for any of the relevant metrics for their", ""],
        ["  position profile are excluded from the position profile tabs. This ensures that", ""],
        ["  only players with meaningful data are included in the analysis.", ""],
        ["• These players may appear in team tabs if they have other position profiles,", ""],
        ["  but will be excluded from position-specific tabs where they have no data.", ""],
        ["", ""],
        ["Empty Metric Cells:", ""],
        ["• If a player has a metric column that is empty (no data recorded), it is filled", ""],
        ["  with 0 in the report. This distinguishes between 'no data' (0) and actual", ""],
        ["  zero performance (which would also show as 0).", ""],
        ["", ""],
        ["DATA FILTERING", ""],
        ["", ""],
        ["• Only players from championship game teams are included in this report.", ""],
        ["• Only intra-conference games from the 2025 season are included.", ""],
        ["• Players are filtered by position to ensure they match the position profile", ""],
        ["  requirements (e.g., Center Backs must have played as CB, LCB, or RCB).", ""],
        ["• Players must have at least one relevant metric populated to be included.", ""],
        ["", ""],
        ["REPORT STRUCTURE", ""],
        ["", ""],
        ["Position Profile Tabs:", ""],
        ["• Each position profile (Hybrid CB, DM Box-To-Box, AM Advanced Playmaker,", ""],
        ["  Right Touchline Winger) has its own tab showing all players in that profile.", ""],
        ["• Players are sorted by 2025 Total Score (highest to lowest).", ""],
        ["• Only metrics relevant to that position profile are shown.", ""],
        ["", ""],
        ["Team Tabs:", ""],
        ["• Each championship team has its own tab showing all players from that team.", ""],
        ["• Players are grouped by position profile within each team tab.", ""],
        ["• All metrics relevant to each position profile are shown, so different players", ""],
        ["  may have different metric columns depending on their position profile.", ""],
        ["", ""],
        ["ADDITIONAL NOTES", ""],
        ["", ""],
        ["• All metrics are normalized to 'per 90 minutes' where applicable, allowing fair", ""],
        ["  comparison between players regardless of playing time.", ""],
        ["• Percentages are displayed to 1 decimal place, other numerical values to", ""],
        ["  2 decimal places.", ""],
        ["• Grade colors: A = Dark Red, B = Medium Red, C = Light Red, D = Light Blue,", ""],
        ["  F = Dark Blue.", ""],
        ["• The scoring system is designed to identify players who excel in the metrics", ""],
        ["  most important for their position, regardless of team context or playing time.", ""],
        ["", ""],
    ]
    
    # Write notes
    for row_idx, (note, value) in enumerate(notes, 1):
        cell = ws.cell(row=row_idx, column=1, value=note)
        if row_idx == 1:  # Title row
            cell.font = Font(bold=True, size=14)
        elif note and note.isupper() and not note.startswith("•"):  # Section headers
            cell.font = Font(bold=True, size=12)
        elif note.startswith("•"):  # Bullet points
            pass  # Keep default font
        else:
            cell.font = Font(bold=True)
        
        if value:
            ws.cell(row=row_idx, column=2, value=value)
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 80
    
    # Set row heights for better readability
    for row in range(1, len(notes) + 1):
        ws.row_dimensions[row].height = 15


def main():
    """Main function."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    print("="*80)
    print("UPDATING MIKE NORRIS REPORTS WITH LATEST DATA")
    print("="*80)
    
    # Process all requested conferences
    conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
    for conference in conferences:
        create_report_for_conference(base_dir, conference)
    
    print("\n" + "="*80)
    print("✅ ALL REPORTS UPDATED")
    print("="*80)
    print("\nNext steps:")
    print("1. Run reorganize_team_sheets_by_position.py to organize team sheets")
    print("2. Run reorder_sheets_alphabetically.py to order sheets")


if __name__ == "__main__":
    main()

