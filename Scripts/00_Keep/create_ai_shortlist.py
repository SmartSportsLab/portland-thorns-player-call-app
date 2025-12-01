#!/usr/bin/env python3
"""
Create an "AI Shortlist" report with top 15 players per position profile.
Uses multi-dimensional analysis to select the best players based on:
- Total Score
- Consistency (metrics above/below average)
- Style Fits (Portland alignment)
- Top 15s (Power Five elite metrics)
- Progression (improvement from previous season)
- Grades
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import sys

# Add path for imports
sys.path.insert(0, str(Path(__file__).parent))

from create_top_15_report import load_all_players_from_reports

# Grade colors
GRADE_COLORS = {
    'A': '8B0000',  # Dark red
    'B': 'C5504B',  # Red
    'C': 'F2A2A2',  # Light red
    'D': '8FAADC',  # Light blue
    'F': '1F4E79'   # Dark blue
}

POSITION_PROFILE_MAP = {
    'Hybrid CB': 'Center Back',
    'DM Box-To-Box': 'Centre Midfielder',
    'AM Advanced Playmaker': 'Attacking Midfielder',
    'Right Touchline Winger': 'Winger'
}

def calculate_ai_score(player_row, position_profile):
    """
    Calculate a composite AI score based on multiple factors.
    Returns: (score, reasoning_dict)
    """
    reasoning = {
        'factors': [],
        'strengths': [],
        'concerns': []
    }
    
    score = 0.0
    
    # 1. Total Score (0-10 scale, weight: 30%)
    total_score = 0
    for col in ['Total_Score_1_10', '2025 Total Score', 'Total Score']:
        if col in player_row.index:
            val = player_row.get(col)
            if pd.notna(val) and val != '':
                try:
                    total_score = float(val)
                    break
                except:
                    pass
    
    score += total_score * 0.30
    if total_score >= 9.0:
        reasoning['strengths'].append(f"Elite total score ({total_score:.2f})")
    elif total_score >= 8.0:
        reasoning['factors'].append(f"Strong total score ({total_score:.2f})")
    elif total_score < 7.0:
        reasoning['concerns'].append(f"Lower total score ({total_score:.2f})")
    
    # 2. Consistency Score (0-100 scale, weight: 20%)
    consistency_score = player_row.get('Consistency Score', 0)
    if pd.notna(consistency_score):
        consistency_score = float(consistency_score)
        # Normalize to 0-10 scale
        consistency_normalized = consistency_score / 10.0
        score += consistency_normalized * 0.20
        
        metrics_below = player_row.get('Metrics Below Avg', 0)
        if pd.notna(metrics_below):
            if isinstance(metrics_below, str) and '/' in metrics_below:
                below_count = int(metrics_below.split('/')[0])
            else:
                below_count = int(metrics_below) if pd.notna(metrics_below) else 0
            
            if below_count == 0:
                reasoning['strengths'].append("No below-average metrics (perfect consistency)")
            elif below_count <= 2:
                reasoning['factors'].append(f"Only {below_count} below-average metrics")
            elif below_count >= 5:
                reasoning['concerns'].append(f"{below_count} below-average metrics")
    
    # 3. Style Fits (count, weight: 25%)
    style_fits = player_row.get('Style Fits', 0)
    if pd.notna(style_fits):
        style_fits = int(style_fits)
        # Normalize: 0 fits = 0, 1-2 fits = 5, 3+ fits = 10
        if style_fits == 0:
            style_score = 0
        elif style_fits <= 2:
            style_score = 5
        else:
            style_score = 10
        
        score += style_score * 0.25
        
        if style_fits >= 3:
            reasoning['strengths'].append(f"{style_fits} style fits (strong Portland alignment)")
        elif style_fits >= 1:
            reasoning['factors'].append(f"{style_fits} style fit(s)")
        else:
            reasoning['concerns'].append("No style fits (may not align with Portland's style)")
    
    # 4. Top 15s (Power Five) (count, weight: 15%)
    top15s = player_row.get('Top 15s (Power Five)', 0)
    if pd.notna(top15s):
        top15s = int(top15s)
        # Normalize: 0 = 0, 1-2 = 3, 3-4 = 7, 5+ = 10
        if top15s == 0:
            top15_score = 0
        elif top15s <= 2:
            top15_score = 3
        elif top15s <= 4:
            top15_score = 7
        else:
            top15_score = 10
        
        score += top15_score * 0.15
        
        if top15s >= 5:
            reasoning['strengths'].append(f"{top15s} Top 15s (elite in multiple metrics)")
        elif top15s >= 3:
            reasoning['factors'].append(f"{top15s} Top 15s")
        elif top15s == 0:
            reasoning['concerns'].append("No Top 15s (not elite in any metric)")
    
    # 5. Progression (Change From Previous, weight: 5%)
    change_from_prev = player_row.get('Change From Previous', None)
    if pd.notna(change_from_prev) and change_from_prev != '':
        try:
            change_val = float(change_from_prev)
            if change_val > 1.0:
                score += 0.5  # Bonus for significant improvement
                reasoning['strengths'].append(f"Strong progression (+{change_val:.2f})")
            elif change_val > 0:
                score += 0.25
                reasoning['factors'].append(f"Positive progression (+{change_val:.2f})")
            elif change_val < -1.0:
                score -= 0.5  # Penalty for significant regression
                reasoning['concerns'].append(f"Regression ({change_val:.2f})")
        except:
            pass
    
    # 6. Grades (weight: 5%)
    conf_grade = str(player_row.get('Conference Grade', '')).strip().upper()
    power_grade = str(player_row.get('Power Five Grade', '')).strip().upper()
    
    if conf_grade == 'A' and power_grade == 'A':
        score += 0.5
        reasoning['strengths'].append("Double A grades")
    elif conf_grade == 'A' or power_grade == 'A':
        score += 0.25
        reasoning['factors'].append(f"One A grade ({conf_grade}/{power_grade})")
    
    return score, reasoning

def select_top_15_players(df, position_profile):
    """
    Select top 15 players based on AI score and additional criteria.
    """
    if len(df) == 0:
        return pd.DataFrame()
    
    # Calculate AI score for each player
    ai_scores = []
    reasoning_list = []
    
    for idx, row in df.iterrows():
        score, reasoning = calculate_ai_score(row, position_profile)
        ai_scores.append(score)
        reasoning_list.append(reasoning)
    
    df['AI_Score'] = ai_scores
    df['AI_Reasoning'] = reasoning_list
    
    # Sort by AI score (descending)
    df_sorted = df.sort_values('AI_Score', ascending=False).reset_index(drop=True)
    
    # Select top 15
    top_15 = df_sorted.head(15).copy()
    
    return top_15

def format_reasoning(reasoning_dict):
    """Format reasoning dictionary into readable text."""
    parts = []
    
    if reasoning_dict['strengths']:
        parts.append("STRENGTHS: " + "; ".join(reasoning_dict['strengths']))
    
    if reasoning_dict['factors']:
        parts.append("FACTORS: " + "; ".join(reasoning_dict['factors']))
    
    if reasoning_dict['concerns']:
        parts.append("CONCERNS: " + "; ".join(reasoning_dict['concerns']))
    
    return " | ".join(parts)

def create_ai_shortlist_report(base_dir, output_file):
    """Create the AI Shortlist report."""
    base_dir = Path(base_dir)
    output_file = Path(output_file)
    
    print("="*70)
    print("AI SHORTLIST REPORT GENERATION")
    print("="*70)
    
    # Load shortlist data
    shortlist_file = base_dir / 'Portland Thorns 2025 Shortlist.xlsx'
    if not shortlist_file.exists():
        print(f"❌ Shortlist file not found: {shortlist_file}")
        return
    
    print(f"\n📊 Loading shortlist data...")
    wb_shortlist = load_workbook(shortlist_file, data_only=True)
    
    position_profiles = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    # Create output workbook
    wb_output = Workbook()
    wb_output.remove(wb_output.active)  # Remove default sheet
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for position_profile in position_profiles:
        print(f"\n📊 Analyzing {position_profile}...")
        
        if position_profile not in wb_shortlist.sheetnames:
            print(f"  ⚠️  Sheet not found: {position_profile}")
            continue
        
        ws_shortlist = wb_shortlist[position_profile]
        
        # Load data from shortlist sheet
        # Find headers (row 3)
        headers = []
        for col_idx in range(1, ws_shortlist.max_column + 1):
            header = ws_shortlist.cell(row=3, column=col_idx).value
            if header:
                headers.append(str(header))
            else:
                headers.append(f"Column_{col_idx}")
        
        # Load data rows
        data_rows = []
        for row_idx in range(4, ws_shortlist.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                value = ws_shortlist.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
            data_rows.append(row_data)
        
        if not data_rows:
            print(f"  ⚠️  No data found for {position_profile}")
            continue
        
        df = pd.DataFrame(data_rows)
        
        # Select top 15
        top_15 = select_top_15_players(df, position_profile)
        
        if len(top_15) == 0:
            print(f"  ⚠️  No players selected for {position_profile}")
            continue
        
        print(f"  ✅ Selected {len(top_15)} players")
        
        # Create sheet
        ws = wb_output.create_sheet(title=position_profile)
        
        # Write title
        ws.merge_cells('A1:Z1')
        title_cell = ws.cell(row=1, column=1, value=f"AI Shortlist: Top 15 {position_profile} Players")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write note
        ws.merge_cells('A2:Z2')
        note_cell = ws.cell(row=2, column=1, value="Selected based on: Total Score (30%), Consistency (20%), Style Fits (25%), Top 15s (15%), Progression (5%), Grades (5%)")
        note_cell.font = Font(italic=True, size=9)
        note_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        output_headers = [
            'Rank', 'Player', 'Team', 'Conference', 'Position',
            'Total Score', 'Conference Grade', 'Power Five Grade',
            'AI Score', 'Consistency Score', 'Metrics Above Avg', 'Metrics Below Avg',
            'Style Fits', 'Top 15s (Power Five)', 'Change From Previous',
            'AI Reasoning'
        ]
        
        for col_idx, header in enumerate(output_headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Write data
        for row_idx, (idx, player_row) in enumerate(top_15.iterrows(), 4):
            # Rank
            ws.cell(row=row_idx, column=1, value=row_idx - 3).border = border
            
            # Player
            ws.cell(row=row_idx, column=2, value=player_row.get('Player', '')).border = border
            
            # Team
            ws.cell(row=row_idx, column=3, value=player_row.get('Team', '')).border = border
            
            # Conference
            ws.cell(row=row_idx, column=4, value=player_row.get('Conference', '')).border = border
            
            # Position
            ws.cell(row=row_idx, column=5, value=player_row.get('Position', '')).border = border
            
            # Total Score
            score = player_row.get('Total_Score_1_10', player_row.get('2025 Total Score', player_row.get('Total Score', '')))
            if pd.notna(score) and score != '':
                try:
                    ws.cell(row=row_idx, column=6, value=round(float(score), 2)).border = border
                except:
                    ws.cell(row=row_idx, column=6, value=score).border = border
            
            # Conference Grade
            conf_grade = player_row.get('Conference Grade', '')
            conf_cell = ws.cell(row=row_idx, column=7, value=str(conf_grade) if pd.notna(conf_grade) else '')
            if conf_grade and str(conf_grade).strip().upper() in GRADE_COLORS:
                conf_cell.fill = PatternFill(start_color=GRADE_COLORS[str(conf_grade).strip().upper()], 
                                             end_color=GRADE_COLORS[str(conf_grade).strip().upper()], 
                                             fill_type='solid')
                conf_cell.font = Font(bold=True, color="FFFFFF")
            conf_cell.border = border
            
            # Power Five Grade
            power_grade = player_row.get('Power Five Grade', '')
            power_cell = ws.cell(row=row_idx, column=8, value=str(power_grade) if pd.notna(power_grade) else '')
            if power_grade and str(power_grade).strip().upper() in GRADE_COLORS:
                power_cell.fill = PatternFill(start_color=GRADE_COLORS[str(power_grade).strip().upper()], 
                                              end_color=GRADE_COLORS[str(power_grade).strip().upper()], 
                                              fill_type='solid')
                power_cell.font = Font(bold=True, color="FFFFFF")
            power_cell.border = border
            
            # AI Score
            ai_score = player_row.get('AI_Score', 0)
            ws.cell(row=row_idx, column=9, value=round(float(ai_score), 2) if pd.notna(ai_score) else 0).border = border
            
            # Consistency Score
            consistency = player_row.get('Consistency Score', 0)
            ws.cell(row=row_idx, column=10, value=round(float(consistency), 1) if pd.notna(consistency) else 0).border = border
            
            # Metrics Above Avg
            metrics_above = player_row.get('Metrics Above Avg', '0/0')
            ws.cell(row=row_idx, column=11, value=str(metrics_above)).border = border
            
            # Metrics Below Avg
            metrics_below = player_row.get('Metrics Below Avg', '0/0')
            ws.cell(row=row_idx, column=12, value=str(metrics_below)).border = border
            
            # Style Fits
            style_fits = player_row.get('Style Fits', 0)
            ws.cell(row=row_idx, column=13, value=int(style_fits) if pd.notna(style_fits) else 0).border = border
            
            # Top 15s
            top15s = player_row.get('Top 15s (Power Five)', 0)
            ws.cell(row=row_idx, column=14, value=int(top15s) if pd.notna(top15s) else 0).border = border
            
            # Change From Previous
            change = player_row.get('Change From Previous', '')
            ws.cell(row=row_idx, column=15, value=change if pd.notna(change) and change != '' else '').border = border
            
            # AI Reasoning
            reasoning = player_row.get('AI_Reasoning', {})
            reasoning_text = format_reasoning(reasoning) if isinstance(reasoning, dict) else str(reasoning)
            reasoning_cell = ws.cell(row=row_idx, column=16, value=reasoning_text)
            reasoning_cell.border = border
            reasoning_cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Set column widths
        column_widths = {
            'A': 8,   # Rank
            'B': 20,  # Player
            'C': 25,  # Team
            'D': 15,  # Conference
            'E': 20,  # Position
            'F': 12,  # Total Score
            'G': 15,  # Conference Grade
            'H': 15,  # Power Five Grade
            'I': 10,  # AI Score
            'J': 15,  # Consistency Score
            'K': 18,  # Metrics Above Avg
            'L': 18,  # Metrics Below Avg
            'M': 12,  # Style Fits
            'N': 18,  # Top 15s
            'O': 18,  # Change From Previous
            'P': 60   # AI Reasoning
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A4'
    
    wb_shortlist.close()
    
    # Save output
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb_output.save(output_file)
    
    print(f"\n✅ AI Shortlist report saved: {output_file}")
    print(f"\nSummary:")
    for position_profile in position_profiles:
        if position_profile in wb_output.sheetnames:
            ws = wb_output[position_profile]
            player_count = ws.max_row - 3  # Subtract header rows
            print(f"  - {position_profile}: {player_count} players selected")

if __name__ == '__main__':
    base_dir = Path('/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search')
    output_file = base_dir / 'AI Shortlist.xlsx'
    
    create_ai_shortlist_report(base_dir, output_file)

