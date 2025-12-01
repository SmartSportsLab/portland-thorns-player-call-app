#!/usr/bin/env python3
"""
Create Experience Rankings Report.
Shows top players by experience level (Rookies, Sophomores, Juniors, Seniors)
with dual comparison: Overall vs All Players, Within-Experience Class.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def load_2025_player_data(base_dir):
    """Load all 2025 player data with scores and experience."""
    
    # Position mappings
    position_map = {
        'AM Advanced Playmaker': 'Attacking Midfielder',
        'CB Hybrid': 'Center Back',
        'DM Box-To-Box': 'Centre Midfielder',
        'W Touchline Winger': 'Winger'
    }
    
    folder_map = {
        'Attacking Midfielder': '03_Advanced_Playmaker',
        'Center Back': '01_Hybrid_CB',
        'Centre Midfielder': '02_DM_Box_To_Box',
        'Winger': '04_Touchline_Winger'
    }
    
    conferences = ['ACC', 'BIG10', 'BIG12', 'IVY', 'SEC']
    
    print("📊 Loading 2025 player data with scores and experience...")
    
    # Load all 2025 data with scores
    all_players_2025 = []
    
    for pos_file_prefix, position_name in position_map.items():
        folder_name = folder_map[position_name]
        
        for conference in conferences:
            # Load Export file for NCAA_Seasons
            export_file = base_dir / "Exports" / f"{pos_file_prefix} {conference} 2025.xlsx"
            
            # Load By_Team file for scores
            by_team_file = base_dir / folder_name / f"{conference}_2025_{position_name.replace(' ', '_')}_Intent_Focused_By_Team.xlsx"
            
            if export_file.exists() and by_team_file.exists():
                try:
                    # Load scores
                    df_scores = pd.read_excel(by_team_file, sheet_name='All Players')
                    
                    # Load experience data
                    df_export = pd.read_excel(export_file, sheet_name=0)
                    
                    # Merge on Player name
                    df_merged = df_scores.merge(
                        df_export[['Player', 'NCAA_Seasons']],
                        on='Player',
                        how='left'
                    )
                    
                    # Add metadata
                    df_merged['Position'] = position_name
                    df_merged['Conference'] = conference
                    
                    all_players_2025.append(df_merged)
                except Exception as e:
                    print(f"  ⚠️  Error loading {conference} {position_name}: {e}")
    
    if not all_players_2025:
        print("  ⚠️  No 2025 data loaded")
        return None
    
    # Combine all
    df_all = pd.concat(all_players_2025, ignore_index=True)
    print(f"  ✅ Loaded {len(df_all)} players with scores and experience")
    
    return df_all

def create_experience_rankings_report(df_all, output_path):
    """Create experience rankings Excel report."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Experience Rankings"
    
    # Create rankings by experience level and position
    row = 1
    
    for exp_level, exp_name in [(1, 'Rookies'), (2, 'Sophomores'), (3, 'Juniors'), (4, 'Seniors')]:
        # Title
        ws.cell(row=row, column=1, value=f"TOP {exp_name.upper()} (NCAA_Seasons = {exp_level})")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        row += 1
        
        # Subdivide by position
        for position in sorted(df_all['Position'].unique()):
            ws.cell(row=row, column=1, value=f"  {position}")
            ws.cell(row=row, column=1).font = Font(size=12, bold=True, italic=True)
            row += 1
            
            # Get players at this experience level and position
            df_subset = df_all[
                (df_all['NCAA_Seasons'] == exp_level) & 
                (df_all['Position'] == position)
            ].copy()
            
            if len(df_subset) > 0:
                # Calculate percentile WITHIN this experience level
                df_subset['Percentile_Within_Exp'] = df_subset['Total_Score_1_10'].rank(pct=True) * 100
                
                # Assign grade based on percentile within experience level
                def assign_grade_within_exp(pct):
                    if pd.isna(pct):
                        return 'F'
                    elif pct >= 90:
                        return 'A'
                    elif pct >= 80:
                        return 'B'
                    elif pct >= 70:
                        return 'C'
                    elif pct >= 60:
                        return 'D'
                    else:
                        return 'F'
                
                df_subset['Grade_Within_Exp'] = df_subset['Percentile_Within_Exp'].apply(assign_grade_within_exp)
                
                # Sort by Total_Score descending
                df_subset = df_subset.sort_values('Total_Score_1_10', ascending=False)
                
                # Headers
                headers = ['Rank', 'Player', 'Team', 'Conference', 'NCAA_Seasons', 'Total_Score', 'Overall %ile', 'Overall Grade', 'Within Exp %ile', 'Within Exp Grade']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center")
                row += 1
                
                # Top 10 players
                for rank_idx, (idx, player_row) in enumerate(df_subset.head(10).iterrows(), 1):
                    ws.cell(row=row, column=1, value=rank_idx)
                    ws.cell(row=row, column=2, value=player_row['Player'])
                    ws.cell(row=row, column=3, value=player_row['Team'])
                    ws.cell(row=row, column=4, value=player_row['Conference'])
                    ws.cell(row=row, column=5, value=int(player_row['NCAA_Seasons']))
                    ws.cell(row=row, column=6, value=f"{player_row['Total_Score_1_10']:.2f}")
                    ws.cell(row=row, column=7, value=f"{player_row['Total_Percentile']:.1f}%")
                    ws.cell(row=row, column=8, value=player_row['Total_Grade'])
                    ws.cell(row=row, column=9, value=f"{player_row['Percentile_Within_Exp']:.1f}%")
                    ws.cell(row=row, column=10, value=player_row['Grade_Within_Exp'])
                    row += 1
            
            row += 1  # Blank line between positions
        
        row += 2  # Blank line between experience levels
    
    # Adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Remove default sheet if exists
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb['Sheet'])
    
    # Save workbook
    wb.save(output_path)

def main():
    """Main function."""
    
    base_dir = Path(__file__).parent.parent.parent
    output_path = base_dir / "Experience_Rankings_2025.xlsx"
    
    print("=" * 70)
    print("🏈 EXPERIENCE RANKINGS REPORT")
    print("=" * 70)
    
    # Load 2025 player data
    df_all = load_2025_player_data(base_dir)
    
    if df_all is not None and len(df_all) > 0:
        # Create report
        create_experience_rankings_report(df_all, output_path)
        print(f"\n✅ Report saved: {output_path}")
    else:
        print("\n⚠️  No player data found to analyze")
    
    print("=" * 70)

if __name__ == "__main__":
    main()



















