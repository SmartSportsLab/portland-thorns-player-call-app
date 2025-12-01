#!/usr/bin/env python3
"""
Remove conditional formatting from score columns in Mike Norris reports.
Keep only the grade column colors.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def remove_score_formatting(file_path):
    """Remove conditional formatting from score columns, keep grade column formatting."""
    print(f"📄 Processing: {file_path.name}")
    
    wb = load_workbook(file_path)
    
    # Define grade colors (keep these)
    colors = {
        'dark_blue': '1F4E79',
        'dark_red': '8B0000',
        'red': 'C5504B',
        'light_red': 'F2A2A2',
        'light_blue': '8FAADC'
    }
    
    grade_colors = {
        'A': colors['dark_red'],
        'B': colors['red'],
        'C': colors['light_red'],
        'D': colors['light_blue'],
        'F': colors['dark_blue']
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Processing sheet: {sheet_name}")
        
        # Find column indices
        score_2025_col = None
        score_2024_col = None
        grade_col = None
        
        # Check header row
        for col_idx, cell in enumerate(ws[1], 1):
            header_value = str(cell.value).strip() if cell.value else ""
            if '2025 Total Score' in header_value or 'Total Score 1 10' in header_value:
                score_2025_col = col_idx
            elif '2024 Total Score' in header_value:
                score_2024_col = col_idx
            elif 'Total Grade' in header_value or 'Total_Grade' in header_value:
                grade_col = col_idx
        
        # Clear all conditional formatting by removing all ranges
        # Store all ranges first
        all_ranges = []
        for cf in ws.conditional_formatting:
            all_ranges.append(str(cf.sqref))
        
        # Remove each range - we'll need to use a different approach
        # Instead, let's manually set white fill on score columns and keep grade formatting
        # First, clear score column fills
        if score_2025_col and ws.max_row > 1:
            score_2025_letter = get_column_letter(score_2025_col)
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{score_2025_letter}{row}']
                if cell.fill.start_color.index != 'FFFFFF':
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        if score_2024_col and ws.max_row > 1:
            score_2024_letter = get_column_letter(score_2024_col)
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{score_2024_letter}{row}']
                if cell.fill.start_color.index != 'FFFFFF':
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Now rebuild conditional formatting - clear all and re-add only grade
        # Get the type of conditional_formatting and create a new empty one
        cf_type = type(ws.conditional_formatting)
        ws.conditional_formatting = cf_type()
        
        # Re-apply only grade column formatting
        if grade_col and ws.max_row > 1:
            grade_col_letter = get_column_letter(grade_col)
            grade_range = f'{grade_col_letter}2:{grade_col_letter}{ws.max_row}'
            
            for grade, color in grade_colors.items():
                grade_rule = CellIsRule(
                    operator='equal', formula=[f'"{grade}"'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
                ws.conditional_formatting.add(grade_range, grade_rule)
            print(f"    ✅ Applied formatting to Grade column ({grade_col_letter})")
        
        # Score columns will have no formatting (white background)
        if score_2025_col:
            print(f"    ✅ Removed formatting from 2025 Total Score column")
        if score_2024_col:
            print(f"    ✅ Removed formatting from 2024 Total Score column")
    
    wb.save(file_path)
    print(f"✅ Saved: {file_path.name}\n")


def main():
    """Main function."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    # Find Mike Norris report files
    report_files = list(base_dir.glob("Mike_Norris_Scouting_Report_*_IntraConference.xlsx"))
    
    if not report_files:
        print("❌ No Mike Norris report files found")
        print(f"   Looking in: {base_dir}")
        return
    
    print("="*80)
    print("REMOVING SCORE COLUMN COLORS FROM MIKE NORRIS REPORTS")
    print("="*80)
    print(f"\nFound {len(report_files)} report file(s):\n")
    
    for file_path in report_files:
        remove_score_formatting(file_path)
    
    print("="*80)
    print("✅ ALL REPORTS UPDATED")
    print("="*80)
    print("\nScore columns now have white backgrounds.")
    print("Grade column colors are preserved.")


if __name__ == "__main__":
    main()

