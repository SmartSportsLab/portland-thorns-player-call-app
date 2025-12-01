#!/usr/bin/env python3
"""
Generate "Portland Thorns Team Fit Analysis" report.
Lists all players who fit Portland's playing style with their matching metrics.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import sys

# Add path for imports
sys.path.insert(0, str(Path(__file__).parent))

from generate_player_overviews import (
    load_nwsl_data, 
    load_player_data_from_shortlist,
    load_player_data_from_conference_reports,
    calculate_thorns_style_fit,
    format_metric_name
)

def generate_fit_report(base_dir, output_file):
    """Generate the Portland Thorns Team Fit Analysis report."""
    base_dir = Path(base_dir)
    output_file = Path(output_file)
    
    print("="*70)
    print("PORTLAND THORNS TEAM FIT ANALYSIS")
    print("="*70)
    
    # Load data
    print("\n📊 Loading data...")
    
    # Load players from shortlist
    shortlist_file = base_dir / 'Portland Thorns 2025 Shortlist.xlsx'
    players_df = load_player_data_from_shortlist(shortlist_file, base_dir)
    print(f"  ✅ Loaded {len(players_df)} players from shortlist")
    
    # Load all players for ranking
    all_players_df = load_player_data_from_conference_reports(base_dir)
    print(f"  ✅ Loaded {len(all_players_df)} total players for ranking")
    
    # Load NWSL data
    nwsl_dir = base_dir / 'Exports' / 'Team Stats By Conference' / 'NWSL'
    league_df, league_avg, thorns_data, thorns_ranks = load_nwsl_data(nwsl_dir)
    
    if thorns_ranks is None:
        print("  ❌ Could not load Portland Thorns ranks")
        return
    
    # Filter to top 3 metrics only
    top3_thorns_ranks = {k: v for k, v in thorns_ranks.items() if v['rank'] <= 3 and 'Unnamed' not in str(k)}
    print(f"  ✅ Portland ranks top 3 in {len(top3_thorns_ranks)} metrics")
    
    # Analyze each player for style fit
    print("\n📊 Analyzing players for style fit...")
    fit_data = []
    
    for idx, (_, player_row) in enumerate(players_df.iterrows()):
        player_name = str(player_row.get('Player', 'Unknown')).strip()
        team = str(player_row.get('Team', 'Unknown')).strip()
        conference = str(player_row.get('Conference', 'Unknown')).strip()
        position_profile = player_row.get('Position Profile', '')
        
        if not position_profile:
            continue
        
        # Load position configs
        config_file = base_dir / 'Scripts' / '00_Keep' / 'position_metrics_config.json'
        position_configs = {}
        if config_file.exists():
            with open(config_file, 'r') as f:
                position_configs = json.load(f)
        else:
            print(f"  ⚠️  Position config file not found: {config_file}")
            continue
        
        # Get style fits
        style_fits = calculate_thorns_style_fit(player_row, top3_thorns_ranks, all_players_df, position_profile, position_configs)
        
        if style_fits:
            # Style fits are already deduplicated by the function
            # Add a row for each matching metric
            for fit in style_fits:
                fit_data.append({
                    'Player': player_name,
                    'Team': team,
                    'Conference': conference,
                    'Position Profile': position_profile,
                    'Metric': format_metric_name(fit['metric']),
                    'Portland NWSL Rank': fit['thorns_nwsl_rank'],
                    'Player Position Rank': fit['player_ranks']
                })
    
    if not fit_data:
        print("  ⚠️  No players found with style fit")
        return
    
    # Create DataFrame
    fit_df = pd.DataFrame(fit_data)
    
    # Sort by player name, then by metric
    fit_df = fit_df.sort_values(['Player', 'Team', 'Metric'])
    
    print(f"  ✅ Found {len(fit_df)} style fit matches across {fit_df['Player'].nunique()} players")
    
    # Create Excel workbook
    print("\n📝 Creating Excel report...")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        fit_df.to_excel(writer, sheet_name='Style Fit Analysis', index=False)
        
        # Get workbook and worksheet
        wb = writer.book
        ws = writer.sheets['Style Fit Analysis']
        
        # Style the header row
        header_fill = PatternFill(start_color='D4AF37', end_color='D4AF37', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Set column widths
        column_widths = {
            'A': 20,  # Player
            'B': 25,  # Team
            'C': 15,  # Conference
            'D': 25,  # Position Profile
            'E': 30,  # Metric
            'F': 20,  # Portland NWSL Rank
            'G': 30   # Player Position Rank
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add summary sheet
        summary_data = []
        # Deduplicate by Player, Team, and Metric to count unique metrics per player
        unique_fits = fit_df.drop_duplicates(subset=['Player', 'Team', 'Metric'])
        player_summary = unique_fits.groupby(['Player', 'Team', 'Conference', 'Position Profile']).size().reset_index(name='Number of Style Fits')
        player_summary = player_summary.sort_values('Number of Style Fits', ascending=False)
        
        summary_data.append(['PORTLAND THORNS TEAM FIT ANALYSIS SUMMARY', '', '', '', ''])
        summary_data.append(['', '', '', '', ''])
        summary_data.append(['Total Players with Style Fit:', len(player_summary), '', '', ''])
        summary_data.append(['Total Style Fit Matches:', len(fit_df), '', '', ''])
        summary_data.append(['', '', '', '', ''])
        summary_data.append(['Portland Thorns Top 3 Metrics:', '', '', '', ''])
        
        for metric, info in sorted(top3_thorns_ranks.items(), key=lambda x: x[1]['rank']):
            summary_data.append(['', format_metric_name(metric), f"Rank #{info['rank']}/14", '', ''])
        
        summary_data.append(['', '', '', '', ''])
        summary_data.append(['Players by Number of Style Fits:', '', '', '', ''])
        summary_data.append(['Player', 'Team', 'Conference', 'Position', 'Style Fits'])
        
        for _, row in player_summary.iterrows():
            summary_data.append([
                row['Player'],
                row['Team'],
                row['Conference'],
                row['Position Profile'],
                row['Number of Style Fits']
            ])
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
        
        ws_summary = writer.sheets['Summary']
        
        # Style summary sheet
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A3'].font = Font(bold=True)
        ws_summary['A4'].font = Font(bold=True)
        ws_summary['A6'].font = Font(bold=True)
        ws_summary['A12'].font = Font(bold=True)
        
        # Header row for player list
        for cell in ws_summary[12]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Style player data rows
        for row in ws_summary.iter_rows(min_row=13, max_row=ws_summary.max_row):
            for cell in row:
                cell.border = border
        
        # Set column widths for summary
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 25
        ws_summary.column_dimensions['E'].width = 15
    
    print(f"\n✅ Report generated successfully!")
    print(f"📁 File saved to: {output_file}")
    print(f"\nSummary:")
    print(f"  - {len(player_summary)} players with style fit")
    print(f"  - {len(fit_df)} total style fit matches")
    print(f"  - Top player: {player_summary.iloc[0]['Player']} ({player_summary.iloc[0]['Team']}) with {player_summary.iloc[0]['Number of Style Fits']} fits")

if __name__ == '__main__':
    base_dir = Path('/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search')
    output_file = base_dir / 'Portland Thorns Team Fit Analysis.xlsx'
    
    generate_fit_report(base_dir, output_file)

