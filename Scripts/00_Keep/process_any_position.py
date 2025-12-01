#!/usr/bin/env python3
"""
Universal script to process any position across all conferences.
Usage: python process_any_position.py [position_name]

Example: 
  python process_any_position.py "Centre Midfielder"
  python process_any_position.py "Attacking Midfielder"
  python process_any_position.py "Winger"
"""

import pandas as pd
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

def add_conditional_formatting(file_path):
    """Add conditional formatting to Excel file for scoring columns."""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path)
    
    # Define color scheme: dark blue (worst) to dark red (best)
    colors = {
        'dark_blue': '1F4E79',    # Worst scores
        'blue': '4472C4',         # Low scores
        'light_blue': '8FAADC',   # Below average
        'white': 'FFFFFF',        # Average
        'light_red': 'F2A2A2',    # Above average
        'red': 'C5504B',          # High scores
        'dark_red': '8B0000'      # Best scores
    }
    
    # Grade-based colors
    grade_colors = {
        'A': colors['dark_red'],   # Dark red for A grades
        'B': colors['red'],        # Red for B grades
        'C': colors['light_red'],  # Light red for C grades
        'D': colors['light_blue'], # Light blue for D grades
        'F': colors['dark_blue']   # Dark blue for F grades
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Find column indices
        score_col = None
        percentile_col = None
        grade_col = None
        
        for col_idx, cell in enumerate(ws[1], 1):  # Check header row
            if cell.value == 'Total_Score_1_10':
                score_col = col_idx
            elif cell.value == 'Total_Percentile':
                percentile_col = col_idx
            elif cell.value == 'Total_Grade':
                grade_col = col_idx
        
        # No formatting for Total_Score_1_10 and Total_Percentile columns - keep them white
        
        # Apply conditional formatting to Total_Grade column
        if grade_col and ws.max_row > 1:
            for grade, color in grade_colors.items():
                grade_rule = CellIsRule(
                    operator='equal', formula=[f'"{grade}"'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
                ws.conditional_formatting.add(f'{chr(64+grade_col)}2:{chr(64+grade_col)}{ws.max_row}', grade_rule)
    
    wb.save(file_path)

def main():
    """Process any position."""
    
    # Parse command line arguments
    import argparse
    parser = argparse.ArgumentParser(description='Process position data for all conferences')
    parser.add_argument('position_name', help='Position name (e.g., "Centre Midfielder")')
    parser.add_argument('--game-type', choices=['all', 'inter-conference', 'non-conference'], 
                       default='inter-conference',
                       help='Game type to process (default: inter-conference)')
    args = parser.parse_args()
    
    position_name = args.position_name
    game_type = args.game_type
    
    # Configuration
    base_dir = Path(__file__).parent.parent.parent
    config_file = base_dir / "Scripts" / "00_Keep" / "position_metrics_config.json"
    exports_dir = base_dir / "Exports"
    historical_dir = exports_dir / "Past Seasons"
    
    conferences = ['ACC', 'BIG10', 'BIG12', 'IVY', 'SEC']
    
    print(f"🏈 Processing {position_name}")
    print(f"📊 Game Type: {game_type}")
    print("=" * 70)
    
    # Load config
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    if position_name not in config['position_profiles']:
        print(f"❌ Position '{position_name}' not found in config!")
        print(f"Available positions: {list(config['position_profiles'].keys())}")
        sys.exit(1)
    
    position_config = config['position_profiles'][position_name]
    
    # Determine file prefix based on position
    file_prefix = get_file_prefix(position_name, config['conferences'])
    
    # Pre-load ALL 2025 data for this position profile (across all conferences)
    # This will be used to normalize percentiles/grades across all 2025 players
    print(f"\n📊 Pre-loading ALL 2025 Data for {position_name}...")
    all_2025_data = []
    for conference in conferences:
        data_2025 = exports_dir / f"{file_prefix} {conference} 2025.xlsx"
        if data_2025.exists():
            df_conf = pd.read_excel(data_2025)
            df_conf['Year'] = 2025
            df_conf['Season'] = '2025'
            df_conf['Conference'] = conference
            df_conf['Position'] = position_name
            # Filter out players with no stats
            df_conf = filter_players_with_data(df_conf, original_count=len(df_conf))
            all_2025_data.append(df_conf)
            print(f"  ✅ Loaded {len(df_conf)} players from {conference} 2025")
    
    if all_2025_data:
        df_all_2025 = pd.concat(all_2025_data, ignore_index=True)
        print(f"\n📊 Total 2025 players across all conferences: {len(df_all_2025)}")
    else:
        print("  ❌ No 2025 data found!")
        sys.exit(1)
    
    # Process each conference
    for conference in conferences:
        print(f"\n{'='*70}")
        print(f"📊 Processing {conference}")
        print(f"{'='*70}")
        
        try:
            process_conference(
                conference, position_name, position_config, 
                file_prefix, exports_dir, historical_dir, base_dir,
                df_all_2025  # Pass the combined 2025 dataset
            )
        except Exception as e:
            print(f"❌ Error processing {conference}: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n{'='*70}")
    print(f"✅ {position_name} Processing Complete!")
    print(f"{'='*70}")
    
    # Move files to appropriate position folder
    move_to_position_folder(base_dir, position_name, conferences)
    
    # Create consolidated file with all conferences
    create_consolidated_file(base_dir, position_name, position_config, conferences)

def filter_players_with_data(df, original_count):
    """Filter out players with no meaningful stats."""
    # Identify numeric columns (excluding metadata)
    numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns
    numeric_cols = [c for c in numeric_cols if c not in ['Year', 'Season', 'Conference', 'Position']]
    
    # Keep players who have at least one non-zero, non-NaN value
    # Check that at least one column is not NaN AND not zero
    has_data = ((df[numeric_cols].notna()) & (df[numeric_cols] != 0)).any(axis=1)
    
    filtered_df = df[has_data].copy()
    
    return filtered_df

def get_relevant_columns(position_config):
    """Get list of columns relevant to this position profile."""
    relevant_cols = set()
    
    # Always include basic info and minutes played
    relevant_cols.update(['Player', 'Team', 'Minutes played', 'Matches played', 'NCAA_Seasons'])
    
    # Add columns from Core metrics
    for metric_name, metric_config in position_config['metrics']['Core'].items():
        if isinstance(metric_config, dict) and 'components' in metric_config:
            relevant_cols.update(metric_config['components'].keys())
        else:
            # Single metric
            relevant_cols.add(metric_name)
    
    # Add columns from Specific metrics
    for metric_name, metric_config in position_config['metrics']['Specific'].items():
        if isinstance(metric_config, dict) and 'components' in metric_config:
            relevant_cols.update(metric_config['components'].keys())
        else:
            # Single metric
            relevant_cols.add(metric_name)
    
    return list(relevant_cols)

def get_file_prefix(position_name, conferences_config):
    """Determine file prefix based on position name."""
    sample_conf = list(conferences_config.values())[0]
    if position_name in sample_conf['files']:
        filename = sample_conf['files'][position_name]
        # Extract prefix (everything before conference and year)
        if 'ACC' in filename:
            return filename.split(' ACC')[0]
    return None

def move_to_position_folder(base_dir, position_name, conferences):
    """Move generated files to the appropriate position folder."""
    position_folder_map = {
        'Centre Midfielder': '02_DM_Box_To_Box',
        'Attacking Midfielder': '03_Advanced_Playmaker',
        'Winger': '04_Touchline_Winger',
        'Center Back': '01_Hybrid_CB'
    }
    
    folder_name = position_folder_map.get(position_name)
    if not folder_name:
        print(f"⚠️  No folder mapping for {position_name}")
        return
    
    output_dir = base_dir / folder_name
    
    # Find and move files
    for conf in conferences:
        for file in base_dir.glob(f"{conf}_2025_*.xlsx"):
            if file.exists():
                file.rename(output_dir / file.name)
                print(f"  📁 Moved {file.name} to {folder_name}/")
    
    return output_dir

def create_consolidated_file(base_dir, position_name, position_config, conferences):
    """Create a consolidated file with all players from all conferences."""
    position_folder_map = {
        'Centre Midfielder': '02_DM_Box_To_Box',
        'Attacking Midfielder': '03_Advanced_Playmaker',
        'Winger': '04_Touchline_Winger',
        'Center Back': '01_Hybrid_CB'
    }
    
    folder_name = position_folder_map.get(position_name)
    if not folder_name:
        print(f"⚠️  No folder mapping for {position_name}")
        return
    
    output_dir = base_dir / folder_name
    
    # Load all conference files for Intent_Focused variant
    all_players = []
    
    for conference in conferences:
        # Try to find the Intent_Focused file
        pattern = f"{conference}_2025_{position_name.replace(' ', '_')}_Intent_Focused_By_Team.xlsx"
        file_path = output_dir / pattern
        
        if file_path.exists():
            df = pd.read_excel(file_path, sheet_name='All Players')
            df['Conference'] = conference
            all_players.append(df)
            print(f"  ✅ Loaded {len(df)} players from {conference}")
    
    if not all_players:
        print(f"  ⚠️  No data files found to consolidate")
        return
    
    # Combine all data
    combined_df = pd.concat(all_players, ignore_index=True)
    
    # Add Conference column after Team
    cols = list(combined_df.columns)
    if 'Conference' in cols:
        cols.remove('Conference')
        if 'Team' in cols:
            team_idx = cols.index('Team')
            cols.insert(team_idx + 1, 'Conference')
        combined_df = combined_df[cols]
    
    # Sort by grade and percentile
    grade_order = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'F': 4}
    combined_df['_Grade_Sort'] = combined_df['Total_Grade'].map(grade_order)
    combined_df = combined_df.sort_values(['_Grade_Sort', 'Total_Percentile'], ascending=[True, False])
    combined_df = combined_df.drop(columns=['_Grade_Sort'])
    
    # Save consolidated file
    output_filename = f"All_Conferences_2025_{position_name.replace(' ', '_')}_Intent_Focused.xlsx"
    output_path = output_dir / output_filename
    
    # Get relevant columns (same as individual files)
    relevant_cols = get_relevant_columns(position_config)
    relevant_cols.append('Conference')
    scoring_cols = ['Total_Score_1_10', 'Total_Percentile', 'Total_Grade']
    for col in scoring_cols:
        if col not in relevant_cols:
            relevant_cols.append(col)
    
    # Remove internal columns
    internal_cols_to_remove = ['Core_Score_Original', 'Specific_Score_Original', 'Total_Score_Original', 
                              'Year', 'Season', 'Position']
    relevant_cols = [col for col in relevant_cols if col not in internal_cols_to_remove]
    
    # Filter columns
    available_cols = [col for col in relevant_cols if col in combined_df.columns]
    
    # Remove Age if exists
    if 'Age' in available_cols:
        available_cols.remove('Age')
    
    # Reorder columns
    priority_cols = ['Player', 'Team', 'Conference', 'Total_Score_1_10', 'Total_Percentile', 'Total_Grade']
    remaining_cols = [col for col in available_cols if col not in priority_cols]
    remaining_cols_sorted = sorted(remaining_cols)
    
    final_cols = []
    for col in priority_cols:
        if col in available_cols:
            final_cols.append(col)
    final_cols.extend(remaining_cols_sorted)
    
    combined_df = combined_df[final_cols]
    
    # Save to Excel
    combined_df.to_excel(output_path, index=False, sheet_name='All Players')
    
    # Add conditional formatting to the consolidated file
    add_conditional_formatting(output_path)
    
    print(f"\n  ✅ Created consolidated file: {output_filename} ({len(combined_df)} total players)")

def process_conference(conference, position_name, position_config, file_prefix, exports_dir, historical_dir, base_dir, df_all_2025):
    """Process a single conference."""
    
    # Load historical data
    print(f"\n📚 Loading Historical Data (2021-2024)...")
    historical_years = [2021, 2022, 2023, 2024]
    historical_data = []
    
    for year in historical_years:
        hist_file = historical_dir / f"{file_prefix} {conference} {year}.xlsx"
        if hist_file.exists():
            df_hist = pd.read_excel(hist_file)
            df_hist['Year'] = year
            df_hist['Season'] = str(year)
            historical_data.append(df_hist)
            print(f"  ✅ Loaded {len(df_hist)} players from {year}")
    
    if historical_data:
        df_historical = pd.concat(historical_data, ignore_index=True)
        print(f"  📊 Total historical players: {len(df_historical)}")
    else:
        df_historical = pd.DataFrame()
        print(f"  ⚠️  No historical data found")
    
    # Load 2025 data for this conference
    print(f"\n📊 Loading 2025 Data for {conference}...")
    data_2025 = exports_dir / f"{file_prefix} {conference} 2025.xlsx"
    
    if not data_2025.exists():
        print(f"  ❌ {data_2025.name} not found")
        return
    
    df_2025 = pd.read_excel(data_2025)
    df_2025['Year'] = 2025
    df_2025['Season'] = '2025'
    df_2025['Conference'] = conference
    df_2025['Position'] = position_name
    
    # Filter out players with no stats (all numeric columns are NaN or zero)
    df_2025_filtered = filter_players_with_data(df_2025, original_count=len(df_2025))
    removed_count = len(df_2025) - len(df_2025_filtered)
    if removed_count > 0:
        print(f"  ⚠️  Removed {removed_count} players with no stats")
    df_2025 = df_2025_filtered
    
    print(f"  ✅ Loaded {len(df_2025)} players from 2025")
    
    # Combine historical + all 2025 for normalization (min/max ranges)
    # But percentiles/grades will be calculated against df_all_2025 only
    if not df_historical.empty:
        df_all_for_norm = pd.concat([df_historical, df_all_2025], ignore_index=True)
        print(f"\n🔢 Combined Dataset for normalization: {len(df_all_for_norm)} total players (historical + all 2025)")
    else:
        df_all_for_norm = df_all_2025
    
    # Calculate scores for both variants
    variants = [
        ('Intent_Focused', 0.80, 0.20),
        ('Balanced', 0.60, 0.40)
    ]
    
    for variant_name, pass_attempt_weight, pass_accuracy_weight in variants:
        print(f"\n  📋 {variant_name} Variant ({int(pass_attempt_weight*100)}/{int(pass_accuracy_weight*100)})")
        
        df_result = calculate_with_historical_normalization(
            df_2025, df_all_for_norm, position_config, pass_attempt_weight, pass_accuracy_weight, position_name,
            df_all_2025  # Pass all 2025 data for percentile/grade calculation
        )
        
        # Create team-by-team version
        output_filename = f"{conference}_2025_{position_name.replace(' ', '_')}_{variant_name}_By_Team.xlsx"
        create_team_tabs(df_result, conference, variant_name, position_name, output_filename, base_dir, position_config)
        
        print(f"  ✅ Complete")

def calculate_with_historical_normalization(df_2025, df_all, config, pass_attempt_weight, pass_accuracy_weight, position_name, df_all_2025=None):
    """Calculate scores normalizing against combined historical + current data.
    
    Args:
        df_2025: Current conference's 2025 data
        df_all: Historical + all 2025 data (for min/max normalization)
        config: Position metrics configuration
        pass_attempt_weight: Weight for pass attempts
        pass_accuracy_weight: Weight for pass accuracy
        position_name: Name of the position
        df_all_2025: All 2025 data across all conferences (for percentile/grade calculation)
    """
    
    df_work = df_2025.copy()
    
    # If df_all_2025 not provided, use df_all (backward compatibility)
    if df_all_2025 is None:
        df_all_2025 = df_all
    
    # Adjust pass weights
    for metric_name, metric_config in config['metrics']['Specific'].items():
        if isinstance(metric_config, dict) and 'components' in metric_config:
            if any('pass' in k.lower() and 'accurate' in k.lower() for k in metric_config['components'].keys()):
                new_components = {}
                for comp, weight in metric_config['components'].items():
                    if 'accurate' in comp.lower():
                        new_components[comp] = pass_accuracy_weight
                    else:
                        new_components[comp] = pass_attempt_weight
                total = sum(new_components.values())
                metric_config['components'] = {k: v/total for k, v in new_components.items()}
    
    # Calculate scores - now all metrics have equal weights, so just sum everything
    total_score = pd.Series(0.0, index=df_work.index)
    
    # Calculate all Core metrics
    core_score = pd.Series(0.0, index=df_work.index)
    for metric_name, metric_config in config['metrics']['Core'].items():
        if isinstance(metric_config, dict):
            weight = metric_config['weight']
            metric_score = calculate_combined_metric_historical(df_work, df_all, metric_config)
            core_score += metric_score * weight
    
    # Calculate all Specific metrics
    specific_score = pd.Series(0.0, index=df_work.index)
    for metric_name, metric_config in config['metrics']['Specific'].items():
        if isinstance(metric_config, dict):
            weight = metric_config['weight']
            metric_score = calculate_combined_metric_historical(df_work, df_all, metric_config)
            specific_score += metric_score * weight
    
    # Total score is sum of all metrics (they already have weights)
    total_score = core_score + specific_score
    
    # Add to dataframe
    df_work['Core_Score_Original'] = core_score
    df_work['Specific_Score_Original'] = specific_score
    df_work['Total_Score_Original'] = total_score
    
    # Calculate scores for ALL 2025 players to determine percentiles/grades
    # This ensures scores are comparable across all conferences
    print(f"    📊 Calculating scores for all {len(df_all_2025)} 2025 players for percentile normalization...")
    
    all_2025_scores = pd.Series(0.0, index=df_all_2025.index)
    all_2025_core = pd.Series(0.0, index=df_all_2025.index)
    all_2025_specific = pd.Series(0.0, index=df_all_2025.index)
    
    for metric_name, metric_config in config['metrics']['Core'].items():
        if isinstance(metric_config, dict):
            weight = metric_config['weight']
            metric_score = calculate_combined_metric_historical(df_all_2025, df_all, metric_config)
            all_2025_core += metric_score * weight
    
    for metric_name, metric_config in config['metrics']['Specific'].items():
        if isinstance(metric_config, dict):
            weight = metric_config['weight']
            metric_score = calculate_combined_metric_historical(df_all_2025, df_all, metric_config)
            all_2025_specific += metric_score * weight
    
    all_2025_scores = all_2025_core + all_2025_specific
    
    # Calculate percentiles and grades based on ALL 2025 players
    # Match each player's score to the full 2025 distribution
    df_work['Total_Percentile'] = df_work['Total_Score_Original'].apply(
        lambda x: calculate_percentile_against_distribution(x, all_2025_scores)
    )
    df_work['Total_Score_1_10'] = df_work['Total_Percentile'].apply(convert_percentile_to_1_10)
    df_work['Total_Grade'] = df_work['Total_Score_1_10'].apply(lambda x: assign_grade_single(x))
    
    # Round all numeric columns
    numeric_cols = df_work.select_dtypes(include=['float64', 'int64']).columns
    for col in numeric_cols:
        df_work[col] = df_work[col].round(2)
    
    return df_work

def create_team_tabs(df, conference, variant_name, position_name, output_filename, base_dir, position_config):
    """Create team-by-team Excel file."""
    
    output_file = base_dir / output_filename
    
    # Get relevant columns for this position
    relevant_cols = get_relevant_columns(position_config)
    
    # Add scoring columns
    scoring_cols = ['Total_Score_1_10', 'Total_Percentile', 'Total_Grade']
    relevant_cols.extend(scoring_cols)
    
    # Remove internal scoring columns that shouldn't be in final output
    internal_cols_to_remove = ['Core_Score_Original', 'Specific_Score_Original', 'Total_Score_Original', 
                              'Year', 'Season', 'Conference', 'Position']
    relevant_cols = [col for col in relevant_cols if col not in internal_cols_to_remove]
    
    # Filter to only include columns that exist in the dataframe
    available_cols = [col for col in relevant_cols if col in df.columns]
    
    # Remove Age column if exists
    if 'Age' in available_cols:
        available_cols.remove('Age')
    
    # Reorder columns: Player, Team, then scoring columns first
    priority_cols = ['Player', 'Team', 'Total_Score_1_10', 'Total_Percentile', 'Total_Grade']
    remaining_cols = [col for col in available_cols if col not in priority_cols]
    
    # Sort remaining columns alphabetically
    remaining_cols_sorted = sorted(remaining_cols)
    
    # Build final column order: priority cols first, then remaining in alphabetical order
    final_cols = []
    for col in priority_cols:
        if col in available_cols:
            final_cols.append(col)
    final_cols.extend(remaining_cols_sorted)
    
    df = df[final_cols]
    
    # Sort by Total_Grade (A first, F last), then by Total_Percentile (highest first)
    grade_order = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'F': 4}
    df['_Grade_Sort'] = df['Total_Grade'].map(grade_order)
    df = df.sort_values(['_Grade_Sort', 'Total_Percentile'], ascending=[True, False])
    df = df.drop(columns=['_Grade_Sort'])
    
    # Get unique teams (filter out NaN values)
    teams = sorted([t for t in df['Team'].unique() if pd.notna(t)])
    
    # Create Excel file with conditional formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # All players sheet - already sorted
        df.to_excel(writer, sheet_name='All Players', index=False)
        
        # Team sheets - sort each team's players the same way
        for team in teams:
            team_df = df[df['Team'] == team].copy()
            team_df['_Grade_Sort'] = team_df['Total_Grade'].map(grade_order)
            team_df = team_df.sort_values(['_Grade_Sort', 'Total_Percentile'], ascending=[True, False])
            team_df = team_df.drop(columns=['_Grade_Sort'])
            sheet_name = team[:31] if len(team) > 31 else team
            team_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Add conditional formatting to the Excel file
    add_conditional_formatting(output_file)
    
    print(f"    ✅ Saved: {output_file.name} ({len(teams)} teams)")

def calculate_scores_historical(df_2025, df_all, config):
    """Calculate scores normalizing against all historical data."""
    core_score = pd.Series(0.0, index=df_2025.index)
    specific_score = pd.Series(0.0, index=df_2025.index)
    
    for metric_name, metric_config in config['metrics']['Core'].items():
        if isinstance(metric_config, dict):
            weight = metric_config['weight']
            metric_score = calculate_combined_metric_historical(df_2025, df_all, metric_config)
            core_score += metric_score * weight
    
    for metric_name, metric_config in config['metrics']['Specific'].items():
        if isinstance(metric_config, dict):
            weight = metric_config['weight']
            metric_score = calculate_combined_metric_historical(df_2025, df_all, metric_config)
            specific_score += metric_score * weight
    
    return core_score, specific_score

def calculate_combined_metric_historical(df_2025, df_all, metric_config):
    """Calculate metric using historical normalization."""
    total_score = pd.Series(0.0, index=df_2025.index)
    total_weight = 0.0
    
    if 'components' in metric_config:
        for component, weight in metric_config['components'].items():
            if component in df_2025.columns and component in df_all.columns:
                normalized = normalize_metric_historical(df_2025[component], df_all[component])
                total_score += normalized * weight
                total_weight += weight
    
    if total_weight > 0:
        total_score = total_score / total_weight
    
    return total_score

def normalize_metric_historical(series_2025, series_all):
    """Normalize using combined historical + current min/max."""
    if series_all.isna().all():
        return pd.Series([0.1] * len(series_2025), index=series_2025.index)
    
    if series_all.max() <= 100 and series_all.min() >= 0:
        normalized = series_2025 / 100
    else:
        min_val, max_val = series_all.min(), series_all.max()
        if max_val == min_val:
            normalized = pd.Series([0.5] * len(series_2025), index=series_2025.index)
        else:
            normalized = (series_2025 - min_val) / (max_val - min_val)
    
    return normalized.fillna(0.1)

def calculate_percentile_against_distribution(score, distribution_series):
    """Calculate percentile rank of a single score against a distribution."""
    if pd.isna(score):
        return 0.0
    valid_scores = distribution_series.dropna()
    if len(valid_scores) == 0:
        return 0.0
    # Calculate percentile: (number of scores <= this score) / total * 100
    percentile = (valid_scores <= score).sum() / len(valid_scores) * 100
    return percentile

def convert_percentile_to_1_10(percentile):
    """Convert a single percentile value to 1-10 scale."""
    if pd.isna(percentile):
        return 1.0
    if percentile >= 90:
        return 9 + (percentile - 90) / 10
    elif percentile >= 80:
        return 8 + (percentile - 80) / 10
    elif percentile >= 70:
        return 7 + (percentile - 70) / 10
    elif percentile >= 60:
        return 6 + (percentile - 60) / 10
    elif percentile >= 50:
        return 5 + (percentile - 50) / 10
    elif percentile >= 40:
        return 4 + (percentile - 40) / 10
    elif percentile >= 30:
        return 3 + (percentile - 30) / 10
    elif percentile >= 20:
        return 2 + (percentile - 20) / 10
    elif percentile >= 10:
        return 1 + (percentile - 10) / 10
    else:
        return 1.0

def convert_to_1_10_scale(score_series):
    scaled_scores = pd.Series(index=score_series.index, dtype=float)
    scaled_scores[score_series.isna()] = 1.0
    valid_scores = score_series.dropna()
    if len(valid_scores) == 0:
        return scaled_scores
    percentiles = valid_scores.rank(pct=True) * 100
    for i, percentile in percentiles.items():
        if percentile >= 90:
            scaled_scores[i] = 9 + (percentile - 90) / 10
        elif percentile >= 80:
            scaled_scores[i] = 8 + (percentile - 80) / 10
        elif percentile >= 70:
            scaled_scores[i] = 7 + (percentile - 70) / 10
        elif percentile >= 60:
            scaled_scores[i] = 6 + (percentile - 60) / 10
        elif percentile >= 50:
            scaled_scores[i] = 5 + (percentile - 50) / 10
        elif percentile >= 40:
            scaled_scores[i] = 4 + (percentile - 40) / 10
        elif percentile >= 30:
            scaled_scores[i] = 3 + (percentile - 30) / 10
        elif percentile >= 20:
            scaled_scores[i] = 2 + (percentile - 20) / 10
        elif percentile >= 10:
            scaled_scores[i] = 1 + (percentile - 10) / 10
        else:
            scaled_scores[i] = 1.0
    return scaled_scores

def calculate_percentile(score_series):
    percentiles = pd.Series(index=score_series.index, dtype=float)
    percentiles[score_series.isna()] = 0.0
    valid_scores = score_series.dropna()
    if len(valid_scores) > 0:
        valid_percentiles = valid_scores.rank(pct=True) * 100
        for i, pct in valid_percentiles.items():
            percentiles[i] = pct
    return percentiles

def assign_grade_single(score):
    """Assign grade for a single score value."""
    if pd.isna(score):
        return 'F'
    elif score >= 9:
        return 'A'
    elif score >= 8:
        return 'B'
    elif score >= 7:
        return 'C'
    elif score >= 6:
        return 'D'
    else:
        return 'F'

def assign_grade(score_series):
    """Assign grade for a Series of scores."""
    return score_series.apply(assign_grade_single)

if __name__ == "__main__":
    main()
