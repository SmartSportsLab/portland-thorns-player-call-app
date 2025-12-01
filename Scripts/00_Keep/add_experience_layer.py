#!/usr/bin/env python3
"""
Add NCAA Experience (seasons played) to player rankings based on Past Seasons data.
This counts how many seasons each player appears in our historical dataset.
"""

import pandas as pd
import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

def get_player_experience(player_name, historical_dir, conference, position_files):
    """
    Count how many seasons a player has appeared in our historical data.
    
    Args:
        player_name: The player's name to search for
        historical_dir: Directory containing Past Seasons files
        conference: Conference abbreviation (ACC, BIG10, etc.)
        position_files: List of position-specific file patterns to search
    
    Returns:
        int: Number of seasons the player appears in historical data
    """
    seasons_count = 0
    
    for file_pattern in position_files:
        # Search through all past seasons for this conference and position
        pattern = f"{file_pattern} {conference} 20*.xlsx"
        for file_path in historical_dir.glob(pattern):
            try:
                # Read the Excel file
                xls = pd.ExcelFile(file_path)
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Check if player appears in any season
                    if 'Player' in df.columns:
                        if player_name in df['Player'].values:
                            seasons_count += 1
                            break  # Only count once per file
            except Exception as e:
                print(f"  Warning: Could not read {file_path.name}: {e}")
                continue
    
    return seasons_count

def add_experience_to_file(file_path, historical_dir, player_experience_cache):
    """
    Add NCAA Experience column to an Excel file.
    
    Args:
        file_path: Path to the Excel file to update
        historical_dir: Directory containing Past Seasons files
        player_experience_cache: Dictionary to cache player experience lookups
    """
    print(f"\n📊 Processing: {file_path.name}")
    
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        all_data = []
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                if 'Player' not in df.columns:
                    print(f"  ⚠️  Sheet '{sheet_name}' missing 'Player' column, skipping")
                    continue
                
                # Initialize experience column if it doesn't exist
                if 'NCAA_Seasons' not in df.columns:
                    df['NCAA_Seasons'] = 0
                
                # Extract conference and position from filename
                filename = file_path.name
                if "AM Advanced Playmaker" in filename:
                    position_pattern = "AM Advanced Playmaker"
                elif "CB Hybrid" in filename:
                    position_pattern = "CB Hybrid"
                elif "DM Box-To-Box" in filename:
                    position_pattern = "DM Box-To-Box"
                elif "W Touchline Winger" in filename:
                    position_pattern = "W Touchline Winger"
                else:
                    print(f"  ⚠️  Could not identify position from filename")
                    all_data.append((sheet_name, df))
                    continue
                
                # Extract conference
                conference = None
                for conf in ['ACC', 'BIG10', 'BIG12', 'IVY', 'SEC', 'NWSL']:
                    if conf in filename:
                        conference = conf
                        break
                
                if not conference:
                    print(f"  ⚠️  Could not identify conference from filename")
                    all_data.append((sheet_name, df))
                    continue
                
                # Calculate experience for each player
                print(f"  📈 Analyzing {len(df)} players in sheet '{sheet_name}'...")
                
                for idx, player_name in enumerate(df['Player']):
                    # Check cache first
                    cache_key = f"{player_name}|{conference}|{position_pattern}"
                    
                    if cache_key not in player_experience_cache:
                        # Count seasons in historical data
                        seasons = get_player_experience(
                            player_name, 
                            historical_dir, 
                            conference, 
                            [position_pattern]
                        )
                        # Add 1 for the current season (2025)
                        seasons += 1
                        player_experience_cache[cache_key] = seasons
                    
                    df.at[idx, 'NCAA_Seasons'] = player_experience_cache[cache_key]
                
                # Show experience distribution
                experience_counts = df['NCAA_Seasons'].value_counts().sort_index()
                print(f"    Experience distribution:")
                for exp, count in experience_counts.items():
                    print(f"      {int(exp)} seasons: {count} players")
                
                all_data.append((sheet_name, df))
                
            except Exception as e:
                print(f"  ❌ Error processing sheet '{sheet_name}': {e}")
                continue
        
        # Save the updated workbook
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
            for sheet_name, df in all_data:
                # Reorder columns: keep existing order but move NCAA_Seasons after Player, Team
                cols = df.columns.tolist()
                if 'NCAA_Seasons' in cols:
                    cols.remove('NCAA_Seasons')
                    # Insert after Team column, or after Player if Team doesn't exist
                    if 'Team' in cols:
                        team_idx = cols.index('Team')
                        cols.insert(team_idx + 1, 'NCAA_Seasons')
                    elif 'Player' in cols:
                        player_idx = cols.index('Player')
                        cols.insert(player_idx + 1, 'NCAA_Seasons')
                    else:
                        cols.insert(0, 'NCAA_Seasons')
                    df = df[cols]
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"  ✅ Successfully updated {file_path.name}")
        
    except Exception as e:
        print(f"  ❌ Error processing file: {e}")

def main():
    """Main function to add experience data to all player ranking files."""
    
    base_dir = Path(__file__).parent.parent.parent
    exports_dir = base_dir / "Exports"
    historical_dir = exports_dir / "Past Seasons"
    
    # Get all 2025 files
    files_to_process = list(exports_dir.glob("*2025.xlsx"))
    
    if not files_to_process:
        print("❌ No 2025 files found to process")
        sys.exit(1)
    
    print("=" * 70)
    print("🏈 ADDING NCAA EXPERIENCE LAYER")
    print("=" * 70)
    print(f"📁 Processing {len(files_to_process)} files from {exports_dir}")
    print(f"📚 Historical data from {historical_dir}")
    
    # Cache for player experience to avoid redundant calculations
    player_experience_cache = {}
    
    # Process each file
    for file_path in sorted(files_to_process):
        add_experience_to_file(file_path, historical_dir, player_experience_cache)
    
    print("\n" + "=" * 70)
    print("✅ COMPLETE: Experience layer added to all files")
    print("=" * 70)
    
    # Summary statistics
    print("\n📊 Experience Cache Summary:")
    total_unique_players = len(set(key.split('|')[0] for key in player_experience_cache.keys()))
    print(f"   Total unique players analyzed: {total_unique_players}")
    
    experience_distribution = defaultdict(int)
    for cache_key, seasons in player_experience_cache.items():
        experience_distribution[seasons] += 1
    
    print("\n   Overall experience distribution:")
    for seasons in sorted(experience_distribution.keys()):
        count = experience_distribution[seasons]
        print(f"     {seasons} seasons: {count} player-records")

if __name__ == "__main__":
    main()

