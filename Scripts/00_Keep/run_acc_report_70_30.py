#!/usr/bin/env python3
"""
Run ACC Report with 70/30 Config
=================================

Generates the ACC Championship Scouting Report using the 70/30 weightings config.
"""

import sys
import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Import necessary functions from update_mike_norris_reports
sys.path.insert(0, str(Path(__file__).parent))
import update_mike_norris_reports as reports_module

def main():
    """Main function."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    print("="*80)
    print("GENERATING ACC REPORT WITH 70/30 WEIGHTINGS")
    print("="*80)
    
    conference = 'ACC'
    
    # Load 70/30 config
    config_file = base_dir / "Scripts" / "00_Keep" / "position_metrics_config_70_30.json"
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    print(f"✅ Loaded 70/30 config: {config_file.name}")
    
    # Get championship teams
    championship_teams = reports_module.get_championship_teams(conference)
    
    # Calculate team possessions for PAdj metrics (must be done early for Power Five scoring)
    print(f"\n📊 Calculating team possessions for PAdj metrics...")
    team_data_dir = base_dir / "Team Analysis" / "ACC"
    team_possessions = reports_module.calculate_team_possessions(team_data_dir)
    league_avg_possessions = reports_module.calculate_league_avg_possessions(team_possessions)
    print(f"  ✅ Calculated possessions for {len(team_possessions)} teams")
    print(f"  ✅ League average possessions: {league_avg_possessions:.2f} per game")
    
    # Load and score all Power Five historical data (2021-2025)
    print(f"\n🌐 Loading and scoring all Power Five historical data (2021-2025)...")
    df_all_power_five = reports_module.load_all_power_five_historical_data(base_dir)
    
    all_power_five_scored = {}
    if len(df_all_power_five) > 0:
        print(f"  ✅ Loaded {len(df_all_power_five)} total Power Five rows")
        
        position_profiles = ['Center Back', 'Centre Midfielder', 'Attacking Midfielder', 'Winger']
        pass_attempt_weight = 0.7  # 70/30 instead of 80/20
        pass_accuracy_weight = 0.3
        
        for position_name in position_profiles:
            if position_name not in config['position_profiles']:
                continue
            
            position_config = config['position_profiles'][position_name]
            display_name = reports_module.POSITION_PROFILE_MAP[position_name]
            
            df_power_five_pos = reports_module.filter_by_position(df_all_power_five, position_name)
            if len(df_power_five_pos) == 0:
                continue
            
            # Add PAdj metrics to Power Five data
            df_power_five_pos = reports_module.add_padj_metrics_to_dataframe(
                df_power_five_pos, team_possessions, league_avg_possessions, team_col='Team'
            )
            
            scored_by_year = []
            for year in range(2021, 2026):
                df_year = df_power_five_pos[df_power_five_pos['Year'] == year].copy()
                if len(df_year) == 0:
                    continue
                
                df_historical = df_power_five_pos[df_power_five_pos['Year'] < year].copy()
                if len(df_historical) == 0:
                    df_historical = df_year.copy()
                
                # Update config to use PAdj metrics
                position_config_padj = reports_module.update_config_to_use_padj(position_config, df_year)
                
                df_scored = reports_module.calculate_with_historical_normalization(
                    df_year, df_historical, position_config_padj,
                    pass_attempt_weight, pass_accuracy_weight, position_name
                )
                scored_by_year.append(df_scored)
            
            if scored_by_year:
                df_power_five_scored = pd.concat(scored_by_year, ignore_index=True)
                df_power_five_scored['Position_Profile'] = display_name
                all_power_five_scored[display_name] = df_power_five_scored
                print(f"  ✅ Scored {display_name}: {len(all_power_five_scored[display_name])} players")
    else:
        print(f"  ⚠️  No Power Five historical data found")
    
    # Load 2025 data
    acc_sec_dir = base_dir / "ACC:SEC Championship Reports"
    file_2025 = acc_sec_dir / f"{conference} All Positions 2025 Intra-Conference.xlsx"
    
    if not file_2025.exists():
        print(f"❌ File not found: {file_2025}")
        return
    
    print(f"📄 Loading 2025 data: {file_2025.name}")
    df_2025 = pd.read_excel(file_2025)
    print(f"  ✅ Loaded {len(df_2025)} players")
    
    # Load historical data
    print(f"\n📚 Loading historical data (2021-2025)...")
    df_all = reports_module.load_historical_data(base_dir, conference)
    
    if len(df_all) == 0:
        print(f"  ⚠️  No historical data found")
        return
    
    print(f"  ✅ Total historical rows: {len(df_all)}")
    
    # Load 2024 data
    print(f"\n📊 Loading 2024 data for comparison...")
    df_2024 = reports_module.load_2024_data_for_comparison(base_dir, conference)
    
    # Load team total minutes
    print(f"\n⏱️  Loading team total minutes from team stats files...")
    team_minutes_dict = reports_module.load_team_total_minutes(base_dir)
    
    # Load team averages for team-relative analysis
    print(f"\n📊 Loading team averages for team-relative analysis...")
    team_averages_dict = reports_module.load_team_averages(base_dir)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    position_profiles = ['Center Back', 'Centre Midfielder', 'Attacking Midfielder', 'Winger']
    all_scored_data = []
    
    # Use 70/30 weights
    pass_attempt_weight = 0.7
    pass_accuracy_weight = 0.3
    
    for position_name in position_profiles:
        if position_name not in config['position_profiles']:
            continue
        
        position_config = config['position_profiles'][position_name]
        display_name = reports_module.POSITION_PROFILE_MAP[position_name]
        
        print(f"\n📋 Processing {display_name}...")
        
        df_2025_pos = reports_module.filter_by_position(df_2025, position_name)
        if len(df_2025_pos) == 0:
            print(f"  ⚠️  No players found for {position_name}")
            continue
        
        df_2025_pos = reports_module.filter_to_championship_teams(df_2025_pos, championship_teams)
        if len(df_2025_pos) == 0:
            print(f"  ⚠️  No players from championship teams for {position_name}")
            continue
        print(f"  📋 Filtered to {len(df_2025_pos)} players from championship teams")
        
        df_all_pos = reports_module.filter_by_position(df_all, position_name)
        if len(df_all_pos) == 0:
            print(f"  ⚠️  No historical data for {position_name}")
            continue
        
        # Add PAdj metrics BEFORE scoring so they can be used in scoring
        print(f"  📊 Adding PAdj metrics...")
        df_2025_pos = reports_module.add_padj_metrics_to_dataframe(
            df_2025_pos, team_possessions, league_avg_possessions, team_col='Team'
        )
        # Also add PAdj to historical data for normalization
        df_all_pos = reports_module.add_padj_metrics_to_dataframe(
            df_all_pos, team_possessions, league_avg_possessions, team_col='Team'
        )
        
        # Update position config to use PAdj versions where applicable
        print(f"  🔄 Updating config to use PAdj metrics where available...")
        position_config_padj = reports_module.update_config_to_use_padj(position_config, df_2025_pos)
        
        # Calculate scores with 70/30 weights using PAdj metrics
        print(f"  🧮 Calculating scores with 70/30 weightings (using PAdj metrics)...")
        df_scored = reports_module.calculate_with_historical_normalization(
            df_2025_pos, df_all_pos, position_config_padj,
            pass_attempt_weight, pass_accuracy_weight, position_name
        )
        
        # Calculate percentiles
        print(f"  📊 Calculating percentiles...")
        reference_distribution = df_scored['Total_Score_1_10']
        df_scored['Total_Percentile'] = df_scored['Total_Score_1_10'].apply(
            lambda x: reports_module.calculate_percentile_against_distribution(x, reference_distribution)
        )
        
        # Assign grades
        print(f"  🎓 Assigning grades...")
        df_scored['Total_Grade'] = df_scored['Total_Score_1_10'].apply(reports_module.assign_grade_single)
        
        # Calculate 2024 scores for comparison
        if len(df_2024) > 0:
            df_2024_pos = reports_module.filter_by_position(df_2024, position_name)
            if len(df_2024_pos) > 0:
                df_2024_scored = reports_module.calculate_with_historical_normalization(
                    df_2024_pos, df_all_pos[df_all_pos['Year'] < 2025], position_config,
                    pass_attempt_weight, pass_accuracy_weight, position_name
                )
                scores_2024_dict = dict(zip(
                    df_2024_scored['Player'].astype(str),
                    df_2024_scored['Total_Score_1_10']
                ))
                df_scored['2024_Total_Score'] = df_scored['Player'].astype(str).map(scores_2024_dict)
                df_scored['Change_From_2024'] = df_scored['Total_Score_1_10'] - df_scored['2024_Total_Score']
            else:
                df_scored['2024_Total_Score'] = None
                df_scored['Change_From_2024'] = None
        else:
            df_scored['2024_Total_Score'] = None
            df_scored['Change_From_2024'] = None
        
        df_scored['Position_Profile'] = display_name
        
        if 'Total_Grade' in df_scored.columns:
            df_scored['Total_Grade'] = df_scored['Total_Grade'].fillna('')
        else:
            df_scored['Total_Grade'] = ''
        
        # Calculate grades
        print(f"  🏆 Calculating Team Grades...")
        df_scored['Team_Grade'] = df_scored.apply(
            lambda row: reports_module.calculate_team_grade(row, df_scored), axis=1
        )
        
        print(f"  🏆 Calculating Conference Grades...")
        df_scored['Conference_Grade'] = df_scored.apply(
            lambda row: reports_module.calculate_conference_grade(row, df_scored, conference), axis=1
        )
        
        print(f"  🏆 Calculating Power Five Grades...")
        if display_name in all_power_five_scored:
            df_power_five_pos = all_power_five_scored[display_name]
            df_scored['Power_Five_Grade'] = df_scored.apply(
                lambda row: reports_module.calculate_power_five_grade(row, df_power_five_pos, display_name), axis=1
            )
        else:
            df_scored['Power_Five_Grade'] = ''
        
        # Filter players with metrics
        relevant_metrics = reports_module.get_relevant_metrics_for_position(position_config)
        df_scored = reports_module.filter_players_with_metrics(df_scored, relevant_metrics)
        
        if len(df_scored) == 0:
            print(f"  ⚠️  No players with metric data for {display_name}, skipping sheet")
            continue
        
        all_scored_data.append(df_scored)
        
        # Calculate percentage of team minutes
        percentages_pos = reports_module.calculate_player_percentage_of_team_minutes(df_scored, team_minutes_dict)
        df_scored['Pct_Of_Team_Minutes'] = percentages_pos
        
        # Create position profile sheet
        reports_module.create_position_profile_sheet(wb, df_scored, display_name, team_minutes_dict, position_config, team_averages_dict)
        
        print(f"  ✅ {display_name}: {len(df_scored)} players processed")
    
    # Create team summary sheets
    if all_scored_data:
        df_all_scored = pd.concat(all_scored_data, ignore_index=True)
        percentages = reports_module.calculate_player_percentage_of_team_minutes(df_all_scored, team_minutes_dict)
        df_all_scored['Pct_Of_Team_Minutes'] = percentages
        reports_module.create_team_summary_sheets(wb, df_all_scored, conference, team_minutes_dict, config, team_averages_dict)
    
    # Create Data Summary sheet
    reports_module.create_data_notes_sheet(wb, conference)
    
    # Set print settings
    print(f"  📄 Setting print areas and page layout...")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 0 and ws.max_column > 0:
            print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            ws.print_area = print_area
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = 100
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
    
    # Save with the correct filename for ACC Championship Report with PAdj
    output_file = base_dir / "Mike_Norris_Scouting_Report_ACC_IntraConference_70_30_PAdj.xlsx"
    wb.save(output_file)
    print(f"\n✅ Report saved: {output_file.name}")
    print(f"   Using 70/30 weightings (70% intent, 30% accuracy)")
    print(f"   Includes PAdj (Possession-Adjusted) metrics")
    
    print("\n" + "="*80)
    print("✅ ACC REPORT GENERATED WITH 70/30 WEIGHTINGS")
    print("="*80)


if __name__ == "__main__":
    main()

