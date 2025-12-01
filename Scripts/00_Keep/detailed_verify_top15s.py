#!/usr/bin/env python3
"""
Detailed verification that exactly matches the shortlist script logic.
"""

import sys
from pathlib import Path
import pandas as pd
import json

sys.path.insert(0, str(Path(__file__).parent))

from update_mike_norris_reports import load_conference_season_data
from create_top_15_report import load_all_players_from_reports

def detailed_verify(player_name, team_name, position_profile, base_dir):
    """Verify Top 15s using the exact same logic as the shortlist script."""
    
    print("="*80)
    print(f"DETAILED VERIFICATION: {player_name} ({team_name})")
    print(f"Position: {position_profile}")
    print("="*80)
    
    # Load config
    config_file = Path(__file__).parent / "position_metrics_config.json"
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    # Get desired_metric_order (same as shortlist)
    desired_metric_order_map = {
        'Hybrid CB': [
            'Defensive duels per 90',
            'Defensive duels won, %',
            'Aerial duels per 90',
            'Aerial duels won, %',
            'Interceptions + Sliding Tackles',
            'Shots blocked per 90',
            'Passes per 90',
            'Accurate passes, %',
            'Progressive passes per 90',
            'Accurate progressive passes, %',
            'Long passes per 90',
            'Accurate long passes, %',
            'Short / medium passes per 90',
            'Accurate short / medium passes, %'
        ]
    }
    
    metric_headers = desired_metric_order_map.get(position_profile, [])
    
    # Load shortlist data (same as shortlist script)
    all_data = load_all_players_from_reports(base_dir)
    df_shortlist = all_data.get(position_profile, pd.DataFrame())
    
    # Find player in shortlist
    player_mask = (
        df_shortlist['Player'].str.contains(player_name, case=False, na=False) &
        df_shortlist['Team'].str.contains(team_name, case=False, na=False)
    )
    player_data_shortlist = df_shortlist[player_mask]
    
    if len(player_data_shortlist) == 0:
        print(f"❌ Player not found in shortlist data")
        return
    
    player_row = player_data_shortlist.iloc[0]
    print(f"\n✅ Found in shortlist: {player_row['Player']} - {player_row['Team']}")
    
    # Load Power Five raw data (same as shortlist script)
    print(f"\n📥 Loading Power Five 2025 raw data...")
    raw_data_frames = []
    power_five_conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
    
    for conference in power_five_conferences:
        df = load_conference_season_data(base_dir, conference, 2025)
        if not df.empty:
            df['Conference'] = conference
            df['Year'] = 2025
            raw_data_frames.append(df)
    
    if not raw_data_frames:
        print("❌ No raw data found")
        return
    
    all_players_raw = pd.concat(raw_data_frames, ignore_index=True)
    power_five_players = all_players_raw[
        (all_players_raw['Position_Profile'] == position_profile) &
        (all_players_raw['Conference'].isin(power_five_conferences))
    ].copy()
    
    print(f"   ✅ Loaded {len(power_five_players)} Power Five players")
    
    # Build metric_column_map (simplified version of shortlist logic)
    metric_column_map = {}
    for metric_header in metric_headers:
        if metric_header == "Interceptions + Sliding Tackles":
            # Find interceptions column
            interceptions_col = None
            for col in player_row.index:
                if 'interception' in str(col).lower() and 'per 90' in str(col).lower():
                    interceptions_col = col
                    break
            if interceptions_col:
                metric_column_map[metric_header] = {'interceptions': interceptions_col}
            continue
        
        # Try to find matching column
        matching_col = None
        metric_lower = str(metric_header).lower()
        
        for col in player_row.index:
            if str(col).lower() == metric_lower:
                matching_col = col
                break
        
        if matching_col:
            metric_column_map[metric_header] = matching_col
    
    print(f"\n📊 Metric column mapping:")
    for header, col in metric_column_map.items():
        if isinstance(col, dict):
            print(f"   {header}: {col}")
        else:
            print(f"   {header}: {col}")
    
    # Build metric_cols_to_check (same as shortlist script)
    print(f"\n🔍 Building metric columns to check...")
    metric_cols_to_check = []
    
    for metric_header in metric_headers:
        if metric_header == "Interceptions + Sliding Tackles":
            continue
        
        # Find matching column in raw data
        matching_col = None
        metric_lower = str(metric_header).lower()
        
        # Try exact match
        for col in power_five_players.columns:
            if str(col).lower().strip() == metric_lower.strip():
                matching_col = col
                break
        
        # Try base name match
        if not matching_col:
            if 'per 90' in metric_lower and '%' not in metric_lower:
                metric_base = metric_lower.replace(' per 90', '').replace(' per90', '').strip()
                for col in power_five_players.columns:
                    col_lower = str(col).lower()
                    if 'per 90' in col_lower and '%' not in col_lower:
                        col_base = col_lower.replace(' per 90', '').replace(' per90', '').strip()
                        if ' '.join(metric_base.split()) == ' '.join(col_base.split()):
                            matching_col = col
                            break
            elif '%' in metric_lower:
                metric_base = metric_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                for col in power_five_players.columns:
                    col_lower = str(col).lower()
                    if ('%' in col_lower or 'percent' in col_lower) and 'per 90' not in col_lower:
                        col_base = col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                        if ' '.join(metric_base.split()) == ' '.join(col_base.split()):
                            matching_col = col
                            break
        
        if matching_col:
            metric_cols_to_check.append((metric_header, matching_col))
            print(f"   ✅ {metric_header} -> {matching_col}")
        else:
            print(f"   ⚠️  {metric_header} -> NOT FOUND")
    
    # Now calculate Top 15s (same logic as shortlist script)
    print(f"\n{'='*80}")
    print("CALCULATING TOP 15s (using shortlist script logic)")
    print(f"{'='*80}\n")
    
    top15_count = 0
    top15_details = []
    
    # Handle combined metric
    combined_metric_handled = False
    for metric_header in metric_headers:
        if metric_header == "Interceptions + Sliding Tackles":
            combined_cols = metric_column_map.get(metric_header)
            if isinstance(combined_cols, dict):
                interceptions_col = combined_cols.get('interceptions')
                if interceptions_col and interceptions_col in player_row.index:
                    player_val = player_row.get(interceptions_col)
                    if pd.notna(player_val) and player_val != '':
                        try:
                            player_val = float(player_val)
                            # Find matching column in power_five_players
                            matching_col = None
                            for col in power_five_players.columns:
                                if 'interception' in str(col).lower() and 'per 90' in str(col).lower():
                                    matching_col = col
                                    break
                            
                            if matching_col:
                                pf_values = pd.to_numeric(power_five_players[matching_col], errors='coerce').dropna()
                                if len(pf_values) > 0:
                                    pf_rank = (pf_values >= player_val).sum()
                                    is_top15 = pf_rank <= 15
                                    
                                    print(f"{'✅' if is_top15 else '❌'} {metric_header} (Interceptions per 90)")
                                    print(f"      Value: {player_val:.2f}, Rank: {pf_rank}, Top 15: {is_top15}")
                                    
                                    if is_top15:
                                        top15_count += 1
                                        top15_details.append(f"{metric_header} (rank {pf_rank})")
                                        combined_metric_handled = True
                        except:
                            pass
            break
    
    # Check regular metrics
    for metric_header, raw_data_col in metric_cols_to_check:
        if combined_metric_handled and ('interception' in str(metric_header).lower() or 'sliding' in str(metric_header).lower()):
            continue
        
        # Get player value (same logic as shortlist script)
        player_value = None
        shortlist_col = metric_column_map.get(metric_header)
        if shortlist_col and shortlist_col in player_row.index:
            player_value = player_row.get(shortlist_col)
        
        # Fallback to raw data
        if pd.isna(player_value) or player_value == '':
            player_name_str = str(player_row.get('Player', '')).strip()
            player_team_str = str(player_row.get('Team', '')).strip()
            if player_name_str and player_team_str:
                player_raw = power_five_players[
                    (power_five_players['Player'].str.contains(player_name_str, case=False, na=False)) &
                    (power_five_players['Team'].str.contains(player_team_str, case=False, na=False))
                ]
                if len(player_raw) > 0 and raw_data_col in player_raw.columns:
                    player_value = player_raw[raw_data_col].iloc[0]
        
        if pd.isna(player_value) or player_value == '':
            print(f"⏭️  {metric_header}: No value")
            continue
        
        try:
            player_val = float(player_value)
        except:
            print(f"⚠️  {metric_header}: Invalid value ({player_value})")
            continue
        
        # Calculate rank
        pf_values = pd.to_numeric(power_five_players[raw_data_col], errors='coerce').dropna()
        if len(pf_values) == 0:
            print(f"⚠️  {metric_header}: No values in Power Five data")
            continue
        
        pf_rank = (pf_values >= player_val).sum()
        is_top15 = pf_rank <= 15
        
        # Show which value was used
        value_source = "shortlist" if shortlist_col and shortlist_col in player_row.index else "raw_data"
        
        print(f"{'✅' if is_top15 else '❌'} {metric_header}")
        print(f"      Value: {player_val:.2f} (from {value_source})")
        print(f"      Rank: {pf_rank} (out of {len(pf_values)})")
        print(f"      Top 15: {is_top15}")
        
        if is_top15:
            top15_count += 1
            top15_details.append(f"{metric_header} (rank {pf_rank})")
        
        print()
    
    # Summary
    print(f"\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}\n")
    print(f"Calculated Top 15s: {top15_count}")
    print(f"Top 15 metrics: {', '.join(top15_details)}")
    
    # Compare with shortlist
    from openpyxl import load_workbook
    shortlist_file = base_dir / "Portland Thorns 2025 Shortlist.xlsx"
    if shortlist_file.exists():
        wb = load_workbook(shortlist_file, data_only=True)
        if position_profile in wb.sheetnames:
            ws = wb[position_profile]
            for row_idx in range(4, ws.max_row + 1):
                ws_player = ws.cell(row=row_idx, column=2).value
                ws_team = ws.cell(row=row_idx, column=3).value
                if ws_player and player_name.lower() in str(ws_player).lower() and team_name.lower() in str(ws_team).lower():
                    # Find Top 15s column
                    top15_col = None
                    for col_idx in range(1, ws.max_column + 1):
                        header = ws.cell(row=3, column=col_idx).value
                        if header and 'Top 15s' in str(header):
                            top15_col = col_idx
                            break
                    shortlist_top15 = ws.cell(row=row_idx, column=top15_col).value if top15_col else None
                    print(f"\nShortlist Top 15s: {shortlist_top15}")
                    if shortlist_top15 == top15_count:
                        print("✅ MATCH!")
                    else:
                        print(f"⚠️  MISMATCH! Difference: {abs(shortlist_top15 - top15_count)}")
                    break
        wb.close()


if __name__ == "__main__":
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    if len(sys.argv) >= 4:
        player_name = sys.argv[1]
        team_name = sys.argv[2]
        position_profile = sys.argv[3]
    else:
        player_name = "H. McLaughlin"
        team_name = "Vanderbilt"
        position_profile = "Hybrid CB"
    
    detailed_verify(player_name, team_name, position_profile, base_dir)

