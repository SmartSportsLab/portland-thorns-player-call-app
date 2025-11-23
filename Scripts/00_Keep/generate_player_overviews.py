#!/usr/bin/env python3
"""
Generate one-page player overviews for all B+ grade players.
Includes key strengths, weaknesses, and Portland Thorns fit analysis.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import json
import sys
from datetime import datetime
import html as html_escape
import base64
try:
    import markdown
    MARKDOWN_AVAILABLE = True
except ImportError:
    MARKDOWN_AVAILABLE = False

try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

# Add path for imports
sys.path.insert(0, str(Path(__file__).parent))

# Import chart generation
try:
    from generate_player_charts import generate_all_charts
    CHARTS_AVAILABLE = True
except ImportError:
    CHARTS_AVAILABLE = False
    print("  ⚠️  Chart generation not available")

# Position profile mapping
POSITION_PROFILE_MAP = {
    'Hybrid CB': 'Center Back',
    'DM Box-To-Box': 'Centre Midfielder',
    'AM Advanced Playmaker': 'Attacking Midfielder',
    'Right Touchline Winger': 'Winger'
}

def load_nwsl_data(nwsl_dir):
    """Load NWSL team stats, calculate league averages, and Portland Thorns ranks."""
    nwsl_dir = Path(nwsl_dir)
    all_teams_data = []
    thorns_data = None
    thorns_team_name = None
    
    for file in nwsl_dir.glob('Team Stats *.xlsx'):
        team_name = file.stem.replace('Team Stats ', '')
        try:
            df = pd.read_excel(file, sheet_name='TeamStats')
            # Skip header rows and get match data
            df = df[df['Date'].notna()].copy()
            
            # Get numeric columns (excluding Date, Match, Competition, Scheme)
            exclude_cols = ['Date', 'Match', 'Competition', 'Scheme']
            numeric_cols = [col for col in df.columns if col not in exclude_cols]
            
            # Calculate totals (sum across all matches)
            team_totals = df[numeric_cols].sum()
            team_totals['Team'] = team_name
            team_totals['Matches'] = len(df)
            all_teams_data.append(team_totals)
            
            if 'Portland' in team_name or 'Thorns' in team_name:
                thorns_data = team_totals.copy()
                thorns_team_name = team_name
                
        except Exception as e:
            print(f"  ⚠️  Error loading {team_name}: {e}")
    
    if all_teams_data:
        league_df = pd.DataFrame(all_teams_data).set_index('Team')
        # Calculate league averages (excluding non-numeric columns)
        league_avg = league_df.select_dtypes(include=['number']).mean()
        
        # Calculate Portland Thorns' ranks among NWSL teams for each metric
        thorns_ranks = {}
        if thorns_team_name and thorns_team_name in league_df.index:
            numeric_cols = league_df.select_dtypes(include=['number']).columns
            
            # Metrics where lower is better (defensive metrics)
            lower_is_better = {'goals conceded', 'shots against', 'ppda', 'xcg', 'expected goals conceded'}
            
            for col in numeric_cols:
                thorns_val = league_df.loc[thorns_team_name, col]
                if pd.notna(thorns_val):
                    all_values = pd.to_numeric(league_df[col], errors='coerce').dropna()
                    if len(all_values) > 0:
                        col_lower = str(col).lower()
                        is_lower_better = any(metric in col_lower for metric in lower_is_better)
                        
                        if is_lower_better:
                            # Rank in ascending order (lower is better)
                            rank = (all_values <= thorns_val).sum()
                        else:
                            # Rank in descending order (higher is better)
                            rank = (all_values >= thorns_val).sum()
                        
                        total_teams = len(all_values)
                        thorns_ranks[col] = {
                            'rank': rank,
                            'total': total_teams,
                            'value': thorns_val,
                            'lower_is_better': is_lower_better
                        }
        
        return league_df, league_avg, thorns_data, thorns_ranks
    return None, None, None, None

def load_player_data_from_shortlist(shortlist_file, base_dir):
    """Load players from shortlist file and enrich with '% better than position' data from conference reports."""
    wb_shortlist = load_workbook(shortlist_file, data_only=True)
    all_players = []
    
    position_profiles = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    # Load conference reports for '% better than position' data
    reports_dir = base_dir
    conference_reports = list(reports_dir.glob('Portland Thorns 2025 * Championship Scouting Report.xlsx'))
    
    # Build a lookup dictionary: (Player, Team, Position Profile) -> player data with '% better than position'
    conference_data_lookup = {}
    for report_file in conference_reports:
        conference = report_file.stem.replace('Portland Thorns 2025 ', '').replace(' Championship Scouting Report', '').replace(' Scouting Report', '')
        try:
            wb = load_workbook(report_file, data_only=True)
            for position_profile in position_profiles:
                if position_profile not in wb.sheetnames:
                    continue
                ws = wb[position_profile]
                
                # Read headers (same logic as before)
                metric_names_row1 = []
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    metric_names_row1.append(str(cell_value) if cell_value else '')
                
                sub_headers_row2 = []
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=2, column=col_idx).value
                    sub_headers_row2.append(str(cell_value) if cell_value else '')
                
                column_mapping = {}
                current_metric = None
                for col_idx in range(1, min(len(metric_names_row1), len(sub_headers_row2)) + 1):
                    metric_name_raw = metric_names_row1[col_idx - 1] if col_idx <= len(metric_names_row1) else ''
                    sub_header = sub_headers_row2[col_idx - 1] if col_idx <= len(sub_headers_row2) else ''
                    metric_name = str(metric_name_raw).strip() if metric_name_raw and str(metric_name_raw).strip() != 'None' else ''
                    if metric_name:
                        current_metric = metric_name
                    if metric_name and not sub_header:
                        column_mapping[col_idx] = metric_name
                    elif sub_header:
                        if '% better than position' in sub_header:
                            if current_metric:
                                column_mapping[col_idx] = f"{current_metric} % better than position"
                            else:
                                column_mapping[col_idx] = f"{metric_name} % better than position" if metric_name else f"Column_{col_idx} % better than position"
                        elif 'per 90' in sub_header:
                            if current_metric:
                                column_mapping[col_idx] = f"{current_metric} per 90"
                            else:
                                column_mapping[col_idx] = f"{metric_name} per 90" if metric_name else f"Column_{col_idx} per 90"
                        else:
                            if current_metric:
                                column_mapping[col_idx] = f"{current_metric} {sub_header}"
                            else:
                                column_mapping[col_idx] = f"{metric_name} {sub_header}" if metric_name else sub_header
                    else:
                        column_mapping[col_idx] = f"Column_{col_idx}"
                
                # Read data
                for row_idx in range(3, ws.max_row + 1):
                    player_data = {}
                    for col_idx, col_name in column_mapping.items():
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        player_data[col_name] = cell_value
                    
                    player_name = str(player_data.get('Player', '')).strip()
                    team = str(player_data.get('Team', '')).strip()
                    if player_name and team:
                        key = (player_name, team, position_profile)
                        conference_data_lookup[key] = player_data
            
            wb.close()
        except Exception as e:
            print(f"  ⚠️  Error loading {report_file.name}: {e}")
    
    # Now load from shortlist and merge with conference data
    for position_profile in position_profiles:
        if position_profile not in wb_shortlist.sheetnames:
            continue
        
        ws = wb_shortlist[position_profile]
        
        # Read headers from row 3
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=3, column=col_idx).value
            if cell_value:
                headers.append(str(cell_value))
        
        # Read data starting from row 4
        for row_idx in range(4, ws.max_row + 1):
            player_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                player_data[header] = cell_value
            
            # Skip empty rows
            if not player_data.get('Player') or pd.isna(player_data.get('Player')):
                continue
            
            player_name = str(player_data.get('Player', '')).strip()
            team = str(player_data.get('Team', '')).strip()
            
            # Merge with conference report data (for '% better than position' columns)
            key = (player_name, team, position_profile)
            if key in conference_data_lookup:
                conference_data = conference_data_lookup[key]
                # Add '% better than position' columns
                for col_name, col_value in conference_data.items():
                    if '% better than position' in col_name or 'per 90' in col_name:
                        player_data[col_name] = col_value
            
            player_data['Position Profile'] = position_profile
            all_players.append(player_data)
    
    wb_shortlist.close()
    return pd.DataFrame(all_players)

def load_player_data_from_conference_reports(base_dir):
    """Load ALL players from conference reports (for ranking calculations)."""
    # Try the Conference Reports folder first, then fall back to base directory
    reports_dir = base_dir / 'Conference Reports'
    if not reports_dir.exists():
        reports_dir = base_dir
    
    # Try multiple naming patterns
    conference_reports = list(reports_dir.glob('Portland Thorns 2025 * Championship Scouting Report.xlsx'))
    if not conference_reports:
        conference_reports = list(reports_dir.glob('Portland Thorns 2025 * Scouting Report.xlsx'))
    if not conference_reports:
        conference_reports = list(reports_dir.glob('* Championship Scouting Report.xlsx'))
    if not conference_reports:
        conference_reports = list(reports_dir.glob('* Scouting Report.xlsx'))
    if not conference_reports:
        conference_reports = list(reports_dir.glob('*.xlsx'))
    
    all_players = []
    position_profiles = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    for report_file in conference_reports:
        conference = report_file.stem.replace('Portland Thorns 2025 ', '').replace(' Championship Scouting Report', '').replace(' Scouting Report', '')
        
        try:
            wb = load_workbook(report_file, data_only=True)
            
            for position_profile in position_profiles:
                if position_profile not in wb.sheetnames:
                    continue
                
                ws = wb[position_profile]
                
                # Read headers (same logic as before)
                metric_names_row1 = []
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    metric_names_row1.append(str(cell_value) if cell_value else '')
                
                sub_headers_row2 = []
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=2, column=col_idx).value
                    sub_headers_row2.append(str(cell_value) if cell_value else '')
                
                column_mapping = {}
                current_metric = None
                for col_idx in range(1, min(len(metric_names_row1), len(sub_headers_row2)) + 1):
                    metric_name_raw = metric_names_row1[col_idx - 1] if col_idx <= len(metric_names_row1) else ''
                    sub_header = sub_headers_row2[col_idx - 1] if col_idx <= len(sub_headers_row2) else ''
                    metric_name = str(metric_name_raw).strip() if metric_name_raw and str(metric_name_raw).strip() != 'None' else ''
                    if metric_name:
                        current_metric = metric_name
                    if metric_name and not sub_header:
                        column_mapping[col_idx] = metric_name
                    elif sub_header:
                        if '% better than position' in sub_header:
                            if current_metric:
                                column_mapping[col_idx] = f"{current_metric} % better than position"
                            else:
                                column_mapping[col_idx] = f"{metric_name} % better than position" if metric_name else f"Column_{col_idx} % better than position"
                        elif 'per 90' in sub_header:
                            if current_metric:
                                column_mapping[col_idx] = f"{current_metric} per 90"
                            else:
                                column_mapping[col_idx] = f"{metric_name} per 90" if metric_name else f"Column_{col_idx} per 90"
                        else:
                            if current_metric:
                                column_mapping[col_idx] = f"{current_metric} {sub_header}"
                            else:
                                column_mapping[col_idx] = f"{metric_name} {sub_header}" if metric_name else sub_header
                    else:
                        column_mapping[col_idx] = f"Column_{col_idx}"
                
                # Read data starting from row 3 (ALL players, not just B+)
                for row_idx in range(3, ws.max_row + 1):
                    player_data = {}
                    for col_idx, col_name in column_mapping.items():
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        player_data[col_name] = cell_value
                    
                    # Skip empty rows
                    if not player_data.get('Player') or pd.isna(player_data.get('Player')):
                        continue
                    
                    player_data['Position Profile'] = position_profile
                    player_data['Conference'] = conference
                    all_players.append(player_data)
            
            wb.close()
        except Exception as e:
            print(f"  ⚠️  Error loading {report_file.name}: {e}")
    
    return pd.DataFrame(all_players)

def find_column_variations(base_name, player_row):
    """Find column variations for a metric name (with/without 'per 90', different capitalization)."""
    variations = []
    
    # Try exact match
    if base_name in player_row.index:
        variations.append(base_name)
    
    # Try with "per 90" suffix
    per90_name = f"{base_name} per 90"
    if per90_name in player_row.index:
        variations.append(per90_name)
    
    # Try without "per 90" (for reports that don't include it)
    if " per 90" in base_name:
        base_without_per90 = base_name.replace(" per 90", "")
        if base_without_per90 in player_row.index:
            variations.append(base_without_per90)
    
    # Try case-insensitive matching
    base_lower = base_name.lower()
    for col in player_row.index:
        if col.lower() == base_lower or col.lower() == base_lower.replace(" per 90", ""):
            if col not in variations:
                variations.append(col)
    
    return variations

def combine_assist_metrics(player_row, position_config):
    """Combine individual assist metrics into combined metrics as defined in config."""
    combined_metrics = {}
    
    if not position_config:
        return combined_metrics
    
    # Check for combined assist metrics in config
    for category in ['Core', 'Specific']:
        if category in position_config.get('metrics', {}):
            for metric_name, metric_data in position_config['metrics'][category].items():
                # Check if this is a combined assist metric
                if isinstance(metric_data, dict) and 'components' in metric_data:
                    components = metric_data['components']
                    # Check if it contains assist components
                    has_assists = any('assist' in str(k).lower() for k in components.keys())
                    
                    # Include if it has assists (with or without goals)
                    if has_assists:
                        # Collect component values and "% better than position" values
                        component_values = {}
                        component_pct_better = {}
                        all_found = True
                        
                        for comp_name_config, comp_weight in components.items():
                            # Find the actual column name in the player_row (handle variations)
                            pct_col_variations = find_column_variations(
                                f"{comp_name_config} % better than position", 
                                player_row
                            )
                            
                            # Also try without "per 90" in the column name
                            comp_name_simple = comp_name_config.replace(" per 90", "")
                            pct_col_variations.extend(find_column_variations(
                                f"{comp_name_simple} % better than position",
                                player_row
                            ))
                            
                            found_pct_col = None
                            for var in pct_col_variations:
                                if '% better than position' in var:
                                    found_pct_col = var
                                    break
                            
                            if found_pct_col:
                                value = player_row.get(found_pct_col)
                                if pd.notna(value) and value != '':
                                    try:
                                        component_pct_better[comp_name_config] = float(value)
                                        
                                        # Also get the actual metric value
                                        per90_variations = find_column_variations(comp_name_config, player_row)
                                        per90_variations.extend(find_column_variations(comp_name_simple, player_row))
                                        
                                        found_value_col = None
                                        for var in per90_variations:
                                            if var in player_row.index:
                                                found_value_col = var
                                                break
                                        
                                        if found_value_col:
                                            val = player_row.get(found_value_col, 0)
                                            component_values[comp_name_config] = val if pd.notna(val) else 0
                                        else:
                                            component_values[comp_name_config] = 0
                                    except (ValueError, TypeError):
                                        all_found = False
                                        break
                                else:
                                    all_found = False
                                    break
                            else:
                                all_found = False
                                break
                        
                        if all_found and component_pct_better:
                            # Calculate weighted average of "% better than position"
                            total_weight = sum(components.values())
                            weighted_pct = sum(
                                component_pct_better.get(comp, 0) * weight 
                                for comp, weight in components.items()
                            ) / total_weight if total_weight > 0 else 0
                            
                            # Create combined metric value string
                            value_parts = []
                            for comp_name, comp_weight in components.items():
                                val = component_values.get(comp_name, 0)
                                if pd.notna(val) and val != 0:
                                    # Format the component name nicely
                                    comp_display = comp_name.replace(" per 90", "")
                                    value_parts.append(f"{comp_display}: {val:.2f}")
                            
                            combined_metrics[metric_name] = {
                                'pct_better': weighted_pct,
                                'value': ' | '.join(value_parts) if value_parts else 'N/A',
                                'weight': metric_data.get('weight', 1.0),
                                'components': components  # Store for rank lookup
                            }
    
    return combined_metrics

def calculate_metric_ranks(all_players_df, position_profile, player_row):
    """Calculate conference and power five ranks for each metric based on per 90 values."""
    ranks = {}
    
    # Check if DataFrame is valid and has required column
    if all_players_df is None or all_players_df.empty or 'Position Profile' not in all_players_df.columns:
        return ranks
    
    # Filter to same position profile
    pos_players = all_players_df[all_players_df['Position Profile'] == position_profile].copy()
    if len(pos_players) == 0:
        return ranks
    
    # Get player's conference
    player_conference = str(player_row.get('Conference', '')).strip()
    
    # Power Five conferences
    power_five = {'ACC', 'SEC', 'BIG10', 'Big Ten', 'BIG12', 'Big 12', 'PAC12', 'Pac-12'}
    
    # Filter conference players
    conf_players = pos_players[pos_players['Conference'] == player_conference].copy() if player_conference else pd.DataFrame()
    
    # Filter power five players
    power_five_players = pos_players[pos_players['Conference'].isin(power_five)].copy()
    
    # Calculate ranks for each metric based on per 90 values
    for col in player_row.index:
        if '% better than position' in str(col):
            metric_name = str(col).replace(' % better than position', '').strip()
            
            # Find the corresponding per 90 column
            per90_col = None
            # Try exact match with "per 90" suffix
            per90_col_name = f"{metric_name} per 90"
            if per90_col_name in player_row.index:
                per90_col = per90_col_name
            # Try without "per 90" (some metrics might not have it)
            elif metric_name in player_row.index:
                # Check if it's a per 90 metric (not a percentage)
                val = player_row.get(metric_name)
                if pd.notna(val):
                    try:
                        float_val = float(val)
                        # If it's a reasonable per 90 value (not a percentage > 100)
                        if float_val < 1000:  # Reasonable threshold for per 90 stats
                            per90_col = metric_name
                    except (ValueError, TypeError):
                        pass
            
            # Skip if we can't find a per 90 value
            if per90_col is None:
                continue
            
            player_value = player_row.get(per90_col)
            if pd.isna(player_value) or player_value == '':
                continue
            
            try:
                player_val = float(player_value)
                
                # Conference rank based on per 90 values
                conf_rank = None
                conf_total = None
                if len(conf_players) > 0 and per90_col in conf_players.columns:
                    conf_values = pd.to_numeric(conf_players[per90_col], errors='coerce').dropna()
                    if len(conf_values) > 0:
                        # Rank in descending order (higher is better for most metrics)
                        # For percentage metrics, we need to check if higher is better
                        # Most metrics: higher is better, but some like "Goals Conceded" lower is better
                        # For now, assume higher is better (can be refined later)
                        conf_rank = (conf_values >= player_val).sum()
                        conf_total = len(conf_values)
                
                # Power Five rank based on per 90 values
                pf_rank = None
                pf_total = None
                if len(power_five_players) > 0 and per90_col in power_five_players.columns:
                    pf_values = pd.to_numeric(power_five_players[per90_col], errors='coerce').dropna()
                    if len(pf_values) > 0:
                        pf_rank = (pf_values >= player_val).sum()
                        pf_total = len(pf_values)
                
                ranks[metric_name] = {
                    'conference_rank': conf_rank,
                    'conference_total': conf_total,
                    'power_five_rank': pf_rank,
                    'power_five_total': pf_total
                }
            except (ValueError, TypeError):
                pass
    
    return ranks

def identify_strengths(player_row, position_config, all_players_df=None):
    """Identify top 3-5 strengths based on ranks within position."""
    strengths = []
    
    # Calculate ranks if we have all players data
    metric_ranks = {}
    if all_players_df is not None and len(all_players_df) > 0:
        position_profile = player_row.get('Position Profile', '')
        metric_ranks = calculate_metric_ranks(all_players_df, position_profile, player_row)
    
    # First, combine assist metrics if applicable
    combined_assist_metrics = combine_assist_metrics(player_row, position_config)
    
    # Exclude administrative/non-metric columns
    exclude_cols = {
        'player', 'team', 'conference', 'position', 'rank', 'total score', 
        'conference grade', 'power five grade', 'previous year', 'previous score',
        'change from previous', 'changed position', 'total minutes', 
        '% of team minutes', 'seasons played', 'column_'
    }
    
    # Add combined assist metrics first
    for metric_name, metric_info in combined_assist_metrics.items():
        if metric_info['pct_better'] > 15:
            # Only include if metric is relevant to position profile
            if not is_metric_in_position_profile(metric_name, position_config):
                continue
            
            # Get ranks for combined metric (use first component's rank as proxy)
            rank_info = {}
            if metric_ranks:
                # Try to find rank for one of the components
                components = metric_info.get('components', {})
                for comp_name in components.keys():
                    comp_simple = comp_name.replace(' per 90', '')
                    if comp_simple in metric_ranks:
                        rank_info = metric_ranks[comp_simple]
                        break
            
            strengths.append({
                'metric': metric_name,
                'pct_better': metric_info['pct_better'],
                'weight': metric_info['weight'],
                'value': metric_info['value'],
                'conference_rank': rank_info.get('conference_rank'),
                'conference_total': rank_info.get('conference_total'),
                'power_five_rank': rank_info.get('power_five_rank'),
                'power_five_total': rank_info.get('power_five_total')
            })
    
    # Look for columns with '% better than position' suffix
    # Include both combined AND individual assist metrics
    for col in player_row.index:
        if '% better than position' in str(col):
            # Extract metric name (everything before " % better than position")
            metric_name = str(col).replace(' % better than position', '').strip()
            
            # Skip if it's an administrative column
            if any(excl in metric_name.lower() for excl in exclude_cols):
                continue
            
            value = player_row.get(col)
            
            if pd.notna(value) and value != '':
                try:
                    pct_diff = float(value)
                    # Consider >15% above average as a strength
                    if pct_diff > 15:
                        # Get the metric weight from config if available
                        weight = 1.0
                        if position_config:
                            for category in ['Core', 'Specific']:
                                if category in position_config.get('metrics', {}):
                                    for metric, metric_data in position_config['metrics'][category].items():
                                        # Try to match metric name (handle variations)
                                        metric_lower_config = metric.lower().replace(' ', '').replace('_', '')
                                        metric_name_lower = metric_name.lower().replace(' ', '').replace('_', '')
                                        if metric_lower_config in metric_name_lower or metric_name_lower in metric_lower_config:
                                            if isinstance(metric_data, dict) and 'weight' in metric_data:
                                                weight = metric_data['weight']
                                            elif isinstance(metric_data, (int, float)):
                                                weight = metric_data
                                            break
                        
                        # Try to find the actual metric value (per 90 or %)
                        metric_value = None
                        # Look for per 90 version
                        per90_col = f"{metric_name} per 90"
                        if per90_col in player_row.index:
                            val = player_row.get(per90_col)
                            if pd.notna(val) and val != '':
                                metric_value = val
                        # Or look for % version
                        if metric_value is None:
                            pct_col = f"{metric_name}, %"
                            if pct_col in player_row.index:
                                val = player_row.get(pct_col)
                                if pd.notna(val) and val != '':
                                    metric_value = val
                        # Or try without "per 90" suffix
                        if metric_value is None:
                            if metric_name in player_row.index:
                                val = player_row.get(metric_name)
                                if pd.notna(val) and val != '':
                                    metric_value = val
                        
                        # Only include if we have a valid metric value
                        if metric_value is None or (isinstance(metric_value, str) and metric_value.strip() == 'N/A'):
                            continue
                        
                        # Only include if metric is relevant to position profile
                        if not is_metric_in_position_profile(metric_name, position_config):
                            continue
                        
                        # Get ranks
                        rank_info = metric_ranks.get(metric_name, {})
                        
                        strengths.append({
                            'metric': metric_name,
                            'pct_better': pct_diff,
                            'weight': weight,
                            'value': metric_value,
                            'conference_rank': rank_info.get('conference_rank'),
                            'conference_total': rank_info.get('conference_total'),
                            'power_five_rank': rank_info.get('power_five_rank'),
                            'power_five_total': rank_info.get('power_five_total')
                        })
                except (ValueError, TypeError):
                    pass
    
    # Sort by weighted importance (weight * pct_better)
    strengths.sort(key=lambda x: x['weight'] * x['pct_better'], reverse=True)
    
    # Return top 5
    return strengths[:5]

def identify_weaknesses(player_row, position_config, all_players_df=None):
    """Identify top 3-5 weaknesses based on ranks within position."""
    weaknesses = []
    
    # Calculate ranks if we have all players data
    metric_ranks = {}
    if all_players_df is not None and len(all_players_df) > 0:
        position_profile = player_row.get('Position Profile', '')
        metric_ranks = calculate_metric_ranks(all_players_df, position_profile, player_row)
    
    # First, combine assist metrics if applicable
    combined_assist_metrics = combine_assist_metrics(player_row, position_config)
    
    # Exclude administrative/non-metric columns and individual assist metrics
    exclude_cols = {
        'player', 'team', 'conference', 'position', 'rank', 'total score', 
        'conference grade', 'power five grade', 'previous year', 'previous score',
        'change from previous', 'changed position', 'total minutes', 
        '% of team minutes', 'seasons played', 'column_'
    }
    
    # Add combined assist metrics first (if they're weaknesses)
    for metric_name, metric_info in combined_assist_metrics.items():
        if metric_info['pct_better'] < -10:
            # Only include if metric is relevant to position profile
            if not is_metric_in_position_profile(metric_name, position_config):
                continue
            
            # Get ranks for combined metric
            rank_info = {}
            if metric_ranks:
                components = metric_info.get('components', {})
                for comp_name in components.keys():
                    comp_simple = comp_name.replace(' per 90', '')
                    if comp_simple in metric_ranks:
                        rank_info = metric_ranks[comp_simple]
                        break
            
            weaknesses.append({
                'metric': metric_name,
                'pct_below': abs(metric_info['pct_better']),
                'weight': metric_info['weight'],
                'value': metric_info['value'],
                'conference_rank': rank_info.get('conference_rank'),
                'conference_total': rank_info.get('conference_total'),
                'power_five_rank': rank_info.get('power_five_rank'),
                'power_five_total': rank_info.get('power_five_total')
            })
    
    # Look for columns with '% better than position' suffix
    # Include both combined AND individual assist metrics
    for col in player_row.index:
        if '% better than position' in str(col):
            # Extract metric name (everything before " % better than position")
            metric_name = str(col).replace(' % better than position', '').strip()
            
            # Skip if it's an administrative column
            if any(excl in metric_name.lower() for excl in exclude_cols):
                continue
            
            value = player_row.get(col)
            
            if pd.notna(value) and value != '':
                try:
                    pct_diff = float(value)
                    # Consider <-10% below average as a weakness
                    if pct_diff < -10:
                        # Get the metric weight from config if available
                        weight = 1.0
                        if position_config:
                            for category in ['Core', 'Specific']:
                                if category in position_config.get('metrics', {}):
                                    for metric, metric_data in position_config['metrics'][category].items():
                                        # Try to match metric name (handle variations)
                                        metric_lower_config = metric.lower().replace(' ', '').replace('_', '')
                                        metric_name_lower = metric_name.lower().replace(' ', '').replace('_', '')
                                        if metric_lower_config in metric_name_lower or metric_name_lower in metric_lower_config:
                                            if isinstance(metric_data, dict) and 'weight' in metric_data:
                                                weight = metric_data['weight']
                                            elif isinstance(metric_data, (int, float)):
                                                weight = metric_data
                                            break
                        
                        # Try to find the actual metric value (per 90 or %)
                        metric_value = None
                        # Look for per 90 version
                        per90_col = f"{metric_name} per 90"
                        if per90_col in player_row.index:
                            val = player_row.get(per90_col)
                            if pd.notna(val) and val != '':
                                metric_value = val
                        # Or look for % version
                        if metric_value is None:
                            pct_col = f"{metric_name}, %"
                            if pct_col in player_row.index:
                                val = player_row.get(pct_col)
                                if pd.notna(val) and val != '':
                                    metric_value = val
                        # Or try without "per 90" suffix
                        if metric_value is None:
                            if metric_name in player_row.index:
                                val = player_row.get(metric_name)
                                if pd.notna(val) and val != '':
                                    metric_value = val
                        
                        # Only include if we have a valid metric value
                        if metric_value is None or (isinstance(metric_value, str) and metric_value.strip() == 'N/A'):
                            continue
                        
                        # Only include if metric is relevant to position profile
                        if not is_metric_in_position_profile(metric_name, position_config):
                            continue
                        
                        # Get ranks
                        rank_info = metric_ranks.get(metric_name, {})
                        
                        weaknesses.append({
                            'metric': metric_name,
                            'pct_below': abs(pct_diff),
                            'weight': weight,
                            'value': metric_value,
                            'conference_rank': rank_info.get('conference_rank'),
                            'conference_total': rank_info.get('conference_total'),
                            'power_five_rank': rank_info.get('power_five_rank'),
                            'power_five_total': rank_info.get('power_five_total')
                        })
                except (ValueError, TypeError):
                    pass
    
    # Sort by weighted importance (weight * pct_below)
    weaknesses.sort(key=lambda x: x['weight'] * x['pct_below'], reverse=True)
    
    # Return top 5
    return weaknesses[:5]

def calculate_thorns_style_fit(player_row, thorns_ranks, all_players_df, position_profile, position_configs=None):
    """Calculate style fit using metrics from JSON config and top 20% rank threshold."""
    style_fits = []
    
    if thorns_ranks is None or len(thorns_ranks) == 0:
        return style_fits
    
    if position_configs is None:
        return style_fits
    
    # Get position name from profile mapping
    position_name = POSITION_PROFILE_MAP.get(position_profile, position_profile)
    
    # Get style fit metrics for this position from config
    style_fit_metrics = position_configs.get('style_fit_metrics', {}).get(position_name, {}).get('metrics', [])
    
    if not style_fit_metrics:
        return style_fits
    
    # Calculate player's ranks within their position
    player_metric_ranks = calculate_metric_ranks(all_players_df, position_profile, player_row)
    
    # Helper function to find matching metric column in player data
    def find_metric_column(metric_name, player_row):
        """Find the matching metric column in player data."""
        metric_lower = metric_name.lower().strip()
        
        # Get all per 90 columns
        all_cols = [(col, str(col).lower()) for col in player_row.index if 'per 90' in str(col).lower() and '%' not in str(col).lower()]
        
        # Define exact matches first (most specific)
        exact_matches = {
            'interceptions': 'interceptions',
            'progressive passes': 'progressive passes',
            'long passes': 'long passes',  # Not "received long passes"
            'defensive duels': 'defensive duels',
            'aerial duels': 'aerial duels',
            'goals': 'goals',
            'assists': 'assists',
            'smart passes': 'smart passes',
            'through passes': 'through passes',
            'deep completions': 'deep completions',
            'dribbles': 'dribbles',
            'offensive duels': 'offensive duels',
            'crosses': 'crosses',
            'progressive runs': 'progressive runs',
            'received long passes': 'received long passes',
            'received passes': 'received passes',
            'passes': 'passes'  # Generic - must be exact, not "short/medium" or "received"
        }
        
        # Try exact match first
        if metric_lower in exact_matches:
            target = exact_matches[metric_lower]
            for col, col_lower in all_cols:
                col_base = col_lower.replace(' per 90', '').replace(' per90', '').strip()
                # Exact match (word boundary check)
                if col_base == target:
                    return str(col).replace(' per 90', '').replace(' per90', '').strip()
                # Also check if it's the exact phrase (for multi-word metrics)
                if target in col_base and len(target) > 5:
                    # Make sure it's not a substring of a longer metric
                    # e.g., "passes" should match "passes" not "short/medium passes"
                    if col_base == target or col_base.startswith(target + ' ') or col_base.endswith(' ' + target):
                        return str(col).replace(' per 90', '').replace(' per90', '').strip()
        
        return None
    
    # Map NWSL metric names to NCAA metric names
    def find_nwsl_matching_metric(ncaa_metric_name, thorns_ranks):
        """Find NWSL metric that matches the NCAA metric."""
        ncaa_lower = ncaa_metric_name.lower()
        
        # Map common NCAA to NWSL metric names (with priority order)
        ncaa_to_nwsl_mappings = {
            'interceptions': ['interceptions'],
            'progressive passes': ['deep completed passes', 'forward passes', 'progressive passes'],
            'long passes': ['long passes', 'long pass'],
            'passes': ['deep completed passes', 'forward passes', 'passes'],  # Generic passes - try progressive first
            'defensive duels': ['duels', 'defensive duels'],
            'aerial duels': ['aerial duels', 'aerial'],
            'goals': ['goals'],
            'assists': ['assists'],
            'smart passes': ['smart passes', 'smart'],
            'through passes': ['through passes', 'through'],
            'deep completions': ['deep completed passes', 'deep completions'],
            'dribbles': ['dribbles'],
            'offensive duels': ['offensive duels', 'offensive'],
            'crosses': ['crosses'],
            'progressive runs': ['progressive runs'],
            'received long passes': ['received long passes'],
            'received passes': ['received passes']
        }
        
        # Find the best matching NWSL metric
        best_match = None
        best_priority = float('inf')
        
        for ncaa_key, nwsl_keys in ncaa_to_nwsl_mappings.items():
            if ncaa_key in ncaa_lower:
                # Try each NWSL key in priority order
                for priority, nwsl_key in enumerate(nwsl_keys):
                    for nwsl_metric in thorns_ranks.keys():
                        nwsl_lower = str(nwsl_metric).lower()
                        if 'unnamed' not in nwsl_lower:
                            # Check for match
                            if nwsl_key in nwsl_lower or nwsl_lower in nwsl_key:
                                if priority < best_priority:
                                    best_match = nwsl_metric
                                    best_priority = priority
                                break  # Found a match for this key, move to next
        
        return best_match
    
    # Track which metrics we've already added (to avoid duplicates)
    seen_metrics = set()
    
    # Check each style fit metric from JSON config
    for style_metric in style_fit_metrics:
        # Find the matching column in player data
        ncaa_metric = find_metric_column(style_metric, player_row)
        
        if not ncaa_metric:
            continue
        
        # Find matching key in player_metric_ranks (case-insensitive)
        matching_rank_key = None
        ncaa_lower = ncaa_metric.lower()
        for rank_key in player_metric_ranks.keys():
            if rank_key.lower() == ncaa_lower:
                matching_rank_key = rank_key
                break
        
        if not matching_rank_key:
            continue
        
        # Skip if we've already added this metric
        if ncaa_metric in seen_metrics:
            continue
        
        # Find matching NWSL metric (try to find the best match)
        nwsl_metric = find_nwsl_matching_metric(ncaa_metric, thorns_ranks)
        
        if not nwsl_metric or nwsl_metric not in thorns_ranks:
            continue
        
        thorns_rank_info = thorns_ranks[nwsl_metric]
        thorns_rank = thorns_rank_info['rank']
        total_nwsl_teams = thorns_rank_info['total']
        
        # Only consider if Portland ranks top 3 in NWSL
        if thorns_rank > 3:
            continue
        
        # Get player's rank (use the matching key with correct capitalization)
        player_rank_info = player_metric_ranks[matching_rank_key]
        player_pf_rank = player_rank_info.get('power_five_rank')
        player_conf_rank = player_rank_info.get('conference_rank')
        
        # Only count as style fit if player ranks in top 20% of Power Five (all five conferences)
        # This matches Portland's threshold: top 3 out of 14 NWSL teams = ~21%
        is_fit = False
        fit_reason = []
        
        if player_pf_rank is not None and player_rank_info.get('power_five_total'):
            pf_total = player_rank_info['power_five_total']
            # Calculate top 20% threshold (round up to be inclusive)
            top_20_percent_threshold = max(1, round(pf_total * 0.20))
            if player_pf_rank <= top_20_percent_threshold:
                is_fit = True
                fit_reason.append(f"#{player_pf_rank}/{pf_total} Power Five (top {round((top_20_percent_threshold/pf_total)*100)}%)")
        
        # Conference rank is shown for reference only, but doesn't count as a style fit
        if player_conf_rank is not None and player_rank_info.get('conference_total'):
            conf_total = player_rank_info['conference_total']
            fit_reason.append(f"#{player_conf_rank}/{conf_total} in conference")
        
        if is_fit:
            # Use the original style_metric name for display, but check duplicates by the actual metric found
            seen_metrics.add(matching_rank_key.lower())  # Mark as seen to avoid duplicates (use rank key)
            style_fits.append({
                'metric': matching_rank_key,  # Use the properly capitalized metric name
                'thorns_nwsl_rank': f"#{thorns_rank}/{total_nwsl_teams}",
                'player_ranks': ', '.join(fit_reason),
                'thorns_value': thorns_rank_info['value']
            })
    
    # Sort by Thorns rank (best fits first)
    style_fits.sort(key=lambda x: int(x['thorns_nwsl_rank'].split('/')[0].replace('#', '')))
    
    return style_fits  # Return all fits (not limited to 5)

def analyze_thorns_fit(player_row, position_profile, thorns_ranks, all_players_df, full_position_configs):
    """Analyze why player might be a good fit for Portland Thorns based on playing style alignment."""
    fit_points = []
    
    # Get player's key metrics
    player_score = player_row.get('Total Score', player_row.get('2025 Total Score', 0))
    conf_grade = player_row.get('Conference Grade', '')
    power_grade = player_row.get('Power Five Grade', '')
    
    # 1. Overall performance
    if pd.notna(player_score) and float(player_score) > 7.0:
        fit_points.append(f"Strong overall performance score ({player_score:.2f}) indicating NWSL readiness")
    
    if conf_grade in ['A', 'B'] and power_grade in ['A', 'B']:
        fit_points.append(f"Consistent high grades (Conference: {conf_grade}, Power Five: {power_grade})")
    
    # 2. Playing style alignment: Where Portland ranks highly among NWSL teams
    style_fits = calculate_thorns_style_fit(player_row, thorns_ranks, all_players_df, position_profile, full_position_configs)
    
    if style_fits:
        fit_points.append("**Playing Style Alignment:**")
        for fit in style_fits:
            metric_display = format_metric_name(fit['metric'])
            fit_points.append(f"- {metric_display}: Portland ranks {fit['thorns_nwsl_rank']} in NWSL, player ranks {fit['player_ranks']} in position")
    else:
        fit_points.append("Limited playing style alignment identified (Portland's top-ranked metrics don't match player's strengths)")
    
    # 3. Playing time and consistency
    pct_minutes = player_row.get('% of Team Minutes', 0)
    if pd.notna(pct_minutes) and float(pct_minutes) >= 70:
        fit_points.append(f"High playing time ({pct_minutes:.1f}% of team minutes) demonstrates coach trust and consistency")
    
    # 4. Progression/improvement
    change_from_prev = player_row.get('Change From Previous', None)
    if pd.notna(change_from_prev) and change_from_prev != '':
        try:
            change = float(change_from_prev)
            if change > 0:
                fit_points.append(f"Positive progression (+{change:.2f} score improvement from previous season)")
        except (ValueError, TypeError):
            pass
    
    # 5. Position versatility (if changed position)
    changed_pos = player_row.get('Changed Position', '')
    if changed_pos and changed_pos != '':
        prev_pos = player_row.get('Previous Position', '')
        fit_points.append(f"Positional versatility (transitioned from {prev_pos} to {position_profile})")
    
    return fit_points

def generate_player_summary(player_row, strengths, weaknesses, style_fits, fit_points, position_profile):
    """Generate a 1-2 line summary of the player's strengths, weaknesses, and Portland fit."""
    summary_parts = []
    
    # Strengths summary
    if strengths:
        top_strength = strengths[0]
        strength_metric = format_metric_name(top_strength['metric'], add_per90=False)
        if top_strength.get('power_five_rank') and top_strength.get('power_five_total'):
            rank_str = f"#{top_strength['power_five_rank']}/{top_strength['power_five_total']}"
        elif top_strength.get('conference_rank') and top_strength.get('conference_total'):
            rank_str = f"#{top_strength['conference_rank']}/{top_strength['conference_total']}"
        else:
            rank_str = "elite"
        summary_parts.append(f"Elite {strength_metric.lower()} ({rank_str} Power Five)")
    
    # Weaknesses summary (if significant)
    if weaknesses and len(weaknesses) > 0:
        top_weakness = weaknesses[0]
        weakness_metric = format_metric_name(top_weakness['metric'], add_per90=False)
        summary_parts.append(f"Needs improvement in {weakness_metric.lower()}")
    
    # Portland fit summary
    if style_fits > 0:
        if style_fits >= 3:
            fit_desc = "Excellent"
        elif style_fits >= 2:
            fit_desc = "Strong"
        else:
            fit_desc = "Good"
        summary_parts.append(f"{fit_desc} Portland Thorns style fit ({style_fits} metric{'s' if style_fits > 1 else ''})")
    elif len(fit_points) > 0:
        summary_parts.append("Potential Portland Thorns fit")
    else:
        summary_parts.append("Limited Portland Thorns style alignment")
    
    # Combine into 1-2 lines and ensure proper capitalization
    if len(summary_parts) >= 2:
        # Split into two lines: strengths/weaknesses on first line, fit on second
        first_line = ". ".join(summary_parts[:2]) + "."
        # Capitalize first letter of first line
        first_line = first_line[0].upper() + first_line[1:] if len(first_line) > 1 else first_line.upper()
        
        if len(summary_parts) > 2:
            second_line = summary_parts[2] + "."
            # Capitalize first letter of second line
            second_line = second_line[0].upper() + second_line[1:] if len(second_line) > 1 else second_line.upper()
            return f"{first_line} {second_line}"
        else:
            return first_line
    elif len(summary_parts) == 1:
        result = summary_parts[0] + "."
        # Capitalize first letter
        result = result[0].upper() + result[1:] if len(result) > 1 else result.upper()
        return result
    else:
        return None

def get_position_profile_metrics(position_config):
    """Extract all metric names from position config (including components)."""
    metrics_set = set()
    
    if not position_config:
        return metrics_set
    
    for category in ['Core', 'Specific']:
        if category in position_config.get('metrics', {}):
            for metric, metric_data in position_config['metrics'][category].items():
                # Add the combined metric name
                metrics_set.add(metric.lower().strip())
                
                # If it's a combined metric, add components too
                if isinstance(metric_data, dict) and 'components' in metric_data:
                    for comp_name in metric_data['components'].keys():
                        comp_clean = comp_name.replace(' per 90', '').replace(', %', '').replace(' %', '').strip().lower()
                        metrics_set.add(comp_clean)
    
    return metrics_set

def is_metric_in_position_profile(metric_name, position_config):
    """Check if a metric is relevant to the position profile."""
    if not position_config:
        return True  # If no config, include all metrics
    
    profile_metrics = get_position_profile_metrics(position_config)
    
    # Normalize metric name for comparison
    metric_normalized = metric_name.lower().strip()
    
    # Remove common suffixes for base comparison
    metric_base = metric_normalized.replace(' per 90', '').replace(' per90', '').replace(', %', '').replace(' %', '').replace(' won', '').replace('accurate', '').strip()
    
    # Check exact match first
    if metric_normalized in profile_metrics:
        return True
    
    # Check base match (exact match after removing suffixes)
    for profile_metric in profile_metrics:
        profile_base = profile_metric.replace(' per 90', '').replace(' per90', '').replace(', %', '').replace(' %', '').replace(' won', '').replace('accurate', '').strip()
        
        # Exact base match (most strict)
        if profile_base == metric_base:
            return True
        
        # For combined metrics, check if metric_base matches any component word
        # But be careful - "shots" should NOT match "shots blocked"
        # Only match if the full base name is contained, not just a word
        if ' ' in profile_base:
            # Multi-word metric - require full match or metric_base is the full phrase
            if metric_base == profile_base:
                return True
            # Check if metric_base is a complete word/phrase within profile_base
            # e.g., "defensive duels" matches "defensive duels per 90"
            if metric_base in profile_base and len(metric_base) > 5:  # Require at least 5 chars to avoid false matches
                # Additional check: ensure it's not a partial word match
                # e.g., "shots" should NOT match "shots blocked"
                words_in_profile = set(profile_base.split())
                words_in_metric = set(metric_base.split())
                # If metric_base is a subset of words, check if it's a meaningful match
                if words_in_metric.issubset(words_in_profile):
                    # Check if the order matches (e.g., "defensive duels" matches "defensive duels won")
                    profile_words = profile_base.split()
                    metric_words = metric_base.split()
                    if len(metric_words) > 0:
                        # Check if metric words appear in order at the start of profile
                        if profile_words[:len(metric_words)] == metric_words:
                            return True
        else:
            # Single word metric - exact match only
            if metric_base == profile_base:
                return True
    
    return False

def format_metric_name(metric_name, add_per90=True):
    """Format metric name for display, adding ' Per 90' if not already present."""
    # Clean up common patterns but preserve structure
    name = metric_name
    
    # Check if it already has "per 90" or is a percentage metric
    has_per90 = 'per 90' in name.lower() or 'per90' in name.lower()
    is_percentage = ', %' in name or ' %' in name or 'won, %' in name.lower() or 'won %' in name.lower()
    
    # Remove existing per 90 for processing
    name_clean = name.replace(' per 90', '').replace(' per90', '').replace(', %', '').replace(' %', '').replace(' won, %', '').replace(' won %', '').strip()
    
    # Capitalize words
    words = name_clean.split()
    formatted = ' '.join(word.capitalize() for word in words)
    
    # Add " Per 90" if requested and not already present and not a percentage metric
    if add_per90 and not has_per90 and not is_percentage:
        formatted += " Per 90"
    
    return formatted

def convert_markdown_to_html(markdown_text):
    """Simple markdown to HTML converter (basic support)."""
    lines = markdown_text.split('\n')
    result_lines = []
    in_list = False
    list_type = None
    
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        
        # Headers (process before other formatting)
        if stripped.startswith('### '):
            if in_list:
                result_lines.append(f'</{list_type}>')
                in_list = False
            result_lines.append(f'<h3>{stripped[4:]}</h3>')
        elif stripped.startswith('## '):
            if in_list:
                result_lines.append(f'</{list_type}>')
                in_list = False
            result_lines.append(f'<h2>{stripped[3:]}</h2>')
        elif stripped.startswith('# '):
            if in_list:
                result_lines.append(f'</{list_type}>')
                in_list = False
            result_lines.append(f'<h1>{stripped[2:]}</h1>')
        # Horizontal rule
        elif stripped == '---':
            if in_list:
                result_lines.append(f'</{list_type}>')
                in_list = False
            result_lines.append('<hr>')
        # Lists
        elif stripped.startswith('- '):
            if not in_list:
                result_lines.append('<ul>')
                in_list = True
                list_type = 'ul'
            content = stripped[2:].strip()
            # Process bold and italic in list items
            content = process_formatting(content)
            result_lines.append(f'<li>{content}</li>')
        elif stripped and stripped[0].isdigit() and '. ' in stripped:
            if not in_list:
                result_lines.append('<ol>')
                in_list = True
                list_type = 'ol'
            # Extract number and content
            parts = stripped.split('. ', 1)
            content = parts[1] if len(parts) > 1 else ''
            content = process_formatting(content)
            result_lines.append(f'<li>{content}</li>')
        # Empty line
        elif not stripped:
            if in_list:
                result_lines.append(f'</{list_type}>')
                in_list = False
            result_lines.append('')
        # Regular paragraph
        else:
            if in_list:
                result_lines.append(f'</{list_type}>')
                in_list = False
            content = process_formatting(stripped)
            if content:
                result_lines.append(f'<p>{content}</p>')
        
        i += 1
    
    if in_list:
        result_lines.append(f'</{list_type}>')
    
    return '\n'.join(result_lines)

def process_formatting(text):
    """Process bold and italic formatting in text."""
    # Handle bold (**text**)
    while '**' in text:
        parts = text.split('**', 2)
        if len(parts) >= 3:
            text = parts[0] + '<strong>' + parts[1] + '</strong>' + ''.join(parts[2:])
        else:
            break
    
    # Handle italic (*text*) - but be careful not to break list markers
    # Only process italic if it's not at the start of a line followed by space or number
    parts = text.split('*')
    if len(parts) > 1:
        result = []
        for i, part in enumerate(parts):
            if i == 0:
                result.append(part)
            elif i % 2 == 1:
                # Check if this looks like a list marker
                if part.strip() and (part.strip()[0].isdigit() or part.strip().startswith('-')):
                    result.append('*' + part)
                else:
                    result.append('<em>' + part)
            else:
                result.append('</em>' + part)
        text = ''.join(result)
    
    return text

def generate_pdf_overview(player_row, position_profile, thorns_ranks, full_position_configs, output_dir, all_players_df=None, is_top15=False):
    """Generate PDF file for a single player."""
    player_name = str(player_row.get('Player', 'Unknown')).strip()
    team = str(player_row.get('Team', 'Unknown')).strip()
    conference = str(player_row.get('Conference', 'Unknown')).strip()
    
    # Create safe filename
    safe_name = player_name.replace(' ', '_').replace('.', '').replace('/', '_')
    filename = f"{safe_name}_{team.replace(' ', '_')}.pdf"
    
    # Determine subfolder based on position and top 15 status
    position_folder_map = {
        'Hybrid CB': 'Hybrid_CB',
        'DM Box-To-Box': 'DM_Box_To_Box',
        'AM Advanced Playmaker': 'AM_Advanced_Playmaker',
        'Right Touchline Winger': 'Right_Touchline_Winger'
    }
    folder_name = position_folder_map.get(position_profile, position_profile.replace(' ', '_'))
    
    # Place in Top 15 or Other subfolder
    if is_top15:
        subfolder = output_dir / 'Top 15' / folder_name
    else:
        subfolder = output_dir / 'Other' / folder_name
    
    subfolder.mkdir(parents=True, exist_ok=True)
    filepath = subfolder / filename
    
    # Get key stats
    total_score = player_row.get('Total Score', player_row.get('2025 Total Score', 'N/A'))
    conf_grade = player_row.get('Conference Grade', 'N/A')
    power_grade = player_row.get('Power Five Grade', 'N/A')
    prev_score = player_row.get('Previous Score', 'N/A')
    prev_year = player_row.get('Previous Year', 'N/A')
    change = player_row.get('Change From Previous', 'N/A')
    minutes = player_row.get('Total Minutes', 'N/A')
    pct_minutes = player_row.get('% of Team Minutes', 'N/A')
    seasons = player_row.get('Seasons Played', 'N/A')
    if seasons == 'N/A' or pd.isna(seasons) or str(seasons).strip() == '':
        seasons = 'Rookie'
    
    # Get new consistency metrics
    consistency_score = player_row.get('Consistency Score', 'N/A')
    metrics_above = player_row.get('Metrics Above Avg', 'N/A')
    metrics_below = player_row.get('Metrics Below Avg', 'N/A')
    metrics_at = player_row.get('Metrics At Avg', 'N/A')
    consistency_pct = player_row.get('Consistency %', 'N/A')
    
    # Get style fits and top 15s
    style_fits = player_row.get('Style Fits', 0)
    if pd.isna(style_fits):
        style_fits = 0
    else:
        style_fits = int(style_fits)
    
    top15s = player_row.get('Top 15s (Power Five)', 0)
    if pd.isna(top15s):
        top15s = 0
    else:
        top15s = int(top15s)
    
    # Get position-specific config for strengths/weaknesses
    position_name = POSITION_PROFILE_MAP.get(position_profile, position_profile)
    position_config = full_position_configs.get('position_profiles', {}).get(position_name, None) if full_position_configs else None
    
    # Identify strengths and weaknesses (with ranks)
    strengths = identify_strengths(player_row, position_config, all_players_df)
    weaknesses = identify_weaknesses(player_row, position_config, all_players_df)
    
    # Analyze Thorns fit (pass full configs for style fit metrics)
    fit_points = analyze_thorns_fit(player_row, position_profile, thorns_ranks, all_players_df, full_position_configs)
    
    # Generate summary
    summary = generate_player_summary(player_row, strengths, weaknesses, style_fits, fit_points, position_profile)
    
    # Generate markdown content
    lines = []
    lines.append(f"# {player_name}")
    lines.append("")
    lines.append(f"**{position_profile}** | {team} | {conference}")
    lines.append("")
    if summary:
        lines.append(f"*{summary}*")
        lines.append("")
    lines.append("## Overview")
    lines.append("")
    # Condense overview into fewer lines
    lines.append(f"**Score**: {total_score} | **Grades**: {conf_grade}/{power_grade} (Conference/Power Five) | **Minutes**: {pct_minutes}% (of team) | **Seasons**: {seasons}")
    if prev_score != 'N/A' and pd.notna(prev_score) and change != 'N/A' and pd.notna(change):
        lines.append(f"**Previous Score**: {prev_score} ({prev_year}) | **Change**: {change:+.2f}")
    lines.append("")
    
    # Performance Metrics Section (condensed)
    lines.append("## Performance Metrics")
    lines.append("")
    # Combine consistency metrics into one line
    consistency_line = []
    if consistency_score != 'N/A' and pd.notna(consistency_score):
        consistency_line.append(f"Score: {consistency_score:.1f}/100")
    if metrics_above != 'N/A' and pd.notna(metrics_above):
        consistency_line.append(f"Above average metrics: {metrics_above}")
    if metrics_below != 'N/A' and pd.notna(metrics_below):
        consistency_line.append(f"Below average metrics: {metrics_below}")
    if consistency_line:
        lines.append(f"**Consistency**: {' | '.join(consistency_line)}")
    
    # Combine elite performance into one line
    elite_line = []
    if top15s > 0:
        elite_line.append(f"Top 15s: {top15s}")
    if style_fits > 0:
        elite_line.append(f"PT Style Fits: {style_fits}")
    if elite_line:
        lines.append(f"**Elite Performance**: {' | '.join(elite_line)}")
    lines.append("")
    
    # Key Strengths (limit to top 3-4 to save space)
    lines.append("## Key Strengths")
    lines.append("")
    if strengths:
        for strength in strengths[:4]:  # Limit to top 4
            metric_display = format_metric_name(strength['metric'], add_per90=True)
            rank_parts = []
            if strength.get('power_five_rank') is not None and strength.get('power_five_total'):
                rank_parts.append(f"#{strength['power_five_rank']}/{strength['power_five_total']} Power Five")
            elif strength.get('conference_rank') is not None and strength.get('conference_total'):
                rank_parts.append(f"#{strength['conference_rank']}/{strength['conference_total']} in {conference}")
            rank_str = f" ({rank_parts[0]})" if rank_parts else ""
            lines.append(f"- **{metric_display}**: {strength['value']}{rank_str}")
    else:
        lines.append("*No significant strengths identified*")
    lines.append("")
    
    # Key Weaknesses (limit to top 3 to save space)
    lines.append("## Key Weaknesses")
    lines.append("")
    if weaknesses:
        for weakness in weaknesses[:3]:  # Limit to top 3
            metric_display = format_metric_name(weakness['metric'], add_per90=True)
            rank_parts = []
            if weakness.get('power_five_rank') is not None and weakness.get('power_five_total'):
                rank_parts.append(f"#{weakness['power_five_rank']}/{weakness['power_five_total']} Power Five")
            elif weakness.get('conference_rank') is not None and weakness.get('conference_total'):
                rank_parts.append(f"#{weakness['conference_rank']}/{weakness['conference_total']} in {conference}")
            rank_str = f" ({rank_parts[0]})" if rank_parts else ""
            lines.append(f"- **{metric_display}**: {weakness['value']}{rank_str}")
    else:
        lines.append("*No significant weaknesses identified*")
    lines.append("")
    
    # Portland Thorns Fit Analysis (condensed)
    lines.append("## Portland Thorns Fit Analysis")
    lines.append("")
    if fit_points:
        # Limit fit points and make them more concise
        for point in fit_points[:5]:  # Limit to 5 points
            # Remove numbering and make more compact
            clean_point = point.replace("**", "").replace("1. ", "").replace("2. ", "").replace("3. ", "").replace("4. ", "").replace("5. ", "")
            lines.append(f"- {clean_point}")
    else:
        lines.append("*Fit analysis based on available data*")
    lines.append("")
    
    # Add Call Notes & Assessment section if call data exists
    try:
        from enhance_player_overview_with_calls import get_player_call_data, generate_call_notes_section
        call_entries = get_player_call_data(player_name, team)
        if call_entries:
            call_section = generate_call_notes_section(call_entries)
            if call_section:
                lines.append(call_section)
                lines.append("")
    except Exception as e:
        # Silently fail if call data not available
        pass
    
    # Remove timestamp to save space
    # lines.append("---")
    # lines.append("")
    # lines.append(f"*Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
    # lines.append("")
    
    # Generate charts and comparison table
    charts_html = ""
    comparison_table_html = ""  # Extract comparison table to add to first page
    if CHARTS_AVAILABLE:
        try:
            # Get shortlist data for percentage metric averages
            # base_dir is the parent of output_dir
            base_dir_for_shortlist = output_dir.parent if output_dir else Path('/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search')
            shortlist_df = None
            try:
                # Try to find and load shortlist file
                possible_files = [
                    base_dir_for_shortlist / 'Portland Thorns 2025 Long Shortlist.xlsx',
                    base_dir_for_shortlist / 'Portland Thorns 2025 Short Shortlist.xlsx',
                    base_dir_for_shortlist / 'Portland Thorns 2025 Shortlist.xlsx',
                    base_dir_for_shortlist / 'AI Shortlist.xlsx'
                ]
                for shortlist_file in possible_files:
                    if shortlist_file.exists():
                        shortlist_df = load_player_data_from_shortlist(shortlist_file, base_dir_for_shortlist)
                        break
            except:
                pass
            
            charts = generate_all_charts(
                player_row, 
                position_profile, 
                position_config, 
                thorns_ranks, 
                full_position_configs, 
                strengths, 
                all_players_df,
                shortlist_df  # This should be passed and used for scatterplots
            )
            
            if charts:
                # Extract comparison table HTML for first page (before processing other charts)
                if charts.get('comparison_table'):
                    comparison_table_html = charts['comparison_table']
                # Convert charts to base64 images, prioritizing radar charts, then scatterplots
                chart_images = []
                # Add radar charts first
                for chart_name in ['radar_combined']:
                    if chart_name in charts:
                        chart_base64 = base64.b64encode(charts[chart_name].getvalue()).decode('utf-8')
                        chart_images.append(f'<img src="data:image/png;base64,{chart_base64}" style="max-width: 100%; margin: 10px 0;" />')
                # Add scatterplots (in order)
                for chart_name in sorted([k for k in charts.keys() if k.startswith('scatter_')]):
                    chart_base64 = base64.b64encode(charts[chart_name].getvalue()).decode('utf-8')
                    chart_images.append(f'<img src="data:image/png;base64,{chart_base64}" style="max-width: 100%; margin: 10px 0;" />')
                # Add other charts (skip comparison_table as it's HTML, not an image)
                for chart_name, chart_buffer in charts.items():
                    if chart_name not in ['radar_combined', 'comparison_table'] and not chart_name.startswith('scatter_'):
                        if hasattr(chart_buffer, 'getvalue'):
                            chart_base64 = base64.b64encode(chart_buffer.getvalue()).decode('utf-8')
                            chart_images.append(f'<img src="data:image/png;base64,{chart_base64}" style="max-width: 100%; margin: 10px 0;" />')
                
                if chart_images:
                    # Separate radar charts from other charts
                    radar_charts = [img for img in chart_images if 'radar' in str(img)]
                    other_charts = [img for img in chart_images if 'radar' not in str(img)]
                    
                    # Organize charts: radar charts first, then scatterplots in grid, then other charts
                    scatter_charts = [img for img in chart_images if 'scatter' in str(img)]
                    other_charts_no_scatter = [img for img in other_charts if 'scatter' not in str(img)]
                    
                    # Build radar charts HTML (inline, no wrapper div) with summary
                    radar_html = ''
                    if radar_charts:
                        radar_summary = '<p style="font-size: 7pt; color: #666; margin: 0.1em 0 0.2em 0;">Performance across 8 key metrics compared to Power Five average. Larger polygon indicates stronger performance.</p>'
                        radar_html = f'<div style="max-width: 100%;">{radar_summary}{"".join([f"<div style=\"max-width: 100%;\">{img}</div>" for img in radar_charts[:2]])}</div>'
                    
                    # Build scatterplots HTML (inline, no wrapper div)
                    scatter_html = ''
                    if scatter_charts:
                        scatter_html = ''.join([f'<div style="max-width: 100%;">{img}</div>' for img in scatter_charts[:3]])
                    
                    # Build other charts HTML (inline, no wrapper div)
                    other_html = ''
                    if other_charts_no_scatter:
                        other_html = ''.join([f'<div style="max-width: 100%;">{img}</div>' for img in other_charts_no_scatter[:4]])
                    
                    # Build beeswarm chart HTML
                    beeswarm_html = ''
                    if charts.get('total_score_beeswarm'):
                        beeswarm_buffer = charts['total_score_beeswarm']
                        if hasattr(beeswarm_buffer, 'getvalue'):
                            beeswarm_img = base64.b64encode(beeswarm_buffer.getvalue()).decode("utf-8")
                            beeswarm_html = f'<img src="data:image/png;base64,{beeswarm_img}" style="max-width: 100%; height: auto;" />'
                    
                    charts_html = f"""
        <div style="page-break-before: always;">
            <h2 style="color: #8B0000; font-size: 10pt; margin-top: 0.1em; margin-bottom: 0.1em; border-bottom: 2px solid #8B0000; padding-bottom: 0.05em;">
                Performance Visualizations
            </h2>
            <div style="display: grid; grid-template-columns: 1fr; gap: 2px; margin-top: 0.1em; margin-bottom: 0.1em;">
                {radar_html}
            </div>
            <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 3px; margin-top: 0.1em; margin-bottom: 0.1em;">
                {scatter_html}
            </div>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 3px; margin-top: 0.1em; margin-bottom: 0.1em;">
                {other_html}
            </div>
            <div style="margin-top: 0.2em; margin-bottom: 0.1em;">
                {beeswarm_html}
            </div>
        </div>
        """
        except Exception as e:
            print(f"  ⚠️  Error generating charts: {e}")
            charts_html = ""
    
    # Convert markdown to HTML
    markdown_content = '\n'.join(lines)
    if MARKDOWN_AVAILABLE:
        html_content = markdown.markdown(markdown_content, extensions=['extra', 'nl2br'])
    else:
        # Simple markdown to HTML conversion (basic support)
        html_content = convert_markdown_to_html(markdown_content)
    
    # Add comparison table to the bottom of first page (before page break)
    # Add a large spacer to push the table to the bottom of the page
    if comparison_table_html:
        # Use a large margin-top to push the table to the bottom of the first page
        html_content += f'''
        <div style="margin-top: 500px; page-break-after: always; page-break-inside: avoid;">
            {comparison_table_html}
        </div>
        '''
    
    # Create styled HTML document
    html_document = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: letter;
                margin: 0.5in;
            }}
            * {{
                page-break-inside: avoid;
            }}
            body {{
                font-family: 'Helvetica', 'Arial', sans-serif;
                font-size: 9pt;
                line-height: 1.4;
                color: #333;
            }}
            h1 {{
                color: #1a1a1a;
                font-size: 18pt;
                margin-bottom: 0.2em;
                margin-top: 0;
                border-bottom: 2px solid #8B0000;
                padding-bottom: 0.15em;
            }}
            h2 {{
                color: #8B0000;
                font-size: 12pt;
                margin-top: 0.5em;
                margin-bottom: 0.25em;
                border-bottom: 1px solid #e0e0e0;
                padding-bottom: 0.1em;
                page-break-after: avoid;
            }}
            h3 {{
                color: #555;
                font-size: 10pt;
                margin-top: 0.5em;
                margin-bottom: 0.25em;
            }}
            p {{
                margin: 0.3em 0;
            }}
            ul, ol {{
                margin: 0.3em 0;
                padding-left: 1.2em;
            }}
            li {{
                margin: 0.15em 0;
            }}
            strong {{
                color: #1a1a1a;
                font-weight: 600;
            }}
            hr {{
                border: none;
                border-top: 1px solid #e0e0e0;
                margin: 0.8em 0;
            }}
            .header-info {{
                font-size: 12pt;
                color: #666;
                margin-bottom: 1em;
            }}
            .overview-section {{
                background-color: #f9f9f9;
                padding: 1em;
                border-left: 4px solid #8B0000;
                margin: 1em 0;
            }}
            .strengths-section {{
                background-color: #f0f8f0;
                padding: 0.8em;
                border-left: 4px solid #4CAF50;
                margin: 1em 0;
            }}
            .weaknesses-section {{
                background-color: #fff5f5;
                padding: 0.8em;
                border-left: 4px solid #f44336;
                margin: 1em 0;
            }}
            .fit-section {{
                background-color: #f5f5ff;
                padding: 0.8em;
                border-left: 4px solid #2196F3;
                margin: 1em 0;
            }}
            .footer {{
                font-size: 9pt;
                color: #999;
                margin-top: 2em;
                text-align: center;
            }}
        </style>
    </head>
    <body>
        {html_content}
        {charts_html}
    </body>
    </html>
    """
    
    # Generate PDF using weasyprint
    if WEASYPRINT_AVAILABLE:
        try:
            HTML(string=html_document).write_pdf(filepath)
            return filepath
        except Exception as e:
            print(f"  ⚠️  Error generating PDF for {player_name}: {e}")
            # Fallback: save as HTML
            html_filepath = filepath.with_suffix('.html')
            html_filepath.write_text(html_document)
            return html_filepath
    else:
        # Fallback: save as HTML if weasyprint not available
        html_filepath = filepath.with_suffix('.html')
        html_filepath.write_text(html_document)
        print(f"  ⚠️  PDF generation not available, saved as HTML: {html_filepath.name}")
        return html_filepath

def main():
    base_dir = Path('/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search')
    
    # Load shortlist file (try multiple possible files)
    possible_shortlist_files = [
        base_dir / 'Portland Thorns 2025 Shortlist.xlsx',
        base_dir / 'Portland Thorns 2025 Long Shortlist.xlsx',
        base_dir / 'Portland Thorns 2025 Short Shortlist.xlsx',
        base_dir / 'AI Shortlist.xlsx',
    ]
    
    shortlist_file = None
    for file_path in possible_shortlist_files:
        if file_path.exists():
            shortlist_file = file_path
            break
    
    if not shortlist_file:
        print(f"❌ No shortlist file found. Tried:")
        for f in possible_shortlist_files:
            print(f"   - {f}")
        return
    
    print(f"✅ Using shortlist file: {shortlist_file.name}")
    
    print("="*80)
    print("GENERATING PLAYER OVERVIEWS FOR B+ PLAYERS")
    print("="*80)
    
    print("\n📊 Loading player data from shortlist...")
    players_df = load_player_data_from_shortlist(shortlist_file, base_dir)
    print(f"  ✅ Loaded {len(players_df)} players from shortlist")
    
    # Also load ALL players from conference reports for ranking calculations
    print("\n📊 Loading all players from conference reports for ranking...")
    all_players_for_ranking = load_player_data_from_conference_reports(base_dir)
    print(f"  ✅ Loaded {len(all_players_for_ranking)} total players for ranking")
    
    # Load NWSL data
    nwsl_dir = base_dir / 'Exports' / 'Team Stats By Conference' / 'NWSL'
    print(f"\n📊 Loading NWSL data from {nwsl_dir}...")
    league_df, league_avg, thorns_data, thorns_ranks = load_nwsl_data(nwsl_dir)
    if league_df is not None:
        print(f"  ✅ Loaded {len(league_df)} NWSL teams")
        if thorns_data is not None:
            print(f"  ✅ Found Portland Thorns data")
        if thorns_ranks:
            print(f"  ✅ Calculated Portland Thorns ranks for {len(thorns_ranks)} metrics")
        else:
            print(f"  ⚠️  Could not calculate Portland Thorns ranks")
    else:
        print(f"  ⚠️  Could not load NWSL data")
        thorns_ranks = None
    
    # Load position config
    config_file = base_dir / 'Scripts' / '00_Keep' / 'position_metrics_config.json'
    position_configs_dict = {}
    full_position_configs = None
    if config_file.exists():
        with open(config_file, 'r') as f:
            full_position_configs = json.load(f)
            for pos_name, pos_config in full_position_configs.get('position_profiles', {}).items():
                position_configs_dict[pos_name] = pos_config
        print(f"\n📊 Loaded position configs for {len(position_configs_dict)} positions")
    else:
        print(f"\n⚠️  Position config file not found: {config_file}")
    
    # Create output directory structure
    output_dir = base_dir / 'Player Overviews'
    output_dir.mkdir(exist_ok=True)
    
    # Create Top 15 and Other subfolders
    position_folder_map = {
        'Hybrid CB': 'Hybrid_CB',
        'DM Box-To-Box': 'DM_Box_To_Box',
        'AM Advanced Playmaker': 'AM_Advanced_Playmaker',
        'Right Touchline Winger': 'Right_Touchline_Winger'
    }
    
    for position_profile in ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']:
        folder_name = position_folder_map[position_profile]
        (output_dir / 'Top 15' / folder_name).mkdir(parents=True, exist_ok=True)
        (output_dir / 'Other' / folder_name).mkdir(parents=True, exist_ok=True)
    
    print(f"\n📁 Output directory: {output_dir}")
    print(f"📁 Created Top 15/ and Other/ subfolders for each position")
    
    # Load AI Shortlist to determine top 15 players
    ai_shortlist_file = base_dir / 'AI Shortlist.xlsx'
    top_15_players = set()
    
    if ai_shortlist_file.exists():
        print(f"\n📊 Loading AI Shortlist to identify Top 15 players...")
        from openpyxl import load_workbook
        wb_ai = load_workbook(ai_shortlist_file, data_only=True)
        
        for position_profile in ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']:
            if position_profile in wb_ai.sheetnames:
                ws = wb_ai[position_profile]
                for row_idx in range(4, ws.max_row + 1):
                    player = ws.cell(row=row_idx, column=2).value
                    team = ws.cell(row=row_idx, column=3).value
                    if player and team:
                        safe_name = str(player).replace(' ', '_').replace('.', '').replace('/', '_')
                        safe_team = str(team).replace(' ', '_')
                        filename = f"{safe_name}_{safe_team}.pdf"
                        top_15_players.add(filename)
        
        wb_ai.close()
        print(f"  ✅ Identified {len(top_15_players)} Top 15 players")
    else:
        print(f"  ⚠️  AI Shortlist not found - all players will go to Other/")
    
    # Generate overviews for each player
    print(f"\n📝 Generating player overviews...")
    generated_count = 0
    top15_count = 0
    other_count = 0
    
    for idx, (_, player_row) in enumerate(players_df.iterrows(), 1):
        player_name = str(player_row.get('Player', 'Unknown')).strip()
        position_profile = player_row.get('Position Profile', 'Unknown')
        team = str(player_row.get('Team', 'Unknown')).strip()
        
        # Determine if this player is in Top 15
        safe_name = player_name.replace(' ', '_').replace('.', '').replace('/', '_')
        safe_team = team.replace(' ', '_')
        filename = f"{safe_name}_{safe_team}.pdf"
        is_top15 = filename in top_15_players
        
        # Get position config
        position_name = POSITION_PROFILE_MAP.get(position_profile, position_profile)
        position_config = position_configs_dict.get(position_name, None)
        
        try:
            filepath = generate_pdf_overview(
                player_row, position_profile, thorns_ranks, 
                full_position_configs, output_dir, all_players_df=all_players_for_ranking,
                is_top15=is_top15
            )
            generated_count += 1
            if is_top15:
                top15_count += 1
            else:
                other_count += 1
            if idx % 10 == 0:
                print(f"  ✅ Generated {idx}/{len(players_df)} overviews... (Top 15: {top15_count}, Other: {other_count})")
        except Exception as e:
            print(f"  ❌ Error generating overview for {player_name}: {e}")
    
    print(f"\n✅ Generated {generated_count} player overviews")
    print(f"  - Top 15: {top15_count} files")
    print(f"  - Other: {other_count} files")
    print(f"📁 Files saved to: {output_dir}/Top 15/ and {output_dir}/Other/")
    if not WEASYPRINT_AVAILABLE:
        print(f"\n⚠️  Note: weasyprint not installed. To generate PDFs instead of HTML:")
        print(f"   Install with: pip3 install --user weasyprint")
        print(f"   Or: pip3 install weasyprint (if using virtual environment)")
        print(f"   Files were saved as HTML instead of PDF.")
        print(f"   You can convert HTML to PDF manually or install weasyprint and rerun.")

if __name__ == '__main__':
    main()
