#!/usr/bin/env python3
"""
Create PAdj Score Comparison Spreadsheet
========================================

Generates a comparison spreadsheet showing pre-PAdj vs post-PAdj scores
for all players from the four championship teams.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_championship_teams():
    """Get the four ACC championship teams"""
    return ['Duke Blue Devils', 'Notre Dame Fighting Irish', 'Stanford Cardinal', 'Virginia Cavaliers']


def load_scores_from_report(report_path, sheet_name):
    """Load player scores from a report sheet"""
    wb = load_workbook(report_path, keep_vba=False, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        return []
    
    ws = wb[sheet_name]
    players = []
    
    # Headers are in row 2, data starts at row 3
    for row in range(3, ws.max_row + 1):
        player = ws.cell(row=row, column=1).value  # Column A: Player
        team = ws.cell(row=row, column=2).value if ws.max_column >= 2 else None  # May not be in position tabs
        position = ws.cell(row=row, column=3).value if ws.max_column >= 3 else None
        position_profile = ws.cell(row=row, column=4).value if ws.max_column >= 4 else None
        score = ws.cell(row=row, column=6).value  # Column F: 2025 Total Score
        
        if player and score is not None:
            players.append({
                'Player': player,
                'Team': team,
                'Position': position,
                'Position_Profile': position_profile,
                'Score': float(score) if score else None
            })
    
    return players


def load_all_team_scores(report_path):
    """Load scores from all team tabs"""
    wb = load_workbook(report_path, keep_vba=False, data_only=True)
    
    championship_teams = get_championship_teams()
    all_players = []
    
    for team in championship_teams:
        if team in wb.sheetnames:
            ws = wb[team]
            
            # Headers are in row 2, data starts at row 3
            for row in range(3, ws.max_row + 1):
                player = ws.cell(row=row, column=1).value  # Column A: Player
                position = ws.cell(row=row, column=2).value  # Column B: Position
                position_profile = ws.cell(row=row, column=3).value  # Column C: Position Profile
                score = ws.cell(row=row, column=6).value  # Column F: 2025 Total Score
                
                if player and score is not None:
                    all_players.append({
                        'Player': player,
                        'Team': team,
                        'Position': position,
                        'Position_Profile': position_profile,
                        'Score': float(score) if score else None
                    })
    
    return all_players


def create_comparison_spreadsheet(base_dir):
    """Create comparison spreadsheet of pre-PAdj vs post-PAdj scores"""
    
    regular_report = base_dir / "Mike_Norris_Scouting_Report_ACC_IntraConference_70_30.xlsx"
    padj_report = base_dir / "Mike_Norris_Scouting_Report_ACC_IntraConference_70_30_PAdj.xlsx"
    
    if not regular_report.exists():
        print(f"❌ Regular report not found: {regular_report}")
        return
    
    if not padj_report.exists():
        print(f"❌ PAdj report not found: {padj_report}")
        return
    
    print("Loading scores from reports...")
    
    # Load scores from team tabs (more comprehensive)
    print("  Loading pre-PAdj scores...")
    pre_padj_players = load_all_team_scores(regular_report)
    print(f"    ✅ Found {len(pre_padj_players)} players")
    
    print("  Loading post-PAdj scores...")
    post_padj_players = load_all_team_scores(padj_report)
    print(f"    ✅ Found {len(post_padj_players)} players")
    
    # Create dictionaries for easier lookup
    pre_dict = {}
    for p in pre_padj_players:
        key = (p['Player'], p['Team'])
        if key not in pre_dict:
            pre_dict[key] = p
    
    post_dict = {}
    for p in post_padj_players:
        key = (p['Player'], p['Team'])
        if key not in post_dict:
            post_dict[key] = p
    
    # Combine data
    comparison_data = []
    all_keys = set(pre_dict.keys()) | set(post_dict.keys())
    
    for key in sorted(all_keys):
        player_name, team = key
        pre_data = pre_dict.get(key, {})
        post_data = post_dict.get(key, {})
        
        pre_score = pre_data.get('Score', None)
        post_score = post_data.get('Score', None)
        
        # Use data from whichever source has it
        position = pre_data.get('Position') or post_data.get('Position', '')
        position_profile = pre_data.get('Position_Profile') or post_data.get('Position_Profile', '')
        
        if pre_score is not None and post_score is not None:
            difference = post_score - pre_score
            percent_change = (difference / pre_score * 100) if pre_score != 0 else 0
        else:
            difference = None
            percent_change = None
        
        comparison_data.append({
            'Player': player_name,
            'Team': team,
            'Position': position,
            'Position Profile': position_profile,
            'Pre-PAdj Score': pre_score,
            'Post-PAdj Score': post_score,
            'Difference': difference,
            '% Change': percent_change
        })
    
    # Create DataFrame
    df = pd.DataFrame(comparison_data)
    
    # Sort by team, then by position profile, then by difference (descending)
    df = df.sort_values(['Team', 'Position Profile', 'Difference'], ascending=[True, True, False])
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PAdj Score Comparison"
    
    # Headers
    headers = ['Player', 'Team', 'Position', 'Position Profile', 
               'Pre-PAdj Score', 'Post-PAdj Score', 'Difference', '% Change']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format numeric columns
            if col_idx >= 5:  # Score columns
                if value is not None:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Highlight positive/negative differences
            if col_idx == 7 and value is not None:  # Difference column
                if value > 0:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif value < 0:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
            
            # Format percentage column
            if col_idx == 8 and value is not None:  # % Change column
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Auto-adjust column widths
    column_widths = {
        'A': 25,  # Player
        'B': 30,  # Team
        'C': 15,  # Position
        'D': 25,  # Position Profile
        'E': 15,  # Pre-PAdj Score
        'F': 15,  # Post-PAdj Score
        'G': 12,  # Difference
        'H': 12   # % Change
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 20
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Add summary statistics by team (on a separate sheet)
    ws_summary = wb.create_sheet(title="Summary by Team")
    
    # Headers
    summary_headers = ['Team', 'Players', 'Avg Pre-PAdj', 'Avg Post-PAdj', 'Avg Difference', 'Max Increase', 'Max Decrease']
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary data
    row = 2
    for team in get_championship_teams():
        team_df = df[df['Team'] == team]
        if len(team_df) > 0:
            avg_pre = team_df['Pre-PAdj Score'].mean()
            avg_post = team_df['Post-PAdj Score'].mean()
            avg_diff = team_df['Difference'].mean()
            max_increase = team_df['Difference'].max()
            max_decrease = team_df['Difference'].min()
            
            ws_summary.cell(row=row, column=1, value=team)
            ws_summary.cell(row=row, column=2, value=len(team_df))
            ws_summary.cell(row=row, column=3, value=avg_pre).number_format = '0.00'
            ws_summary.cell(row=row, column=4, value=avg_post).number_format = '0.00'
            ws_summary.cell(row=row, column=5, value=avg_diff).number_format = '0.00'
            ws_summary.cell(row=row, column=6, value=max_increase).number_format = '0.00'
            ws_summary.cell(row=row, column=7, value=max_decrease).number_format = '0.00'
            
            # Format difference column
            diff_cell = ws_summary.cell(row=row, column=5)
            if avg_diff > 0:
                diff_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif avg_diff < 0:
                diff_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
    
    # Auto-adjust summary column widths
    ws_summary.column_dimensions['A'].width = 30
    for col in range(2, 8):
        ws_summary.column_dimensions[get_column_letter(col)].width = 15
    
    # Save
    output_file = base_dir / "PAdj_Score_Comparison.xlsx"
    wb.save(output_file)
    print(f"\n✅ Comparison spreadsheet saved: {output_file.name}")
    print(f"   Total players: {len(df)}")
    print(f"   Players with score changes: {len(df[df['Difference'].notna() & (df['Difference'] != 0)])}")
    
    return output_file


def main():
    """Main function"""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    print("="*80)
    print("CREATING PADJ SCORE COMPARISON SPREADSHEET")
    print("="*80)
    
    create_comparison_spreadsheet(base_dir)
    
    print("\n" + "="*80)
    print("✅ COMPARISON SPREADSHEET CREATED")
    print("="*80)


if __name__ == "__main__":
    main()

