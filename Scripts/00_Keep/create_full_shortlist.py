#!/usr/bin/env python3
"""
Create a Full Shortlist report with ALL players from all conferences.
No filters applied - includes all players with their scores, grades, and metrics.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import sys

# Add path for imports
sys.path.insert(0, str(Path(__file__).parent))

from update_mike_norris_reports import (
    load_conference_season_data,
    get_relevant_metrics_for_position,
    POSITION_PROFILE_MAP,
    PLAYER_FILE_PREFIXES
)
from create_top_15_report import (
    load_seasons_data,
    get_relevant_metrics_for_position as get_relevant_metrics_from_config
)

# Grade colors (matching update_mike_norris_reports.py)
GRADE_COLORS = {
    'A': '8B0000',  # Dark red
    'B': 'C5504B',  # Red
    'C': 'F2A2A2',  # Light red
    'D': '8FAADC',  # Light blue
    'F': '1F4E79'   # Dark blue
}

# Position profile to internal name mapping
POSITION_PROFILE_MAP_FULL = {
    'Hybrid CB': 'Center Back',
    'DM Box-To-Box': 'Centre Midfielder',
    'AM Advanced Playmaker': 'Attacking Midfielder',
    'Right Touchline Winger': 'Winger'
}

# Position profile to initials mapping (for Changed Position column)
POSITION_INITIALS = {
    'Hybrid CB': 'HCB',
    'DM Box-To-Box': 'DM',
    'AM Advanced Playmaker': 'AM',
    'Right Touchline Winger': 'RTW'
}

def load_all_players_from_reports(base_dir):
    """Load ALL players from conference reports (which already have scores and grades)."""
    from create_top_15_report import load_all_players_from_reports as load_from_reports
    
    print("\n📥 Loading ALL players from conference reports...")
    print("   (Conference reports include all players with scores and grades)")
    
    # Use the existing function from create_top_15_report
    all_data = load_from_reports(base_dir)
    
    return all_data


def create_full_shortlist(base_dir):
    """Create Excel report with ALL players per position profile."""
    print("="*80)
    print("CREATING PORTLAND THORNS 2025 FULL SHORTLIST REPORT")
    print("="*80)
    
    # Load config to get relevant metrics
    config_file = base_dir / "Scripts" / "00_Keep" / "position_metrics_config.json"
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    # Load ALL players from conference reports (which already have scores/grades)
    all_data = load_all_players_from_reports(base_dir)
    
    if not all_data:
        print("❌ No data loaded!")
        return
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Position profiles in order
    position_profiles = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    note_font = Font(italic=True, size=9)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for position_profile in position_profiles:
        if position_profile not in all_data:
            print(f"  ⚠️  No data for {position_profile}")
            continue
        
        df = all_data[position_profile]
        if df.empty:
            print(f"  ⚠️  No data for {position_profile}")
            continue
        
        print(f"\n📊 Processing {position_profile}...")
        print(f"   Total players: {len(df)}")
        
        # Find the score column
        score_col = None
        for col in ['Total_Score_1_10', '2025 Total Score', 'Total Score', 'Score']:
            if col in df.columns:
                score_col = col
                break
        
        if score_col is None:
            print(f"  ⚠️  No score column found for {position_profile}")
            continue
        
        # Convert score column to numeric
        df[score_col] = pd.to_numeric(df[score_col], errors='coerce')
        df = df[df[score_col].notna()].copy()
        
        # Sort by score descending
        df = df.sort_values(score_col, ascending=False).reset_index(drop=True)
        
        # Add rank
        df['Rank'] = range(1, len(df) + 1)
        
        # Find grade columns
        conf_grade_col = None
        power_grade_col = None
        
        for col in ['Conference_Grade', 'Conference Grade']:
            if col in df.columns:
                conf_grade_col = col
                break
        
        for col in ['Power_Five_Grade', 'Power Five Grade']:
            if col in df.columns:
                power_grade_col = col
                break
        
        # Get position name for config
        position_name = POSITION_PROFILE_MAP_FULL.get(position_profile, position_profile)
        
        # Get desired metric order (same as shortlist)
        desired_metric_order_map = {
            'Hybrid CB': [
                'Defensive duels per 90',
                'Defensive duels won, %',
                'Aerial duels per 90',
                'Aerial duels won, %',
                'Interceptions + Sliding Tackles',
                'Shots blocked per 90',
                'Passes per 90',
                'Accurate passes, %',
                'Progressive passes per 90',
                'Accurate progressive passes, %',
                'Long passes per 90',
                'Accurate long passes, %',
                'Short / medium passes per 90',
                'Accurate short / medium passes, %'
            ],
            'DM Box-To-Box': [
                'Defensive duels per 90',
                'Defensive duels won, %',
                'Interceptions per 90',
                'Passes per 90',
                'Accurate passes, %',
                'Progressive passes per 90',
                'Accurate progressive passes, %',
                'Received passes per 90',
                'Dribbles per 90',
                'Successful dribbles, %',
                'Offensive duels per 90',
                'Offensive duels won, %'
            ],
            'AM Advanced Playmaker': [
                'Goals per 90',
                'Assists per 90',
                'Smart passes per 90',
                'Accurate smart passes, %',
                'Through passes per 90',
                'Accurate through passes, %',
                'Deep completions per 90',
                'Dribbles per 90',
                'Successful dribbles, %',
                'Offensive duels per 90',
                'Offensive duels won, %',
                'Passes per 90',
                'Accurate passes, %',
                'xG per 90 - Goals per 90'
            ],
            'Right Touchline Winger': [
                'Assists per 90',
                'Crosses per 90',
                'Accurate crosses, %',
                'Dribbles per 90',
                'Successful dribbles, %',
                'Offensive duels per 90',
                'Offensive duels won, %',
                'Progressive runs per 90',
                'Received long passes per 90'
            ]
        }
        
        metric_headers = desired_metric_order_map.get(position_profile, [])
        
        # Create worksheet
        ws = wb.create_sheet(title=position_profile)
        
        # Title row
        ws.merge_cells('A1:Z1')
        title_cell = ws['A1']
        title_cell.value = f"Portland Thorns 2025 Full Shortlist - {position_profile}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Note row
        ws.merge_cells('A2:Z2')
        note_cell = ws['A2']
        note_cell.value = f"All players from ACC, SEC, BIG10, BIG12, and IVY conferences. Total: {len(df)} players."
        note_cell.font = note_font
        note_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Base headers
        base_headers = [
            'Rank', 'Player', 'Team', 'Conference', 'Position', 
            'Total Score', 'Conference Grade', 'Power Five Grade',
            'Previous Year', 'Previous Score', 'Change From Previous', 'Changed Position', 
            'Total Minutes', '% of Team Minutes', 'Seasons Played', 'Top 15s (Power Five)'
        ]
        
        # Add metric headers
        all_headers = base_headers + metric_headers
        
        # Write headers (row 3)
        for col_idx, header in enumerate(all_headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Column widths
        column_widths = {
            'A': 8,   # Rank
            'B': 20,  # Player
            'C': 25,  # Team
            'D': 12,  # Conference
            'E': 15,  # Position
            'F': 12,  # Total Score
            'G': 18,  # Conference Grade
            'H': 18,  # Power Five Grade
            'I': 15,  # Previous Year
            'J': 12,  # Previous Score
            'K': 15,  # Change
            'L': 18,  # Changed Position
            'M': 15,  # Minutes
            'N': 18,  # % of Team Minutes
            'O': 15,  # Seasons Played
            'P': 10   # Top 15s
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set metric column widths
        for col_idx in range(len(base_headers) + 1, len(all_headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
        
        # Find column names in dataframe (handle variations)
        score_col = None
        conf_grade_col = None
        power_grade_col = None
        
        for col in df.columns:
            col_lower = str(col).lower()
            if 'total score' in col_lower and ('2025' in col_lower or 'score_1_10' in col_lower):
                score_col = col
            elif 'conference grade' in col_lower or 'conference_grade' in col_lower:
                conf_grade_col = col
            elif 'power five grade' in col_lower or 'power_five_grade' in col_lower:
                power_grade_col = col
        
        # Write data rows
        for row_idx, (_, player_row) in enumerate(df.iterrows(), start=4):
            # Base columns
            ws.cell(row=row_idx, column=1, value=player_row.get('Rank', ''))
            ws.cell(row=row_idx, column=2, value=player_row.get('Player', ''))
            ws.cell(row=row_idx, column=3, value=player_row.get('Team', ''))
            ws.cell(row=row_idx, column=4, value=player_row.get('Conference', ''))
            ws.cell(row=row_idx, column=5, value=position_profile)
            
            # Score - try multiple column name variations
            score_val = None
            if score_col:
                score_val = player_row.get(score_col, '')
            else:
                # Try alternative column names
                for alt_col in ['Total_Score_1_10', '2025 Total Score', 'Total Score', 'Score']:
                    if alt_col in player_row.index:
                        score_val = player_row.get(alt_col, '')
                        break
            
            if pd.notna(score_val) and score_val != '':
                try:
                    ws.cell(row=row_idx, column=6, value=round(float(score_val), 2))
                except:
                    ws.cell(row=row_idx, column=6, value=score_val)
            else:
                ws.cell(row=row_idx, column=6, value='')
            
            # Grades
            conf_grade = player_row.get(conf_grade_col, '') if conf_grade_col else player_row.get('Conference Grade', player_row.get('Conference_Grade', ''))
            power_grade = player_row.get(power_grade_col, '') if power_grade_col else player_row.get('Power Five Grade', player_row.get('Power_Five_Grade', ''))
            
            ws.cell(row=row_idx, column=7, value=conf_grade)
            if conf_grade and conf_grade in GRADE_COLORS:
                ws.cell(row=row_idx, column=7).fill = PatternFill(
                    start_color=GRADE_COLORS[conf_grade], 
                    end_color=GRADE_COLORS[conf_grade], 
                    fill_type="solid"
                )
            
            ws.cell(row=row_idx, column=8, value=power_grade)
            if power_grade and power_grade in GRADE_COLORS:
                ws.cell(row=row_idx, column=8).fill = PatternFill(
                    start_color=GRADE_COLORS[power_grade], 
                    end_color=GRADE_COLORS[power_grade], 
                    fill_type="solid"
                )
            
            # Previous score info
            prev_year = player_row.get('Previous Year', '')
            prev_score = player_row.get('Previous Score', '')
            
            # Calculate Change From Previous if not already present
            change_from_prev = player_row.get('Change From Previous', None)
            if pd.isna(change_from_prev) or change_from_prev == '':
                # Try to calculate it
                if pd.notna(score_val) and score_val != '' and pd.notna(prev_score) and prev_score != '':
                    try:
                        score_float = float(score_val)
                        prev_float = float(prev_score)
                        change_from_prev = round(score_float - prev_float, 2)
                    except:
                        change_from_prev = ''
                else:
                    change_from_prev = ''
            
            ws.cell(row=row_idx, column=9, value=prev_year)
            
            # Previous Score
            if pd.notna(prev_score) and prev_score != '':
                try:
                    ws.cell(row=row_idx, column=10, value=round(float(prev_score), 2))
                except:
                    ws.cell(row=row_idx, column=10, value=prev_score)
            else:
                ws.cell(row=row_idx, column=10, value='')
            
            # Change From Previous
            if pd.notna(change_from_prev) and change_from_prev != '':
                try:
                    ws.cell(row=row_idx, column=11, value=round(float(change_from_prev), 2))
                except:
                    ws.cell(row=row_idx, column=11, value=change_from_prev)
            else:
                ws.cell(row=row_idx, column=11, value='')
            ws.cell(row=row_idx, column=12, value=player_row.get('Changed Position', ''))
            
            # Minutes
            minutes_col = None
            for col in ['Total Minutes', 'Minutes played', 'Minutes']:
                if col in player_row.index:
                    minutes_col = col
                    break
            ws.cell(row=row_idx, column=13, value=player_row.get(minutes_col, '') if minutes_col else '')
            
            # % of Team Minutes
            pct_col = None
            for col in ['% of Team Minutes', 'Pct_Of_Team_Minutes', '% of team minutes']:
                if col in player_row.index:
                    pct_col = col
                    break
            ws.cell(row=row_idx, column=14, value=player_row.get(pct_col, '') if pct_col else '')
            
            # Seasons Played
            ws.cell(row=row_idx, column=15, value=player_row.get('Seasons Played', ''))
            
            # Top 15s (Power Five) - will calculate after all rows are written
            ws.cell(row=row_idx, column=16, value=0)  # Placeholder
            
            # Metric columns
            for metric_idx, metric_header in enumerate(metric_headers, start=len(base_headers) + 1):
                # Find matching column in player_row
                metric_value = 0  # Default to 0 instead of None
                
                if metric_header == "Interceptions + Sliding Tackles":
                    # Combined metric - try to find components
                    interceptions_col = None
                    sliding_col = None
                    for col in player_row.index:
                        if 'interception' in str(col).lower() and 'per 90' in str(col).lower():
                            interceptions_col = col
                        if 'sliding' in str(col).lower() and 'per 90' in str(col).lower():
                            sliding_col = col
                    
                    metric_value = 0  # Default to 0
                    if interceptions_col and sliding_col:
                        try:
                            inter_val = float(player_row.get(interceptions_col, 0) or 0)
                            slide_val = float(player_row.get(sliding_col, 0) or 0)
                            metric_value = inter_val + slide_val
                        except:
                            metric_value = 0
                    elif interceptions_col:
                        try:
                            inter_val = float(player_row.get(interceptions_col, 0) or 0)
                            metric_value = inter_val
                        except:
                            metric_value = 0
                    elif sliding_col:
                        try:
                            slide_val = float(player_row.get(sliding_col, 0) or 0)
                            metric_value = slide_val
                        except:
                            metric_value = 0
                    # metric_value is already set to 0 if no components found
                    # Skip the rest of the matching logic for this metric
                elif metric_header == "xG per 90 - Goals per 90":
                    # Calculate difference between xG per 90 and Goals per 90
                    xg_val = None
                    goals_val = None
                    
                    # Find xG per 90 column - try multiple variations
                    for col in player_row.index:
                        col_lower = str(col).lower()
                        if ('xg per 90' in col_lower or 'expected goals per 90' in col_lower) and 'goals' not in col_lower and 'against' not in col_lower:
                            test_val = player_row.get(col)
                            if pd.notna(test_val) and test_val != '':
                                xg_val = test_val
                                break
                    
                    # Find Goals per 90 column - try multiple variations
                    for col in player_row.index:
                        col_lower = str(col).lower()
                        if 'goals per 90' in col_lower and 'xg' not in col_lower and 'non-penalty' not in col_lower and 'head' not in col_lower:
                            test_val = player_row.get(col)
                            if pd.notna(test_val) and test_val != '':
                                goals_val = test_val
                                break
                    
                    # If still not found, try loading from raw data
                    if (pd.isna(xg_val) or xg_val == '') and position_profile == 'AM Advanced Playmaker':
                        try:
                            from update_mike_norris_reports import load_conference_season_data
                            player_name = str(player_row.get('Player', '')).strip()
                            team_name = str(player_row.get('Team', '')).strip()
                            
                            # Try to load from raw data
                            raw_data_frames = []
                            for conf in ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']:
                                try:
                                    raw_df = load_conference_season_data(base_dir, conf, 2025, 'Attacking Midfielder')
                                    if raw_df is not None and not raw_df.empty:
                                        raw_data_frames.append(raw_df)
                                except:
                                    pass
                            
                            if raw_data_frames:
                                raw_combined = pd.concat(raw_data_frames, ignore_index=True)
                                player_match = raw_combined[
                                    (raw_combined['Player'].astype(str).str.strip().str.lower() == player_name.lower()) &
                                    (raw_combined['Team'].astype(str).str.strip().str.lower() == team_name.lower())
                                ]
                                
                                if len(player_match) > 0:
                                    player_raw = player_match.iloc[0]
                                    # Try to find xG per 90
                                    for col in player_raw.index:
                                        col_lower = str(col).lower()
                                        if ('xg per 90' in col_lower or 'expected goals per 90' in col_lower) and 'goals' not in col_lower and 'against' not in col_lower:
                                            test_val = player_raw.get(col)
                                            if pd.notna(test_val) and test_val != '':
                                                xg_val = test_val
                                                break
                                    # Try to find Goals per 90
                                    for col in player_raw.index:
                                        col_lower = str(col).lower()
                                        if 'goals per 90' in col_lower and 'xg' not in col_lower and 'non-penalty' not in col_lower and 'head' not in col_lower:
                                            test_val = player_raw.get(col)
                                            if pd.notna(test_val) and test_val != '':
                                                goals_val = test_val
                                                break
                        except Exception as e:
                            pass  # Silently fail if can't load raw data
                    
                    # Calculate difference if both exist
                    if pd.notna(xg_val) and pd.notna(goals_val) and xg_val != '' and goals_val != '':
                        try:
                            metric_value = round(float(xg_val) - float(goals_val), 2)
                        except:
                            metric_value = 0  # Fill with 0 if calculation fails
                    else:
                        metric_value = 0  # Fill with 0 if either value is missing
                else:
                    # Try exact match
                    for col in player_row.index:
                        if str(col).lower() == str(metric_header).lower():
                            metric_value = player_row.get(col)
                            break
                    
                    # If not found, try improved matching logic (same as create_top_15_report.py)
                    if pd.isna(metric_value) or metric_value == '':
                        metric_lower = str(metric_header).lower()
                        
                        # Strategy 1: Try with _report suffix
                        report_col = metric_header + '_report'
                        if report_col in player_row.index:
                            test_val = player_row.get(report_col, '')
                            if pd.notna(test_val) and test_val != '':
                                metric_value = test_val
                        
                        # Strategy 2: Try base name without "per 90" or "%"
                        if pd.isna(metric_value) or metric_value == '':
                            base_name = metric_lower.replace(' per 90', '').replace(' per90', '').replace(', %', '').replace(' %', '').replace(' won', '').replace('accurate', '').strip()
                            for col in player_row.index:
                                col_lower = str(col).lower()
                                col_base = col_lower.replace(' per 90', '').replace(' per90', '').replace(', %', '').replace(' %', '').replace(' won', '').replace('accurate', '').replace('_report', '').strip()
                                if base_name == col_base and '_vs_' not in col_lower:
                                    test_val = player_row.get(col, '')
                                    if pd.notna(test_val) and test_val != '':
                                        metric_value = test_val
                                        break
                        
                        # Strategy 3: Try partial/fuzzy match
                        if pd.isna(metric_value) or metric_value == '':
                            for col in player_row.index:
                                col_lower = str(col).lower()
                                if (metric_lower in col_lower or col_lower in metric_lower) and '_vs_' not in col_lower:
                                    test_val = player_row.get(col, '')
                                    if pd.notna(test_val) and test_val != '':
                                        metric_value = test_val
                                        break
                    
                    # Fill missing values with 0
                    if pd.isna(metric_value) or metric_value == '' or metric_value is None:
                        metric_value = 0
                
                # Ensure we always write a numeric value (0 if still missing)
                final_value = 0
                if metric_value is not None:
                    try:
                        if isinstance(metric_value, (int, float)):
                            final_value = float(metric_value)
                        elif isinstance(metric_value, str) and metric_value.strip() != '':
                            final_value = float(metric_value)
                        else:
                            final_value = 0
                    except (ValueError, TypeError):
                        final_value = 0
                
                # Write the value (0 if missing)
                ws.cell(row=row_idx, column=metric_idx, value=final_value)
            
            # Apply borders to all cells in row
            for col_idx in range(1, len(all_headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = border
        
        # Calculate Top 15s (Power Five) for all players
        print(f"  📊 Calculating Top 15s (Power Five) for {len(df)} players...")
        try:
            from update_mike_norris_reports import load_conference_season_data
            
            # Load Power Five 2025 data for ranking
            raw_data_frames = []
            power_five_conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
            
            for conference in power_five_conferences:
                df_raw = load_conference_season_data(base_dir, conference, 2025)
                if not df_raw.empty:
                    df_raw['Conference'] = conference
                    df_raw['Year'] = 2025
                    raw_data_frames.append(df_raw)
            
            if raw_data_frames:
                all_players_raw = pd.concat(raw_data_frames, ignore_index=True)
                power_five_players = all_players_raw[
                    (all_players_raw['Position_Profile'] == position_profile) &
                    (all_players_raw['Conference'].isin(power_five_conferences))
                ].copy()
                
                print(f"     Loaded {len(power_five_players)} Power Five players for ranking")
                
                # Build metric columns to check
                metric_cols_to_check = []
                for metric_header in metric_headers:
                    if metric_header == "Interceptions + Sliding Tackles":
                        continue
                    
                    # Find matching column in power_five_players
                    matching_col = None
                    metric_lower = str(metric_header).lower()
                    
                    for col in power_five_players.columns:
                        if str(col).lower().strip() == metric_lower.strip():
                            matching_col = col
                            break
                    
                    if not matching_col:
                        # Try base name matching
                        if 'per 90' in metric_lower and '%' not in metric_lower:
                            metric_base = metric_lower.replace(' per 90', '').replace(' per90', '').strip()
                            for col in power_five_players.columns:
                                col_lower = str(col).lower()
                                if 'per 90' in col_lower and '%' not in col_lower:
                                    col_base = col_lower.replace(' per 90', '').replace(' per90', '').strip()
                                    if ' '.join(metric_base.split()) == ' '.join(col_base.split()):
                                        matching_col = col
                                        break
                        elif '%' in metric_lower:
                            metric_base = metric_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                            for col in power_five_players.columns:
                                col_lower = str(col).lower()
                                if ('%' in col_lower or 'percent' in col_lower) and 'per 90' not in col_lower:
                                    col_base = col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                                    if ' '.join(metric_base.split()) == ' '.join(col_base.split()):
                                        matching_col = col
                                        break
                    
                    if matching_col:
                        metric_cols_to_check.append((metric_header, matching_col))
                
                # Calculate Top 15s for each player
                for row_idx, (_, player_row) in enumerate(df.iterrows(), start=4):
                    top15_count = 0
                    
                    # Handle combined metric
                    combined_metric_handled = False
                    if "Interceptions + Sliding Tackles" in metric_headers:
                        interceptions_col = None
                        for col in player_row.index:
                            if 'interception' in str(col).lower() and 'per 90' in str(col).lower():
                                interceptions_col = col
                                break
                        
                        if interceptions_col:
                            player_val = player_row.get(interceptions_col)
                            if pd.notna(player_val) and player_val != '':
                                try:
                                    player_val = float(player_val)
                                    pf_col = None
                                    for col in power_five_players.columns:
                                        if 'interception' in str(col).lower() and 'per 90' in str(col).lower():
                                            pf_col = col
                                            break
                                    
                                    if pf_col:
                                        pf_values = pd.to_numeric(power_five_players[pf_col], errors='coerce').dropna()
                                        if len(pf_values) > 0:
                                            # Count how many players have >= this value
                                            pf_rank = (pf_values >= player_val).sum()
                                            # Get the 15th highest value (or lowest if we need at least 15 players)
                                            sorted_values = pf_values.sort_values(ascending=False)
                                            if len(sorted_values) >= 15:
                                                value_at_15th = sorted_values.iloc[14]  # 0-indexed, so 14 = 15th
                                                # Top 15 if: rank <= 15 OR value >= 15th highest value (handles ties)
                                                if pf_rank <= 15 or player_val >= value_at_15th:
                                                    top15_count += 1
                                                    combined_metric_handled = True
                                            elif pf_rank <= 15:
                                                # If less than 15 players total, just check rank
                                                top15_count += 1
                                                combined_metric_handled = True
                                except:
                                    pass
                    
                    # Check regular metrics
                    for metric_header, raw_data_col in metric_cols_to_check:
                        if combined_metric_handled and ('interception' in str(metric_header).lower() or 'sliding' in str(metric_header).lower()):
                            continue
                        
                        player_value = None
                        for col in player_row.index:
                            if str(col).lower() == str(metric_header).lower():
                                player_value = player_row.get(col)
                                break
                        
                        if pd.isna(player_value) or player_value == '':
                            continue
                        
                        try:
                            player_val = float(player_value)
                            if raw_data_col in power_five_players.columns:
                                pf_values = pd.to_numeric(power_five_players[raw_data_col], errors='coerce').dropna()
                                if len(pf_values) > 0:
                                    # Count how many players have >= this value
                                    pf_rank = (pf_values >= player_val).sum()
                                    # Get the 15th highest value (for handling ties)
                                    sorted_values = pf_values.sort_values(ascending=False)
                                    if len(sorted_values) >= 15:
                                        value_at_15th = sorted_values.iloc[14]  # 0-indexed, so 14 = 15th
                                        # Top 15 if: rank <= 15 OR value >= 15th highest value (handles ties)
                                        if pf_rank <= 15 or player_val >= value_at_15th:
                                            top15_count += 1
                                    elif pf_rank <= 15:
                                        # If less than 15 players total, just check rank
                                        top15_count += 1
                        except:
                            continue
                    
                    # Write Top 15s (Power Five) value
                    ws.cell(row=row_idx, column=16, value=int(top15_count))
                
                print(f"     ✅ Calculated Top 15s (Power Five) for all players")
        except Exception as e:
            print(f"     ⚠️  Could not calculate Top 15s: {e}")
            import traceback
            traceback.print_exc()
        
        print(f"  ✅ Created sheet with {len(df)} players")
    
    # Save workbook
    output_file = base_dir / "Portland Thorns 2025 Full Shortlist.xlsx"
    wb.save(output_file)
    print(f"\n✅ Full Shortlist saved to: {output_file}")
    print(f"   Total players across all positions: {sum(len(all_data[pos]) for pos in position_profiles if pos in all_data)}")


if __name__ == "__main__":
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    create_full_shortlist(base_dir)

