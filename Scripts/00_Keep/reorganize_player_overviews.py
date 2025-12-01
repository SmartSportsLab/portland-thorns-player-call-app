#!/usr/bin/env python3
"""
Reorganize Player Overviews folder into:
- Top 15/[Position Profile]/ (for AI Shortlist players)
- Other/[Position Profile]/ (for all other players)
"""

from pathlib import Path
from openpyxl import load_workbook
import shutil

def reorganize_overviews(base_dir):
    """Reorganize player overview PDFs into Top 15 and Other folders by position."""
    base_dir = Path(base_dir)
    
    overviews_dir = base_dir / 'Player Overviews'
    ai_shortlist_file = base_dir / 'AI Shortlist.xlsx'
    
    if not overviews_dir.exists():
        print(f"❌ Player Overviews directory not found: {overviews_dir}")
        return
    
    if not ai_shortlist_file.exists():
        print(f"❌ AI Shortlist file not found: {ai_shortlist_file}")
        return
    
    print("="*70)
    print("REORGANIZING PLAYER OVERVIEWS")
    print("="*70)
    
    # Load AI Shortlist to identify top 15 players
    print("\n📊 Loading AI Shortlist...")
    wb = load_workbook(ai_shortlist_file, data_only=True)
    
    position_profiles = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    # Map position profile names to folder-safe names
    position_folder_map = {
        'Hybrid CB': 'Hybrid_CB',
        'DM Box-To-Box': 'DM_Box_To_Box',
        'AM Advanced Playmaker': 'AM_Advanced_Playmaker',
        'Right Touchline Winger': 'Right_Touchline_Winger'
    }
    
    top_15_filenames = set()
    position_to_top15 = {}
    
    for position in position_profiles:
        if position in wb.sheetnames:
            ws = wb[position]
            position_top15 = []
            
            for row_idx in range(4, ws.max_row + 1):
                player = ws.cell(row=row_idx, column=2).value
                team = ws.cell(row=row_idx, column=3).value
                
                if player and team:
                    # Create filename format (matching overview naming)
                    safe_name = str(player).replace(' ', '_').replace('.', '').replace('/', '_')
                    safe_team = str(team).replace(' ', '_')
                    filename = f"{safe_name}_{safe_team}.pdf"
                    position_top15.append(filename)
                    top_15_filenames.add(filename)
            
            position_to_top15[position] = position_top15
            print(f"  ✅ {position}: {len(position_top15)} top 15 players")
    
    wb.close()
    
    # Create folder structure
    top15_dir = overviews_dir / 'Top 15'
    other_dir = overviews_dir / 'Other'
    
    for position in position_profiles:
        folder_name = position_folder_map[position]
        (top15_dir / folder_name).mkdir(parents=True, exist_ok=True)
        (other_dir / folder_name).mkdir(parents=True, exist_ok=True)
    
    print(f"\n📁 Created folder structure:")
    print(f"  - Top 15/[Position]/")
    print(f"  - Other/[Position]/")
    
    # Get all PDF files
    all_pdfs = list(overviews_dir.glob('*.pdf'))
    print(f"\n📄 Found {len(all_pdfs)} PDF files")
    
    # Determine position for each PDF by checking shortlist
    print("\n📊 Determining player positions from shortlist...")
    shortlist_file = base_dir / 'Portland Thorns 2025 Shortlist.xlsx'
    
    pdf_to_position = {}
    if shortlist_file.exists():
        wb_shortlist = load_workbook(shortlist_file, data_only=True)
        
        for position in position_profiles:
            if position in wb_shortlist.sheetnames:
                ws = wb_shortlist[position]
                
                # Find Player column
                player_col = None
                team_col = None
                for col_idx in range(1, ws.max_column + 1):
                    header = ws.cell(row=3, column=col_idx).value
                    if header and 'Player' in str(header):
                        player_col = col_idx
                    if header and 'Team' in str(header) and 'Minutes' not in str(header):
                        team_col = col_idx
                
                if player_col and team_col:
                    for row_idx in range(4, ws.max_row + 1):
                        player = ws.cell(row=row_idx, column=player_col).value
                        team = ws.cell(row=row_idx, column=team_col).value
                        
                        if player and team:
                            safe_name = str(player).replace(' ', '_').replace('.', '').replace('/', '_')
                            safe_team = str(team).replace(' ', '_')
                            filename = f"{safe_name}_{safe_team}.pdf"
                            pdf_to_position[filename] = position
        
        wb_shortlist.close()
    
    # Move files
    print("\n📦 Moving files...")
    moved_top15 = 0
    moved_other = 0
    not_found = []
    
    for pdf_file in all_pdfs:
        filename = pdf_file.name
        
        # Determine position
        position = pdf_to_position.get(filename)
        if not position:
            # Try to infer from filename patterns or check all positions
            # For now, we'll skip files we can't identify
            not_found.append(filename)
            continue
        
        folder_name = position_folder_map[position]
        
        # Determine if top 15 or other
        if filename in top_15_filenames:
            dest_dir = top15_dir / folder_name
            moved_top15 += 1
        else:
            dest_dir = other_dir / folder_name
            moved_other += 1
        
        # Move file
        dest_path = dest_dir / filename
        try:
            shutil.move(str(pdf_file), str(dest_path))
        except Exception as e:
            print(f"  ⚠️  Error moving {filename}: {e}")
    
    print(f"\n✅ Reorganization complete!")
    print(f"  - Moved {moved_top15} files to Top 15/")
    print(f"  - Moved {moved_other} files to Other/")
    if not_found:
        print(f"  - ⚠️  {len(not_found)} files could not be categorized (left in root)")
        if len(not_found) <= 10:
            for f in not_found:
                print(f"      - {f}")

if __name__ == '__main__':
    base_dir = Path('/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search')
    reorganize_overviews(base_dir)

