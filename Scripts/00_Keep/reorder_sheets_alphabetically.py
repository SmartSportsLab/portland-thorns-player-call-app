#!/usr/bin/env python3
"""
Reorder sheets in Mike Norris reports so team tabs are alphabetical after position profile tabs.
Order: Data Notes, Position Profile tabs (Hybrid CB, DM Box-To-Box, AM Advanced Playmaker, Right Touchline Winger), then team tabs alphabetically.
"""

from pathlib import Path
from openpyxl import load_workbook

# Position profile order
POSITION_PROFILE_ORDER = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']


def reorder_sheets(file_path):
    """Reorder sheets in the workbook."""
    print(f"📄 Processing: {file_path.name}")
    
    wb = load_workbook(file_path)
    
    # Separate sheets into categories
    data_notes_sheet = None
    position_profile_sheets = []
    team_sheets = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Data Notes':
            data_notes_sheet = sheet_name
        elif sheet_name in POSITION_PROFILE_ORDER:
            position_profile_sheets.append(sheet_name)
        else:
            # Assume it's a team sheet
            team_sheets.append(sheet_name)
    
    # Sort position profile sheets by our defined order
    position_profile_sheets_sorted = []
    for pos in POSITION_PROFILE_ORDER:
        if pos in position_profile_sheets:
            position_profile_sheets_sorted.append(pos)
    
    # Sort team sheets alphabetically
    team_sheets_sorted = sorted(team_sheets)
    
    # Create new sheet order
    new_order = []
    if data_notes_sheet:
        new_order.append(data_notes_sheet)
    new_order.extend(position_profile_sheets_sorted)
    new_order.extend(team_sheets_sorted)
    
    print(f"  📋 New sheet order:")
    for i, sheet_name in enumerate(new_order, 1):
        print(f"     {i}. {sheet_name}")
    
    # Reorder sheets by moving them to their target positions
    # We'll move sheets one by one, starting from the end of the desired order
    # This ensures we don't mess up indices while moving
    
    # Create a mapping of desired positions
    target_positions = {name: idx for idx, name in enumerate(new_order)}
    
    # Move sheets to their target positions, working backwards through the desired order
    # This way each move doesn't affect the positions of sheets we've already positioned
    for target_idx in range(len(new_order) - 1, -1, -1):
        sheet_name = new_order[target_idx]
        current_idx = wb.sheetnames.index(sheet_name)
        
        if current_idx != target_idx:
            # Calculate offset needed to move to target position
            offset = target_idx - current_idx
            wb.move_sheet(wb[sheet_name], offset=offset)
    
    wb.save(file_path)
    print(f"  ✅ Sheets reordered successfully\n")


def main():
    """Main function."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    # Find Mike Norris report files
    report_files = list(base_dir.glob("Mike_Norris_Scouting_Report_*_IntraConference.xlsx"))
    
    if not report_files:
        print("❌ No Mike Norris report files found")
        return
    
    print("="*80)
    print("REORDERING SHEETS ALPHABETICALLY")
    print("="*80)
    
    for file_path in report_files:
        reorder_sheets(file_path)
    
    print("="*80)
    print("✅ ALL REPORTS REORDERED")
    print("="*80)
    print("\nSheet order:")
    print("  1. Data Notes")
    print("  2. Hybrid CB")
    print("  3. DM Box-To-Box")
    print("  4. AM Advanced Playmaker")
    print("  5. Right Touchline Winger")
    print("  6. Team tabs (alphabetically)")


if __name__ == "__main__":
    main()

