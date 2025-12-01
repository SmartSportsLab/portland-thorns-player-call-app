#!/usr/bin/env python3
"""
Create Scatterplot Visualizations for Mike Norris Scouting Reports
==================================================================

Generates scatterplots with:
- X-axis: Total Score (1-10)
- Y-axis: Grade (converted to numeric)
- Circle size: Minutes played
- Color: Grade or Team

Saves as PNG images and optionally embeds in Excel.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import seaborn as sns
from matplotlib import font_manager

# Set style
sns.set_style("whitegrid")
plt.rcParams['figure.dpi'] = 300
plt.rcParams['savefig.dpi'] = 300
plt.rcParams['font.size'] = 10

# Grade to numeric mapping
GRADE_VALUES = {'A': 5, 'B': 4, 'C': 3, 'D': 2, 'F': 1, 'A+': 5.5, '': 0}
GRADE_COLORS = {
    'A': '#8B0000',  # Dark red
    'A+': '#8B0000',
    'B': '#C5504B',  # Medium red
    'C': '#F2A2A2',  # Light red
    'D': '#8FAADC',  # Light blue
    'F': '#1F4E79'   # Dark blue
}


def create_scatterplot_per_position(df, position_name, output_dir, grade_col='Conference_Grade'):
    """
    Create scatterplot for a specific position profile.
    
    Args:
        df: DataFrame with player data
        position_name: Name of position profile (e.g., 'Hybrid CB')
        output_dir: Directory to save images
        grade_col: Column name for grade ('Conference_Grade' or 'Power_Five_Grade')
    """
    if len(df) == 0:
        print(f"  ⚠️  No data for {position_name}")
        return None
    
    # Prepare data
    df_plot = df.copy()
    
    # Convert grade to numeric
    df_plot['Grade_Value'] = df_plot[grade_col].map(GRADE_VALUES).fillna(0)
    
    # Get minutes (use for sizing)
    minutes_col = 'Total Minutes' if 'Total Minutes' in df_plot.columns else 'Minutes played'
    if minutes_col not in df_plot.columns:
        print(f"  ⚠️  No minutes column found for {position_name}")
        return None
    
    df_plot['Minutes'] = pd.to_numeric(df_plot[minutes_col], errors='coerce').fillna(0)
    
    # Get score
    score_col = '2025 Total Score' if '2025 Total Score' in df_plot.columns else 'Total_Score_1_10'
    if score_col not in df_plot.columns:
        print(f"  ⚠️  No score column found for {position_name}")
        return None
    
    df_plot['Score'] = pd.to_numeric(df_plot[score_col], errors='coerce').fillna(0)
    
    # Filter out invalid data
    df_plot = df_plot[
        (df_plot['Score'] > 0) & 
        (df_plot['Grade_Value'] > 0) & 
        (df_plot['Minutes'] > 0)
    ].copy()
    
    if len(df_plot) == 0:
        print(f"  ⚠️  No valid data points for {position_name}")
        return None
    
    # Create figure
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Normalize minutes for circle size (min 50, max 500)
    min_minutes = df_plot['Minutes'].min()
    max_minutes = df_plot['Minutes'].max()
    if max_minutes > min_minutes:
        # Scale to 50-500 range
        df_plot['Circle_Size'] = 50 + (df_plot['Minutes'] - min_minutes) / (max_minutes - min_minutes) * 450
    else:
        df_plot['Circle_Size'] = 200
    
    # Color by grade
    df_plot['Color'] = df_plot[grade_col].map(GRADE_COLORS).fillna('#808080')
    
    # Create scatter plot
    for grade in ['A+', 'A', 'B', 'C', 'D', 'F']:
        if grade not in df_plot[grade_col].values:
            continue
        
        grade_data = df_plot[df_plot[grade_col] == grade]
        if len(grade_data) == 0:
            continue
        
        ax.scatter(
            grade_data['Score'],
            grade_data['Grade_Value'],
            s=grade_data['Circle_Size'],
            c=grade_data['Color'],
            alpha=0.6,
            edgecolors='black',
            linewidths=1.5,
            label=f'Grade {grade}',
            zorder=3
        )
    
    # Add player names for top performers or high minutes
    # Show top 10 by score or minutes > 700
    top_players = df_plot.nlargest(10, 'Score')
    high_minutes_players = df_plot[df_plot['Minutes'] >= 700]
    
    for _, player in pd.concat([top_players, high_minutes_players]).drop_duplicates().iterrows():
        player_name = str(player.get('Player', ''))[:20]  # Truncate long names
        ax.annotate(
            player_name,
            (player['Score'], player['Grade_Value']),
            fontsize=8,
            alpha=0.7,
            xytext=(5, 5),
            textcoords='offset points',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7, edgecolor='gray', linewidth=0.5)
        )
    
    # Labels and title
    ax.set_xlabel('Total Score (1-10 scale)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Grade Value', fontsize=12, fontweight='bold')
    ax.set_title(f'{position_name} - Performance vs Grade\n(Circle size = Minutes played)', 
                fontsize=14, fontweight='bold', pad=20)
    
    # Set axis ranges
    ax.set_xlim(0, 10.5)
    ax.set_ylim(0.5, 5.5)
    
    # Set y-axis labels to show grades
    ax.set_yticks([1, 2, 3, 4, 5])
    ax.set_yticklabels(['F', 'D', 'C', 'B', 'A'])
    
    # Add grid
    ax.grid(True, alpha=0.3, linestyle='--')
    
    # Add quadrant labels
    ax.axhline(y=3, color='gray', linestyle='--', alpha=0.3, linewidth=1)
    ax.axvline(x=7, color='gray', linestyle='--', alpha=0.3, linewidth=1)
    ax.text(8.5, 4.5, 'High Score\nHigh Grade', fontsize=9, alpha=0.5, 
            bbox=dict(boxstyle='round', facecolor='lightgreen', alpha=0.3))
    ax.text(1, 1.5, 'Low Score\nLow Grade', fontsize=9, alpha=0.5,
            bbox=dict(boxstyle='round', facecolor='lightcoral', alpha=0.3))
    
    # Legend
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        legend = ax.legend(handles, labels, loc='upper left', framealpha=0.9, 
                          title='Grades', title_fontsize=10)
        legend.get_frame().set_facecolor('white')
    
    # Add minutes legend (size reference)
    minutes_legend_elements = [
        mpatches.Circle((0, 0), radius=50, facecolor='gray', alpha=0.6, 
                       label=f'Min: {int(min_minutes)} min'),
        mpatches.Circle((0, 0), radius=250, facecolor='gray', alpha=0.6,
                       label=f'Max: {int(max_minutes)} min')
    ]
    minutes_legend = ax.legend(handles=minutes_legend_elements, loc='lower right',
                               title='Circle Size = Minutes', title_fontsize=9, framealpha=0.9)
    ax.add_artist(legend)  # Re-add grade legend
    
    plt.tight_layout()
    
    # Save
    output_path = output_dir / f"{position_name.replace(' ', '_')}_scatterplot.png"
    plt.savefig(output_path, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"  ✅ Created scatterplot: {output_path.name}")
    return output_path


def create_team_scatterplot(df, team_name, output_dir, grade_col='Conference_Grade'):
    """Create scatterplot for a specific team across all positions."""
    if len(df) == 0:
        return None
    
    df_plot = df.copy()
    
    # Convert grade to numeric
    df_plot['Grade_Value'] = df_plot[grade_col].map(GRADE_VALUES).fillna(0)
    
    # Get minutes
    minutes_col = 'Total Minutes' if 'Total Minutes' in df_plot.columns else 'Minutes played'
    df_plot['Minutes'] = pd.to_numeric(df_plot[minutes_col], errors='coerce').fillna(0)
    
    # Get score
    score_col = '2025 Total Score' if '2025 Total Score' in df_plot.columns else 'Total_Score_1_10'
    df_plot['Score'] = pd.to_numeric(df_plot[score_col], errors='coerce').fillna(0)
    
    # Filter
    df_plot = df_plot[
        (df_plot['Score'] > 0) & 
        (df_plot['Grade_Value'] > 0) & 
        (df_plot['Minutes'] > 0)
    ].copy()
    
    if len(df_plot) == 0:
        return None
    
    # Create figure
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Normalize circle size
    min_minutes = df_plot['Minutes'].min()
    max_minutes = df_plot['Minutes'].max()
    if max_minutes > min_minutes:
        df_plot['Circle_Size'] = 50 + (df_plot['Minutes'] - min_minutes) / (max_minutes - min_minutes) * 450
    else:
        df_plot['Circle_Size'] = 200
    
    # Color by position profile
    position_profiles = df_plot['Position_Profile'].unique() if 'Position_Profile' in df_plot.columns else []
    colors = sns.color_palette("husl", len(position_profiles))
    position_colors = dict(zip(position_profiles, colors))
    df_plot['Color'] = df_plot['Position_Profile'].map(position_colors).fillna('#808080')
    
    # Plot by position profile
    for position in position_profiles:
        pos_data = df_plot[df_plot['Position_Profile'] == position]
        if len(pos_data) == 0:
            continue
        
        ax.scatter(
            pos_data['Score'],
            pos_data['Grade_Value'],
            s=pos_data['Circle_Size'],
            c=[position_colors[position]],
            alpha=0.6,
            edgecolors='black',
            linewidths=1.5,
            label=position,
            zorder=3
        )
    
    # Add player names
    top_players = df_plot.nlargest(10, 'Score')
    for _, player in top_players.iterrows():
        player_name = str(player.get('Player', ''))[:20]
        ax.annotate(
            player_name,
            (player['Score'], player['Grade_Value']),
            fontsize=8,
            alpha=0.7,
            xytext=(5, 5),
            textcoords='offset points',
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7, edgecolor='gray', linewidth=0.5)
        )
    
    # Labels
    ax.set_xlabel('Total Score (1-10 scale)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Grade Value', fontsize=12, fontweight='bold')
    ax.set_title(f'{team_name} - All Positions\n(Circle size = Minutes played)', 
                fontsize=14, fontweight='bold', pad=20)
    
    ax.set_xlim(0, 10.5)
    ax.set_ylim(0.5, 5.5)
    ax.set_yticks([1, 2, 3, 4, 5])
    ax.set_yticklabels(['F', 'D', 'C', 'B', 'A'])
    
    ax.grid(True, alpha=0.3, linestyle='--')
    
    # Legend
    if position_profiles:
        legend = ax.legend(title='Position Profiles', loc='upper left', framealpha=0.9)
        legend.get_frame().set_facecolor('white')
    
    plt.tight_layout()
    
    # Save
    safe_team_name = team_name.replace(' ', '_').replace('.', '')
    output_path = output_dir / f"{safe_team_name}_scatterplot.png"
    plt.savefig(output_path, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"  ✅ Created team scatterplot: {output_path.name}")
    return output_path


def create_all_scatterplots(excel_file, output_dir=None, grade_col='Conference_Grade'):
    """
    Create scatterplots for all position profiles and teams in the Excel file.
    
    Args:
        excel_file: Path to Excel file
        output_dir: Directory to save images (default: same as Excel file)
        grade_col: 'Conference_Grade' or 'Power_Five_Grade'
    """
    excel_path = Path(excel_file)
    if output_dir is None:
        output_dir = excel_path.parent / "Scatterplots"
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(exist_ok=True)
    
    print(f"\n{'='*80}")
    print(f"CREATING SCATTERPLOT VISUALIZATIONS")
    print(f"{'='*80}")
    print(f"Excel file: {excel_path.name}")
    print(f"Output directory: {output_dir}")
    print(f"Grade column: {grade_col}")
    
    # Load Excel
    xl = pd.ExcelFile(excel_path)
    
    # Position profiles
    position_profiles = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    created_images = []
    
    # Create position profile scatterplots
    print(f"\n📊 Creating position profile scatterplots...")
    for position in position_profiles:
        if position not in xl.sheet_names:
            continue
        
        df = pd.read_excel(excel_path, sheet_name=position)
        image_path = create_scatterplot_per_position(df, position, output_dir, grade_col)
        if image_path:
            created_images.append(image_path)
    
    # Create team scatterplots
    print(f"\n🏆 Creating team scatterplots...")
    for sheet_name in xl.sheet_names:
        if sheet_name == 'Data Summary' or sheet_name in position_profiles:
            continue
        
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        if 'Player' not in df.columns:
            continue
        
        image_path = create_team_scatterplot(df, sheet_name, output_dir, grade_col)
        if image_path:
            created_images.append(image_path)
    
    print(f"\n✅ Created {len(created_images)} scatterplot images")
    return created_images


def embed_images_in_excel(excel_file, images_dir):
    """
    Embed scatterplot images into Excel sheets.
    Note: This requires images to be in the images_dir.
    """
    excel_path = Path(excel_file)
    images_path = Path(images_dir)
    
    if not images_path.exists():
        print(f"  ⚠️  Images directory not found: {images_path}")
        return
    
    print(f"\n📎 Embedding images into Excel...")
    
    wb = load_workbook(excel_path)
    
    # Find matching images for each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Look for matching image
        safe_name = sheet_name.replace(' ', '_').replace('.', '')
        image_file = images_path / f"{safe_name}_scatterplot.png"
        
        if image_file.exists():
            try:
                img = Image(str(image_file))
                # Resize to fit (adjust as needed)
                img.width = 800
                img.height = 600
                # Insert at column A, after last row
                ws.add_image(img, f'A{ws.max_row + 3}')
                print(f"  ✅ Embedded image in {sheet_name}")
            except Exception as e:
                print(f"  ⚠️  Could not embed image in {sheet_name}: {e}")
    
    wb.save(excel_path)
    print(f"  ✅ Excel file updated with embedded images")


def main():
    """Main function."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python create_scatterplot_visualizations.py [ACC|SEC] [grade_col] [embed]")
        print("\nOptions:")
        print("  grade_col: 'Conference_Grade' (default) or 'Power_Five_Grade'")
        print("  embed: 'yes' to embed images in Excel (optional)")
        sys.exit(1)
    
    conference = sys.argv[1].upper()
    grade_col = sys.argv[2] if len(sys.argv) > 2 else 'Conference_Grade'
    embed = sys.argv[3].lower() == 'yes' if len(sys.argv) > 3 else False
    
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    excel_file = base_dir / f"Mike_Norris_Scouting_Report_{conference}_IntraConference.xlsx"
    
    if not excel_file.exists():
        print(f"❌ File not found: {excel_file}")
        sys.exit(1)
    
    # Create scatterplots
    images = create_all_scatterplots(excel_file, grade_col=grade_col)
    
    # Optionally embed in Excel
    if embed and images:
        images_dir = excel_file.parent / "Scatterplots"
        embed_images_in_excel(excel_file, images_dir)
    
    print(f"\n✅ Done! Scatterplots saved to: {excel_file.parent / 'Scatterplots'}")


if __name__ == "__main__":
    main()

