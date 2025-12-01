#!/usr/bin/env python3
"""
Reorganize team summary sheets to group by position profile instead of grade.
Each position profile section will have players sorted by grade (A-F).
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import pandas as pd

# Position profile display names
POSITION_DISPLAY_NAMES = {
    'Center Back': 'Hybrid CB',
    'Centre Midfielder': 'DM Box-To-Box',
    'Attacking Midfielder': 'AM Advanced Playmaker',
    'Winger': 'Right Touchline Winger'
}

# Position profile order
POSITION_ORDER = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']

# Grade order (A first, F last)
GRADE_ORDER = ['A', 'B', 'C', 'D', 'F']


def get_position_profile_display_name(position_profile):
    """Get display name for position profile."""
    if not position_profile:
        return ''
    return POSITION_DISPLAY_NAMES.get(position_profile, position_profile)


def add_conditional_formatting_to_team_sheet(ws, headers, max_row):
    """Add conditional formatting to Grade and Minutes columns."""
    if max_row < 2:  # Need at least header row and one data row
        return
    
    # Color definitions
    colors = {
        'dark_red': '8B0000',
        'red': 'C5504B',
        'light_red': 'F2A2A2',
        'light_blue': '8FAADC',
        'blue': '4472C4',
        'dark_blue': '1F4E79'
    }
    
    # Grade colors
    grade_colors = {
        'A': colors['dark_red'],
        'B': colors['red'],
        'C': colors['light_red'],
        'D': colors['light_blue'],
        'F': colors['dark_blue']
    }
    
    # Find column indices
    team_grade_col = None
    conf_grade_col = None
    power_five_grade_col = None
    minutes_col = None
    
    for col_idx, header in enumerate(headers, 1):
        if 'Team Grade' in str(header):
            team_grade_col = col_idx
        elif 'Conference Grade' in str(header):
            conf_grade_col = col_idx
        elif 'Power Five Grade' in str(header):
            power_five_grade_col = col_idx
        elif 'Minutes' in str(header):
            minutes_col = col_idx
    
    # Apply Team Grade formatting
    if team_grade_col and max_row > 1:
        team_grade_col_letter = get_column_letter(team_grade_col)
        team_grade_range = f'{team_grade_col_letter}2:{team_grade_col_letter}{max_row}'
        
        for grade, color in grade_colors.items():
            team_grade_rule = CellIsRule(
                operator='equal', formula=[f'"{grade}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
            )
            ws.conditional_formatting.add(team_grade_range, team_grade_rule)
    
    # Apply Conference Grade formatting
    if conf_grade_col and max_row > 1:
        conf_grade_col_letter = get_column_letter(conf_grade_col)
        conf_grade_range = f'{conf_grade_col_letter}2:{conf_grade_col_letter}{max_row}'
        
        for grade, color in grade_colors.items():
            conf_grade_rule = CellIsRule(
                operator='equal', formula=[f'"{grade}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
            )
            ws.conditional_formatting.add(conf_grade_range, conf_grade_rule)
    
    # Apply Power Five Grade formatting
    if power_five_grade_col and max_row > 1:
        power_five_grade_col_letter = get_column_letter(power_five_grade_col)
        power_five_grade_range = f'{power_five_grade_col_letter}2:{power_five_grade_col_letter}{max_row}'
        
        for grade, color in grade_colors.items():
            power_five_grade_rule = CellIsRule(
                operator='equal', formula=[f'"{grade}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
            )
            ws.conditional_formatting.add(power_five_grade_range, power_five_grade_rule)
    
    # Apply minutes formatting (discrete ranges) - now column J (10)
    if minutes_col and max_row > 1:
        minutes_col_letter = get_column_letter(minutes_col)
        minutes_range = f'{minutes_col_letter}2:{minutes_col_letter}{max_row}'
        
        # Find max minutes to determine ranges
        max_minutes = 0
        for row_idx in range(2, max_row + 1):
            cell_value = ws.cell(row=row_idx, column=minutes_col).value
            if cell_value and isinstance(cell_value, (int, float)):
                max_minutes = max(max_minutes, cell_value)
        
        # If no max found, use default ranges
        if max_minutes == 0:
            max_minutes = 1119
        
        # Define minute ranges with colors
        minute_ranges = [
            (0, 187, colors['dark_blue']),      # 0 - 187: Dark blue (very small sample)
            (187, 373, colors['blue']),          # 187 - 373: Blue (small sample)
            (373, 560, colors['light_blue']),    # 373 - 560: Light blue (below average)
            (560, 747, colors['light_red']),     # 560 - 747: Light red (above average)
            (747, 933, colors['red']),           # 747 - 933: Red (large sample)
            (933, max_minutes, colors['dark_red'])  # 933 - max_minutes: Dark red (very large sample)
        ]
        
        for min_val, max_val, color in minute_ranges:
            if min_val == 0:
                minute_rule = CellIsRule(
                    operator='lessThanOrEqual', 
                    formula=[f'{max_val}'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
            elif max_val >= max_minutes:
                minute_rule = CellIsRule(
                    operator='greaterThanOrEqual', 
                    formula=[f'{min_val}'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
            else:
                minute_rule = CellIsRule(
                    operator='between', 
                    formula=[f'{min_val}', f'{max_val}'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                )
            ws.conditional_formatting.add(minutes_range, minute_rule)


def reorganize_team_sheet(file_path, team_name):
    """Reorganize a team sheet to group by position profile."""
    print(f"  📄 Reorganizing {team_name} sheet...")
    
    wb = load_workbook(file_path)
    
    # Load grade data from position profile sheets
    grade_data = {}  # {player_name: {position_profile: grade}}
    position_sheets = ['Hybrid CB', 'DM Box-To-Box', 'AM Advanced Playmaker', 'Right Touchline Winger']
    
    for pos_sheet in position_sheets:
        if pos_sheet in wb.sheetnames:
            try:
                df_pos = pd.read_excel(file_path, sheet_name=pos_sheet)
                if 'Player' in df_pos.columns and 'Total Grade' in df_pos.columns:
                    for _, row in df_pos.iterrows():
                        player = row.get('Player')
                        grade = row.get('Total Grade')
                        team = row.get('Team', '')
                        if player and grade and pd.notna(grade):
                            # Check if team matches
                            if team_name in str(team) or str(team) in team_name:
                                if player not in grade_data:
                                    grade_data[player] = {}
                                grade_data[player][pos_sheet] = grade
            except Exception as e:
                print(f"      ⚠️  Could not load grades from {pos_sheet}: {e}")
    
    if team_name not in wb.sheetnames:
        print(f"    ⚠️  Sheet '{team_name}' not found")
        return
    
    ws = wb[team_name]
    
    # Read the current data using pandas to get all players
    try:
        # Read the entire sheet and find all data rows
        # The sheet has multiple "Grade X Players" sections, each with headers
        all_data = []
        current_headers = None
        current_section = None
        
        # Read all rows
        for i in range(1, ws.max_row + 1):
            row_vals = [cell.value for cell in ws[i]]
            
            # Check if this is a section header (e.g., "Grade A Players")
            first_val = str(row_vals[0]) if row_vals and row_vals[0] else ''
            if 'Grade' in first_val and 'Players' in first_val:
                current_section = first_val
                continue
            
            # Check if this is a headers row
            if 'Player' in first_val and 'Position' in ' '.join([str(v) for v in row_vals if v]):
                # Extract headers from this row
                current_headers = [str(v).strip() if v else '' for v in row_vals]
                continue
            
            # If we have headers and this looks like a data row
            if current_headers and row_vals[0] and str(row_vals[0]).strip() and str(row_vals[0]) != 'Player':
                # Create a dict for this row
                row_dict = {}
                for idx, header in enumerate(current_headers):
                    if idx < len(row_vals):
                        row_dict[header] = row_vals[idx]
                if row_dict.get('Player'):
                    all_data.append(row_dict)
        
        if not all_data:
            print(f"    ⚠️  No player data found")
            return
        
        df = pd.DataFrame(all_data)
        
        if len(df) == 0:
            print(f"    ⚠️  No player data found")
            return
        
        # Get column names (handle various possible names)
        player_col = None
        position_profile_col = None
        grade_col = None
        score_col = None
        position_col = None
        minutes_col = None
        score_2024_col = None
        change_col = None
        team_grade_col = None
        conf_grade_col = None
        power_five_grade_col = None
        
        for col in df.columns:
            col_str = str(col).strip()
            if 'Player' in col_str and player_col is None:
                player_col = col
            elif 'Position Profile' in col_str:
                position_profile_col = col
            elif 'Team Grade' in col_str:
                team_grade_col = col
            elif 'Conference Grade' in col_str:
                conf_grade_col = col
            elif 'Power Five Grade' in col_str:
                power_five_grade_col = col
            elif '2025 Total Score' in col_str or (col_str == 'Score' and score_col is None):
                score_col = col
            elif 'Position' in col_str and 'Profile' not in col_str:
                position_col = col
            elif 'Minutes' in col_str:
                minutes_col = col
            elif '2024 Total Score' in col_str:
                score_2024_col = col
            elif 'Change From' in col_str:
                change_col = col
        
        if player_col is None:
            print(f"    ⚠️  Missing Player column. Available: {list(df.columns)}")
            return
        
        # If we don't have position profile column, we need to infer it or skip
        if position_profile_col is None:
            print(f"    ⚠️  Missing Position Profile column. Available: {list(df.columns)}")
            # Try to infer from position column or skip this sheet
            return
        
        # Clear the sheet - delete all rows (we'll start fresh from row 1)
        print(f"    🗑️  Clearing sheet (had {ws.max_row} rows)")
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        print(f"    ✅ Sheet cleared (now has {ws.max_row} rows)")
        
        row = 1  # Start writing from row 1
        
        # Group by position profile, then sort by grade within each position
        position_profiles = df[position_profile_col].dropna().unique()
        
        # Sort position profiles by our defined order
        # Map position profiles to their display names for matching
        position_profiles_sorted = []
        for pos in POSITION_ORDER:
            for pp in position_profiles:
                pp_display = get_position_profile_display_name(pp)
                pp_str = str(pp).strip()
                if pp_display == pos or pp_str == pos:
                    if pp not in position_profiles_sorted:
                        position_profiles_sorted.append(pp)
                    break
        
        # Add any remaining position profiles not in our list
        for pp in position_profiles:
            if pp not in position_profiles_sorted:
                position_profiles_sorted.append(pp)
        
        # Write data grouped by position profile
        total_players = 0
        print(f"    📊 Found {len(position_profiles_sorted)} position profiles: {position_profiles_sorted}")
        
        for position_profile in position_profiles_sorted:
            position_df = df[df[position_profile_col] == position_profile].copy()
            
            if len(position_df) == 0:
                print(f"      ⚠️  No players for {position_profile}")
                continue
            
            # Get display name
            display_name = get_position_profile_display_name(position_profile)
            print(f"      📝 Writing {len(position_df)} players for {display_name}")
            
            # Write position profile header
            ws.cell(row=row, column=1, value=f"{display_name} Players").font = Font(bold=True, size=12)
            row += 1
            
            # Write column headers
            headers = ['Player', 'Position', 'Position Profile', 'Team Grade', 'Conference Grade', 'Power Five Grade', '2025 Total Score', '2024 Total Score', 'Change From 2024', 'Minutes']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            row += 1
            
            # Sort by Power Five Grade (A first, F last), then by score
            if power_five_grade_col and power_five_grade_col in position_df.columns:
                position_df['Grade_Order'] = position_df[power_five_grade_col].map({g: i for i, g in enumerate(GRADE_ORDER)}).fillna(999)
                sort_cols = ['Grade_Order']
                if score_col and score_col in position_df.columns:
                    sort_cols.append(score_col)
                position_df = position_df.sort_values(sort_cols, ascending=[True, False], na_position='last')
            else:
                # If no grade column, just sort by score
                if score_col and score_col in position_df.columns:
                    position_df = position_df.sort_values(score_col, ascending=False, na_position='last')
                elif player_col and player_col in position_df.columns:
                    position_df = position_df.sort_values(player_col)
            
            # Write players for this position profile
            for _, player_row in position_df.iterrows():
                player_name = player_row.get(player_col, '')
                
                # Get Team Grade value
                team_grade_value = ''
                if team_grade_col and team_grade_col in position_df.columns:
                    try:
                        team_grade_val = player_row[team_grade_col]
                        if pd.notna(team_grade_val) and team_grade_val is not None:
                            team_grade_val_str = str(team_grade_val).strip()
                            if team_grade_val_str and team_grade_val_str != 'None' and team_grade_val_str != 'nan':
                                team_grade_value = team_grade_val_str
                    except (KeyError, IndexError):
                        pass
                
                # Get Conference Grade value
                conf_grade_value = ''
                if conf_grade_col and conf_grade_col in position_df.columns:
                    try:
                        conf_grade_val = player_row[conf_grade_col]
                        if pd.notna(conf_grade_val) and conf_grade_val is not None:
                            conf_grade_val_str = str(conf_grade_val).strip()
                            if conf_grade_val_str and conf_grade_val_str != 'None' and conf_grade_val_str != 'nan':
                                conf_grade_value = conf_grade_val_str
                    except (KeyError, IndexError):
                        pass
                
                # Get Power Five Grade value
                power_five_grade_value = ''
                if power_five_grade_col and power_five_grade_col in position_df.columns:
                    try:
                        power_five_grade_val = player_row[power_five_grade_col]
                        if pd.notna(power_five_grade_val) and power_five_grade_val is not None:
                            power_five_grade_val_str = str(power_five_grade_val).strip()
                            if power_five_grade_val_str and power_five_grade_val_str != 'None' and power_five_grade_val_str != 'nan':
                                power_five_grade_value = power_five_grade_val_str
                    except (KeyError, IndexError):
                        pass
                
                ws.cell(row=row, column=1, value=player_name)
                ws.cell(row=row, column=2, value=player_row.get(position_col, ''))
                ws.cell(row=row, column=3, value=display_name)
                ws.cell(row=row, column=4, value=team_grade_value)  # Team Grade
                ws.cell(row=row, column=5, value=conf_grade_value)  # Conference Grade
                ws.cell(row=row, column=6, value=power_five_grade_value)  # Power Five Grade
                ws.cell(row=row, column=7, value=player_row.get(score_col, ''))
                ws.cell(row=row, column=8, value=player_row.get(score_2024_col, ''))
                ws.cell(row=row, column=9, value=player_row.get(change_col, ''))
                ws.cell(row=row, column=10, value=player_row.get(minutes_col, ''))
                row += 1
                total_players += 1
            
            # Add spacing between position profiles
            row += 1
        
        print(f"    ✅ Reorganized {total_players} players into {len(position_profiles_sorted)} position profiles")
        
        # Auto-adjust column widths
        headers = ['Player', 'Position', 'Position Profile', '2025 Grade', '2025 Total Score', '2024 Total Score', 'Change From 2024', 'Minutes']
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
        
        # Add conditional formatting for Grade and Minutes columns
        add_conditional_formatting_to_team_sheet(ws, headers, row - 1)
        
    except Exception as e:
        print(f"    ⚠️  Error reorganizing sheet: {e}")
        import traceback
        traceback.print_exc()
        return
    
    try:
        wb.save(file_path)
        print(f"    💾 Saved successfully")
    except Exception as e:
        print(f"    ❌ Error saving file: {e}")
        import traceback
        traceback.print_exc()


def main():
    """Main function."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    # Find Mike Norris report files
    report_files = list(base_dir.glob("Mike_Norris_Scouting_Report_*_IntraConference.xlsx"))
    
    if not report_files:
        print("❌ No Mike Norris report files found")
        return
    
    print("="*80)
    print("REORGANIZING TEAM SHEETS BY POSITION PROFILE")
    print("="*80)
    
    # Team names for each conference
    acc_teams = ['Stanford', 'Virginia', 'Notre Dame', 'Duke']
    sec_teams = ['Vanderbilt', 'Alabama', 'Georgia', 'Kentucky', 'Arkansas', 'Mississippi State', 'Tennessee', 'LSU']
    
    for file_path in report_files:
        print(f"\n📄 Processing: {file_path.name}")
        
        # Determine which teams are in this report
        wb = load_workbook(file_path)
        team_sheets = [s for s in wb.sheetnames if s not in ['Data Notes', 'Hybrid CB', 'DM Box-To-Box', 
                                                             'AM Advanced Playmaker', 'Right Touchline Winger']]
        
        for team_name in team_sheets:
            reorganize_team_sheet(file_path, team_name)
        
        # Don't save here - reorganize_team_sheet already saves
        print(f"✅ Completed: {file_path.name}")
    
    print(f"\n{'='*80}")
    print("✅ ALL TEAM SHEETS REORGANIZED")
    print(f"{'='*80}")
    print("\nTeams are now organized by position profile:")
    print("  • Hybrid CB (sorted A-F)")
    print("  • DM Box-To-Box (sorted A-F)")
    print("  • AM Advanced Playmaker (sorted A-F)")
    print("  • Right Touchline Winger (sorted A-F)")


if __name__ == "__main__":
    main()

