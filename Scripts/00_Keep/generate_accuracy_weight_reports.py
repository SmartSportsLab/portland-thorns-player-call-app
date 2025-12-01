#!/usr/bin/env python3
"""
Generate ACC scouting reports with different attempt vs. accuracy weightings.

This script:
  • Reads the baseline position metrics config
  • Creates adjusted config copies for 80/20, 70/30, and 60/40 attempt:accuracy ratios
  • Runs the full Mike Norris ACC report pipeline (with PAdj metrics) for each weighting
  • Saves both the config copies and the resulting Excel reports

Usage: python generate_accuracy_weight_reports.py
"""

from __future__ import annotations

import copy
import json
from pathlib import Path
from typing import Dict, Iterable, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Bring the main reporting helpers into scope
import update_mike_norris_reports as reports_module


BASE_DIR = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
CONFIG_DIR = BASE_DIR / "Scripts" / "00_Keep"
BASE_CONFIG_PATH = CONFIG_DIR / "position_metrics_config.json"

# Output names
REPORT_BASENAME = "Mike_Norris_Scouting_Report_ACC_IntraConference"
CONFERENCE = "ACC"


def load_base_config() -> Dict:
    """Load the default position metrics configuration."""
    if not BASE_CONFIG_PATH.exists():
        raise FileNotFoundError(f"Base config not found at {BASE_CONFIG_PATH}")
    with BASE_CONFIG_PATH.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def is_accuracy_metric(metric_name: str) -> bool:
    """Heuristic to flag component metrics that represent accuracy/efficiency."""
    name = metric_name.lower()
    return (
        "%" in metric_name
        or "accurate" in name
        or "successful" in name
        or "conversion" in name
        or "won" in name  # e.g., duels won %
        or "converted" in name
    )


def adjust_accuracy_weights(config: Dict, attempt_ratio: float) -> Dict:
    """
    Return a deep-copied config with attempt vs accuracy component weights rebalanced.

    For any metric with components split between attempt- and accuracy-style metrics,
    rescale the weights so that attempts sum to `attempt_ratio` and accuracy-oriented
    metrics sum to `1 - attempt_ratio`, maintaining the relative distribution within
    each group.
    """
    accuracy_ratio = 1.0 - attempt_ratio
    adjusted = copy.deepcopy(config)

    for profile in adjusted.get("position_profiles", {}).values():
        metrics = profile.get("metrics", {})
        for section_name in ("Core", "Specific"):
            section = metrics.get(section_name, {})
            for metric_name, metric_cfg in section.items():
                if not isinstance(metric_cfg, dict) or "components" not in metric_cfg:
                    continue

                components = metric_cfg["components"]
                attempt_keys = [k for k in components if not is_accuracy_metric(k)]
                accuracy_keys = [k for k in components if is_accuracy_metric(k)]

                if not attempt_keys or not accuracy_keys:
                    # Nothing to rebalance if we don't have both sides
                    continue

                attempt_total = sum(components[k] for k in attempt_keys)
                accuracy_total = sum(components[k] for k in accuracy_keys)

                if attempt_total <= 0 or accuracy_total <= 0:
                    continue

                attempt_scale = attempt_ratio / attempt_total
                accuracy_scale = accuracy_ratio / accuracy_total

                for key in attempt_keys:
                    components[key] = round(components[key] * attempt_scale, 6)
                for key in accuracy_keys:
                    components[key] = round(components[key] * accuracy_scale, 6)

                # Small numerical guard to ensure totals sum to ~1.0
                total = sum(components.values())
                if total != 0:
                    correction = round(1.0 - total, 6)
                    # Apply correction to the first attempt metric to keep ratios intact
                    components[attempt_keys[0]] = round(components[attempt_keys[0]] + correction, 6)

    return adjusted


def save_config(config: Dict, output_path: Path) -> None:
    """Persist the adjusted config to disk."""
    with output_path.open("w", encoding="utf-8") as fh:
        json.dump(config, fh, indent=2)
        fh.write("\n")


def generate_report(
    config: Dict,
    attempt_weight: float,
    accuracy_weight: float,
    config_label: str,
) -> Path:
    """Run the full ACC report pipeline with the supplied config and weighting."""
    print("\n" + "=" * 80)
    print(f"GENERATING ACC REPORT ({config_label}) — Attempt:{attempt_weight:.0%} Accuracy:{accuracy_weight:.0%}")
    print("=" * 80)

    championship_teams = reports_module.get_championship_teams(CONFERENCE)

    # Possession metrics for PAdj
    print("\n📊 Calculating team possessions for PAdj metrics...")
    team_data_dir = BASE_DIR / "Team Analysis" / CONFERENCE
    team_possessions = reports_module.calculate_team_possessions(team_data_dir)
    league_avg_possessions = reports_module.calculate_league_avg_possessions(team_possessions)
    print(f"  ✅ Calculated possessions for {len(team_possessions)} teams")
    print(f"  ✅ League average possessions: {league_avg_possessions:.2f} per game")

    # Power Five historical scoring (for Power Five grade)
    print("\n🌐 Loading and scoring all Power Five historical data (2021-2025)...")
    df_all_power_five = reports_module.load_all_power_five_historical_data(BASE_DIR)
    all_power_five_scored = {}

    if len(df_all_power_five) > 0:
        position_profiles = ['Center Back', 'Centre Midfielder', 'Attacking Midfielder', 'Winger']

        for position_name in position_profiles:
            if position_name not in config['position_profiles']:
                continue

            position_config = config['position_profiles'][position_name]
            display_name = reports_module.POSITION_PROFILE_MAP[position_name]

            df_power_five_pos = reports_module.filter_by_position(df_all_power_five, position_name)
            if len(df_power_five_pos) == 0:
                continue

            # Add PAdj metrics for Power Five data
            df_power_five_pos = reports_module.add_padj_metrics_to_dataframe(
                df_power_five_pos, team_possessions, league_avg_possessions, team_col='Team'
            )

            scored_by_year = []
            for year in range(2021, 2026):
                df_year = df_power_five_pos[df_power_five_pos['Year'] == year].copy()
                if len(df_year) == 0:
                    continue

                df_historical = df_power_five_pos[df_power_five_pos['Year'] < year].copy()
                if len(df_historical) == 0:
                    df_historical = df_year.copy()

                position_config_padj = reports_module.update_config_to_use_padj(position_config, df_year)
                df_scored = reports_module.calculate_with_historical_normalization(
                    df_year, df_historical, position_config_padj,
                    attempt_weight, accuracy_weight, position_name
                )
                scored_by_year.append(df_scored)

            if scored_by_year:
                df_power_five_scored = pd.concat(scored_by_year, ignore_index=True)
                df_power_five_scored['Position_Profile'] = display_name
                all_power_five_scored[display_name] = df_power_five_scored
                print(f"  ✅ Scored {display_name}: {len(df_power_five_scored)} players")
    else:
        print("  ⚠️  No Power Five historical data found")

    # Load core datasets
    acc_sec_dir = BASE_DIR / "ACC:SEC Championship Reports"
    file_2025 = acc_sec_dir / f"{CONFERENCE} All Positions 2025 Intra-Conference.xlsx"
    if not file_2025.exists():
        raise FileNotFoundError(f"2025 intra-conference file not found at {file_2025}")

    print(f"\n📄 Loading 2025 data: {file_2025.name}")
    df_2025 = pd.read_excel(file_2025)
    print(f"  ✅ Loaded {len(df_2025)} players")

    print("\n📚 Loading historical data (2021-2025)...")
    df_all = reports_module.load_historical_data(BASE_DIR, CONFERENCE)
    print(f"  ✅ Total historical rows: {len(df_all)}")

    print("\n📊 Loading 2024 data for comparison...")
    df_2024 = reports_module.load_2024_data_for_comparison(BASE_DIR, CONFERENCE)

    print("\n⏱️  Loading team total minutes from team stats files...")
    team_minutes_dict = reports_module.load_team_total_minutes(BASE_DIR)

    print("\n📊 Loading team averages for team-relative analysis...")
    team_averages_dict = reports_module.load_team_averages(BASE_DIR)

    wb = Workbook()
    wb.remove(wb.active)

    position_profiles = ['Center Back', 'Centre Midfielder', 'Attacking Midfielder', 'Winger']
    all_scored_data = []

    for position_name in position_profiles:
        if position_name not in config['position_profiles']:
            continue

        position_config = config['position_profiles'][position_name]
        display_name = reports_module.POSITION_PROFILE_MAP[position_name]

        print(f"\n📋 Processing {display_name}...")

        df_2025_pos = reports_module.filter_by_position(df_2025, position_name)
        df_2025_pos = reports_module.filter_to_championship_teams(df_2025_pos, championship_teams)
        if len(df_2025_pos) == 0:
            print(f"  ⚠️  No players from championship teams for {position_name}")
            continue

        df_all_pos = reports_module.filter_by_position(df_all, position_name)
        if len(df_all_pos) == 0:
            print(f"  ⚠️  No historical data for {position_name}")
            continue

        # Add PAdj metrics
        print("  📊 Adding PAdj metrics...")
        df_2025_pos = reports_module.add_padj_metrics_to_dataframe(
            df_2025_pos, team_possessions, league_avg_possessions, team_col='Team'
        )
        df_all_pos = reports_module.add_padj_metrics_to_dataframe(
            df_all_pos, team_possessions, league_avg_possessions, team_col='Team'
        )

        # Update config to use PAdj versions if present
        print("  🔄 Updating config to use PAdj metrics...")
        position_config_padj = reports_module.update_config_to_use_padj(position_config, df_2025_pos)

        print(f"  🧮 Calculating scores with {attempt_weight:.0%}/{accuracy_weight:.0%} weighting...")
        df_scored = reports_module.calculate_with_historical_normalization(
            df_2025_pos, df_all_pos, position_config_padj,
            attempt_weight, accuracy_weight, position_name
        )

        # Percentiles & grades
        reference_distribution = df_scored['Total_Score_1_10']
        df_scored['Total_Percentile'] = df_scored['Total_Score_1_10'].apply(
            lambda x: reports_module.calculate_percentile_against_distribution(x, reference_distribution)
        )
        df_scored['Total_Grade'] = df_scored['Total_Score_1_10'].apply(reports_module.assign_grade_single)

        # 2024 comparison
        if len(df_2024) > 0:
            df_2024_pos = reports_module.filter_by_position(df_2024, position_name)
            if len(df_2024_pos) > 0:
                df_2024_scored = reports_module.calculate_with_historical_normalization(
                    df_2024_pos, df_all_pos[df_all_pos['Year'] < 2025], position_config,
                    attempt_weight, accuracy_weight, position_name
                )
                scores_2024_dict = dict(zip(
                    df_2024_scored['Player'].astype(str),
                    df_2024_scored['Total_Score_1_10']
                ))
                df_scored['2024_Total_Score'] = df_scored['Player'].astype(str).map(scores_2024_dict)
                df_scored['Change_From_2024'] = df_scored['Total_Score_1_10'] - df_scored['2024_Total_Score']
            else:
                df_scored['2024_Total_Score'] = None
                df_scored['Change_From_2024'] = None
        else:
            df_scored['2024_Total_Score'] = None
            df_scored['Change_From_2024'] = None

        df_scored['Position_Profile'] = display_name
        df_scored['Total_Grade'] = df_scored['Total_Grade'].fillna('')

        print("  🏆 Calculating Team Grades...")
        df_scored['Team_Grade'] = df_scored.apply(
            lambda row: reports_module.calculate_team_grade(row, df_scored), axis=1
        )

        print("  🏆 Calculating Conference Grades...")
        df_scored['Conference_Grade'] = df_scored.apply(
            lambda row: reports_module.calculate_conference_grade(row, df_scored, CONFERENCE), axis=1
        )

        print("  🏆 Calculating Power Five Grades...")
        if display_name in all_power_five_scored:
            df_power_five_pos = all_power_five_scored[display_name]
            df_scored['Power_Five_Grade'] = df_scored.apply(
                lambda row: reports_module.calculate_power_five_grade(row, df_power_five_pos, display_name), axis=1
            )
        else:
            df_scored['Power_Five_Grade'] = ''

        relevant_metrics = reports_module.get_relevant_metrics_for_position(position_config)
        df_scored = reports_module.filter_players_with_metrics(df_scored, relevant_metrics)
        if len(df_scored) == 0:
            print(f"  ⚠️  No players with metric data for {display_name}, skipping sheet")
            continue

        all_scored_data.append(df_scored)

        # Team minutes percentage
        percentages_pos = reports_module.calculate_player_percentage_of_team_minutes(df_scored, team_minutes_dict)
        df_scored['Pct_Of_Team_Minutes'] = percentages_pos

        reports_module.create_position_profile_sheet(
            wb, df_scored, display_name, team_minutes_dict, position_config, team_averages_dict
        )
        print(f"  ✅ {display_name}: {len(df_scored)} players processed")

    if all_scored_data:
        df_all_scored = pd.concat(all_scored_data, ignore_index=True)
        df_all_scored['Pct_Of_Team_Minutes'] = reports_module.calculate_player_percentage_of_team_minutes(
            df_all_scored, team_minutes_dict
        )
        reports_module.create_team_summary_sheets(
            wb, df_all_scored, CONFERENCE, team_minutes_dict, config, team_averages_dict
        )

    reports_module.create_data_notes_sheet(wb, CONFERENCE)

    # Page setup
    for ws in wb.worksheets:
        if ws.max_row > 0 and ws.max_column > 0:
            print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            ws.print_area = print_area
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = 100
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

    attempt_percent = int(round(attempt_weight * 100))
    accuracy_percent = int(round(accuracy_weight * 100))
    suffix = f"{attempt_percent}_{accuracy_percent}"
    output_path = BASE_DIR / f"{REPORT_BASENAME}_{suffix}_PAdj.xlsx"
    wb.save(output_path)

    print(f"\n✅ Report saved: {output_path.name}")
    print(f"   Using weighting: {attempt_percent}% attempts / {accuracy_percent}% accuracy")
    return output_path


def main():
    base_config = load_base_config()

    weighting_specs: Iterable[Tuple[float, str]] = [
        (0.80, "accuracy_80_20"),
        (0.70, "accuracy_70_30"),
        (0.60, "accuracy_60_40"),
    ]

    generated_reports = []

    for attempt_ratio, label in weighting_specs:
        adjusted_config = adjust_accuracy_weights(base_config, attempt_ratio)
        config_path = CONFIG_DIR / f"position_metrics_config_{label}.json"
        save_config(adjusted_config, config_path)
        print(f"\n💾 Saved adjusted config: {config_path.name}")

        report_path = generate_report(
            adjusted_config,
            attempt_weight=attempt_ratio,
            accuracy_weight=1.0 - attempt_ratio,
            config_label=label.upper(),
        )
        generated_reports.append((attempt_ratio, report_path))

    print("\n" + "=" * 80)
    print("SUMMARY — Generated Reports")
    print("=" * 80)
    for attempt_ratio, path in generated_reports:
        print(f"  • {int(round(attempt_ratio * 100))}/{int(round((1 - attempt_ratio) * 100))} weighting → {path.name}")
    print("=" * 80)


if __name__ == "__main__":
    main()


