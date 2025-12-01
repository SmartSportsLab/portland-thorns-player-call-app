#!/usr/bin/env python3
"""
Manual verification tool for Top 15s calculation.
Allows you to check if a player's Top 15s count is correct.
"""

import sys
from pathlib import Path
import pandas as pd
import json

# Add path for imports
sys.path.insert(0, str(Path(__file__).parent))

from update_mike_norris_reports import load_conference_season_data, get_relevant_metrics_for_position

def verify_player_top15s(player_name, team_name, position_profile, base_dir):
    """
    Verify Top 15s calculation for a specific player.
    
    Args:
        player_name: Player name (e.g., "H. McLaughlin")
        team_name: Team name (e.g., "Vanderbilt")
        position_profile: Position profile (e.g., "Hybrid CB")
        base_dir: Base directory for data
    """
    print("="*80)
    print(f"TOP 15s VERIFICATION FOR {player_name.upper()} ({team_name})")
    print(f"Position Profile: {position_profile}")
    print("="*80)
    
    # Load config
    config_file = Path(__file__).parent / "position_metrics_config.json"
    with open(config_file, 'r') as f:
        config = json.load(f)
    
    # Get position name mapping
    POSITION_PROFILE_MAP = {
        'Hybrid CB': 'Center Back',
        'DM Box-To-Box': 'Centre Midfielder',
        'AM Advanced Playmaker': 'Attacking Midfielder',
        'Right Touchline Winger': 'Winger'
    }
    
    position_name = POSITION_PROFILE_MAP.get(position_profile, position_profile)
    
    # Get relevant metrics for this position
    if position_name in config.get('position_profiles', {}):
        position_config = config['position_profiles'][position_name]
        relevant_metrics = get_relevant_metrics_for_position(position_config)
    else:
        print(f"⚠️  Position {position_name} not found in config")
        return
    
    # Use the same desired_metric_order as shortlist script
    desired_metric_order_map = {
        'Hybrid CB': [
            'Defensive duels per 90',
            'Defensive duels won, %',
            'Aerial duels per 90',
            'Aerial duels won, %',
            'Interceptions + Sliding Tackles',  # Combined metric
            'Shots blocked per 90',
            'Passes per 90',
            'Accurate passes, %',
            'Progressive passes per 90',
            'Accurate progressive passes, %',
            'Long passes per 90',
            'Accurate long passes, %',
            'Short / medium passes per 90',
            'Accurate short / medium passes, %'
        ],
        'DM Box-To-Box': [
            # Add other positions as needed
        ],
        'AM Advanced Playmaker': [
            # Add other positions as needed
        ],
        'Right Touchline Winger': [
            # Add other positions as needed
        ]
    }
    
    # Use desired_metric_order if available, otherwise use relevant_metrics
    if position_profile in desired_metric_order_map:
        metrics_to_check = desired_metric_order_map[position_profile]
        print(f"\n📊 Using shortlist metric order for {position_profile}: {len(metrics_to_check)} metrics")
    else:
        metrics_to_check = sorted(relevant_metrics)
        print(f"\n📊 Using configured metrics for {position_profile}: {len(metrics_to_check)} metrics")
    
    print(f"   Metrics: {', '.join(metrics_to_check[:10])}...")
    
    # Load Power Five 2025 data
    print(f"\n📥 Loading Power Five 2025 data...")
    all_data = []
    power_five_conferences = ['ACC', 'SEC', 'BIG10', 'BIG12', 'IVY']
    
    for conference in power_five_conferences:
        df = load_conference_season_data(base_dir, conference, 2025)
        if not df.empty:
            df['Conference'] = conference
            df['Year'] = 2025
            all_data.append(df)
    
    if not all_data:
        print("❌ No Power Five data found!")
        return
    
    all_players_raw = pd.concat(all_data, ignore_index=True)
    power_five_players = all_players_raw[
        (all_players_raw['Position_Profile'] == position_profile) &
        (all_players_raw['Conference'].isin(power_five_conferences))
    ].copy()
    
    print(f"   ✅ Loaded {len(power_five_players)} Power Five players for {position_profile}")
    
    # Find the player
    player_mask = (
        power_five_players['Player'].str.contains(player_name, case=False, na=False) &
        power_five_players['Team'].str.contains(team_name, case=False, na=False)
    )
    
    player_data = power_five_players[player_mask]
    
    if len(player_data) == 0:
        print(f"\n❌ Player '{player_name}' from '{team_name}' not found!")
        print(f"   Available players with similar names:")
        similar_players = power_five_players[
            power_five_players['Player'].str.contains(player_name.split()[0] if ' ' in player_name else player_name, case=False, na=False)
        ][['Player', 'Team']].drop_duplicates()
        for _, row in similar_players.head(10).iterrows():
            print(f"     - {row['Player']} ({row['Team']})")
        return
    
    if len(player_data) > 1:
        print(f"\n⚠️  Multiple players found:")
        for idx, (_, row) in enumerate(player_data.iterrows()):
            print(f"   {idx+1}. {row['Player']} - {row['Team']} ({row.get('Conference', 'N/A')})")
        print(f"\n   Using first match...")
    
    player_row = player_data.iloc[0]
    print(f"\n✅ Found: {player_row['Player']} - {player_row['Team']} ({player_row.get('Conference', 'N/A')})")
    
    # Check each metric
    print(f"\n{'='*80}")
    print("METRIC-BY-METRIC TOP 15s CHECK")
    print(f"{'='*80}\n")
    
    top15_count = 0
    top15_metrics = []
    not_top15_metrics = []
    combined_metric_handled = False
    
    for metric_header in metrics_to_check:
        # Handle combined metric "Interceptions + Sliding Tackles"
        if metric_header == "Interceptions + Sliding Tackles":
            # Check interceptions component
            interceptions_col = None
            for col in player_row.index:
                if 'interception' in str(col).lower() and 'per 90' in str(col).lower():
                    interceptions_col = col
                    break
            
            if interceptions_col:
                player_val = player_row.get(interceptions_col)
                if pd.notna(player_val):
                    try:
                        player_val = float(player_val)
                        # Find matching column in power_five_players
                        pf_col = None
                        for col in power_five_players.columns:
                            if 'interception' in str(col).lower() and 'per 90' in str(col).lower():
                                pf_col = col
                                break
                        
                        if pf_col:
                            pf_values = pd.to_numeric(power_five_players[pf_col], errors='coerce').dropna()
                            if len(pf_values) > 0:
                                pf_rank = (pf_values >= player_val).sum()
                                is_top15 = pf_rank <= 15
                                
                                status = "✅ TOP 15" if is_top15 else "❌"
                                print(f"{status} {metric_header} (using Interceptions per 90)")
                                print(f"      Player Value: {player_val:.2f}")
                                print(f"      Rank: {pf_rank} (out of {len(pf_values)} players)")
                                print()
                                
                                if is_top15:
                                    top15_count += 1
                                    top15_metrics.append((metric_header, pf_rank, player_val))
                                    combined_metric_handled = True
                                else:
                                    not_top15_metrics.append((metric_header, pf_rank, player_val))
                    except:
                        pass
            continue
        
        # Find matching column in player data
        matching_col = None
        metric_lower = str(metric_header).lower()
        
        # Try exact match
        for col in player_row.index:
            if str(col).lower() == metric_lower:
                matching_col = col
                break
        
        # Try base name match
        if not matching_col:
            if 'per 90' in metric_lower:
                metric_base = metric_lower.replace(' per 90', '').replace(' per90', '').strip()
                for col in player_row.index:
                    col_lower = str(col).lower()
                    if 'per 90' in col_lower and '%' not in col_lower:
                        col_base = col_lower.replace(' per 90', '').replace(' per90', '').strip()
                        if ' '.join(metric_base.split()) == ' '.join(col_base.split()):
                            matching_col = col
                            break
            elif '%' in metric_lower:
                metric_base = metric_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                for col in player_row.index:
                    col_lower = str(col).lower()
                    if ('%' in col_lower or 'percent' in col_lower) and 'per 90' not in col_lower:
                        col_base = col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                        if ' '.join(metric_base.split()) == ' '.join(col_base.split()):
                            matching_col = col
                            break
        
        if not matching_col:
            print(f"⚠️  {metric}: Column not found")
            continue
        
        # Get player's value
        player_value = player_row.get(matching_col)
        if pd.isna(player_value) or player_value == '':
            print(f"⏭️  {metric}: No value")
            continue
        
        try:
            player_val = float(player_value)
        except (ValueError, TypeError):
            print(f"⚠️  {metric}: Invalid value ({player_value})")
            continue
        
        # Find matching column in power_five_players
        pf_matching_col = None
        for col in power_five_players.columns:
            if str(col).lower() == str(matching_col).lower():
                pf_matching_col = col
                break
        
        if not pf_matching_col:
            # Try base name match
            col_lower = str(matching_col).lower()
            if 'per 90' in col_lower:
                col_base = col_lower.replace(' per 90', '').replace(' per90', '').strip()
                for pf_col in power_five_players.columns:
                    pf_col_lower = str(pf_col).lower()
                    if 'per 90' in pf_col_lower and '%' not in pf_col_lower:
                        pf_col_base = pf_col_lower.replace(' per 90', '').replace(' per90', '').strip()
                        if ' '.join(col_base.split()) == ' '.join(pf_col_base.split()):
                            pf_matching_col = pf_col
                            break
            elif '%' in col_lower:
                col_base = col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                for pf_col in power_five_players.columns:
                    pf_col_lower = str(pf_col).lower()
                    if ('%' in pf_col_lower or 'percent' in pf_col_lower) and 'per 90' not in pf_col_lower:
                        pf_col_base = pf_col_lower.replace(' won, %', '').replace(' won %', '').replace(' won', '').replace(', %', '').replace(' %', '').replace('accurate, %', '').replace('accurate %', '').strip()
                        if ' '.join(col_base.split()) == ' '.join(pf_col_base.split()):
                            pf_matching_col = pf_col
                            break
        
        if not pf_matching_col:
            print(f"⚠️  {metric}: Could not find matching column in Power Five data")
            continue
        
        # Calculate rank
        pf_values = pd.to_numeric(power_five_players[pf_matching_col], errors='coerce').dropna()
        if len(pf_values) == 0:
            print(f"⚠️  {metric}: No valid values in Power Five data")
            continue
        
        # Rank = how many players have >= this value
        pf_rank = (pf_values >= player_val).sum()
        is_top15 = pf_rank <= 15
        
        # Skip if this is part of combined metric we already handled
        if combined_metric_handled and ('interception' in metric_lower or 'sliding' in metric_lower):
            continue
        
        # Show details
        status = "✅ TOP 15" if is_top15 else "❌"
        print(f"{status} {metric_header}")
        print(f"      Player Value: {player_val:.2f}")
        print(f"      Rank: {pf_rank} (out of {len(pf_values)} players)")
        print(f"      Column Used: {pf_matching_col}")
        
        # Show top 5 values for context
        top_values = pf_values.nlargest(5).tolist()
        print(f"      Top 5 Values: {[f'{v:.2f}' for v in top_values]}")
        
        if is_top15:
            top15_count += 1
            top15_metrics.append((metric_header, pf_rank, player_val))
        else:
            not_top15_metrics.append((metric_header, pf_rank, player_val))
        
        print()
    
    # Summary
    print(f"\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}\n")
    print(f"Total Top 15s: {top15_count}")
    print(f"\n✅ Metrics in Top 15 ({len(top15_metrics)}):")
    for metric, rank, val in sorted(top15_metrics, key=lambda x: x[1]):
        print(f"   {metric}: Rank {rank} (Value: {val:.2f})")
    
    print(f"\n❌ Metrics NOT in Top 15 ({len(not_top15_metrics)}):")
    for metric, rank, val in sorted(not_top15_metrics, key=lambda x: x[1])[:10]:
        print(f"   {metric}: Rank {rank} (Value: {val:.2f})")
    if len(not_top15_metrics) > 10:
        print(f"   ... and {len(not_top15_metrics) - 10} more")
    
    # Compare with shortlist value
    print(f"\n{'='*80}")
    print("COMPARISON WITH SHORTLIST")
    print(f"{'='*80}\n")
    
    shortlist_file = base_dir / "Portland Thorns 2025 Shortlist.xlsx"
    if shortlist_file.exists():
        from openpyxl import load_workbook
        wb = load_workbook(shortlist_file, data_only=True)
        if position_profile in wb.sheetnames:
            ws = wb[position_profile]
            
            # Find Top 15s column
            top15_col = None
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=3, column=col_idx).value
                if header and 'Top 15s' in str(header):
                    top15_col = col_idx
                    break
            
            # Find player
            for row_idx in range(4, ws.max_row + 1):
                ws_player = ws.cell(row=row_idx, column=2).value
                ws_team = ws.cell(row=row_idx, column=3).value
                if ws_player and player_name.lower() in str(ws_player).lower() and team_name.lower() in str(ws_team).lower():
                    shortlist_top15 = ws.cell(row=row_idx, column=top15_col).value if top15_col else None
                    print(f"Shortlist Top 15s: {shortlist_top15}")
                    print(f"Calculated Top 15s: {top15_count}")
                    if shortlist_top15 == top15_count:
                        print("✅ MATCH!")
                    else:
                        print(f"⚠️  MISMATCH! Difference: {abs(shortlist_top15 - top15_count)}")
                    break
        wb.close()
    else:
        print("Shortlist file not found")


def main():
    """Interactive verification tool."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    if len(sys.argv) >= 4:
        # Command line arguments
        player_name = sys.argv[1]
        team_name = sys.argv[2]
        position_profile = sys.argv[3]
    else:
        # Interactive mode
        print("TOP 15s VERIFICATION TOOL")
        print("="*80)
        print("\nEnter player details:")
        player_name = input("Player Name (e.g., H. McLaughlin): ").strip()
        team_name = input("Team Name (e.g., Vanderbilt): ").strip()
        
        print("\nPosition Profile:")
        print("1. Hybrid CB")
        print("2. DM Box-To-Box")
        print("3. AM Advanced Playmaker")
        print("4. Right Touchline Winger")
        choice = input("Select (1-4): ").strip()
        
        position_map = {
            '1': 'Hybrid CB',
            '2': 'DM Box-To-Box',
            '3': 'AM Advanced Playmaker',
            '4': 'Right Touchline Winger'
        }
        position_profile = position_map.get(choice, 'Hybrid CB')
    
    verify_player_top15s(player_name, team_name, position_profile, base_dir)


if __name__ == "__main__":
    main()

