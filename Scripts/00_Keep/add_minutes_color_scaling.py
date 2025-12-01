#!/usr/bin/env python3
"""
Add color scaling to Minutes Played column in Mike Norris reports.
Dark red = most minutes (largest sample size)
Dark blue = least minutes (smallest sample size)
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def find_max_minutes(report_files):
    """Find the maximum minutes played across all reports."""
    max_minutes = 0
    
    for file_path in report_files:
        wb = load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find Minutes Played column
            minutes_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                header_value = str(cell.value).strip() if cell.value else ""
                if 'Minutes Played' in header_value or 'Minutes played' in header_value:
                    minutes_col = col_idx
                    break
            
            if minutes_col and ws.max_row > 1:
                minutes_col_letter = get_column_letter(minutes_col)
                
                # Find max minutes in this sheet
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{minutes_col_letter}{row}']
                    if cell.value:
                        try:
                            minutes = float(cell.value)
                            if minutes > max_minutes:
                                max_minutes = minutes
                        except (ValueError, TypeError):
                            continue
    
    return max_minutes


def add_minutes_color_scaling(file_path, max_minutes):
    """Add color scaling to Minutes Played column."""
    print(f"📄 Processing: {file_path.name}")
    
    wb = load_workbook(file_path)
    
    # Define color scheme
    colors = {
        'dark_blue': '1F4E79',    # Least minutes (smallest sample)
        'blue': '4472C4',         # Low minutes
        'light_blue': '8FAADC',   # Below average
        'white': 'FFFFFF',        # Average
        'light_red': 'F2A2A2',    # Above average
        'red': 'C5504B',          # High minutes
        'dark_red': '8B0000'      # Most minutes (largest sample)
    }
    
    sheets_processed = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Skip Data Notes sheet
        if sheet_name == 'Data Notes':
            continue
        
        # Find Minutes Played column
        minutes_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            header_value = str(cell.value).strip() if cell.value else ""
            if 'Minutes Played' in header_value or 'Minutes played' in header_value:
                minutes_col = col_idx
                break
        
        if minutes_col and ws.max_row > 1:
            minutes_col_letter = get_column_letter(minutes_col)
            minutes_range = f'{minutes_col_letter}2:{minutes_col_letter}{ws.max_row}'
            
            # Define discrete minute ranges with solid colors
            minute_ranges = [
                (0, 187, colors['dark_blue']),      # 0 - 187: Dark blue (very small sample)
                (187, 373, colors['blue']),          # 187 - 373: Blue (small sample)
                (373, 560, colors['light_blue']),    # 373 - 560: Light blue (below average)
                (560, 747, colors['light_red']),     # 560 - 747: Light red (above average)
                (747, 933, colors['red']),           # 747 - 933: Red (large sample)
                (933, max_minutes, colors['dark_red'])  # 933 - 1119: Dark red (very large sample)
            ]
            
            # Clear any existing formatting for this column by rebuilding conditional formatting
            # We'll replace the entire conditional formatting list, keeping non-minutes formatting
            # For simplicity, clear all and rebuild (we'll need to preserve grade formatting separately)
            # Actually, let's just clear the minutes range by creating a new conditional formatting list
            # and only adding back non-minutes ranges, then add our minutes ranges
            
            # Get all existing conditional formatting ranges except minutes
            existing_ranges = []
            for cf in ws.conditional_formatting:
                if str(cf.sqref) != minutes_range:
                    # Keep this range, we'll re-add it
                    existing_ranges.append((str(cf.sqref), cf.rules))
            
            # Clear all and rebuild
            cf_type = type(ws.conditional_formatting)
            ws.conditional_formatting = cf_type()
            
            # Re-add non-minutes ranges (like grade formatting)
            for range_ref, rules in existing_ranges:
                for rule in rules:
                    ws.conditional_formatting.add(range_ref, rule)
            
            # Add discrete color ranges using CellIsRule
            for min_val, max_val, color in minute_ranges:
                if min_val == 0:
                    # For the first range (0 - 187), use lessThanOrEqual
                    minute_rule = CellIsRule(
                        operator='lessThanOrEqual', formula=[f'{max_val}'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                    )
                elif max_val >= max_minutes:
                    # For the last range (933 - 1119), use greaterThanOrEqual
                    minute_rule = CellIsRule(
                        operator='greaterThanOrEqual', formula=[f'{min_val}'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                    )
                else:
                    # For middle ranges, use between
                    minute_rule = CellIsRule(
                        operator='between', formula=[f'{min_val}', f'{max_val}'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type='solid')
                    )
                
                ws.conditional_formatting.add(minutes_range, minute_rule)
            sheets_processed += 1
            print(f"  ✅ Added color scaling to {sheet_name} (Minutes column {minutes_col_letter})")
    
    wb.save(file_path)
    print(f"✅ Saved: {file_path.name} ({sheets_processed} sheets processed)\n")


def main():
    """Main function."""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    # Find Mike Norris report files
    report_files = list(base_dir.glob("Mike_Norris_Scouting_Report_*_IntraConference.xlsx"))
    
    if not report_files:
        print("❌ No Mike Norris report files found")
        print(f"   Looking in: {base_dir}")
        return
    
    print("="*80)
    print("ADDING MINUTES PLAYED COLOR SCALING")
    print("="*80)
    
    # First, find the maximum minutes across all reports
    print(f"\n📊 Finding maximum minutes played across all reports...")
    max_minutes = find_max_minutes(report_files)
    print(f"   ✅ Maximum minutes found: {max_minutes:.0f}")
    
    print(f"\n🎨 Applying color scaling to {len(report_files)} report(s)...\n")
    
    for file_path in report_files:
        add_minutes_color_scaling(file_path, max_minutes)
    
    print("="*80)
    print("✅ ALL REPORTS UPDATED")
    print("="*80)
    print(f"\nDiscrete color ranges applied:")
    print(f"  • 0 - 187 minutes: Dark blue (#1F4E79) - Very small sample")
    print(f"  • 187 - 373 minutes: Blue (#4472C4) - Small sample")
    print(f"  • 373 - 560 minutes: Light blue (#8FAADC) - Below average")
    print(f"  • 560 - 747 minutes: Light red (#F2A2A2) - Above average")
    print(f"  • 747 - 933 minutes: Red (#C5504B) - Large sample")
    print(f"  • 933 - {max_minutes:.0f} minutes: Dark red (#8B0000) - Very large sample")
    print(f"\nThis helps identify players with small sample sizes (e.g., F. Spiekerkoetter: 80 minutes = dark blue)")


if __name__ == "__main__":
    main()

