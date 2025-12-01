#!/usr/bin/env python3
"""
Add Standout Players Analysis to PAdj Report
=============================================

Identifies players who perform well despite their team's playing style.
Highlights players who excel in metrics that aren't emphasized by their team.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def get_championship_teams():
    """Get the four ACC championship teams"""
    return ['Duke Blue Devils', 'Notre Dame Fighting Irish', 'Stanford Cardinal', 'Virginia Cavaliers']


def get_team_possession_context():
    """Get team possession data"""
    import sys
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    sys.path.insert(0, str(base_dir / "Scripts" / "00_Keep"))
    import update_mike_norris_reports as reports_module
    
    team_data_dir = base_dir / "Team Analysis" / "ACC"
    team_possessions = reports_module.calculate_team_possessions(team_data_dir)
    league_avg = reports_module.calculate_league_avg_possessions(team_possessions)
    
    return team_possessions, league_avg


def load_team_tab_data(report_path, team):
    """Load detailed player data from a team tab"""
    wb = load_workbook(report_path, keep_vba=False, data_only=True)
    
    if team not in wb.sheetnames:
        return []
    
    ws = wb[team]
    players = []
    
    # Find column headers (row 2)
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=2, column=col).value
        if header_val:
            headers[header_val] = col
    
    # Read player data (starts at row 3)
    for row in range(3, ws.max_row + 1):
        player = ws.cell(row=row, column=1).value
        if not player:
            continue
        
        player_data = {
            'Player': player,
            'Team': team,
            'Position': ws.cell(row=row, column=2).value,
            'Position_Profile': ws.cell(row=row, column=3).value,
            'Score': ws.cell(row=row, column=6).value,  # 2025 Total Score
        }
        
        # Extract team-relative metrics (% team +/- columns)
        team_relative_metrics = {}
        
        # Find all '% team +/-' columns
        for col_idx in range(1, min(ws.max_column + 1, 150)):
            header2 = ws.cell(row=2, column=col_idx).value
            value = ws.cell(row=row, column=col_idx).value
            
            if header2 and '% team +/-' in str(header2):
                try:
                    if value is not None:
                        float_val = float(value)
                        # Get the metric name from the previous column (usually the % column)
                        prev_header = ws.cell(row=2, column=col_idx - 1).value if col_idx > 1 else None
                        if prev_header:
                            # Clean up the metric name
                            base_metric = str(prev_header).replace(' per 90', '').replace('%', '').strip()
                            if base_metric and base_metric != '% team +/-' and len(base_metric) > 2:
                                team_relative_metrics[base_metric] = float_val
                except (ValueError, TypeError):
                    pass
        
        player_data['Team_Relative_Metrics'] = team_relative_metrics
        
        # Extract PAdj metrics
        padj_metrics = {}
        for header, col_idx in headers.items():
            if header and 'PAdj' in str(header):
                value = ws.cell(row=row, column=col_idx).value
                if value is not None and isinstance(value, (int, float)):
                    padj_metrics[str(header)] = float(value)
        
        player_data['PAdj_Metrics'] = padj_metrics
        
        players.append(player_data)
    
    return players


def identify_standout_players(report_path, team_possessions, league_avg):
    """Identify players who excel despite their team's playing style"""
    
    all_players = []
    for team in get_championship_teams():
        players = load_team_tab_data(report_path, team)
        all_players.extend(players)
    
    standouts = []
    
    # Define which metrics indicate "going against team style"
    # High possession teams: Look for defensive standouts
    # Low possession teams: Look for attacking/passing standouts
    
    for player in all_players:
        team = player['Team']
        if team not in team_possessions:
            continue
        
        team_poss = team_possessions[team]
        team_relative = player.get('Team_Relative_Metrics', {})
        
        # Calculate standout score
        standout_score = 0
        standout_reasons = []
        
        # High possession team (many opportunities for attacking players)
        # Standout: Player excels defensively or in metrics team doesn't emphasize
        if team_poss > league_avg + 3:  # High possession
            # Look for defensive standouts or players with high team-relative metrics
            for metric, pct_diff in team_relative.items():
                if pct_diff >= 20:  # 20%+ above team average
                    standout_score += 1
                    standout_reasons.append(f"{metric}: +{pct_diff:.1f}% vs team")
        
        # Low possession team (fewer opportunities for attacking players)
        # Standout: Player excels in attacking/passing despite fewer opportunities
        elif team_poss < league_avg - 3:  # Low possession
            # Look for attacking/passing standouts
            for metric, pct_diff in team_relative.items():
                metric_lower = metric.lower()
                # Attacking/passing metrics
                if any(keyword in metric_lower for keyword in ['pass', 'progressive', 'shot', 'key', 'assist', 'dribble', 'offensive']):
                    if pct_diff >= 20:
                        standout_score += 1
                        standout_reasons.append(f"{metric}: +{pct_diff:.1f}% vs team")
        
        # Average possession team: Look for any strong team-relative performance
        else:
            for metric, pct_diff in team_relative.items():
                if pct_diff >= 25:  # Higher threshold for average teams
                    standout_score += 1
                    standout_reasons.append(f"{metric}: +{pct_diff:.1f}% vs team")
        
        # Also check if player has high PAdj scores relative to team average
        # This indicates they perform well despite team possession style
        if player['Score'] and player['Score'] > 7.0:  # Good overall score
            standout_score += 1
        
        if standout_score >= 1:  # At least 1 standout metric or good PAdj score
            standouts.append({
                'Player': player['Player'],
                'Team': team,
                'Position': player['Position'],
                'Position_Profile': player['Position_Profile'],
                'Score': player['Score'],
                'Team_Possessions': team_poss,
                'Possession_Style': 'High' if team_poss > league_avg + 3 else 'Low' if team_poss < league_avg - 3 else 'Average',
                'Standout_Score': standout_score,
                'Standout_Reasons': '; '.join(standout_reasons[:5])  # Top 5 reasons
            })
    
    return sorted(standouts, key=lambda x: x['Standout_Score'], reverse=True)


def add_standout_players_sheet(wb, report_path, team_possessions, league_avg):
    """Add sheet highlighting standout players"""
    
    ws = wb.create_sheet(title="Standout Performers", index=1)
    
    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value="STANDOUT PLAYERS: Excellence Despite Team Style")
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Explanation
    row = 3
    explanation = [
        "These players excel in metrics that their team doesn't typically emphasize,",
        "showing they perform well DESPITE their team's playing style, not because of it.",
        "",
        "High Possession Teams (e.g., Notre Dame):",
        "  • Look for defensive standouts who excel despite team's attacking focus",
        "  • Players with high defensive metrics when team emphasizes possession",
        "",
        "Low Possession Teams (e.g., Virginia):",
        "  • Look for attacking/passing standouts who create despite fewer opportunities",
        "  • Players with high creative/attacking metrics when team plays conservatively",
        "",
        "Criteria:",
        "  • ≥20% above team average in relevant metrics, OR",
        "  • High PAdj scores (≥7.0) indicating strong performance despite team style"
    ]
    
    for line in explanation:
        ws.cell(row=row, column=1, value=line)
        row += 1
    
    row += 2
    
    # Headers
    headers = ['Player', 'Team', 'Position', 'Position Profile', 'Score', 'Team Style', 'Standout Metrics']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 1
    
    # Get standout players
    standouts = identify_standout_players(report_path, team_possessions, league_avg)
    
    print(f"  ✅ Identified {len(standouts)} standout players")
    
    # Group by team style
    high_poss = [s for s in standouts if s['Possession_Style'] == 'High']
    low_poss = [s for s in standouts if s['Possession_Style'] == 'Low']
    avg_poss = [s for s in standouts if s['Possession_Style'] == 'Average']
    
    # Write standouts
    for standout in standouts:
        ws.cell(row=row, column=1, value=standout['Player'])
        ws.cell(row=row, column=2, value=standout['Team'])
        ws.cell(row=row, column=3, value=standout['Position'])
        ws.cell(row=row, column=4, value=standout['Position_Profile'])
        ws.cell(row=row, column=5, value=standout['Score']).number_format = '0.00'
        ws.cell(row=row, column=6, value=standout['Possession_Style'])
        ws.cell(row=row, column=7, value=standout['Standout_Reasons'])
        
        # Highlight based on team style
        style_cell = ws.cell(row=row, column=6)
        if standout['Possession_Style'] == 'High':
            style_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        elif standout['Possession_Style'] == 'Low':
            style_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        
        row += 1
    
    # Add summary by team style
    row += 2
    ws.cell(row=row, column=1, value="SUMMARY BY TEAM STYLE").font = Font(bold=True, size=14)
    row += 2
    
    summary_headers = ['Team Style', 'Count', 'Top Player']
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row += 1
    
    if high_poss:
        top_high = max(high_poss, key=lambda x: x['Standout_Score'])
        ws.cell(row=row, column=1, value="High Possession Teams")
        ws.cell(row=row, column=2, value=len(high_poss))
        ws.cell(row=row, column=3, value=f"{top_high['Player']} ({top_high['Team']}) - {top_high['Standout_Score']} standout metrics")
        row += 1
    
    if low_poss:
        top_low = max(low_poss, key=lambda x: x['Standout_Score'])
        ws.cell(row=row, column=1, value="Low Possession Teams")
        ws.cell(row=row, column=2, value=len(low_poss))
        ws.cell(row=row, column=3, value=f"{top_low['Player']} ({top_low['Team']}) - {top_low['Standout_Score']} standout metrics")
        row += 1
    
    if avg_poss:
        top_avg = max(avg_poss, key=lambda x: x['Standout_Score'])
        ws.cell(row=row, column=1, value="Average Possession Teams")
        ws.cell(row=row, column=2, value=len(avg_poss))
        ws.cell(row=row, column=3, value=f"{top_avg['Player']} ({top_avg['Team']}) - {top_avg['Standout_Score']} standout metrics")
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 80


def main():
    """Main function"""
    base_dir = Path("/Users/daniel/Documents/Smart Sports Lab/Football/Sports Data Campus/Portland Thorns/Data/Advanced Search")
    
    padj_report = base_dir / "Mike_Norris_Scouting_Report_ACC_IntraConference_70_30_PAdj.xlsx"
    
    if not padj_report.exists():
        print(f"❌ PAdj report not found: {padj_report}")
        return
    
    print("="*80)
    print("ADDING STANDOUT PLAYERS ANALYSIS")
    print("="*80)
    
    # Get team possession data
    team_possessions, league_avg = get_team_possession_context()
    
    print(f"  ✅ League average possessions: {league_avg:.2f}")
    print(f"  ✅ Team possessions: {len(team_possessions)} teams")
    
    # Open PAdj report workbook
    wb = load_workbook(padj_report, keep_vba=False)
    
    # Add standout players sheet
    print("\nAnalyzing standout players...")
    add_standout_players_sheet(wb, padj_report, team_possessions, league_avg)
    
    # Save
    wb.save(padj_report)
    print(f"\n✅ Enhanced report saved: {padj_report.name}")
    print("   Added: Standout Performers sheet")
    
    print("\n" + "="*80)
    print("✅ STANDOUT PLAYERS ANALYSIS COMPLETE")
    print("="*80)


if __name__ == "__main__":
    main()

