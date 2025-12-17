"""
Generate Excel template with all metrics from radar_chart_metrics_short.json
This creates a template file that includes all required columns for the Performance Metrics page.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def load_metrics_config():
    """Load metrics from JSON config file."""
    config_path = Path(__file__).parent / "radar_chart_metrics_short.json"
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
        return config
    except Exception as e:
        print(f"Error loading JSON config: {e}")
        return {}

def get_all_metrics(metrics_config):
    """Get all unique metrics across all positions."""
    all_metrics = set()
    for position, metrics in metrics_config.items():
        all_metrics.update(metrics)
    return sorted(list(all_metrics))

def create_template_excel(output_path):
    """Create Excel template with all required columns."""
    
    # Load metrics config
    metrics_config = load_metrics_config()
    if not metrics_config:
        print("❌ Could not load metrics config. Exiting.")
        return
    
    # Get all unique metrics
    all_metrics = get_all_metrics(metrics_config)
    
    print(f"📊 Found {len(all_metrics)} unique metrics across {len(metrics_config)} positions")
    print(f"📋 Positions: {', '.join(metrics_config.keys())}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Player Data"
    
    # Define base columns (required for the app)
    base_columns = [
        'Player',
        'Team',
        'Conference',
        'Position Profile',
        'Grade',
        'Position Rank',
        'Minutes played',
        'Minutes',
    ]
    
    # Combine base columns with all metrics
    all_columns = base_columns + all_metrics
    
    # Write header row
    header_row = 1
    for col_idx, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add instructions row
    instructions_row = 2
    ws.cell(row=instructions_row, column=1, value="INSTRUCTIONS:")
    ws.merge_cells(f'A{instructions_row}:{get_column_letter(len(all_columns))}{instructions_row}')
    instruction_cell = ws.cell(row=instructions_row, column=1)
    instruction_cell.font = Font(bold=True, italic=True)
    instruction_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    instructions_row += 1
    ws.cell(row=instructions_row, column=1, value="1. Fill in player information in the base columns (Player, Team, Conference, Position Profile, etc.)")
    ws.merge_cells(f'A{instructions_row}:{get_column_letter(len(all_columns))}{instructions_row}')
    
    instructions_row += 1
    ws.cell(row=instructions_row, column=1, value="2. Fill in metric values for each player. Metric names must match exactly as shown in the headers.")
    ws.merge_cells(f'A{instructions_row}:{get_column_letter(len(all_columns))}{instructions_row}')
    
    instructions_row += 1
    ws.cell(row=instructions_row, column=1, value="3. Position Profile should match one of: Center Back, Centre Midfielder, Attacking Midfielder, Winger")
    ws.merge_cells(f'A{instructions_row}:{get_column_letter(len(all_columns))}{instructions_row}')
    
    instructions_row += 1
    ws.cell(row=instructions_row, column=1, value="4. Metrics per position (from radar_chart_metrics_short.json):")
    ws.merge_cells(f'A{instructions_row}:{get_column_letter(len(all_columns))}{instructions_row}')
    
    # Add position-specific metric lists
    for position, metrics in metrics_config.items():
        instructions_row += 1
        ws.cell(row=instructions_row, column=1, value=f"   {position}: {', '.join(metrics)}")
        ws.merge_cells(f'A{instructions_row}:{get_column_letter(len(all_columns))}{instructions_row}')
    
    # Set column widths
    for col_idx, col_name in enumerate(all_columns, start=1):
        # Base columns get wider width
        if col_name in base_columns:
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        else:
            # Metric columns get appropriate width
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 2, 15)
    
    # Create a summary sheet with position metrics mapping
    summary_ws = wb.create_sheet("Position Metrics Reference")
    summary_ws.cell(row=1, column=1, value="Position").font = Font(bold=True)
    summary_ws.cell(row=1, column=2, value="Required Metrics").font = Font(bold=True)
    
    row = 2
    for position, metrics in metrics_config.items():
        summary_ws.cell(row=row, column=1, value=position)
        summary_ws.cell(row=row, column=2, value=", ".join(metrics))
        row += 1
    
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 80
    
    # Save file
    wb.save(output_path)
    print(f"✅ Template created: {output_path}")
    print(f"📊 Total columns: {len(all_columns)}")
    print(f"   - Base columns: {len(base_columns)}")
    print(f"   - Metric columns: {len(all_metrics)}")

def main():
    """Main function."""
    # Set output path
    script_dir = Path(__file__).parent
    output_path = script_dir.parent.parent / "Portland Thorns 2025 Long Shortlist Template.xlsx"
    
    print("🚀 Generating Excel template with all metrics...")
    print(f"📁 Output: {output_path}")
    print()
    
    create_template_excel(output_path)
    
    print()
    print("✅ Done! You can now:")
    print("   1. Open the template file")
    print("   2. Fill in your player data")
    print("   3. Upload it to the Performance Metrics page")

if __name__ == "__main__":
    main()

